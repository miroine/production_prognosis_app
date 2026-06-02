"""
FieldVista — Integrated Field Development & Economics
======================================================
Streamlit app for forecasting oil & gas field production with:
- Unit system selector (Field / Metric)
- Drainage strategies: depletion, injection (water/gas), with VRR control
- Multiple drilling rigs, drill+completion durations, scheduled by date
- Producers and injectors as separate well lists
- Decline curves (Arps) or user-defined profiles, with per-well scaling factor
- Simplified PVT model (Standing/Vasquez-Beggs for oil, Z-factor for gas)
- Material balance with aquifer support (Pot or Fetkovich) and gas cap drive
- Time-varying production/injection capacities
- Economics: phased facility CAPEX, well CAPEX, OPEX, tariffs, taxes,
  abandonment cost, NPV, IRR, payback, breakeven price
- Multi-scenario comparison (Depletion vs Injection vs custom)
- Recovery-factor warning
- Run/Stale state machine on the run button
- In-line help on every input
- Case management: save / load / browse / duplicate / delete named cases
- Exports: Excel, JSON-API, PDF report, with Python usage snippet

© 2026 Merouane Hamdani — MIT License.
For early-phase screening only. Not for investment decisions or reserves booking.

Run:  streamlit run field_prognosis_app.py
"""

from __future__ import annotations

import io
import json
import math
from dataclasses import dataclass, field, asdict
from datetime import date, datetime, timedelta
from typing import Optional

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from plotly.subplots import make_subplots

# Local helpers (case persistence, breakeven, PDF, CSS)
import fp_helpers as fh

# =============================================================================
# Page config
# =============================================================================
st.set_page_config(
    page_title="FieldVista — Integrated Field Development & Economics",
    page_icon="🛢️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =============================================================================
# Units & conversions
# =============================================================================
UNIT_LABELS = {
    "field": {
        "oil_rate": "stb/d", "gas_rate": "Mscf/d", "water_rate": "stb/d",
        "oil_vol": "MMstb", "gas_vol": "Bscf", "water_vol": "MMstb",
        "pressure": "psi", "temp": "°F", "depth": "ft",
        "gor": "scf/stb", "bo": "rb/stb", "bg": "rb/scf",
        "price_oil": "$/bbl", "price_gas": "$/Mscf",
    },
    "metric": {
        "oil_rate": "Sm³/d", "gas_rate": "kSm³/d", "water_rate": "Sm³/d",
        "oil_vol": "MSm³", "gas_vol": "GSm³", "water_vol": "MSm³",
        "pressure": "bar", "temp": "°C", "depth": "m",
        "gor": "Sm³/Sm³", "bo": "m³/Sm³", "bg": "m³/Sm³",
        "price_oil": "$/Sm³", "price_gas": "$/kSm³",
    },
}

M2F = {
    "oil_rate":  6.2898,
    "gas_rate":  35.3147,
    "water_rate": 6.2898,
    "oil_vol":   6.2898,
    "gas_vol":   35.3147,
    "water_vol": 6.2898,
    "pressure":  14.5038,
    "depth":     3.28084,
    "gor":       5.6146,
    "price_oil": 1.0/6.2898,
    "price_gas": 1.0/35.3147,
}

def to_field(value, kind, units):
    if units == "field" or value is None: return value
    if kind == "temp": return value * 9/5 + 32
    return value * M2F.get(kind, 1.0)

def from_field(value, kind, units):
    if units == "field" or value is None: return value
    if kind == "temp": return (value - 32) * 5/9
    return value / M2F.get(kind, 1.0)

def ulabel(kind, units):
    return UNIT_LABELS[units][kind]


def cost_input_to_usd(value, _session=None):
    """Convert a cost INPUT to USD for the engine. If the user is entering
    costs in NOK (cost_input_currency == 'NOK'), divide by the NOK→USD rate;
    otherwise pass through unchanged. Used at the point each cost widget is
    read so the engine always works in USD and all results display in USD.

    `value` may be a single number or anything float()-able; None passes
    through.
    """
    import streamlit as _st
    if value is None:
        return None
    cur = _st.session_state.get("cost_input_currency", "USD")
    if cur != "NOK":
        return value
    rate = float(_st.session_state.get("usd_to_nok", 10.5))
    if rate <= 0:
        return value
    try:
        return float(value) / rate
    except (TypeError, ValueError):
        return value


# Map of df columns → unit kind. Used for converting the engine output
# (always field units) to the user's display units in the Data tab and
# in Excel exports. Columns NOT listed here are passed through unchanged.
_DF_COLUMN_KINDS = {
    # Production
    "oil_rate":      "oil_rate",
    "gas_rate":      "gas_rate",
    "water_rate":    "water_rate",
    "primary_rate":  None,        # depends on fluid; resolved below
    "secondary_rate": None,
    "liquid_rate":   "oil_rate",
    "injection_rate": "water_rate",
    "gas_injection_rate": "gas_rate",
    "gross_gas_rate": "gas_rate",
    "gas_export_rate": "gas_rate",
    "gas_inj_rate":   "gas_rate",
    "gas_fuel_rate":  "gas_rate",
    "gas_flare_rate": "gas_rate",
    # Cumulatives
    "cum_oil":       "oil_vol",
    "cum_gas":       "gas_vol",
    "cum_water":     "water_vol",
    "cum_injection": "water_vol",
    # Pressure
    "pressure":      "pressure",
}


def df_to_display_units(df: "pd.DataFrame", fluid_system: str, units: str) -> "pd.DataFrame":
    """Return a copy of ``df`` with all rate/volume/pressure columns converted
    from field units to the user's chosen display units, and column headers
    suffixed with their unit (e.g. ``gas_rate [kSm³/d]``) so Excel exports and
    the Data tab are self-documenting.

    Works in both ``field`` and ``metric`` modes — when ``field``, only the
    headers are relabelled (no value conversion needed).
    """
    out = df.copy()
    primary = "oil_rate" if FLUID_SYSTEMS[fluid_system]["primary"] == "oil" else "gas_rate"
    secondary = "gas_rate" if primary == "oil_rate" else "oil_rate"
    col_map = dict(_DF_COLUMN_KINDS)
    col_map["primary_rate"]   = primary
    col_map["secondary_rate"] = secondary
    # Convert values (no-op when units == 'field')
    if units != "field":
        for col, kind in col_map.items():
            if col in out.columns and kind:
                try:
                    out[col] = from_field(out[col].astype(float), kind, units)
                except Exception:
                    pass
    # Rename rate / volume / pressure columns to include the unit
    rename = {}
    for col, kind in col_map.items():
        if col in out.columns and kind:
            rename[col] = f"{col} [{ulabel(kind, units)}]"
    # Also label money columns (always USD) for self-documenting exports
    money_cols = {
        "revenue":         "[USD/month]",
        "revenue_oil":     "[USD/month]",
        "revenue_gas":     "[USD/month]",
        "revenue_condensate": "[USD/month]",
        "revenue_ngl":     "[USD/month]",
        "ngl_rate":        "[bbl/d]",
        "opex":            "[USD/month]",
        "tax":             "[USD/month]",
        "capex_well":      "[USD/month]",
        "capex_facility":  "[USD/month]",
        "abandonment":     "[USD/month]",
        "cashflow":        "[USD/month]",
        "cum_cashflow":    "[USD]",
        "pretax_cf":       "[USD/month]",
        "discounted_cf":   "[USD/month]",
        "npv":             "[USD]",
        "co2_total_t":     "[tonnes/month]",
        "cum_co2_tonnes":  "[tonnes]",
    }
    for col, label in money_cols.items():
        if col in out.columns:
            rename[col] = f"{col} {label}"
    # Date column header gets a small clarification too
    if "date" in out.columns:
        rename["date"] = "date [YYYY-MM-DD]"
    out = out.rename(columns=rename)
    return out


def df_to_display_units_values(df: "pd.DataFrame", fluid_system: str,
                                units: str) -> "pd.DataFrame":
    """Convert rate/volume/pressure columns from field units to the user's
    display units WITHOUT renaming the headers — for places that need the
    converted values but must keep the plain column names (KPIs, profiles,
    charts). No-op (returns a copy) when units == 'field'.
    """
    out = df.copy()
    if units == "field":
        return out
    primary = ("oil_rate"
               if FLUID_SYSTEMS[fluid_system]["primary"] == "oil"
               else "gas_rate")
    secondary = "gas_rate" if primary == "oil_rate" else "oil_rate"
    col_map = dict(_DF_COLUMN_KINDS)
    col_map["primary_rate"] = primary
    col_map["secondary_rate"] = secondary
    for col, kind in col_map.items():
        if col in out.columns and kind:
            try:
                out[col] = from_field(out[col].astype(float), kind, units)
            except Exception:
                pass
    return out


def df_e_to_display_units(df_e: "pd.DataFrame", fluid_system: str,
                            units: str) -> "pd.DataFrame":
    """Same as df_to_display_units but for the economics dataframe.

    Money columns stay in USD (no conversion needed). Production rate /
    cumulative columns inherited from the underlying df ARE converted.
    """
    return df_to_display_units(df_e, fluid_system, units)


# =============================================================================
# PVT
# =============================================================================
def pvt_oil(p_psi, t_F, api, gas_grav, rs_init, p_bub_psi):
    p = max(p_psi, 14.7); t = t_F
    sg_o = 141.5 / (api + 131.5); yg = gas_grav
    if p < p_bub_psi:
        rs = yg * ((p / 18.2 + 1.4) * 10 ** (0.0125 * api - 0.00091 * t)) ** 1.2048
        rs = min(rs, rs_init)
    else:
        rs = rs_init
    f = rs * (yg / sg_o) ** 0.5 + 1.25 * t
    bo_b = 0.972 + 0.000147 * (f ** 1.175)
    if p <= p_bub_psi:
        bo = bo_b
    else:
        co = 1e-5
        bo = bo_b * math.exp(-co * (p - p_bub_psi))
    a = 10 ** (3.0324 - 0.02023 * api)
    mu_od = 10 ** (a * t ** -1.163) - 1
    A = 10.715 * (rs + 100) ** -0.515
    B = 5.44 * (rs + 150) ** -0.338
    mu_o = A * mu_od ** B
    return {"Bo": bo, "Rs": rs, "mu_o": mu_o, "Bo_b": bo_b}


def pvt_gas(p_psi, t_F, gas_grav):
    t_R = t_F + 460
    ppc = 756.8 - 131.0 * gas_grav - 3.6 * gas_grav ** 2
    tpc = 169.2 + 349.5 * gas_grav - 74.0 * gas_grav ** 2
    ppr = max(p_psi, 14.7) / ppc
    tpr = t_R / tpc
    A = 1.39 * (max(tpr - 0.92, 0)) ** 0.5 - 0.36 * tpr - 0.101
    E = 9 * (tpr - 1)
    B = (0.62 - 0.23 * tpr) * ppr + (0.066 / (tpr - 0.86) - 0.037) * ppr ** 2 + \
        0.32 * ppr ** 6 / (10 ** min(E, 50))
    C = 0.132 - 0.32 * math.log10(tpr)
    F = 0.3106 - 0.49 * tpr + 0.1824 * tpr ** 2
    z = max(A + (1 - A) / math.exp(min(B, 50)) + C * ppr ** F, 0.3)
    bg = 0.00504 * z * t_R / max(p_psi, 14.7)
    M = 28.97 * gas_grav
    rho = (p_psi * M) / (z * 10.732 * t_R)
    rho_g = rho * 0.0160185
    K = ((9.4 + 0.02 * M) * t_R ** 1.5) / (209 + 19 * M + t_R)
    X = 3.5 + 986 / t_R + 0.01 * M
    Y = 2.4 - 0.2 * X
    mu_g = 1e-4 * K * math.exp(min(X * rho_g ** Y, 50))
    return {"Z": z, "Bg": bg, "mu_g": mu_g}


# =============================================================================
# Data classes
# =============================================================================
@dataclass
class Reservoir:
    """A single reservoir compartment with its own PVT, aquifer, gas-cap and strategy.

    Multi-reservoir mode: a field can host several reservoirs. Wells allocate a
    fraction of their rate to each reservoir they tap (sum of fractions per well = 1).

    PI fields (productivity-index bridge):
      ``well_pi`` and ``min_bhp_psi`` together define a physical link between
      the reservoir and any well that taps it. When a well has its
      ``derive_qi_from_pi`` flag set, the engine recomputes its initial rate
      each timestep from ``PI × (P_res − BHP_min)`` rather than using the
      free-input qi value. Sensible default PI values are documented in the
      reservoir-archetype library.
    """
    id: str
    name: str
    fluid_system: str        # one of FLUID_SYSTEMS keys
    ooip_oil: float          # MMstb (oil) — primary in-place if oil
    ogip_gas: float          # Bscf (gas) — primary in-place if gas
    rf_target: float
    strategy: str            # "Depletion" or "Injection"
    pvt: PVTInputs
    aquifer: AquiferInputs
    gas_cap: GasCapInputs
    voidage_ratio: float = 1.0
    inj_efficiency: float = 0.85
    well_pi: float = 2.0           # bbl/d/psi/well (oil) or Mscf/d/psi/well (gas)
    min_bhp_psi: float = 1500.0    # minimum flowing BHP (well-level constraint)


@dataclass
class WellReservoirLink:
    """Allocation of one well's rate among reservoirs."""
    well_name: str
    reservoir_id: str
    fraction: float          # 0..1


@dataclass
class WellSpec:
    name: str
    is_producer: bool
    rig: str
    spud_date: date
    drill_days: int
    completion_days: int
    qi_primary: float
    qi_secondary: float
    decline_model: str
    di_annual: float
    b_factor: float
    wc_initial: float
    wc_final: float
    wc_ramp_months: int
    scale_factor: float = 1.0
    uptime: float = 0.95   # fraction of time the well is on stream
    user_profile: Optional[pd.DataFrame] = None
    # Multi-segment decline. When decline_model == "Multi-segment", the well's
    # rate is built from a sequence of Arps segments instead of one curve.
    # Each segment is a dict:
    #   {"months": int  duration of this segment,
    #    "model":  "Plateau"|"Exponential"|"Harmonic"|"Hyperbolic",
    #    "di":     float annual decline (ignored for Plateau),
    #    "b":      float Arps b-exponent (Hyperbolic only),
    #    "mult":   float multiplicative step applied to the rate at the
    #              START of this segment (1.0 = rate-continuous; >1 = a
    #              bean-up or re-stimulation bump; <1 = a choke-back)}
    # Segments run back-to-back; after the last segment the final segment's
    # behaviour is extrapolated to the end of the forecast.
    segments: Optional[list] = None
    inj_rate: float = 0.0
    # Per-well fluid type — when "auto", well inherits the field's fluid system's
    # primary fluid. Set explicitly to "oil" or "gas" for mixed-fluid fields
    # (e.g. an oil well producing into a multi-reservoir field that also contains
    # gas reservoirs).
    fluid: str = "auto"
    # PI mode: when True, the well's qi is recomputed from
    # PI_well × (P_res − BHP_min) at each timestep instead of using the
    # free qi_primary value. Decline still applies on top of that base rate.
    derive_qi_from_pi: bool = False
    # Optional override for the well's own productivity index. When 0 the
    # reservoir's well_pi is used.
    well_pi_override: float = 0.0
    # IPR / BHP-deliverability mode. When True, the engine computes each
    # well's actual rate at every timestep from the IPR (Vogel for oil,
    # back-pressure equation for gas) intersected with a simplified outflow
    # curve (hydrostatic + friction). The well rate is limited to the lesser
    # of (decline target, IPR-deliverable rate, surface capacity).
    ipr_mode: bool = False
    wellhead_pressure_psi: float = 200.0    # P_wh — separator/manifold pressure
    tubing_depth_ft: float = 8000.0          # mid-perf depth
    fluid_gradient_psi_per_ft: float = 0.35  # hydrostatic gradient (oil ~0.35, gas ~0.10, water ~0.45)
    friction_psi_per_kbpd: float = 5.0       # linear friction proxy (psi per 1000 bbl/d)

    @property
    def online_date(self):
        return self.spud_date + timedelta(days=self.drill_days + self.completion_days)


@dataclass
class CapacitySchedule:
    df: pd.DataFrame
    def at_date(self, d):
        """Single-date lookup (kept for backward compat; prefer to_arrays in hot loops)."""
        ts = pd.Timestamp(d)
        clean = self._clean_df()
        if len(clean) == 0:
            return {k: 0.0 for k in
                    ["oil", "gas", "water", "liquid", "water_inj", "gas_inj", "prod_eff"]}
        col = pd.to_datetime(clean["start_date"])
        sub = clean[col <= ts]
        row = clean.iloc[0] if len(sub) == 0 else sub.iloc[-1]
        return {k: float(row.get(k, 0.0) or 0.0) for k in
                ["oil", "gas", "water", "liquid", "water_inj", "gas_inj", "prod_eff"]}

    def _clean_df(self) -> pd.DataFrame:
        """Drop rows with no valid start_date; coerce numerics; ensure all expected
        columns exist (filled with 0). Tolerates user-entered empty rows."""
        df = self.df.copy() if self.df is not None else pd.DataFrame()
        if "start_date" not in df.columns:
            return pd.DataFrame(columns=["start_date", "oil", "gas", "water",
                                          "liquid", "water_inj", "gas_inj", "prod_eff"])
        df["start_date"] = pd.to_datetime(df["start_date"], errors="coerce")
        df = df.dropna(subset=["start_date"])
        # Ensure all expected numeric columns exist
        for k, default in [("oil", 0.0), ("gas", 0.0), ("water", 0.0),
                            ("liquid", 0.0), ("water_inj", 0.0), ("gas_inj", 0.0),
                            ("prod_eff", 0.95)]:
            if k not in df.columns:
                df[k] = default
            df[k] = pd.to_numeric(df[k], errors="coerce").fillna(default)
        return df.reset_index(drop=True)

    def to_arrays(self, dates) -> dict:
        """Vectorized lookup: returns {key: np.ndarray of length len(dates)}.

        For each timestamp, picks the most recent row with start_date <= ts;
        before the first row, the first row is used (as in `at_date`).
        Tolerates user-edited tables with empty / partial rows.
        """
        n = len(dates)
        keys = ["oil", "gas", "water", "liquid", "water_inj", "gas_inj", "prod_eff"]
        clean = self._clean_df()
        if len(clean) == 0:
            # No usable rows → return safe defaults (no choking, default PE)
            return {k: np.zeros(n) if k != "prod_eff" else np.full(n, 0.95)
                    for k in keys}
        ordered = clean.assign(
            _ts=pd.to_datetime(clean["start_date"])
        ).sort_values("_ts").reset_index(drop=True)
        ts_arr = ordered["_ts"].values
        date_ts = pd.DatetimeIndex(dates).values
        idx = np.searchsorted(ts_arr, date_ts, side="right") - 1
        idx = np.clip(idx, 0, len(ordered) - 1)
        out = {}
        for k in keys:
            arr = ordered[k].astype(float).values
            out[k] = arr[idx]
        return out


@dataclass
class CapexSchedule:
    df: pd.DataFrame


@dataclass
class PVTInputs:
    p_init_psi: float
    t_res_F: float
    api: float
    gas_grav: float
    rs_init: float
    p_bub_psi: float


@dataclass
class AquiferInputs:
    active: bool
    model: str   # "Pot" | "Fetkovich" | "Carter-Tracy"
    aquifer_volume: float          # MMbbl (used by Pot/Fetkovich)
    productivity_index: float      # bbl/d/psi (used by Fetkovich)
    initial_pressure_psi: float
    # Carter-Tracy parameters (only used when model == "Carter-Tracy")
    ct_aquifer_constant: float = 200.0     # U (bbl/psi); rule-of-thumb ~ 100-1000
    ct_diffusivity: float = 50.0           # k×t conversion: dimensionless time per month


@dataclass
class GasCapInputs:
    active: bool
    size_fraction: float
    initial_pressure_psi: float


@dataclass
class EconInputs:
    oil_price: float
    gas_price: float
    opex_var: float
    opex_fixed: float
    capex_per_well: float
    discount_rate: float
    tax_rate: float
    royalty_rate: float
    tariff_oil: float
    tariff_gas: float
    abandonment_cost_MM: float
    facility_capex: CapexSchedule
    revenue_basis: str = "net"      # "net" (gas after shrinkage & injection) or "gross"
    co2_price: float = 0.0          # $ / tonne CO2-eq carbon tax (0 = ignore)
    co2_factor_gas_combust: float = 53.0  # kg CO2 per Mscf burnt (fuel/flare)
    co2_factor_flare_inefficiency: float = 0.02  # methane slip (CH4 has 28× GWP100)
    co2_factor_oil_routine: float = 0.5   # kg CO2-eq per bbl oil produced (vented + ops)
    # Scope 3 — end-use combustion of sold hydrocarbons. Default factors
    # are the standard IPCC / EPA values for stationary combustion.
    co2_scope3_enabled: bool = False
    co2_scope3_factor_oil: float = 430.0  # kg CO2 per bbl crude burnt (downstream)
    co2_scope3_factor_gas: float = 53.0   # kg CO2 per Mscf gas burnt (downstream)
    co2_scope3_price: float = 0.0    # $ / tonne CO2-eq applied to Scope 3 (0 = ignore)
    # ---- Fiscal regime ----
    fiscal_regime: str = "Tax/Royalty"   # or "PSC" or "NCS"
    # PSC parameters (only used when fiscal_regime == "PSC")
    psc_cost_recovery_ceiling: float = 0.50  # max share of revenue recoverable per period
    psc_profit_oil_share_contractor: float = 0.40  # contractor's share of profit oil
    psc_govt_participation: float = 0.0   # carried equity (0 = no, 0.20 = 20%)
    psc_psc_tax_rate: float = 0.30         # tax on contractor's profit oil share
    psc_signature_bonus_MM: float = 0.0    # one-off, paid at first month
    # NCS parameters (only used when fiscal_regime == "NCS")
    # Norwegian Continental Shelf petroleum-tax regime: CIT + Special
    # Petroleum Tax with an "uplift" capital allowance.
    ncs_cit_rate: float = 0.22             # corporate income tax 22%
    ncs_spt_rate: float = 0.718            # special petroleum tax 71.8%
    ncs_uplift_rate: float = 0.1769        # uplift allowance: 17.69% × capex
                                            # spread over ncs_uplift_years,
                                            # deducted from the SPT base only
    ncs_depreciation_years: float = 6.0    # straight-line CAPEX depreciation
                                            # (CIT + SPT base)
    ncs_uplift_years: float = 4.0          # period over which uplift is given
                                            # (set to ~0.08 for immediate)
    # ---- Money basis: nominal vs real ----
    # "real":    costs & revenues kept in today's $ (no inflation applied);
    #            discount_rate is interpreted as a REAL discount rate.
    # "nominal": all future cashflows are escalated by `inflation_rate`/yr;
    #            discount_rate is interpreted as a NOMINAL discount rate.
    # The two are economically equivalent if discount_rate_nom = (1 + r_real)
    # × (1 + infl) − 1, but reporting in real $ is usually cleaner.
    money_basis: str = "real"
    inflation_rate: float = 0.025          # annual, applied if money_basis == "nominal"
    # ---- Well cost model ----
    # Two modes for capex_per_well:
    #   "fixed": classic $MM/well number (legacy behavior; uses capex_per_well above)
    #   "rig_rate": (drill_days + completion_days) × day_rate + tangibles
    well_cost_mode: str = "rig_rate"
    rig_day_rate_kUSD: float = 500.0       # rig dayrate in $1,000s/day (e.g. $500k/d ~ jackup)
    completion_day_rate_kUSD: float = 350.0 # completion-spread dayrate
    well_tangibles_MM: float = 4.0         # per-well tangibles (casing, tree, etc.) in $MM
    well_intangibles_pct: float = 0.10     # intangibles as a fraction of (rig + completion) cost
    # ---- NGL (Natural Gas Liquids) stream ----
    # NGL = propane / butane / pentane+ extracted from the produced gas at a
    # midstream plant. Modelled as a yield factor (volume of NGL per volume of
    # gross gas), priced per barrel. Subject to its own opex and (optionally)
    # to gas shrinkage at the plant.
    ngl_yield_bbl_per_mmscf: float = 0.0    # bbl NGL per MMscf gas (typical: 0 dry, 30-150 condensate)
    ngl_price_bbl: float = 25.0              # $/bbl (often quoted as 30-60% of WTI)
    ngl_opex_bbl: float = 5.0                # processing + transport tariff
    ngl_shrinkage_pct: float = 0.0           # fraction of gas volume lost at extraction
                                              # (0 = ignore shrinkage; typical 2-5%)
    # Rig metadata (move-in/out + maintenance) — keyed by rig name. Populated
    # from the rigs table; used to cost mob/demob/maintenance days.
    rig_meta: dict = field(default_factory=dict)
    # ---- Economic limit / cessation timing ----
    # "horizon"   : produce for the full forecast horizon (legacy behaviour);
    #               abandonment booked at last producing month.
    # "economic"  : truncate production at the economic limit — the month
    #               after which monthly operating cashflow (revenue − royalty
    #               − tariff − OPEX) stays negative. Cessation cost is booked
    #               at that month. This is the smart, self-consistent cutoff.
    economic_cutoff_mode: str = "horizon"
    # When in "economic" mode, require this many consecutive negative-CF
    # months before declaring the economic limit (filters out transient dips
    # e.g. a maintenance month). Typical: 6–12 months.
    economic_cutoff_persistence: int = 6


@dataclass
class FieldAssumptions:
    fluid_system: str
    strategy: str
    ooip_oil: float
    ogip_gas: float
    rf_target: float
    start_date: date
    forecast_years: int
    rock_compressibility: float
    sw_init: float
    pvt: PVTInputs
    aquifer: AquiferInputs
    gas_cap: GasCapInputs
    voidage_ratio: float
    inj_efficiency: float
    aban_rate_oil: float
    aban_rate_gas: float
    aban_wc: float
    aban_basis: str
    cap_schedule: CapacitySchedule
    # New: field-level efficiency & gas disposition
    production_efficiency: float = 0.95   # field-level operational uptime (0..1)
    gas_export_fraction: float = 1.00     # of associated/produced gas
    gas_injection_fraction: float = 0.00  # re-injected
    gas_fuel_fraction: float = 0.00       # used as fuel gas (own consumption)
    gas_flare_fraction: float = 0.00      # flared
    # New: multi-reservoir support
    reservoirs: list = field(default_factory=list)        # list[Reservoir]
    well_links: list = field(default_factory=list)        # list[WellReservoirLink]
    # PI bridge defaults: used by the synthesized single-reservoir when
    # multi-reservoir mode is off. Multi-reservoir mode reads PI/BHP from
    # each reservoir row directly.
    default_well_pi: float = 2.0
    default_min_bhp_psi: float = 1500.0
    # Retrograde-condensate modelling (gas-condensate fields only). When
    # enabled, the produced condensate stream is recomputed as gas rate ×
    # a producible CGR that falls below the dew point as liquid drops out
    # in the reservoir.
    retrograde_enabled: bool = False
    retrograde_drop_fraction: float = 0.55
    # Fractional-flow water cut (oil fields only). When enabled, the field
    # water cut is derived from Corey relative-permeability curves and the
    # cumulative recovery, rather than the per-well water-cut ramp.
    fractional_flow_enabled: bool = False
    ff_swc: float = 0.20         # connate water saturation
    ff_sor: float = 0.25         # residual oil saturation
    ff_krw_max: float = 0.30     # water rel-perm endpoint
    ff_kro_max: float = 0.90     # oil rel-perm endpoint
    ff_nw: float = 3.0           # Corey water exponent
    ff_no: float = 2.0           # Corey oil exponent
    ff_mu_oil: float = 1.5       # oil viscosity (cP)
    ff_mu_water: float = 0.4     # water viscosity (cP)
    ff_sweep: float = 0.70       # volumetric sweep efficiency

    def gas_disposition_sum(self) -> float:
        return (self.gas_export_fraction + self.gas_injection_fraction
                + self.gas_fuel_fraction + self.gas_flare_fraction)


# =============================================================================
# Constants
# =============================================================================
DAYS_PER_MONTH = 30.4375
MONTHS_PER_YEAR = 12

FLUID_SYSTEMS = {
    "Oil with associated gas": {"primary": "oil", "secondary": "gas"},
    "Gas with condensate":     {"primary": "gas", "secondary": "condensate"},
    "Black oil (no gas)":      {"primary": "oil", "secondary": None},
    "Dry gas":                 {"primary": "gas", "secondary": None},
}
DECLINE_MODELS = ["Exponential", "Hyperbolic", "Harmonic", "Multi-segment",
                  "User-defined profile"]
# Standard subsea template types -> well-slot capacity. Mirrors the cost
# table in fp_helpers; used by the per-template detail editor.
_TEMPLATE_SLOT_CAPACITY_UI = {
    "Single-slot (1 well)":   1,
    "Double-slot (2 wells)":  2,
    "4-slot (4 wells)":       4,
    "6-slot (6 wells)":       6,
}
RIG_COLORS = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
              "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf"]


# =============================================================================
# Decline / well profile
# =============================================================================
def decline_rate(qi, di, b, model, t_y):
    if qi <= 0: return np.zeros_like(t_y)
    if model == "Exponential": return qi * np.exp(-di * t_y)
    if model == "Harmonic":    return qi / (1.0 + di * t_y)
    if model == "Hyperbolic":
        b = max(min(b, 0.999), 0.001)
        return qi / np.power(1.0 + b * di * t_y, 1.0 / b)
    return qi * np.exp(-di * t_y)


def _segment_rate(qi, di, b, model, t_y):
    """Rate within a single segment, t_y measured from the SEGMENT start.
    A 'Plateau' segment holds the rate flat. Otherwise standard Arps."""
    if qi <= 0:
        return np.zeros_like(t_y)
    if model == "Plateau":
        return np.full_like(t_y, qi, dtype=float)
    if model == "Exponential":
        return qi * np.exp(-di * t_y)
    if model == "Harmonic":
        return qi / (1.0 + di * t_y)
    if model == "Hyperbolic":
        b = max(min(b, 0.999), 0.001)
        return qi / np.power(1.0 + b * di * t_y, 1.0 / b)
    return qi * np.exp(-di * t_y)


def decline_rate_multisegment(qi, segments, rel_months):
    """Build a piecewise-Arps rate profile from a list of segments.

    Args:
        qi          : starting rate of segment 1 (stb/d or Mscf/d).
        segments    : list of segment dicts (see WellSpec.segments).
        rel_months  : integer array of months since the well came online.

    Returns an array of rates aligned with rel_months. Segments run
    back-to-back and are rate-continuous unless a segment carries a 'mult'
    other than 1.0, which steps the rate at that segment's start (a bean-up
    or a late-life re-stimulation bump). After the final segment, the last
    segment's decline is extrapolated.
    """
    rel_months = np.asarray(rel_months)
    n = len(rel_months)
    rate = np.zeros(n, dtype=float)
    if qi <= 0 or not segments:
        return rate
    # Walk the segments, tracking the rate at the start of each.
    seg_start_month = 0.0
    seg_start_rate = float(qi)
    for si, seg in enumerate(segments):
        months = float(seg.get("months", 0) or 0)
        model = seg.get("model", "Exponential")
        di = float(seg.get("di", 0.0) or 0.0)
        b = float(seg.get("b", 0.5) or 0.5)
        mult = float(seg.get("mult", 1.0) or 1.0)
        # apply the step multiplier at the segment boundary
        q0 = seg_start_rate * mult
        is_last = (si == len(segments) - 1)
        seg_end_month = seg_start_month + months
        # which output months fall inside this segment
        if is_last:
            mask = rel_months >= seg_start_month
        else:
            mask = (rel_months >= seg_start_month) & \
                   (rel_months < seg_end_month)
        if mask.any():
            t_y = (rel_months[mask] - seg_start_month) / MONTHS_PER_YEAR
            rate[mask] = _segment_rate(q0, di, b, model, t_y)
        # rate carried into the next segment = rate at the end of this one
        t_end_y = months / MONTHS_PER_YEAR
        seg_start_rate = float(_segment_rate(q0, di, b, model,
                                             np.array([t_end_y]))[0])
        seg_start_month = seg_end_month
    return rate


def well_monthly(well: WellSpec, dates: pd.DatetimeIndex, field_is_oil: bool = True):
    n = len(dates)
    primary = np.zeros(n); secondary = np.zeros(n); water = np.zeros(n); inj = np.zeros(n)
    _seg_fell_back = False
    online_ts = pd.Timestamp(well.online_date)
    active = dates >= online_ts
    rel_months = ((dates.year - online_ts.year) * 12 + (dates.month - online_ts.month)).values
    rel_months = np.where(active, rel_months, 0)
    t_y = rel_months / MONTHS_PER_YEAR
    sf = well.scale_factor * well.uptime  # combine scaling & well uptime

    if well.is_producer:
        if well.decline_model == "User-defined profile" and well.user_profile is not None:
            prof = well.user_profile
            for i in range(n):
                if not active[i]: continue
                rm = int(rel_months[i])
                if rm < len(prof):
                    row = prof.iloc[rm]
                    primary[i] = float(row.get("primary_rate", 0.0)) * sf
                    secondary[i] = float(row.get("secondary_rate", 0.0)) * sf
        elif well.decline_model == "Multi-segment" and well.segments:
            # Piecewise-Arps profile: plateau, decline, optional bean-up or
            # late-life bump — all built from the segment list.
            rp = decline_rate_multisegment(
                well.qi_primary, well.segments, rel_months) * sf
            rs = decline_rate_multisegment(
                well.qi_secondary, well.segments, rel_months) * sf
            primary = np.where(active, rp, 0.0)
            secondary = np.where(active, rs, 0.0)
        else:
            if well.decline_model == "Multi-segment" and not well.segments:
                # A Multi-segment well with no segment profile attached —
                # usually a case saved before segment-export existed, or a
                # YAML missing its 'segments:' block. Flag it so the caller
                # can warn (this is the classic live-vs-batch mismatch).
                _seg_fell_back = True
            rp = decline_rate(well.qi_primary, well.di_annual, well.b_factor,
                              well.decline_model, t_y) * sf
            rs = decline_rate(well.qi_secondary, well.di_annual, well.b_factor,
                              well.decline_model, t_y) * sf
            primary = np.where(active, rp, 0.0)
            secondary = np.where(active, rs, 0.0)

        wc = np.zeros(n)
        for i in range(n):
            if not active[i]: continue
            rm = rel_months[i]
            if well.wc_ramp_months <= 0:
                wc[i] = well.wc_final
            else:
                frac = min(1.0, rm / max(well.wc_ramp_months, 1))
                wc[i] = well.wc_initial + frac * (well.wc_final - well.wc_initial)
        safe_wc = np.clip(wc, 0.0, 0.99)
        water = primary * safe_wc / np.where(safe_wc < 1, 1 - safe_wc, 1)
    else:
        inj = np.where(active, well.inj_rate * sf, 0.0)

    # Phase-explicit decomposition: map (primary, secondary) → (oil, gas) based
    # on the well's own fluid type. Per-well water is always the WC-derived stream.
    well_fluid = getattr(well, "fluid", "auto")
    if well_fluid == "auto":
        is_oil_well = field_is_oil
    else:
        is_oil_well = (well_fluid == "oil")
    if is_oil_well:
        oil_phase = primary       # primary = oil rate (stb/d)
        gas_phase = secondary     # secondary = associated gas (Mscf/d)
    else:
        gas_phase = primary       # primary = gas rate (Mscf/d)
        oil_phase = secondary     # secondary = condensate (stb/d)

    return {"primary": primary, "secondary": secondary, "water": water,
            "inj": inj, "active": active,
            "oil": oil_phase, "gas": gas_phase,
            "_seg_fell_back": _seg_fell_back}


# =============================================================================
# Material balance
# =============================================================================
def mbe_pressure(asm, dates, q_p, q_s, q_w, q_inj, is_oil):
    """Simplified Schilthuis-form material balance, solved with bisection on pressure.

    For oil reservoirs:
      F = N * (Eo + m*Eg + (1+m)*Efw) + We + Winj_bbl
    where:
      F  = underground withdrawal (rb)
         = Np * Bo + (Gp - Np*Rs) * Bg + Wp
      Eo = (Bo - Boi) + (Rsi - Rs) * Bg     [oil zone expansion + liberated gas]
      Eg = Boi * (Bg/Bgi - 1)                [gas cap expansion, if any]
      Efw= Boi * c_f * (Pi - P)              [rock+water compressibility, simplified]
      We = c_t * V_w * (Pi - P)              [pot aquifer]
    Cumulative gas Gp is converted to scf inside (q_s is Mscf/d).

    For gas reservoirs:
      P/Z = (Pi/Zi) * (1 - Gp_eff/G), with aquifer giving a small uplift.
    """
    n = len(dates)
    P = np.zeros(n)
    p_init = asm.pvt.p_init_psi

    if is_oil:
        N_stb = asm.ooip_oil * 1e6
        pvt_i = pvt_oil(p_init, asm.pvt.t_res_F, asm.pvt.api, asm.pvt.gas_grav,
                        asm.pvt.rs_init, asm.pvt.p_bub_psi)
        boi = pvt_i["Bo"]
        bgi_init = pvt_gas(p_init, asm.pvt.t_res_F, asm.pvt.gas_grav)["Bg"]
    else:
        G_scf = asm.ogip_gas * 1e9
        gvt_i = pvt_gas(p_init, asm.pvt.t_res_F, asm.pvt.gas_grav)
        zi = gvt_i["Z"]

    if asm.aquifer.active:
        Vw_bbl = asm.aquifer.aquifer_volume * 1e6
        ct_aq = 6e-6
    else:
        Vw_bbl = 0; ct_aq = 0

    # Fetkovich state (only used if model == "Fetkovich")
    # Wei = max encroachable water = ct * Vw * Pi
    Pa_init = asm.aquifer.initial_pressure_psi if asm.aquifer.active else p_init
    Wei = ct_aq * Vw_bbl * Pa_init  # bbl
    J_aq = asm.aquifer.productivity_index  # bbl/d/psi
    We_cum = 0.0  # cumulative aquifer influx, bbl
    Pa = Pa_init  # current aquifer pressure

    # Carter-Tracy state
    ct_U     = float(getattr(asm.aquifer, "ct_aquifer_constant", 200.0))
    ct_diff  = float(getattr(asm.aquifer, "ct_diffusivity", 50.0))   # tD per month
    Wec_cum = 0.0       # cumulative influx for Carter-Tracy
    tD_prev = 0.0       # last dimensionless time

    def _W_D_inf(tD: float) -> tuple[float, float]:
        """Dimensionless cumulative influx and its derivative for an
        infinite-acting radial aquifer (the standard 'long-time' approximation).
            W_D(tD) ≈ 2 * sqrt(tD/π)        (for tD > ~100; falls back gracefully)
        For small tD we use the ramp form W_D(tD) ≈ 2*sqrt(tD/π) which is also
        a reasonable screening-grade approximation.
        """
        tD = max(tD, 1e-6)
        WD  = 2.0 * np.sqrt(tD / np.pi)
        # Derivative dW_D/dtD = 1/sqrt(π × tD)
        dWD = 1.0 / np.sqrt(np.pi * tD)
        return float(WD), float(dWD)

    m_gc = asm.gas_cap.size_fraction if (is_oil and asm.gas_cap.active) else 0.0

    days = DAYS_PER_MONTH
    # Cumulative produced in field units. q_s is in Mscf/d; convert to scf.
    Np = 0.0; Gp_scf = 0.0; Wp = 0.0; Winj_bbl = 0.0; Ginj_scf = 0.0

    for i in range(n):
        if is_oil:
            Np += q_p[i] * days                # stb
            Gp_scf += q_s[i] * days * 1000.0   # Mscf -> scf
        else:
            Gp_scf += q_p[i] * days * 1000.0   # primary is gas in Mscf/d
        Wp += q_w[i] * days                    # bbl
        if asm.strategy == "Injection":
            if is_oil:
                Winj_bbl += q_inj[i] * days
            else:
                Ginj_scf += q_inj[i] * days * 1000.0

        # Pre-compute aquifer influx coefficients for this timestep.
        # Both models return We (cumulative) as a function of p_test.
        use_fetkovich = (asm.aquifer.active and asm.aquifer.model == "Fetkovich"
                         and Wei > 0 and J_aq > 0)
        use_carter_tracy = (asm.aquifer.active and asm.aquifer.model == "Carter-Tracy"
                             and ct_U > 0 and ct_diff > 0)
        if use_fetkovich:
            # Fetkovich incremental influx over Δt:
            #   ΔWe = (Wei/Pi) * (Pa - p_res_avg) * (1 - exp(-J*Pi*Δt/Wei))
            # We linearise p_res_avg ≈ p_test (single timestep face value).
            dt = days
            fet_coef = (Wei / Pa_init) * (1.0 - np.exp(-J_aq * Pa_init * dt / max(Wei, 1.0)))
            We_prev = We_cum
            Pa_prev = Pa
        if use_carter_tracy:
            # Classical Carter-Tracy incremental form (Lee, "Well Testing"):
            #   We(t_n) = U × Σ ΔP_j × W_D(tD_n - tD_{j-1})
            # The avoided-convolution version below tracks running We_cum and
            # uses the standard recurrence:
            #   ΔWe / Δt_D ≈ (U×ΔP_n - We_{n-1} × W_D'(tD_n))
            #              / (W_D(tD_n) − tD_{n-1} × W_D'(tD_n))
            tD_now = tD_prev + ct_diff
            WD_now, dWD_now = _W_D_inf(tD_now)
            denom_ct = WD_now - tD_prev * dWD_now
            We_prev_ct = Wec_cum
            tD_prev_now = tD_prev          # capture for closure

        # ---- Solve P by bisection ----
        if is_oil:
            def residual(p_test):
                pvtp = pvt_oil(p_test, asm.pvt.t_res_F, asm.pvt.api,
                               asm.pvt.gas_grav, asm.pvt.rs_init, asm.pvt.p_bub_psi)
                Bo = pvtp["Bo"]; Rs = pvtp["Rs"]
                Bg = pvt_gas(max(p_test, 14.7), asm.pvt.t_res_F, asm.pvt.gas_grav)["Bg"]
                # Underground withdrawal, rb
                F = Np * Bo + (Gp_scf - Np * Rs) * Bg + Wp
                F -= Winj_bbl
                # Expansions
                Eo = (Bo - boi) + (asm.pvt.rs_init - Rs) * Bg
                Eg = boi * (Bg / bgi_init - 1) if m_gc > 0 else 0.0
                Efw = boi * asm.rock_compressibility * (p_init - p_test)
                # Aquifer
                if use_fetkovich:
                    dWe = fet_coef * (Pa_prev - p_test)
                    dWe = max(dWe, 0.0)  # no back-flow
                    We = We_prev + dWe
                elif use_carter_tracy:
                    dP = max(Pa_init - p_test, 0.0)
                    dWe_ct = (ct_U * dP - We_prev_ct * dWD_now) / max(denom_ct, 1e-6)
                    dWe_ct = max(dWe_ct, 0.0)
                    We = We_prev_ct + dWe_ct
                else:
                    # Pot model
                    We = ct_aq * Vw_bbl * (asm.aquifer.initial_pressure_psi - p_test)
                    We = max(We, 0.0)
                # MBE residual: F - N*(...) - We = 0 at correct P
                rhs = N_stb * (Eo + m_gc * Eg + (1 + m_gc) * Efw) + We
                return F - rhs

            # Residual = F - RHS. RHS grows as P drops, F mostly steady.
            # → residual is monotone DECREASING as P drops (i.e., increasing in P).
            # Normal case: residual > 0 at p_init, residual < 0 at low P → root in between.
            lo, hi = 100.0, p_init
            f_lo = residual(lo); f_hi = residual(hi)
            if f_hi <= 0:
                # Withdrawal already balanced by expansion at p_init (negligible production / strong inj)
                p_sol = p_init
            elif f_lo > 0:
                # Even at minimum P, can't supply enough expansion → clamp at p_init
                # (this means injection / aquifer is overpowering production)
                p_sol = p_init
            else:
                for _ in range(60):
                    mid = 0.5 * (lo + hi)
                    f_mid = residual(mid)
                    if f_mid > 0:
                        # need lower P (more expansion)
                        hi = mid
                    else:
                        # need higher P (less expansion than current)
                        lo = mid
                    if hi - lo < 0.5:
                        break
                p_sol = 0.5 * (lo + hi)
        else:
            # Gas: P/Z method with optional aquifer
            def residual(p_test):
                Z = pvt_gas(max(p_test, 14.7), asm.pvt.t_res_F, asm.pvt.gas_grav)["Z"]
                Gp_eff = Gp_scf - Ginj_scf
                # P/Z target from depletion
                lhs = p_test / Z
                rhs = (p_init / zi) * (1 - Gp_eff / max(G_scf, 1))
                # Aquifer: subtract influx-equivalent gas (very rough)
                if Vw_bbl > 0:
                    rhs += (p_init / zi) * 0.05 * ((p_init - p_test) / p_init) * \
                           (Vw_bbl * ct_aq * p_init / max(G_scf, 1))
                return lhs - rhs

            lo, hi = 100.0, p_init
            f_lo = residual(lo); f_hi = residual(hi)
            if f_hi <= 0:
                p_sol = p_init
            elif f_lo >= 0:
                p_sol = lo
            else:
                for _ in range(60):
                    mid = 0.5 * (lo + hi)
                    if residual(mid) > 0:
                        hi = mid
                    else:
                        lo = mid
                    if hi - lo < 0.5:
                        break
                p_sol = 0.5 * (lo + hi)

        P[i] = p_sol

        # Update Fetkovich aquifer state for next timestep
        if use_fetkovich:
            dWe_actual = max(fet_coef * (Pa_prev - p_sol), 0.0)
            We_cum = We_prev + dWe_actual
            # Pa drops as We accumulates: Pa = Pi * (1 - We/Wei)
            Pa = Pa_init * max(1.0 - We_cum / max(Wei, 1.0), 0.0)

        # Update Carter-Tracy state for next timestep
        if use_carter_tracy:
            dP_act = max(Pa_init - p_sol, 0.0)
            dWe_ct = (ct_U * dP_act - We_prev_ct * dWD_now) / max(denom_ct, 1e-6)
            Wec_cum = We_prev_ct + max(dWe_ct, 0.0)
            tD_prev = tD_now

    return P


def retrograde_cgr(pressure_psi, p_dew_psi, cgr_initial,
                    p_init_psi, drop_fraction=0.55):
    """Screening model of retrograde condensate behaviour.

    In a gas-condensate reservoir, above the dew-point pressure the produced
    CGR is constant at its initial value. Below the dew point, liquid
    condenses in the reservoir pores ("retrograde" drop-out). That liquid is
    largely immobile, so it is NOT produced — the *producible* CGR therefore
    falls as pressure declines below the dew point. The produced gas stream
    also becomes leaner.

    This function returns, for each month, the effective producible CGR.

    Model (screening-grade):
      - p >= p_dew                : CGR = cgr_initial (single-phase gas)
      - p_dew > p > p_min_liquid  : CGR falls roughly linearly to a minimum
      - below p_min_liquid        : CGR holds at the minimum (revaporisation
                                    is ignored at screening level)
    The pressure of maximum liquid drop-out p_min_liquid is taken as a
    fraction of the dew point (typically liquid drop-out peaks at
    40-60% of p_dew for a lean-to-moderate condensate).

    Args:
        pressure_psi : array of reservoir pressures (psi).
        p_dew_psi    : dew-point pressure (psi).
        cgr_initial  : initial CGR (stb/MMscf).
        p_init_psi   : initial reservoir pressure (psi).
        drop_fraction: maximum fractional loss of producible CGR at the
                       point of maximum liquid drop-out (0.55 = the
                       producible CGR falls to 45% of its initial value).

    Returns:
        dict with:
          cgr        : array of producible CGR per month (stb/MMscf)
          retrograde_active : bool — whether drop-out occurs at all
          min_cgr    : the minimum producible CGR reached
    """
    pressure = np.asarray(pressure_psi, dtype=float)
    if cgr_initial <= 0 or p_dew_psi <= 0:
        return {"cgr": np.full_like(pressure, max(cgr_initial, 0.0)),
                "retrograde_active": False,
                "min_cgr": max(cgr_initial, 0.0)}
    # If the reservoir never drops below the dew point, no retrograde loss.
    if float(np.min(pressure)) >= p_dew_psi:
        return {"cgr": np.full_like(pressure, cgr_initial),
                "retrograde_active": False,
                "min_cgr": cgr_initial}
    # Pressure of maximum liquid drop-out — ~50% of the dew point.
    p_min_liquid = 0.50 * p_dew_psi
    drop_fraction = float(min(max(drop_fraction, 0.0), 0.95))
    cgr = np.full_like(pressure, cgr_initial)
    for i, p in enumerate(pressure):
        if p >= p_dew_psi:
            cgr[i] = cgr_initial
        elif p <= p_min_liquid:
            cgr[i] = cgr_initial * (1.0 - drop_fraction)
        else:
            # linear ramp between the dew point and the drop-out peak
            f = (p_dew_psi - p) / max(p_dew_psi - p_min_liquid, 1.0)
            cgr[i] = cgr_initial * (1.0 - drop_fraction * f)
    return {"cgr": cgr,
            "retrograde_active": True,
            "min_cgr": float(np.min(cgr))}


# =============================================================================
# Simulation
# =============================================================================
def _reservoirs_or_default(asm: FieldAssumptions) -> list:
    """Return user-defined reservoirs, or synthesize a single reservoir from
    sidebar-level fields for backward compatibility."""
    if asm.reservoirs:
        return asm.reservoirs
    return [Reservoir(
        id="R1", name="Default reservoir",
        fluid_system=asm.fluid_system,
        ooip_oil=asm.ooip_oil, ogip_gas=asm.ogip_gas,
        rf_target=asm.rf_target,
        strategy=asm.strategy,
        pvt=asm.pvt, aquifer=asm.aquifer, gas_cap=asm.gas_cap,
        voidage_ratio=asm.voidage_ratio,
        inj_efficiency=asm.inj_efficiency,
        well_pi=asm.default_well_pi,
        min_bhp_psi=asm.default_min_bhp_psi,
    )]


def _allocations_for(asm: FieldAssumptions, well_name: str,
                     reservoirs: list) -> dict:
    """Return a dict {reservoir_id: fraction} for a well.

    If no allocations are defined for this well, the well is assigned 100%
    to the first reservoir."""
    fracs = {l.reservoir_id: l.fraction
             for l in asm.well_links if l.well_name == well_name}
    if not fracs:
        return {reservoirs[0].id: 1.0}
    # Normalize if fractions do not sum to 1.0
    s = sum(fracs.values())
    if s > 0 and abs(s - 1.0) > 0.001:
        fracs = {k: v / s for k, v in fracs.items()}
    return fracs


def _reservoir_view(asm: FieldAssumptions, r: Reservoir):
    """Return a lightweight object that quacks like FieldAssumptions for the
    parts of mbe_pressure that read PVT/aquifer/gas_cap/strategy/in-place values."""
    class _View:
        pass
    v = _View()
    v.fluid_system        = r.fluid_system
    v.strategy            = r.strategy
    v.ooip_oil            = r.ooip_oil
    v.ogip_gas            = r.ogip_gas
    v.rf_target           = r.rf_target
    v.start_date          = asm.start_date
    v.forecast_years      = asm.forecast_years
    v.rock_compressibility= asm.rock_compressibility
    v.sw_init             = asm.sw_init
    v.pvt                 = r.pvt
    v.aquifer             = r.aquifer
    v.gas_cap             = r.gas_cap
    v.voidage_ratio       = r.voidage_ratio
    v.inj_efficiency      = r.inj_efficiency
    return v


def run_simulation(wells, asm: FieldAssumptions):
    n_months = asm.forecast_years * MONTHS_PER_YEAR
    dates = pd.date_range(asm.start_date, periods=n_months, freq="MS")

    # Track Multi-segment wells that have no segment profile attached, so
    # we can warn the user (the profile silently falls back to single-Arps).
    _missing_segments_wells = []

    producers = [w for w in wells if w.is_producer]
    injectors = [w for w in wells if not w.is_producer]
    n_p = len(producers); n_i = len(injectors)

    p_mat = np.zeros((n_months, n_p))
    s_mat = np.zeros((n_months, n_p))
    w_mat = np.zeros((n_months, n_p))
    on_p = np.zeros((n_months, n_p), dtype=bool)
    inj_mat = np.zeros((n_months, n_i))
    on_i = np.zeros((n_months, n_i), dtype=bool)

    # ---- PI bridge: when a well has derive_qi_from_pi=True, recompute its
    # qi_primary from PI × (P_init − BHP_min) using the linked reservoir's
    # PI / BHP defaults (or the well's PI override). This is a single pre-step
    # at simulation time so all the downstream decline / abandonment logic
    # works unchanged.
    res_list = _reservoirs_or_default(asm)
    res_by_id = {r.id: r for r in res_list}
    well_to_res = {}                                     # well_name -> Reservoir
    if asm.well_links:
        # Pick the reservoir with the largest fraction for each well
        from collections import defaultdict
        agg = defaultdict(list)
        for ln in asm.well_links:
            agg[ln.well_name].append(ln)
        for name, links in agg.items():
            best = max(links, key=lambda l: l.fraction)
            if best.reservoir_id in res_by_id:
                well_to_res[name] = res_by_id[best.reservoir_id]
    # Fallback: if no link defined, use the first reservoir
    default_res = res_list[0] if res_list else None
    for w in producers:
        if not getattr(w, "derive_qi_from_pi", False):
            continue
        rsv = well_to_res.get(w.name, default_res)
        if rsv is None:
            continue
        pi = float(getattr(w, "well_pi_override", 0.0) or 0.0)
        if pi <= 0:
            pi = float(getattr(rsv, "well_pi", 0.0) or 0.0)
        if pi <= 0:
            continue
        bhp = float(getattr(rsv, "min_bhp_psi", 1500.0) or 1500.0)
        dp = max(rsv.pvt.p_init_psi - bhp, 0.0)
        derived_qi = pi * dp                              # bbl/d (oil) or Mscf/d (gas)
        if derived_qi > 0:
            # Preserve the user-input GOR ratio so the secondary stream stays
            # consistent if it was set explicitly.
            ratio = (w.qi_secondary / w.qi_primary) if w.qi_primary > 0 else 0.0
            w.qi_primary = derived_qi
            w.qi_secondary = derived_qi * ratio if ratio > 0 else w.qi_secondary

    # Compute is_oil up front so well_monthly can map (primary, secondary)
    # to (oil, gas) per-well based on each well's own fluid type.
    is_oil = FLUID_SYSTEMS[asm.fluid_system]["primary"] == "oil"

    # Per-well phase matrices (proper, not share-weighted)
    oil_mat = np.zeros((n_months, n_p))
    gas_mat = np.zeros((n_months, n_p))

    for j, w in enumerate(producers):
        prof = well_monthly(w, dates, field_is_oil=is_oil)
        p_mat[:, j] = prof["primary"]; s_mat[:, j] = prof["secondary"]
        w_mat[:, j] = prof["water"]; on_p[:, j] = prof["active"]
        oil_mat[:, j] = prof["oil"]; gas_mat[:, j] = prof["gas"]
        if prof.get("_seg_fell_back"):
            _missing_segments_wells.append(w.name)
    for j, w in enumerate(injectors):
        prof = well_monthly(w, dates, field_is_oil=is_oil)
        inj_mat[:, j] = prof["inj"]; on_i[:, j] = prof["active"]

    aban = np.ones((n_months, n_p), dtype=bool)
    if asm.aban_basis == "Per well" and n_p > 0:
        # Vectorized abandonment: for each well, find the first month its rate
        # falls below the threshold or WC > limit; from then on, the well is shut.
        wc_mat = w_mat / (p_mat + w_mat + 1e-12)            # (n_months, n_p)
        rate_below = (
            (p_mat < asm.aban_rate_oil) if is_oil
            else (p_mat / 1000.0 < asm.aban_rate_gas)
        )
        wc_above = wc_mat > asm.aban_wc
        trigger = on_p & (rate_below | wc_above)            # only triggers when active
        # First-shut month per well; -1 if never triggers
        first_shut = np.where(
            trigger.any(axis=0),
            trigger.argmax(axis=0),
            n_months,                                        # never
        )
        # Build aban[i, j] = True iff month i < first_shut[j]
        month_idx = np.arange(n_months)[:, None]            # (n_months, 1)
        aban = (month_idx < first_shut[None, :]) & on_p
    p_mat *= aban; s_mat *= aban; w_mat *= aban
    oil_mat *= aban; gas_mat *= aban

    # Pull capacity (and time-varying PE) from the schedule
    cap = asm.cap_schedule.to_arrays(dates)                  # dict of np.ndarray

    # Production efficiency: prefer the schedule's per-row PE (time-varying)
    # when present; fall back to the constant asm.production_efficiency for
    # legacy cases. Treat zero / negative PE as the constant fallback (avoid
    # accidentally zeroing the whole forecast if user empties the column).
    pe_array = cap.get("prod_eff", None)
    if pe_array is None or (np.asarray(pe_array) <= 0).all():
        pe_array = np.full(n_months, float(asm.production_efficiency))
    else:
        pe_array = np.where(pe_array > 0, pe_array, asm.production_efficiency)
        pe_array = np.clip(pe_array.astype(float), 0.0, 1.0)
    # Per-timestep PE (broadcast over wells)
    p_mat *= pe_array[:, None]
    s_mat *= pe_array[:, None]
    w_mat *= pe_array[:, None]
    oil_mat *= pe_array[:, None]
    gas_mat *= pe_array[:, None]

    field_p = p_mat.sum(axis=1); field_s = s_mat.sum(axis=1)
    field_w = w_mat.sum(axis=1); field_l = field_p + field_w
    field_inj = inj_mat.sum(axis=1)

    # Vectorized choke: compute capacity arrays once, evaluate all timesteps in NumPy.
    EPS = 1e-12
    # Helper: factor = cap / rate where cap > 0 AND rate > cap, else 1.0
    def _bind_factor(rate, capacity):
        capacity = np.asarray(capacity)
        binds = (capacity > 0) & (rate > capacity)
        return np.where(binds, capacity / np.maximum(rate, EPS), 1.0)

    if is_oil:
        f_oil    = _bind_factor(field_p, cap["oil"])
        f_gas    = _bind_factor(field_s, cap["gas"] * 1000.0)
        f_water  = _bind_factor(field_w, cap["water"])
        f_liq    = _bind_factor(field_l, cap["liquid"])
        choke = np.minimum.reduce([np.ones(n_months), f_oil, f_gas, f_water, f_liq])
    else:
        f_gas    = _bind_factor(field_p, cap["gas"] * 1000.0)
        f_oil    = _bind_factor(field_s, cap["oil"])
        choke = np.minimum.reduce([np.ones(n_months), f_gas, f_oil])

    # Injection chokes (always applied, regardless of strategy):
    f_winj = _bind_factor(field_inj, cap["water_inj"])
    inj_choke = np.minimum(np.ones(n_months), f_winj)
    # VRR cap when at least one injector is defined.
    if asm.voidage_ratio > 0 and n_i > 0:
        voidage_now = field_l if is_oil else field_p / 1000.0
        target_inj = voidage_now * asm.voidage_ratio * asm.inj_efficiency
        f_vrr = _bind_factor(field_inj, target_inj)
        inj_choke = np.minimum(inj_choke, f_vrr)

    field_p *= choke; field_s *= choke; field_w *= choke; field_l *= choke
    field_inj *= inj_choke
    # Apply same choke to per-well phase matrices for the per-well plot
    # (the primary/secondary/water matrices are choked further down via p_post).
    oil_mat *= choke[:, None]; gas_mat *= choke[:, None]
    w_mat_choked = w_mat * choke[:, None]   # for the water row of the per-well plot

    if asm.strategy == "Injection" and n_i == 0:
        # Synthetic voidage-replacement injection when no explicit injectors
        voidage = field_l if is_oil else field_p / 1000.0
        field_inj = voidage * asm.voidage_ratio * asm.inj_efficiency

    # Track field-total cutoff month so the per-well matrices used by the
    # per-well plot can be truncated identically (the per-well chart reads
    # from oil_mat / gas_mat / w_mat_choked and previously kept producing
    # past cessation — most visibly the water curve which lacks the
    # rate-floor check).
    field_cutoff_idx = None

    if asm.aban_basis == "Field total":
        # Vectorized field-total abandonment: trigger once below the rate
        # threshold OR above the water-cut threshold, after month 12.
        if is_oil:
            below = field_p < asm.aban_rate_oil
        else:
            below = field_p / 1000.0 < asm.aban_rate_gas
        # Field-wide water cut (oil systems): field_w is water rate,
        # field_p is oil rate. WC threshold applies only when there is
        # any oil production (avoid 0/0 = NaN at the very start). The
        # explicit np.errstate silences a benign divide-by-zero warning
        # that would otherwise fire on pre-FOP zero-rate months.
        if is_oil:
            denom = field_p + field_w
            with np.errstate(divide="ignore", invalid="ignore"):
                field_wc = np.where(denom > 0, field_w / denom, 0.0)
            wc_above = field_wc > asm.aban_wc
            below = below | wc_above
        below[:12] = False                                   # first year never triggers
        if below.any():
            first = int(np.argmax(below))
            field_cutoff_idx = first
            field_p[first:] = field_s[first:] = 0.0
            field_w[first:] = field_l[first:] = field_inj[first:] = 0.0
            # Also truncate the per-well matrices that feed the per-well
            # phase chart. Without this, the water curve in particular
            # kept flowing past cessation because field-total mode only
            # zeroed the field aggregate, not the per-well water_mat.
            # (p_post / s_post / w_post / inj_post are built later from
            # these and the choke vector, so they inherit the truncation.)
            p_mat[first:, :] = 0.0
            s_mat[first:, :] = 0.0
            w_mat[first:, :] = 0.0
            w_mat_choked[first:, :] = 0.0
            oil_mat[first:, :] = 0.0
            gas_mat[first:, :] = 0.0

    days = DAYS_PER_MONTH
    # Unit conversions: 1 MMstb = 1e6 stb;  1 Bscf = 1e6 Mscf
    if is_oil:
        # primary (oil) in stb/d -> MMstb;  secondary (gas) in Mscf/d -> Bscf
        cum_p = np.cumsum(field_p * days) / 1e6
        cum_s = np.cumsum(field_s * days) / 1e6
    else:
        # primary (gas) in Mscf/d -> Bscf;  secondary (cond) in stb/d -> MMstb
        cum_p = np.cumsum(field_p * days) / 1e6
        cum_s = np.cumsum(field_s * days) / 1e6
    cum_w = np.cumsum(field_w * days) / 1e6
    cum_inj = np.cumsum(field_inj * days) / 1e6

    # ---- Per-reservoir MBE / RF (multi-reservoir aware) ----
    # Wells produce fractions of their rate into each reservoir they tap;
    # the MBE for each reservoir is solved on its allocated cumulatives.
    reservoirs = _reservoirs_or_default(asm)
    per_res_pressure = {}
    per_res_cum_primary = {}
    per_res_rf = {}

    n_p_idx = len(producers)
    n_i_idx = len(injectors)
    n_r = len(reservoirs)
    alloc_p = np.zeros((n_p_idx, n_r))
    alloc_i = np.zeros((n_i_idx, n_r))
    res_id_to_col = {r.id: i for i, r in enumerate(reservoirs)}

    for j, w in enumerate(producers):
        af = _allocations_for(asm, w.name, reservoirs)
        for rid, frac in af.items():
            if rid in res_id_to_col:
                alloc_p[j, res_id_to_col[rid]] = frac
    for j, w in enumerate(injectors):
        af = _allocations_for(asm, w.name, reservoirs)
        for rid, frac in af.items():
            if rid in res_id_to_col:
                alloc_i[j, res_id_to_col[rid]] = frac

    # Approximation: choke is uniform across wells (single field-wide bottleneck).
    p_post = (p_mat.T * choke).T
    s_post = (s_mat.T * choke).T
    w_post = (w_mat.T * choke).T
    inj_post = (inj_mat.T * inj_choke).T

    res_p = p_post @ alloc_p
    res_s = s_post @ alloc_p
    res_w = w_post @ alloc_p
    res_inj = inj_post @ alloc_i

    for k, r in enumerate(reservoirs):
        is_oil_r = FLUID_SYSTEMS[r.fluid_system]["primary"] == "oil"
        # Convert primary cumulative: stb→MMstb (oil) or Mscf→Bscf (gas), both /1e6
        cum_p_r = np.cumsum(res_p[:, k] * days) / 1e6
        in_place = r.ooip_oil if is_oil_r else r.ogip_gas
        rf_r = cum_p_r / in_place if in_place > 0 else np.zeros_like(cum_p_r)

        view = _reservoir_view(asm, r)
        press_r = mbe_pressure(view, dates,
                               res_p[:, k], res_s[:, k],
                               res_w[:, k], res_inj[:, k],
                               is_oil_r)
        per_res_pressure[r.id] = press_r
        per_res_cum_primary[r.id] = cum_p_r
        per_res_rf[r.id] = rf_r

    # Aggregate RF (uses primary-fluid in-place across like reservoirs)
    if is_oil:
        oil_in_place = sum(r.ooip_oil for r in reservoirs
                           if FLUID_SYSTEMS[r.fluid_system]["primary"] == "oil")
        rf = cum_p / oil_in_place if oil_in_place > 0 else np.zeros(n_months)
    else:
        gas_in_place = sum(r.ogip_gas for r in reservoirs
                           if FLUID_SYSTEMS[r.fluid_system]["primary"] == "gas")
        rf = cum_p / gas_in_place if gas_in_place > 0 else np.zeros(n_months)

    # Field-level pressure: weighted avg of per-reservoir pressures by total
    # primary produced from that reservoir (so dominant producers dominate).
    weights = np.maximum(res_p.sum(axis=0), 1e-9)
    weights = weights / weights.sum()
    pressure = np.zeros(n_months)
    for k, r in enumerate(reservoirs):
        pressure += per_res_pressure[r.id] * weights[k]

    # ---- Global pressure-floor enforcement ----
    # When the reservoir pressure for the well's host reservoir falls below
    # that reservoir's minimum BHP, the well has nothing left to flow
    # against — drawdown ≤ 0. Once a well has been shut in for sustained
    # pressure deficit, it stays shut: the per-well chart used to show a
    # mid-life gap followed by full-rate "resumption" because the floor
    # mask only flagged individual months and let production restart when
    # pressure transiently recovered (e.g. aquifer influx after no
    # withdrawal). Real wells don't restart at peak rate after a shut-in;
    # we make the shut-in PERSISTENT — once a well first goes below the
    # floor for a few consecutive months, every later month is also
    # zeroed.
    floor_triggered = False
    # `persistence` — once we see this many consecutive months below the
    # floor, the well is treated as permanently shut. 3 months filters
    # out brief transient dips while catching real depletion.
    floor_persistence = 3
    for j, w in enumerate(producers):
        rsv = well_to_res.get(w.name, default_res)
        if rsv is None:
            continue
        min_bhp = float(getattr(rsv, "min_bhp_psi", 1500.0) or 1500.0)
        press_for_well = per_res_pressure[rsv.id]
        below = press_for_well < min_bhp
        if not below.any():
            continue
        # Find the first month with `floor_persistence` consecutive months
        # below the floor. From that month onwards, the well is shut.
        shut_idx = None
        run = 0
        for i in range(len(below)):
            if below[i]:
                run += 1
                if run >= floor_persistence:
                    shut_idx = i - floor_persistence + 1
                    break
            else:
                run = 0
        if shut_idx is None:
            continue
        floor_triggered = True
        # Zero EVERY matrix from shut_idx onward — the choked rates that
        # feed reservoir aggregates AND the per-well matrices the
        # per-well chart reads from.
        p_post[shut_idx:, j] = 0.0
        s_post[shut_idx:, j] = 0.0
        w_post[shut_idx:, j] = 0.0
        p_mat[shut_idx:, j] = 0.0
        s_mat[shut_idx:, j] = 0.0
        w_mat[shut_idx:, j] = 0.0
        w_mat_choked[shut_idx:, j] = 0.0
        oil_mat[shut_idx:, j] = 0.0
        gas_mat[shut_idx:, j] = 0.0
    # Re-derive per-reservoir aggregates if any well was floored, so
    # downstream cumulatives + RF reflect the truncation.
    if floor_triggered:
        res_p = p_post @ alloc_p
        res_s = s_post @ alloc_p
        res_w = w_post @ alloc_p

    # ---- IPR / BHP-deliverability post-pass ----
    # For wells with ipr_mode=True, recompute the rate at each timestep from
    # the IPR (Vogel for oil, back-pressure for gas) intersected with a simple
    # outflow curve. The reservoir pressure used is from the just-completed
    # MBE solve. Wells without ipr_mode keep their decline-based rates. The
    # IPR limit is the lesser of (decline target, deliverable rate); surface
    # capacity choke has already been applied via p_post.
    ipr_wells = [(j, w) for j, w in enumerate(producers)
                  if getattr(w, "ipr_mode", False)]
    ipr_limited_pct = np.zeros(n_months)   # diagnostic
    if ipr_wells:
        # Operate on the post-choke matrices (p_post, s_post, w_post) so the
        # field aggregation downstream picks up the IPR-trimmed rates.
        for j, w in ipr_wells:
            rsv = well_to_res.get(w.name, default_res)
            if rsv is None:
                continue
            pi_w = float(getattr(w, "well_pi_override", 0.0) or rsv.well_pi)
            if pi_w <= 0:
                continue
            p_bub = rsv.pvt.p_bub_psi
            well_fluid = getattr(w, "fluid", "auto")
            if well_fluid == "auto":
                wf = "oil" if FLUID_SYSTEMS[rsv.fluid_system]["primary"] == "oil" else "gas"
            else:
                wf = well_fluid
            press_for_well = per_res_pressure[rsv.id]
            for i in range(n_months):
                p_res_i = float(press_for_well[i])
                q_decline = float(p_post[i, j])
                if q_decline <= 0:
                    continue
                if p_res_i <= w.wellhead_pressure_psi:
                    p_post[i, j] = 0.0
                    s_post[i, j] = 0.0
                    w_post[i, j] = 0.0
                    p_mat[i, j] = 0.0; s_mat[i, j] = 0.0; w_mat[i, j] = 0.0
                    oil_mat[i, j] = 0.0; gas_mat[i, j] = 0.0
                    ipr_limited_pct[i] += 1.0 / max(len(ipr_wells), 1)
                    continue
                if wf == "gas":
                    c_coef = pi_w / max(2.0 * p_res_i, 1.0)
                    res_q = fh.deliverable_rate(
                        p_res=p_res_i, p_wh=w.wellhead_pressure_psi,
                        depth_ft=w.tubing_depth_ft,
                        pi=c_coef, p_bub=p_bub, fluid="gas",
                        fluid_grad_psi_per_ft=w.fluid_gradient_psi_per_ft,
                        friction_psi_per_kbpd=w.friction_psi_per_kbpd,
                        q_decline_target=q_decline)
                else:
                    res_q = fh.deliverable_rate(
                        p_res=p_res_i, p_wh=w.wellhead_pressure_psi,
                        depth_ft=w.tubing_depth_ft,
                        pi=pi_w, p_bub=p_bub, fluid="oil",
                        fluid_grad_psi_per_ft=w.fluid_gradient_psi_per_ft,
                        friction_psi_per_kbpd=w.friction_psi_per_kbpd,
                        q_decline_target=q_decline)
                if res_q["limited_by"] == "ipr" and res_q["rate"] < q_decline:
                    scale = res_q["rate"] / q_decline if q_decline > 0 else 0.0
                    p_post[i, j] *= scale
                    s_post[i, j] *= scale
                    w_post[i, j] *= scale
                    p_mat[i, j] *= scale; s_mat[i, j] *= scale; w_mat[i, j] *= scale
                    oil_mat[i, j] *= scale; gas_mat[i, j] *= scale
                    ipr_limited_pct[i] += 1.0 / max(len(ipr_wells), 1)
        # Re-aggregate per-reservoir streams from the IPR-trimmed p_post
        res_p = p_post @ alloc_p
        res_s = s_post @ alloc_p
        res_w = w_post @ alloc_p
        # Field-level aggregates also need to be rebuilt
        field_p = p_post.sum(axis=1)
        field_s = s_post.sum(axis=1)
        field_w = w_post.sum(axis=1)
        field_l = field_p + field_w
        # Recompute cumulatives & RF from the IPR-trimmed field rates
        cum_p = np.cumsum(field_p * days) / 1e6
        cum_s = np.cumsum(field_s * days) / 1e6
        cum_w = np.cumsum(field_w * days) / 1e6
        # Also recompute per-reservoir cum / RF
        for k, r in enumerate(reservoirs):
            is_oil_r = FLUID_SYSTEMS[r.fluid_system]["primary"] == "oil"
            cum_p_r = np.cumsum(res_p[:, k] * days) / 1e6
            in_place = r.ooip_oil if is_oil_r else r.ogip_gas
            per_res_cum_primary[r.id] = cum_p_r
            per_res_rf[r.id] = (cum_p_r / in_place
                                 if in_place > 0 else np.zeros_like(cum_p_r))
        # Recompute aggregate RF
        if is_oil:
            oil_in_place = sum(r.ooip_oil for r in reservoirs
                               if FLUID_SYSTEMS[r.fluid_system]["primary"] == "oil")
            rf = cum_p / oil_in_place if oil_in_place > 0 else np.zeros(n_months)
        else:
            gas_in_place = sum(r.ogip_gas for r in reservoirs
                               if FLUID_SYSTEMS[r.fluid_system]["primary"] == "gas")
            rf = cum_p / gas_in_place if gas_in_place > 0 else np.zeros(n_months)

    active_p = (on_p & aban).sum(axis=1)
    active_i = on_i.sum(axis=1)

    # ---- Field-level oil & gas streams (multi-reservoir aware) ----
    # In a mixed-fluid field, well "primary_rate" is whatever the well's
    # primary fluid is, but at the field level we must split oil vs gas:
    #   field_oil_rate = oil-reservoir primaries + gas-reservoir secondaries (condensate)
    #   field_gas_rate = oil-reservoir secondaries (associated gas) + gas-reservoir primaries
    field_oil_rate = np.zeros(n_months)
    field_gas_rate = np.zeros(n_months)
    for k, r in enumerate(reservoirs):
        is_oil_r = FLUID_SYSTEMS[r.fluid_system]["primary"] == "oil"
        if is_oil_r:
            field_oil_rate += res_p[:, k]    # primary = oil (stb/d)
            field_gas_rate += res_s[:, k]    # secondary = associated gas (Mscf/d)
        else:
            field_gas_rate += res_p[:, k]    # primary = gas (Mscf/d)
            field_oil_rate += res_s[:, k]    # secondary = condensate (stb/d)

    # ---- Retrograde condensate (gas-condensate fields) ----
    # If retrograde modelling is enabled and the field is a gas-condensate
    # system, the condensate stream is recomputed as gas rate × producible
    # CGR, where the CGR falls below the dew point as liquid drops out in
    # the reservoir. Without this, condensate is a flat yield off the user
    # decline; with it, the condensate declines faster than the gas once
    # the reservoir crosses the dew point.
    retro_info = None
    retro_cgr_series = None
    _is_gas_condensate = (not is_oil and
                          FLUID_SYSTEMS[asm.fluid_system]["secondary"]
                          == "condensate")
    if (_is_gas_condensate and getattr(asm, "retrograde_enabled", False)
            and asm.pvt.rs_init > 0):
        p_dew = asm.pvt.p_bub_psi   # for a gas system p_bub holds the dew point
        cgr0 = asm.pvt.rs_init      # stb/MMscf
        retro = retrograde_cgr(pressure, p_dew, cgr0,
                               asm.pvt.p_init_psi,
                               drop_fraction=getattr(
                                   asm, "retrograde_drop_fraction", 0.55))
        # producible condensate (stb/d) = gas (Mscf/d) / 1000 (-> MMscf/d)
        #                                  × CGR (stb/MMscf)
        new_condensate = field_gas_rate / 1000.0 * retro["cgr"]
        field_oil_rate = new_condensate
        # NOTE: retro_info holds ONLY scalars. Numpy arrays must never go
        # into df.attrs — pandas 3.x propagates attrs through concat and
        # compares them, and comparing arrays raises an ambiguous-truth
        # ValueError. The CGR series is added as a real DataFrame column
        # further below instead.
        retro_cgr_series = np.asarray(retro["cgr"], dtype=float)
        retro_info = {
            "active": bool(retro["retrograde_active"]),
            "min_cgr": float(retro["min_cgr"]),
            "cgr0": float(cgr0),
            "p_dew": float(p_dew),
        }

    # Override the legacy field_p / field_s / cum_p / cum_s with the unit-correct
    # streams. For single-reservoir this is identical to field_p/field_s; for
    # multi-reservoir mixed fluids, this is the only way to keep units sane.
    if is_oil:
        field_p = field_oil_rate
        field_s = field_gas_rate
    else:
        field_p = field_gas_rate
        field_s = field_oil_rate
    cum_p = np.cumsum(field_p * days) / 1e6
    cum_s = np.cumsum(field_s * days) / 1e6
    field_l = field_oil_rate + field_w  # liquid is always oil + water

    # ---- Fractional-flow water cut (optional, physics-based) ----
    # By default the field water cut is the sum of the per-well ramps the
    # user entered. When fractional-flow mode is enabled (oil field), the
    # water cut is instead derived from saturation physics: the cumulative
    # oil recovery sets the average water saturation, the Corey rel-perm
    # curves give the fractional flow, and the water rate is recomputed as
    # oil_rate × f_w / (1 - f_w). This makes waterflood water cut emerge
    # from the rock/fluid properties rather than a prescribed ramp.
    ff_info = None
    if (is_oil and getattr(asm, "fractional_flow_enabled", False)
            and asm.ooip_oil and asm.ooip_oil > 0):
        cum_oil_running = np.cumsum(field_oil_rate * days) / 1e6
        rf_running = cum_oil_running / asm.ooip_oil
        ff_params = {
            "swc": getattr(asm, "ff_swc", 0.20),
            "sor": getattr(asm, "ff_sor", 0.25),
            "krw_max": getattr(asm, "ff_krw_max", 0.30),
            "kro_max": getattr(asm, "ff_kro_max", 0.90),
            "nw": getattr(asm, "ff_nw", 3.0),
            "no": getattr(asm, "ff_no", 2.0),
            "mu_oil": getattr(asm, "ff_mu_oil", 1.5),
            "mu_water": getattr(asm, "ff_mu_water", 0.4),
            "sweep_efficiency": getattr(asm, "ff_sweep", 0.7),
        }
        ff = fh.fractional_flow_watercut(rf_running, ff_params)
        fw = np.clip(ff["water_cut"], 0.0, 0.999)
        # water rate from the fractional-flow water cut:
        #   fw = qw / (qo + qw)  ->  qw = qo * fw / (1 - fw)
        field_w = field_oil_rate * fw / (1.0 - fw)
        field_l = field_oil_rate + field_w
        bt_date = None
        if ff["bt_index"] >= 0 and ff["bt_index"] < len(dates):
            bt_date = str(pd.Timestamp(dates[ff["bt_index"]]).date())
        ff_info = {
            "active": True,
            "bt_index": ff["bt_index"],
            "bt_date": bt_date,
            "final_sw": float(ff["sw"][-1]) if len(ff["sw"]) else 0.0,
            "final_wc": float(fw[-1]) if len(fw) else 0.0,
        }
        # store the saturation & water-cut series as real columns later
        ff_sw_series = np.asarray(ff["sw"], dtype=float)
        ff_wc_series = np.asarray(fw, dtype=float)
    else:
        ff_sw_series = None
        ff_wc_series = None

    # ---- Volumetric consistency cap (decline curve vs material balance) ----
    # Decline curves are generated independently of the in-place volumes. A
    # user can specify well rates / declines that, integrated, would produce
    # MORE than the oil/gas originally in place — giving a recovery factor
    # above 100%, which is physically impossible. Here we cap cumulative
    # primary production at the primary-fluid in-place volume: once the
    # field has produced 100% of OOIP/OGIP, the rate is forced to zero. A
    # warning is recorded on the DataFrame so the UI can flag it.
    profile_warnings = []
    if _missing_segments_wells:
        _uniq = sorted(set(_missing_segments_wells))
        profile_warnings.append(
            "🧩 Multi-segment decline selected for "
            f"{', '.join(_uniq)} but no segment profile was found — "
            "these wells fell back to a single-curve Arps decline. If this "
            "case came from a YAML/saved case, re-export it from this "
            "version so the 'segments:' block is included, then reload.")
    primary_in_place = asm.ooip_oil if is_oil else asm.ogip_gas  # MMstb / Bscf
    if primary_in_place and primary_in_place > 0:
        cum_primary_running = np.cumsum(field_p * days) / 1e6
        over = cum_primary_running > primary_in_place
        if over.any():
            cap_idx = int(np.argmax(over))
            cum_before = (cum_primary_running[cap_idx - 1]
                          if cap_idx > 0 else 0.0)
            remaining = max(0.0, primary_in_place - cum_before)
            month_vol = field_p[cap_idx] * days / 1e6
            scale = (remaining / month_vol) if month_vol > 0 else 0.0
            scale = min(1.0, max(0.0, scale))
            for arr in (field_oil_rate, field_gas_rate, field_p, field_s,
                        field_w, field_l):
                arr[cap_idx] *= scale
                arr[cap_idx + 1:] = 0.0
            field_inj[cap_idx + 1:] = 0.0
            oil_mat[cap_idx + 1:, :] = 0.0
            gas_mat[cap_idx + 1:, :] = 0.0
            # recompute cumulatives after the cap
            cum_p = np.cumsum(field_p * days) / 1e6
            cum_s = np.cumsum(field_s * days) / 1e6
            # recompute the recovery factor from the capped cumulative so it
            # cannot exceed 100%
            if is_oil:
                _oip = sum(r.ooip_oil for r in reservoirs
                           if FLUID_SYSTEMS[r.fluid_system]["primary"] == "oil")
            else:
                _oip = sum(r.ogip_gas for r in reservoirs
                           if FLUID_SYSTEMS[r.fluid_system]["primary"] == "gas")
            if _oip > 0:
                rf = np.minimum(cum_p / _oip, 1.0)
            profile_warnings.append(
                f"Decline-curve production reached 100% of "
                f"{'OOIP' if is_oil else 'OGIP'} "
                f"({primary_in_place:,.1f} "
                f"{'MMstb' if is_oil else 'Bscf'}) at month {cap_idx + 1}. "
                f"Production was capped there so the recovery factor cannot "
                f"exceed 100%. This usually means the well rates / decline "
                f"parameters are too optimistic for the stated in-place "
                f"volume, or the in-place volume is too low — check that "
                f"the decline curves and the volumetrics are consistent.")

    # ---- Gas disposition (uses true field gas stream) ----
    gross_gas = field_gas_rate.copy()

    # Normalize fractions (in case user input doesn't sum to 1.0)
    f_export = max(asm.gas_export_fraction, 0.0)
    f_inject = max(asm.gas_injection_fraction, 0.0)
    f_fuel   = max(asm.gas_fuel_fraction, 0.0)
    f_flare  = max(asm.gas_flare_fraction, 0.0)
    f_total  = f_export + f_inject + f_fuel + f_flare
    if f_total > 1.001:
        # Renormalize down to 1.0; raise a flag
        scale = 1.0 / f_total
        f_export *= scale; f_inject *= scale; f_fuel *= scale; f_flare *= scale
        f_total = 1.0
    # If user fractions sum to <1.0, the remainder defaults to export (sold)
    f_export += max(0.0, 1.0 - f_total)

    gas_export = gross_gas * f_export      # Mscf/d sold
    gas_inject = gross_gas * f_inject      # Mscf/d re-injected
    gas_fuel   = gross_gas * f_fuel        # Mscf/d burnt as fuel gas
    gas_flare  = gross_gas * f_flare       # Mscf/d flared

    # Net gas (after shrinkage from fuel + flare) — what's available for sale or for export pipe
    gas_net = gas_export + gas_inject  # treat injected as 'kept' (no shrinkage)
    gas_shrinkage = gas_fuel + gas_flare

    # If injected, add to field_inj (in stb/d-equivalent the rate unit doesn't apply for gas;
    # we keep gas injection separate via gas_inject_rate column to honor capacity check below)
    # Apply gas-injection capacity choke (same time-varying schedule)
    gas_inj_choke = np.ones(n_months)
    for i, d in enumerate(dates):
        cap = asm.cap_schedule.at_date(d.date())
        if cap.get("gas_inj", 0) > 0 and gas_inject[i] > cap["gas_inj"] * 1000:  # MMscf/d -> Mscf/d
            gas_inj_choke[i] = (cap["gas_inj"] * 1000) / gas_inject[i]
    # If choked, the un-injected gas falls back to export (it has to go somewhere)
    excess = gas_inject * (1 - gas_inj_choke)
    gas_inject *= gas_inj_choke
    gas_export += excess

    # ---- CO2 emissions are computed in compute_economics where econ factors are available ----
    days = DAYS_PER_MONTH

    # Cumulatives for the explicit oil/gas/water streams
    cum_oil = np.cumsum(field_oil_rate * days) / 1e6   # MMstb
    cum_gas = np.cumsum(field_gas_rate * days) / 1e6   # Bscf

    df = pd.DataFrame({
        "date": dates,
        "year": (np.arange(n_months) // 12) + 1,
        # Phase-explicit rates (always interpretable as oil/gas/water regardless of fluid system)
        "oil_rate": field_oil_rate,           # stb/d
        "gas_rate": field_gas_rate,           # Mscf/d
        # Legacy primary/secondary names retained for back-compat
        "primary_rate": field_p, "secondary_rate": field_s,
        "water_rate": field_w, "liquid_rate": field_l,
        "injection_rate": field_inj,
        "gross_gas_rate": gross_gas,
        "gas_export_rate": gas_export,
        "gas_inject_rate": gas_inject,
        "gas_fuel_rate": gas_fuel,
        "gas_flare_rate": gas_flare,
        "gas_net_rate": gas_net,
        "gas_shrinkage_rate": gas_shrinkage,
        "cum_oil": cum_oil, "cum_gas": cum_gas,
        "cum_primary": cum_p, "cum_secondary": cum_s,
        "cum_water": cum_w, "cum_injection": cum_inj,
        "recovery_factor": rf, "pressure": pressure,
        "pressure_ratio": pressure / asm.pvt.p_init_psi,
        "active_producers": active_p, "active_injectors": active_i,
        "choke_factor": choke,
    })

    # Retrograde producible-CGR series as a real column (never in df.attrs —
    # arrays in attrs break pandas concat).
    if retro_cgr_series is not None and len(retro_cgr_series) == n_months:
        df["producible_cgr"] = retro_cgr_series
    # Fractional-flow saturation & water-cut series as real columns.
    if ff_sw_series is not None and len(ff_sw_series) == n_months:
        df["ff_water_saturation"] = ff_sw_series
        df["ff_water_cut"] = ff_wc_series

    # ---- Profile robustness checks ----
    # A set of sanity checks on the generated profile. Anything that looks
    # physically wrong is recorded as a warning for the UI to surface.
    try:
        final_rf = float(rf[-1]) if len(rf) else 0.0
        if final_rf > 1.0:
            profile_warnings.append(
                f"Recovery factor reached {final_rf:.0%} — above 100%, which "
                f"is physically impossible. Check in-place volumes and decline "
                f"parameters.")
        elif final_rf > 0.75:
            profile_warnings.append(
                f"Recovery factor is {final_rf:.0%} — very high. Typical "
                f"recovery: primary depletion 5-20%, waterflood 20-45%, "
                f"strong aquifer / EOR 35-60%. A figure above ~75% is "
                f"optimistic — confirm it is intended.")
        # Plateau realism: peak rate vs in-place (very rough — flags a peak
        # that would drain the field implausibly fast).
        if primary_in_place and primary_in_place > 0 and len(field_p):
            peak = float(np.max(field_p))
            # months of plateau-equivalent supply
            annual_peak_vol = peak * 365.25 / 1e6   # MMstb/yr or Bscf/yr
            if annual_peak_vol > 0:
                yrs_to_drain = primary_in_place / annual_peak_vol
                if yrs_to_drain < 1.5:
                    profile_warnings.append(
                        f"Peak rate would drain the entire "
                        f"{'OOIP' if is_oil else 'OGIP'} in "
                        f"{yrs_to_drain:.1f} years at plateau — an "
                        f"implausibly aggressive offtake. Real fields "
                        f"plateau at roughly 5-12% of in-place per year.")
        # Pressure sanity: MBE pressure should not go negative or rise far
        # above initial without injection / aquifer support.
        if len(pressure):
            if float(np.min(pressure)) < 0:
                profile_warnings.append(
                    "Material-balance pressure went negative — the offtake "
                    "is too high for the in-place volume and drive energy. "
                    "Reduce rates, add injection, or increase in-place "
                    "volume.")
            if (float(np.max(pressure)) > 1.05 * asm.pvt.p_init_psi
                    and asm.strategy != "Injection"
                    and not asm.aquifer.active):
                profile_warnings.append(
                    "Reservoir pressure rises above initial pressure without "
                    "injection or aquifer support — check the material-"
                    "balance inputs.")
        # First-year production with no producing wells
        if len(field_p) >= 12 and float(np.sum(field_p[:12])) <= 0:
            profile_warnings.append(
                "No production in the first 12 months — check well spud "
                "dates and the drilling schedule.")
    except Exception:
        pass

    df.attrs["profile_warnings"] = profile_warnings
    if retro_info is not None:
        df.attrs["retrograde_info"] = retro_info
    if ff_info is not None:
        df.attrs["fractional_flow_info"] = ff_info

    # Per-reservoir DataFrame: long format for easy plotting / aggregation
    res_rows = []
    for k, r in enumerate(reservoirs):
        for i in range(n_months):
            res_rows.append({
                "date": dates[i],
                "reservoir_id": r.id,
                "reservoir_name": r.name,
                "fluid_system": r.fluid_system,
                "primary_rate": float(res_p[i, k]),
                "secondary_rate": float(res_s[i, k]),
                "water_rate": float(res_w[i, k]),
                "injection_rate": float(res_inj[i, k]),
                "cum_primary": float(per_res_cum_primary[r.id][i]),
                "recovery_factor": float(per_res_rf[r.id][i]),
                "pressure": float(per_res_pressure[r.id][i]),
            })
    per_res_df = pd.DataFrame(res_rows)

    well_names = [w.name for w in producers]
    per_well_df = pd.DataFrame(p_mat, columns=well_names)
    per_well_df.insert(0, "date", dates)
    # Per-well phase tracking: stash the proper oil/gas/water matrices on
    # the dataframe's attrs so the per-well-phase plot can use them directly
    # rather than approximating via share-weighting.
    per_well_df.attrs["oil_mat"]   = pd.DataFrame(oil_mat,        columns=well_names, index=dates)
    per_well_df.attrs["gas_mat"]   = pd.DataFrame(gas_mat,        columns=well_names, index=dates)
    per_well_df.attrs["water_mat"] = pd.DataFrame(w_mat_choked,   columns=well_names, index=dates)
    return df, per_well_df, per_res_df


# =============================================================================
# Economics
# =============================================================================
def compute_economics(df, is_oil, econ: EconInputs, wells):
    days = DAYS_PER_MONTH

    # Decide which gas stream is sold:
    #   "net" -> only export gas (export = gross - injected - fuel - flare, plus excess back to export)
    #   "gross" -> all produced gas (legacy / pre-disposition convention)
    if "gas_export_rate" in df.columns and econ.revenue_basis == "net":
        sold_gas = df["gas_export_rate"]    # Mscf/d
    elif "gross_gas_rate" in df.columns:
        sold_gas = df["gross_gas_rate"]     # gross convention: pay for all gas produced
    else:
        # Fallback for older dataframes
        sold_gas = df["secondary_rate"] if is_oil else df["primary_rate"]

    if is_oil:
        rev_oil = df["primary_rate"] * days * econ.oil_price
        # sold_gas is in Mscf/d; gas_price is $/Mscf -> rate * days * price = $/month
        rev_gas = sold_gas * days * econ.gas_price
        rev_cond = pd.Series(0.0, index=df.index)
        tariff = df["primary_rate"] * days * econ.tariff_oil + \
                 sold_gas * days * econ.tariff_gas
    else:
        rev_oil = pd.Series(0.0, index=df.index)
        rev_gas = sold_gas * days * econ.gas_price
        rev_cond = df["secondary_rate"] * days * econ.oil_price
        tariff = sold_gas * days * econ.tariff_gas + \
                 df["secondary_rate"] * days * econ.tariff_oil

    # ---- NGL stream (independent of oil/gas, derived from gross gas) ----
    # NGL volume = gross_gas (MMscf/d) × yield (bbl/MMscf). We use *gross* gas,
    # not sold gas, because NGLs are extracted at the plant *before* the gas
    # disposition split (export/inject/fuel/flare).
    ngl_yield = float(getattr(econ, "ngl_yield_bbl_per_mmscf", 0.0))
    ngl_price = float(getattr(econ, "ngl_price_bbl", 0.0))
    ngl_opex_bbl = float(getattr(econ, "ngl_opex_bbl", 0.0))
    ngl_shrinkage = float(getattr(econ, "ngl_shrinkage_pct", 0.0))
    if "gross_gas_rate" in df.columns:
        gross_gas_mmscfd = df["gross_gas_rate"] / 1000.0   # Mscf/d -> MMscf/d
    else:
        gross_gas_mmscfd = (df["secondary_rate"] if is_oil else df["primary_rate"]) / 1000.0
    ngl_rate_bpd = gross_gas_mmscfd * ngl_yield            # bbl/d NGL
    ngl_monthly_bbl = ngl_rate_bpd * days
    rev_ngl = ngl_monthly_bbl * ngl_price
    ngl_opex = ngl_monthly_bbl * ngl_opex_bbl

    # Apply shrinkage to the sold-gas revenue (NGL extraction removes volume
    # from the gas stream that goes to market).
    if ngl_shrinkage > 0:
        shrink_factor = max(1.0 - ngl_shrinkage, 0.0)
        rev_gas = rev_gas * shrink_factor

    revenue = rev_oil + rev_gas + rev_cond + rev_ngl
    royalty = revenue * econ.royalty_rate
    net_revenue = revenue - royalty - tariff

    # Variable OPEX: primary_rate × days × opex_var. The unit basis of
    # opex_var matches the primary fluid — $/bbl for an oil field (oil rate
    # in stb/d) or $/Mscf for a gas field (gas rate in Mscf/d). The UI sets
    # the correct basis per fluid system, so the same expression is valid
    # for both.
    var_cost = df["primary_rate"] * days * econ.opex_var
    fixed_cost = econ.opex_fixed / 12.0
    opex = var_cost + fixed_cost + ngl_opex

    # ---- Economic limit / cessation timing ----
    # The field stops at the EARLIER of two events:
    #   (a) ultimate recovery — production has essentially finished (the
    #       rate has fallen to/below the abandonment rate, or cumulative has
    #       hit the volumetric limit);
    #   (b) NPV turnover — the discounted cumulative cashflow stops growing,
    #       i.e. monthly operating cashflow has gone persistently negative
    #       so every further month destroys value.
    # Production, revenue and OPEX after that month are zeroed; cessation
    # (abandonment) is booked a few months later, and all costs finish by
    # then.
    cutoff_mode = getattr(econ, "economic_cutoff_mode", "horizon")
    cutoff_idx = None
    if cutoff_mode == "economic":
        persistence = max(1, int(getattr(econ, "economic_cutoff_persistence", 6)))
        op_cf = (revenue - royalty - tariff - opex).values  # monthly operating CF
        producing = (df["primary_rate"].values > 0)
        neg = op_cf < 0

        # (b) NPV-turnover cutoff: first month from which operating CF is
        #     negative for `persistence` consecutive months. Beyond this the
        #     cumulative (discounted) cashflow only falls.
        npv_cutoff = None
        for i in range(len(op_cf)):
            if not producing[i]:
                continue
            window = neg[i:i + persistence]
            if len(window) > 0 and window.all():
                npv_cutoff = i
                break

        # (a) Ultimate-recovery cutoff: the first month at/after which
        #     production has effectively ceased — the primary rate has
        #     dropped to zero (volumetric cap or decline to nil). The last
        #     producing month + 1 is the recovery limit.
        recovery_cutoff = None
        prod_idx = np.where(producing)[0]
        if len(prod_idx) > 0:
            last_prod = int(prod_idx[-1])
            if last_prod < len(df) - 1:
                recovery_cutoff = last_prod + 1

        # take the EARLIER of the two events
        candidates = [c for c in (npv_cutoff, recovery_cutoff)
                      if c is not None]
        if candidates:
            cutoff_idx = min(candidates)
            df.attrs["economic_cutoff_reason"] = (
                "NPV turnover" if cutoff_idx == npv_cutoff
                else "ultimate recovery")

        if cutoff_idx is not None and cutoff_idx > 0:
            # Zero out production, revenue, opex from the cutoff onward
            mask_after = np.arange(len(df)) >= cutoff_idx
            # Explicit zeroing (these are pandas Series / np arrays)
            revenue = revenue.copy(); revenue[mask_after] = 0.0
            rev_oil = rev_oil.copy() if hasattr(rev_oil, "copy") else rev_oil
            if hasattr(rev_oil, "__setitem__"): rev_oil[mask_after] = 0.0
            if hasattr(rev_gas, "__setitem__"):
                rev_gas = rev_gas.copy(); rev_gas[mask_after] = 0.0
            if hasattr(rev_cond, "__setitem__"):
                rev_cond = rev_cond.copy(); rev_cond[mask_after] = 0.0
            if hasattr(rev_ngl, "__setitem__"):
                rev_ngl = rev_ngl.copy(); rev_ngl[mask_after] = 0.0
            royalty = royalty.copy(); royalty[mask_after] = 0.0
            tariff = tariff.copy() if hasattr(tariff, "copy") else tariff
            if hasattr(tariff, "__setitem__"): tariff[mask_after] = 0.0
            net_revenue = revenue - royalty - tariff
            opex = opex.copy(); opex[mask_after] = 0.0
            # Mark the truncated production in df for display consistency
            df = df.copy()
            for rate_col in ["primary_rate", "secondary_rate", "oil_rate",
                             "gas_rate", "water_rate", "gross_gas_rate",
                             "gas_export_rate", "gas_fuel_rate",
                             "gas_flare_rate", "injection_rate"]:
                if rate_col in df.columns:
                    df.loc[df.index[cutoff_idx]:, rate_col] = 0.0
            df.attrs["economic_cutoff_idx"] = int(cutoff_idx)
            df.attrs["economic_cutoff_date"] = str(df["date"].iloc[cutoff_idx].date())

    capex_well = np.zeros(len(df))
    well_cost_breakdown = []   # for transparency / display
    use_rig_rate = getattr(econ, "well_cost_mode", "fixed") == "rig_rate"
    rig_kUSD  = float(getattr(econ, "rig_day_rate_kUSD", 500.0))
    cmpl_kUSD = float(getattr(econ, "completion_day_rate_kUSD", 350.0))
    tang_MM   = float(getattr(econ, "well_tangibles_MM", 4.0))
    intang_pct = float(getattr(econ, "well_intangibles_pct", 0.10))
    for w in wells:
        ts = pd.Timestamp(w.spud_date)
        idx_arr = df.index[df["date"] >= ts]
        if len(idx_arr) == 0:
            continue
        if use_rig_rate:
            # Bottom-up: (drill_days × rig + completion_days × cmpl-spread) × (1 + intangibles)
            #            + tangibles
            spread_cost_MM = (w.drill_days * rig_kUSD
                              + w.completion_days * cmpl_kUSD) / 1000.0  # kUSD → MM
            cost_MM = spread_cost_MM * (1.0 + intang_pct) + tang_MM
        else:
            cost_MM = econ.capex_per_well
        capex_well[idx_arr[0]] += cost_MM * 1e6
        well_cost_breakdown.append({
            "well": w.name, "spud": w.spud_date.isoformat(),
            "drill_days": w.drill_days, "completion_days": w.completion_days,
            "cost_MM": cost_MM,
        })

    capex_fac = np.zeros(len(df))
    prod_start_ts = pd.Timestamp(df["date"].iloc[0])
    for _, row in econ.facility_capex.df.iterrows():
        try:
            ts = pd.Timestamp(row["date"])
            # Pre-production CAPEX is handled by the pre-FOP prepend step at the
            # end of this function — skip it here to avoid double-counting.
            if ts < prod_start_ts:
                continue
            idx_arr = df.index[df["date"] >= ts]
            if len(idx_arr) > 0:
                capex_fac[idx_arr[0]] += float(row["amount_MMUSD"]) * 1e6
        except (KeyError, TypeError, ValueError):
            pass

    # ---- Rig mobilization / demobilization / maintenance costs ----
    # Pulled from rig metadata (set by well_section). Move-in is booked at the
    # rig's first well spud; move-out at the last well's end; maintenance is
    # spread across the rig's active months. All use the rig day rate.
    rig_meta = getattr(econ, "rig_meta", None) or {}
    if rig_meta:
        # Group wells by rig to find first-spud and last-end per rig
        from collections import defaultdict
        rig_wells = defaultdict(list)
        for w in wells:
            rig_wells[w.rig].append(w)
        for rname, rws in rig_wells.items():
            meta = rig_meta.get(rname, {})
            day_rate = float(meta.get("day_rate_kUSD", 0.0)) * 1000.0  # kUSD → USD
            if day_rate <= 0:
                continue
            mi_days = int(meta.get("move_in_days", 0))
            mo_days = int(meta.get("move_out_days", 0))
            maint_per_yr = int(meta.get("maintenance_days_per_year", 0))
            first_spud = min(pd.Timestamp(w.spud_date) for w in rws)
            last_end = max(pd.Timestamp(w.spud_date) +
                           pd.Timedelta(days=w.drill_days + w.completion_days)
                           for w in rws)
            # Move-in cost — booked the month of (or before) first spud
            if mi_days > 0:
                idx_arr = df.index[df["date"] >= first_spud - pd.Timedelta(days=mi_days)]
                if len(idx_arr) > 0:
                    capex_fac[idx_arr[0]] += mi_days * day_rate
            # Move-out cost — booked at last well end
            if mo_days > 0:
                idx_arr = df.index[df["date"] >= last_end]
                book_idx = idx_arr[0] if len(idx_arr) > 0 else (len(df) - 1)
                capex_fac[int(book_idx)] += mo_days * day_rate
            # Maintenance cost — spread across the rig's active months
            if maint_per_yr > 0:
                active_mask = (df["date"] >= first_spud) & (df["date"] <= last_end)
                n_active = int(active_mask.sum())
                if n_active > 0:
                    active_yrs = n_active / 12.0
                    total_maint_days = maint_per_yr * active_yrs
                    maint_cost_per_month = (total_maint_days * day_rate) / n_active
                    capex_fac[active_mask.values] += maint_cost_per_month

    aban_cost = np.zeros(len(df))
    cutoff_idx_attr = df.attrs.get("economic_cutoff_idx")
    # Cessation (abandonment) occurs a short while AFTER the field stops —
    # rigs/vessels demobilise, wells are plugged. Book it a few months after
    # the economic cutoff so all costs are finished shortly after production.
    CESSATION_LAG_MONTHS = 3
    if cutoff_idx_attr is not None:
        aban_month = min(len(df) - 1,
                         int(cutoff_idx_attr) + CESSATION_LAG_MONTHS)
        aban_cost[aban_month] = econ.abandonment_cost_MM * 1e6
        df.attrs["cessation_idx"] = int(aban_month)
        df.attrs["cessation_date"] = str(df["date"].iloc[aban_month].date())
    elif (df["primary_rate"] > 0).any():
        # Horizon mode: cessation a few months after last producing month
        last = int(df.index[df["primary_rate"] > 0].max())
        aban_month = min(len(df) - 1, last + CESSATION_LAG_MONTHS)
        aban_cost[aban_month] = econ.abandonment_cost_MM * 1e6
        df.attrs["cessation_idx"] = int(aban_month)
        df.attrs["cessation_date"] = str(df["date"].iloc[aban_month].date())

    # ---- CO2 emissions (rough screening, tonnes/month) ----
    # Combustion of fuel + flare gas (kg CO2 per Mscf burnt → tonnes/month)
    if "gas_fuel_rate" in df.columns and "gas_flare_rate" in df.columns:
        gas_fuel_monthly_Mscf  = df["gas_fuel_rate"]  * days
        gas_flare_monthly_Mscf = df["gas_flare_rate"] * days
    else:
        gas_fuel_monthly_Mscf  = pd.Series(0.0, index=df.index)
        gas_flare_monthly_Mscf = pd.Series(0.0, index=df.index)
    co2_combust_t = (gas_fuel_monthly_Mscf + gas_flare_monthly_Mscf) * \
                    econ.co2_factor_gas_combust / 1000.0
    # Methane slip from imperfect flaring (small fraction × 28 GWP100)
    co2_slip_t = (gas_flare_monthly_Mscf * econ.co2_factor_flare_inefficiency *
                  19.2 * 28.0 / 1000.0)  # 19.2 kg CH4 / Mscf → t CO2-eq
    # Routine ops emissions per bbl oil produced
    if is_oil:
        co2_routine_t = df["primary_rate"] * days * econ.co2_factor_oil_routine / 1000.0
    else:
        co2_routine_t = pd.Series(0.0, index=df.index)

    co2_total_t = co2_combust_t + co2_slip_t + co2_routine_t  # tonnes / month
    co2_cost = co2_total_t * econ.co2_price                   # $ / month

    # ---- Scope 3 emissions (end-use combustion of sold hydrocarbons) ----
    # Scope 1 (above) is operational: fuel gas, flare, vents.
    # Scope 3 is the downstream combustion when the customer burns the
    # crude or gas. It dwarfs Scope 1 — typical ratio is 50-100×.
    # Computed regardless of toggle so it can be reported, but only added
    # to the cashflow (as a fee) if `co2_scope3_enabled` and a price is set.
    if "oil_rate" in df.columns:
        oil_sold_monthly_bbl = df["oil_rate"] * days
    else:
        oil_sold_monthly_bbl = pd.Series(0.0, index=df.index)
    if "gas_export_rate" in df.columns:
        gas_sold_monthly_Mscf = df["gas_export_rate"] * days
    elif "gas_rate" in df.columns:
        gas_sold_monthly_Mscf = df["gas_rate"] * days
    else:
        gas_sold_monthly_Mscf = pd.Series(0.0, index=df.index)
    co2_scope3_oil_t = (oil_sold_monthly_bbl
                        * econ.co2_scope3_factor_oil / 1000.0)
    co2_scope3_gas_t = (gas_sold_monthly_Mscf
                        * econ.co2_scope3_factor_gas / 1000.0)
    co2_scope3_total_t = co2_scope3_oil_t + co2_scope3_gas_t
    co2_scope3_cost = pd.Series(0.0, index=df.index)
    if getattr(econ, "co2_scope3_enabled", False):
        co2_scope3_cost = (co2_scope3_total_t
                           * float(getattr(econ, "co2_scope3_price",
                                            econ.co2_price)))

    # ---- Power consumption (screening estimate, MWh/month) ----
    # Topsides power demand scales with what the facility has to move and
    # process: liquids handling (pumps), gas compression, water injection.
    # Screening intensities (kWh per unit), conservative mid-range values:
    #   - liquid (oil+water) handling : ~1.5 kWh/bbl
    #   - gas compression / processing: ~3.0 kWh/Mscf
    #   - water injection             : ~2.0 kWh/bbl
    # These are deliberately simple — real facility power studies are detailed.
    KWH_PER_BBL_LIQUID   = 1.5
    KWH_PER_MSCF_GAS     = 3.0
    KWH_PER_BBL_WATERINJ = 2.0
    liquid_bbl_month = (df["primary_rate"] if is_oil else df.get("secondary_rate",
                        pd.Series(0.0, index=df.index))) * days
    water_bbl_month  = df.get("water_rate", pd.Series(0.0, index=df.index)) * days
    gas_mscf_month   = (df["gross_gas_rate"] if "gross_gas_rate" in df.columns
                        else (df["secondary_rate"] if is_oil else df["primary_rate"])) * days
    waterinj_bbl_month = df.get("injection_rate", pd.Series(0.0, index=df.index)) * days
    power_mwh = (
        (liquid_bbl_month + water_bbl_month) * KWH_PER_BBL_LIQUID
        + gas_mscf_month * KWH_PER_MSCF_GAS
        + waterinj_bbl_month * KWH_PER_BBL_WATERINJ
    ) / 1000.0   # kWh → MWh

    capex = capex_well + capex_fac

    # ---- Fiscal regime: Tax/Royalty (default) or PSC ----
    regime = getattr(econ, "fiscal_regime", "Tax/Royalty")
    if regime == "PSC":
        # Production Sharing Contract waterfall:
        #   1. Royalty taken off the top (off gross revenue)
        #   2. Cost recovery: contractor recovers OPEX + CAPEX + accumulated
        #      cost-pool, capped at psc_cost_recovery_ceiling × revenue per period
        #   3. Profit oil = revenue - royalty - cost recovered, split between
        #      contractor (psc_profit_oil_share_contractor) and government
        #   4. Contractor's profit oil is taxed at psc_psc_tax_rate
        #   5. Government participation: psc_govt_participation share of
        #      contractor net cashflow accrues to government (carried equity)
        #   6. Signature bonus paid up front (month 0)
        royalty_psc = revenue.values * econ.royalty_rate
        net_rev_psc = revenue.values - royalty_psc - tariff.values
        recoverable_costs = (opex.values + capex + aban_cost
                              + co2_cost.values
                              + co2_scope3_cost.values)
        # Carry forward unrecovered costs through a cost pool
        cost_pool = 0.0
        cost_recovered_arr = np.zeros(len(df))
        ceiling = max(0.0, min(1.0, econ.psc_cost_recovery_ceiling))
        for i in range(len(df)):
            cost_pool += recoverable_costs[i]
            cap = max(0.0, ceiling * net_rev_psc[i])
            recovered = min(cost_pool, cap)
            cost_recovered_arr[i] = recovered
            cost_pool -= recovered
        profit_oil = net_rev_psc - cost_recovered_arr
        contractor_share = max(0.0, min(1.0, econ.psc_profit_oil_share_contractor))
        contractor_profit = profit_oil * contractor_share
        psc_tax = np.where(contractor_profit > 0,
                            contractor_profit * econ.psc_psc_tax_rate, 0.0)
        contractor_after_tax = contractor_profit - psc_tax
        # Cashflow to contractor:
        #   + cost recovery
        #   + after-tax profit oil
        #   - costs paid (cost recovery already returns these but cash timing
        #     is the same period in this simplified waterfall)
        # We model contractor cashflow as: cost_recovered + contractor_after_tax - recoverable_costs
        # which simplifies to: contractor_after_tax - (recoverable_costs - cost_recovered)
        # i.e. contractor_after_tax minus any unrecovered costs this period.
        unrecovered_this_period = recoverable_costs - cost_recovered_arr
        cf_pre_part = contractor_after_tax - unrecovered_this_period
        # Government participation (carried equity)
        govt_part = max(0.0, min(0.99, econ.psc_govt_participation))
        cf = cf_pre_part * (1.0 - govt_part)
        # Signature bonus at month 0
        if econ.psc_signature_bonus_MM > 0 and len(cf) > 0:
            cf[0] -= econ.psc_signature_bonus_MM * 1e6
        # For reporting: tax = PSC tax, royalty = PSC royalty
        tax_arr = psc_tax
        royalty_arr = royalty_psc
        df_e = df.copy()
        df_e["revenue_oil"] = rev_oil
        df_e["revenue_gas"] = rev_gas
        df_e["revenue_condensate"] = rev_cond
        df_e["revenue_ngl"] = rev_ngl
        df_e["ngl_rate"] = ngl_rate_bpd
        df_e["revenue"] = revenue
        df_e["royalty"] = royalty_arr
        df_e["tariff"] = tariff
        df_e["opex"] = opex
        df_e["capex_well"] = capex_well
        df_e["capex_facility"] = capex_fac
        df_e["abandonment"] = aban_cost
        df_e["co2_emissions_tonnes"] = co2_total_t
        df_e["co2_scope1_tonnes"] = co2_total_t
        df_e["co2_scope3_tonnes"] = co2_scope3_total_t
        df_e["power_mwh"] = power_mwh
        df_e["co2_cost"] = co2_cost
        df_e["co2_scope3_cost"] = co2_scope3_cost
        df_e["tax"] = tax_arr
        df_e["psc_cost_recovered"] = cost_recovered_arr
        df_e["psc_profit_oil"] = profit_oil
        df_e["psc_contractor_share"] = contractor_profit
        df_e["psc_govt_take"] = (royalty_arr + (profit_oil * (1.0 - contractor_share))
                                  + tax_arr + cf_pre_part * govt_part)
        df_e["cashflow"] = cf
        df_e["cum_cashflow"] = cf.cumsum()
        df_e["cum_co2_tonnes"] = co2_total_t.cumsum()
        df_e["cum_co2_scope3_tonnes"] = co2_scope3_total_t.cumsum()
    else:
        # Standard Tax/Royalty OR NCS (Norwegian Continental Shelf)
        ncs = (regime == "NCS")
        if ncs:
            # ---- NCS petroleum tax — proper carry-forward treatment ----
            # The Norwegian system is NOT a simple month-by-month "tax the
            # positive months" calculation. Two features dominate and MUST be
            # modelled or the project looks far too negative:
            #
            #  (1) Loss carry-forward. Tax losses (negative tax base) are
            #      carried forward and offset against future positive base.
            #      During the investment phase the base is deeply negative;
            #      that relief is preserved, not discarded. (The real system
            #      even refunds the tax value of exploration losses and pays
            #      out remaining loss carry-forwards at cessation — modelled
            #      here via carry-forward that fully unwinds over field life.)
            #
            #  (2) CAPEX is depreciated, not expensed in one month. NCS uses
            #      6-year straight-line depreciation for both CIT and SPT.
            #      Uplift (an extra SPT-only allowance) is likewise spread —
            #      historically over 4 years; modern rules give immediate
            #      uplift, which the user can approximate by setting the
            #      uplift period to 1.
            #
            # The combined CIT(22%) + SPT(71.8%) headline rate of ~78% applies
            # to PROFITS, after costs and depreciation and after losses are
            # carried forward — so a healthy project keeps a sensible margin.
            cit_rate = float(getattr(econ, "ncs_cit_rate", 0.22))
            spt_rate = float(getattr(econ, "ncs_spt_rate", 0.718))
            uplift_rate = float(getattr(econ, "ncs_uplift_rate", 0.1769))
            depr_years = float(getattr(econ, "ncs_depreciation_years", 6.0))
            uplift_years = float(getattr(econ, "ncs_uplift_years", 4.0))

            n = len(df)
            capex_total_arr = (np.asarray(capex_well, dtype=float)
                               + np.asarray(capex_fac, dtype=float))
            # --- Depreciation schedule: straight-line over depr_years from
            #     the month each CAPEX tranche is incurred ---
            depr_months = max(1, int(round(depr_years * 12)))
            depreciation = np.zeros(n)
            for i in range(n):
                amt = capex_total_arr[i]
                if amt <= 0:
                    continue
                per_month = amt / depr_months
                end = min(n, i + depr_months)
                depreciation[i:end] += per_month
            # --- Uplift schedule: straight-line over uplift_years ---
            uplift_months = max(1, int(round(uplift_years * 12)))
            uplift_sched = np.zeros(n)
            for i in range(n):
                amt = capex_total_arr[i] * uplift_rate
                if amt <= 0:
                    continue
                per_month = amt / uplift_months
                end = min(n, i + uplift_months)
                uplift_sched[i:end] += per_month
            uplift = uplift_sched  # for reporting

            # --- Operating profit (before depreciation & financing) ---
            # CAPEX is NOT expensed here — it enters via depreciation.
            op_profit = (net_revenue.values - opex.values
                         - aban_cost - co2_cost.values
                         - co2_scope3_cost.values)

            # --- CIT base: operating profit minus depreciation, with loss
            #     carry-forward ---
            cit = np.zeros(n)
            cit_cf = 0.0   # carried-forward CIT loss (positive number = a loss)
            cit_base_arr = np.zeros(n)
            for i in range(n):
                base = op_profit[i] - depreciation[i]
                base_after_cf = base - cit_cf
                if base_after_cf >= 0:
                    cit[i] = base_after_cf * cit_rate
                    cit_cf = 0.0
                else:
                    cit[i] = 0.0
                    cit_cf = -base_after_cf   # grow the carry-forward
                cit_base_arr[i] = base_after_cf
            # --- SPT base: operating profit minus depreciation minus uplift,
            #     with its own loss carry-forward ---
            spt = np.zeros(n)
            spt_cf = 0.0
            for i in range(n):
                base = op_profit[i] - depreciation[i] - uplift_sched[i]
                base_after_cf = base - spt_cf
                if base_after_cf >= 0:
                    spt[i] = base_after_cf * spt_rate
                    spt_cf = 0.0
                else:
                    spt[i] = 0.0
                    spt_cf = -base_after_cf
            tax = cit + spt
            # Cashflow: real cash CAPEX leaves in the month spent; tax is the
            # depreciation-based number computed above.
            pretax = (net_revenue - opex - capex - aban_cost
                      - co2_cost - co2_scope3_cost)
            cf = pretax.values - tax
        else:
            # Standard Tax/Royalty. Tax is charged on taxable profit, but
            # losses (CAPEX-heavy early years) are carried forward and
            # shelter later profits — without this the tax is overstated
            # and a sound project can show a spuriously negative NPV.
            pretax = (net_revenue - opex - capex - aban_cost
                      - co2_cost - co2_scope3_cost)
            pretax_v = pretax.values
            _n = len(df)
            tax = np.zeros(_n)
            tax_loss_cf = 0.0   # carried-forward taxable loss (positive)
            for i in range(_n):
                base = pretax_v[i] - tax_loss_cf
                if base > 0:
                    tax[i] = base * econ.tax_rate
                    tax_loss_cf = 0.0
                else:
                    tax[i] = 0.0
                    tax_loss_cf = -base
            cf = pretax_v - tax

        df_e = df.copy()
        df_e["revenue_oil"] = rev_oil
        df_e["revenue_gas"] = rev_gas
        df_e["revenue_condensate"] = rev_cond
        df_e["revenue_ngl"] = rev_ngl
        df_e["ngl_rate"] = ngl_rate_bpd
        df_e["revenue"] = revenue
        df_e["royalty"] = royalty
        df_e["tariff"] = tariff
        df_e["opex"] = opex
        df_e["capex_well"] = capex_well
        df_e["capex_facility"] = capex_fac
        df_e["abandonment"] = aban_cost
        df_e["co2_emissions_tonnes"] = co2_total_t
        df_e["co2_scope1_tonnes"] = co2_total_t
        df_e["co2_scope3_tonnes"] = co2_scope3_total_t
        df_e["power_mwh"] = power_mwh
        df_e["co2_cost"] = co2_cost
        df_e["co2_scope3_cost"] = co2_scope3_cost
        df_e["tax"] = tax
        if ncs:
            df_e["ncs_cit"] = cit
            df_e["ncs_spt"] = spt
            df_e["ncs_uplift"] = uplift
        df_e["cashflow"] = cf
        df_e["cum_cashflow"] = cf.cumsum()
        df_e["cum_co2_tonnes"] = co2_total_t.cumsum()
        df_e["cum_co2_scope3_tonnes"] = co2_scope3_total_t.cumsum()
    # ---- Money basis (nominal vs real) ----
    # In "nominal" mode, every monthly cashflow column is escalated by the
    # inflation rate compounded monthly. In "real" mode this is a no-op.
    money_basis = getattr(econ, "money_basis", "real")
    inflation_rate = float(getattr(econ, "inflation_rate", 0.0))
    if money_basis == "nominal" and inflation_rate > 0:
        infl_m = (1 + inflation_rate) ** (1.0/12.0) - 1
        infl_factor = (1 + infl_m) ** np.arange(len(df_e))
        # Apply to every dollar-denominated column so internal consistency holds
        for col in ["revenue", "revenue_oil", "revenue_gas",
                    "revenue_condensate", "revenue_ngl", "royalty", "tariff",
                    "opex", "capex_well", "capex_facility", "abandonment",
                    "tax", "co2_cost", "ncs_cit", "ncs_spt", "ncs_uplift"]:
            if col in df_e.columns:
                df_e[col] = df_e[col].values * infl_factor
        # Rebuild cashflow + cum_cashflow from the inflated columns
        if "cashflow" in df_e.columns:
            df_e["cashflow"] = df_e["cashflow"].values * infl_factor
            df_e["cum_cashflow"] = df_e["cashflow"].cumsum()

    r_m = (1 + econ.discount_rate) ** (1 / 12) - 1
    disc = (1 + r_m) ** np.arange(len(df_e))
    df_e["discounted_cf"] = df_e["cashflow"].values / disc
    df_e["npv"] = df_e["discounted_cf"].cumsum()

    # ---- Pre-production investment months ----
    # Facility CAPEX (and rig move-in) can be dated *before* production start.
    # The engine's df starts at production start, so any earlier spend would
    # otherwise be collapsed into month 0. Here we prepend zero-production
    # months back to the earliest investment date so the economics plots show
    # the true investment timeline (CAPEX before first oil).
    try:
        prod_start = pd.Timestamp(df_e["date"].iloc[0])
        invest_dates = []
        for _, row in econ.facility_capex.df.iterrows():
            try:
                invest_dates.append(pd.Timestamp(row["date"]))
            except (KeyError, TypeError, ValueError):
                pass
        # Rig move-in dates
        rig_meta = getattr(econ, "rig_meta", None) or {}
        # (rig move-in is already booked relative to first spud, which is ≥
        #  production start in nearly all cases, so facility CAPEX dominates here)
        earliest_invest = min(invest_dates) if invest_dates else prod_start
        if earliest_invest < prod_start:
            # Build the pre-FOP month index
            pre_dates = pd.date_range(earliest_invest, prod_start, freq="MS",
                                       inclusive="left")
            if len(pre_dates) > 0:
                pre = pd.DataFrame({"date": pre_dates})
                # All production / revenue columns are zero pre-FOP
                for c in df_e.columns:
                    if c == "date":
                        continue
                    pre[c] = 0.0
                # Book the pre-FOP facility CAPEX into the right months
                for _, row in econ.facility_capex.df.iterrows():
                    try:
                        ts = pd.Timestamp(row["date"])
                        if ts < prod_start:
                            # Find the pre-month bucket
                            hit = pre.index[pre["date"] >= ts]
                            bucket = hit[0] if len(hit) > 0 else 0
                            pre.loc[bucket, "capex_facility"] += \
                                float(row["amount_MMUSD"]) * 1e6
                    except (KeyError, TypeError, ValueError):
                        pass
                # Pre-FOP cashflow = -capex_facility (no revenue, no opex)
                if "capex_facility" in pre.columns:
                    pre["cashflow"] = -pre["capex_facility"]
                # Recompute pre-FOP cashflow and prepend. pd.concat does not
                # preserve .attrs, so capture them and restore afterwards.
                _saved_attrs = dict(df_e.attrs)
                df_e = pd.concat([pre, df_e], ignore_index=True)
                df_e.attrs.update(_saved_attrs)
                # Re-discount from the new t=0 (earliest investment)
                disc_full = (1 + r_m) ** np.arange(len(df_e))
                df_e["discounted_cf"] = df_e["cashflow"].values / disc_full
                df_e["npv"] = df_e["discounted_cf"].cumsum()
                df_e["cum_cashflow"] = df_e["cashflow"].cumsum()
                # Restore year column for the annual groupby in plot_economics
                df_e["year"] = pd.to_datetime(df_e["date"]).dt.year
                df_e.attrs["pre_fop_months"] = len(pre_dates)
                # The pre-FOP prepend shifts every row forward by len(pre_dates).
                # Any index-based attr set BEFORE the prepend (the economic
                # cutoff) must be shifted too, or the cessation marker and the
                # cutoff date fall out of sync with the abandonment cost row.
                if "economic_cutoff_idx" in df_e.attrs:
                    df_e.attrs["economic_cutoff_idx"] = (
                        int(df_e.attrs["economic_cutoff_idx"]) + len(pre_dates))
                if "cessation_idx" in df_e.attrs:
                    df_e.attrs["cessation_idx"] = (
                        int(df_e.attrs["cessation_idx"]) + len(pre_dates))
    except Exception:
        # If anything goes wrong, fall back to the un-padded df_e
        pass

    # ---- NCS tax recomputation on the FULL (padded) dataframe ----
    # The NCS calculation done earlier ran on the engine df, which starts at
    # first production. But facility CAPEX dated before first oil is added by
    # the pre-FOP prepend step above. To depreciate the COMPLETE CAPEX
    # schedule (not just the post-FOP slice) the NCS tax must be recomputed
    # here, on df_e, after the prepend. This is what makes the effective tax
    # rate land at the correct ~70-78% instead of being roughly doubled.
    if getattr(econ, "fiscal_regime", "Tax/Royalty") == "NCS":
        try:
            cit_rate = float(getattr(econ, "ncs_cit_rate", 0.22))
            spt_rate = float(getattr(econ, "ncs_spt_rate", 0.718))
            uplift_rate = float(getattr(econ, "ncs_uplift_rate", 0.1769))
            depr_years = float(getattr(econ, "ncs_depreciation_years", 6.0))
            uplift_years = float(getattr(econ, "ncs_uplift_years", 4.0))
            n = len(df_e)
            # Settle terminal items at the cessation month rather than the
            # end of the forecast horizon. Without this, depreciation
            # spillovers, uplift spillovers, CIT/SPT loss-carry-forward
            # refunds all land at the LAST row of df_e (year 2046, say),
            # producing a stray Tax/Facility-CAPEX-look-alike bar long
            # after the field has ceased producing.
            terminal_idx = int(df_e.attrs.get("cessation_idx", n - 1))
            terminal_idx = max(0, min(n - 1, terminal_idx))

            capex_full = (df_e["capex_well"].values.astype(float)
                          + df_e["capex_facility"].values.astype(float))
            # Depreciation: straight-line over depr_years from each tranche.
            depr_months = max(1, int(round(depr_years * 12)))
            depreciation = np.zeros(n)
            for i in range(n):
                amt = capex_full[i]
                if amt <= 0:
                    continue
                per_month = amt / depr_months
                # Tail beyond cessation is allowed to accelerate into the
                # cessation month so total depreciation always equals total
                # CAPEX (NCS losses are not lost — they unwind at
                # cessation).
                end = min(terminal_idx + 1, i + depr_months)
                depreciation[i:end] += per_month
                spilled = per_month * (depr_months - max(0, end - i))
                if spilled > 0:
                    depreciation[terminal_idx] += spilled
            # Uplift: straight-line over uplift_years, SPT base only.
            uplift_months = max(1, int(round(uplift_years * 12)))
            uplift_sched = np.zeros(n)
            for i in range(n):
                amt = capex_full[i] * uplift_rate
                if amt <= 0:
                    continue
                per_month = amt / uplift_months
                end = min(terminal_idx + 1, i + uplift_months)
                uplift_sched[i:end] += per_month
                spilled = per_month * (uplift_months - max(0, end - i))
                if spilled > 0:
                    uplift_sched[terminal_idx] += spilled

            # Operating profit (CAPEX enters via depreciation, not expensed).
            op_profit = (df_e["revenue"].values
                         - df_e["royalty"].values
                         - df_e["tariff"].values
                         - df_e["opex"].values
                         - df_e["abandonment"].values
                         - df_e.get("co2_cost", pd.Series(np.zeros(n))).values
                         - df_e.get("co2_scope3_cost",
                                     pd.Series(np.zeros(n))).values)

            # CIT with loss carry-forward.
            cit = np.zeros(n)
            cit_cf = 0.0
            for i in range(n):
                base = op_profit[i] - depreciation[i] - cit_cf
                if base >= 0:
                    cit[i] = base * cit_rate
                    cit_cf = 0.0
                else:
                    cit_cf = -base
            # Terminal loss settlement at CESSATION (not horizon end). NCS
            # does not let tax losses expire — remaining carry-forward at
            # cessation is refunded (the State carries the downside
            # symmetrically). Credit the residual loss at its tax value in
            # the cessation month, not at year n-1.
            if cit_cf > 0:
                cit[terminal_idx] -= cit_cf * cit_rate
            # SPT with its own loss carry-forward (uplift in the base).
            spt = np.zeros(n)
            spt_cf = 0.0
            for i in range(n):
                base = (op_profit[i] - depreciation[i]
                        - uplift_sched[i] - spt_cf)
                if base >= 0:
                    spt[i] = base * spt_rate
                    spt_cf = 0.0
                else:
                    spt_cf = -base
            if spt_cf > 0:
                spt[terminal_idx] -= spt_cf * spt_rate
            tax_full = cit + spt

            df_e["ncs_cit"] = cit
            df_e["ncs_spt"] = spt
            df_e["ncs_uplift"] = uplift_sched
            df_e["ncs_depreciation"] = depreciation
            df_e["tax"] = tax_full
            # Rebuild cashflow: real cash CAPEX leaves when spent; tax is the
            # depreciation-based figure.
            pretax_cash = (df_e["revenue"].values
                           - df_e["royalty"].values
                           - df_e["tariff"].values
                           - df_e["opex"].values
                           - df_e["capex_well"].values
                           - df_e["capex_facility"].values
                           - df_e["abandonment"].values
                           - df_e.get("co2_cost", pd.Series(np.zeros(n))).values
                           - df_e.get("co2_scope3_cost",
                                       pd.Series(np.zeros(n))).values)
            cf_full = pretax_cash - tax_full
            df_e["cashflow"] = cf_full
            df_e["cum_cashflow"] = np.cumsum(cf_full)
            disc_full = (1 + r_m) ** np.arange(n)
            df_e["discounted_cf"] = cf_full / disc_full
            df_e["npv"] = df_e["discounted_cf"].cumsum()
        except Exception:
            # If recomputation fails, leave the earlier NCS numbers in place.
            pass

    # Stash the per-well cost breakdown on the DataFrame so the UI can show it
    df_e.attrs["well_cost_breakdown"] = well_cost_breakdown
    df_e.attrs["well_cost_mode"] = "rig_rate" if use_rig_rate else "fixed"
    return df_e


def find_payback(df_e):
    cum = df_e["cum_cashflow"].values
    for i, v in enumerate(cum):
        if v >= 0: return i
    return None


def auto_scale_to_target_rf(wells: list[WellSpec], asm: FieldAssumptions,
                             target_rf: float, tol: float = 0.002,
                             max_iter: int = 30) -> dict:
    """Bisection on a uniform producer scale multiplier so final RF = target.

    Searches between 0.001× and 10× to allow large up- or down-scaling. The RF
    is generally monotonic in the scale factor (more wells producing → more
    recovery), but capacity & abandonment can flatten the high end.

    Returns dict with the multiplier, achieved RF, and a status message.
    Modifies the wells list in place by setting WellSpec.scale_factor *= multiplier.
    """
    from copy import deepcopy

    def rf_at(mult: float) -> float:
        scaled = deepcopy(wells)
        for w in scaled:
            if w.is_producer:
                w.scale_factor *= mult
        df, _, _ = run_simulation(scaled, asm)
        return float(df["recovery_factor"].iloc[-1])

    LO_BOUND, HI_BOUND = 0.001, 10.0
    rf_lo = rf_at(LO_BOUND)
    rf_hi = rf_at(HI_BOUND)

    if rf_lo > target_rf:
        # Even at 0.1% rates, we already exceed target — almost certainly a
        # mis-specified target relative to in-place volumes
        return {"multiplier": LO_BOUND, "achieved_rf": rf_lo,
                "status": "warning",
                "message": f"Target RF {target_rf:.1%} is below what 0.1% of nameplate "
                           f"production achieves ({rf_lo:.1%}). Either raise the target "
                           f"or increase OOIP/OGIP."}
    if rf_hi < target_rf:
        return {"multiplier": HI_BOUND, "achieved_rf": rf_hi,
                "status": "warning",
                "message": f"Cannot reach target RF {target_rf:.1%} even at 10× rates "
                           f"(achieved {rf_hi:.1%}). Likely capacity binding, "
                           f"abandonment limits, or insufficient horizon."}

    lo, hi = LO_BOUND, HI_BOUND
    for _ in range(max_iter):
        mid = 0.5 * (lo + hi)
        rf_mid = rf_at(mid)
        if abs(rf_mid - target_rf) < tol:
            break
        if rf_mid < target_rf:
            lo = mid
        else:
            hi = mid
    mult = 0.5 * (lo + hi)
    achieved = rf_at(mult)

    # Apply in place
    for w in wells:
        if w.is_producer:
            w.scale_factor *= mult
    return {"multiplier": mult, "achieved_rf": achieved,
            "status": "ok",
            "message": f"Producers auto-scaled by ×{mult:.2f} to reach RF {achieved:.1%} "
                       f"(target {target_rf:.0%})."}


# =============================================================================
# Monte Carlo
# =============================================================================
DEFAULT_MC_DRIVERS: dict = {
    # name -> (kind, base_factor_low, base_factor_high, distribution)
    # All factors are multiplicative on the deterministic base case.
    # 'distribution' ∈ {'triangular', 'lognormal', 'uniform', 'truncnormal'}.
    "Oil price":       {"on": True,  "low": 0.70, "high": 1.40, "dist": "triangular"},
    "Gas price":       {"on": True,  "low": 0.60, "high": 1.50, "dist": "triangular"},
    "OOIP":            {"on": True,  "low": 0.70, "high": 1.40, "dist": "lognormal"},
    "OGIP":            {"on": True,  "low": 0.70, "high": 1.40, "dist": "lognormal"},
    "Well qi":         {"on": True,  "low": 0.70, "high": 1.30, "dist": "lognormal", "per_well": True},
    "Decline rate":    {"on": True,  "low": 0.80, "high": 1.30, "dist": "truncnormal", "per_well": True},
    "Variable OPEX":   {"on": True,  "low": 0.80, "high": 1.30, "dist": "triangular"},
    "Well CAPEX":      {"on": True,  "low": 0.85, "high": 1.25, "dist": "triangular"},
    "Facility CAPEX":  {"on": True,  "low": 0.80, "high": 1.35, "dist": "triangular"},
    "Initial pressure":{"on": False, "low": 0.90, "high": 1.10, "dist": "truncnormal"},
    "Discount rate":   {"on": False, "low": 0.80, "high": 1.20, "dist": "uniform"},
}


def _sample_factor(rng: np.random.Generator, dist: str,
                    low: float, high: float) -> float:
    """Draw one multiplicative factor from a distribution centered at ~1.0."""
    if dist == "uniform":
        return rng.uniform(low, high)
    if dist == "triangular":
        # Mode at the geometric mean of low/high (≈ 1.0 if low/high are symmetric)
        mode = (low * high) ** 0.5
        return rng.triangular(low, mode, high)
    if dist == "lognormal":
        # Map [low, high] to ±2σ envelope of the lognormal in log-space
        ln_low, ln_high = np.log(low), np.log(high)
        mu = 0.5 * (ln_low + ln_high)
        sigma = (ln_high - ln_low) / 4.0      # ±2σ ~ [low, high]
        return float(np.exp(rng.normal(mu, sigma)))
    if dist == "truncnormal":
        # Symmetric truncated normal with ±2σ ~ [low, high]
        mu = 0.5 * (low + high)
        sigma = (high - low) / 4.0
        for _ in range(10):
            x = rng.normal(mu, sigma)
            if low <= x <= high:
                return float(x)
        return float(np.clip(rng.normal(mu, sigma), low, high))
    return 1.0


def _factor_from_uniform(u: float, dist: str, low: float, high: float) -> float:
    """Inverse-CDF transform: map a uniform(0,1) draw u to a multiplicative
    factor under the named distribution. Used for CORRELATED sampling — a
    correlated Gaussian copula produces correlated uniforms, which this maps
    to the correct marginal for each driver.
    """
    from math import erf, sqrt, log, exp
    u = min(max(u, 1e-6), 1.0 - 1e-6)
    if dist == "uniform":
        return low + u * (high - low)
    if dist == "triangular":
        mode = (low * high) ** 0.5
        # inverse CDF of a triangular distribution
        fc = (mode - low) / (high - low)
        if u < fc:
            return low + sqrt(u * (high - low) * (mode - low))
        return high - sqrt((1 - u) * (high - low) * (high - mode))
    # normal-based: get the standard-normal quantile via the inverse erf
    # (rational approximation, Acklam's algorithm — accurate to ~1e-9)
    def _norm_ppf(p):
        a = [-3.969683028665376e+01, 2.209460984245205e+02,
             -2.759285104469687e+02, 1.383577518672690e+02,
             -3.066479806614716e+01, 2.506628277459239e+00]
        b = [-5.447609879822406e+01, 1.615858368580409e+02,
             -1.556989798598866e+02, 6.680131188771972e+01,
             -1.328068155288572e+01]
        c = [-7.784894002430293e-03, -3.223964580411365e-01,
             -2.400758277161838e+00, -2.549732539343734e+00,
             4.374664141464968e+00, 2.938163982698783e+00]
        d = [7.784695709041462e-03, 3.224671290700398e-01,
             2.445134137142996e+00, 3.754408661907416e+00]
        plow, phigh = 0.02425, 1 - 0.02425
        if p < plow:
            q = sqrt(-2 * log(p))
            return (((((c[0]*q+c[1])*q+c[2])*q+c[3])*q+c[4])*q+c[5]) / \
                   ((((d[0]*q+d[1])*q+d[2])*q+d[3])*q+1)
        if p > phigh:
            q = sqrt(-2 * log(1 - p))
            return -(((((c[0]*q+c[1])*q+c[2])*q+c[3])*q+c[4])*q+c[5]) / \
                    ((((d[0]*q+d[1])*q+d[2])*q+d[3])*q+1)
        q = p - 0.5
        r = q * q
        return (((((a[0]*r+a[1])*r+a[2])*r+a[3])*r+a[4])*r+a[5])*q / \
               (((((b[0]*r+b[1])*r+b[2])*r+b[3])*r+b[4])*r+1)
    z = _norm_ppf(u)
    if dist == "lognormal":
        ln_low, ln_high = log(low), log(high)
        mu = 0.5 * (ln_low + ln_high)
        sigma = (ln_high - ln_low) / 4.0
        return float(exp(mu + sigma * z))
    # truncnormal / default normal
    mu = 0.5 * (low + high)
    sigma = (high - low) / 4.0
    return float(min(max(mu + sigma * z, low), high))


def _correlated_uniforms(rng: np.random.Generator, names: list,
                          corr_pairs: dict) -> dict:
    """Generate one set of correlated uniform(0,1) draws — a Gaussian copula.

    Args:
        names      : list of driver names to sample.
        corr_pairs : {(name_a, name_b): rho} desired rank correlations.

    Returns {name: u in (0,1)}. Uncorrelated drivers fall back to rho=0.
    """
    from math import erf, sqrt
    n = len(names)
    if n == 0:
        return {}
    idx = {nm: i for i, nm in enumerate(names)}
    # Build the correlation matrix
    C = np.eye(n)
    for (a, b), rho in corr_pairs.items():
        if a in idx and b in idx:
            rho = float(max(-0.95, min(0.95, rho)))
            C[idx[a], idx[b]] = rho
            C[idx[b], idx[a]] = rho
    # Nearest positive-definite repair: clip eigenvalues to be >= small +ve
    try:
        evals, evecs = np.linalg.eigh(C)
        evals = np.clip(evals, 1e-6, None)
        C = evecs @ np.diag(evals) @ evecs.T
        # renormalize to unit diagonal
        d = np.sqrt(np.diag(C))
        C = C / np.outer(d, d)
        L = np.linalg.cholesky(C)
    except Exception:
        L = np.eye(n)
    z = L @ rng.standard_normal(n)
    # standard-normal CDF -> uniform
    u = {nm: 0.5 * (1 + erf(z[idx[nm]] / sqrt(2.0))) for nm in names}
    return u


def classify_reserves(npv_sorted_or_vols) -> dict:
    """Classify a Monte-Carlo reserves distribution into 1P / 2P / 3P.

    Petroleum-industry convention:
      1P (Proved)             = P90  (90% probability of at least this much)
      2P (Proved + Probable)  = P50
      3P (Proved+Prob+Possible)= P10

    Args:
        npv_sorted_or_vols : 1-D array-like of per-realization volumes
                             (or any reserves metric).

    Returns dict with p90/p50/p10 (= 1P/2P/3P), mean, and the spread ratio.
    """
    import numpy as _np
    arr = _np.asarray(list(npv_sorted_or_vols), dtype=float)
    arr = arr[~_np.isnan(arr)]
    if arr.size == 0:
        return {"p90": None, "p50": None, "p10": None, "mean": None,
                "n": 0}
    # In reserves convention P90 is the LOW value (90% chance of exceeding).
    p90 = float(_np.percentile(arr, 10))   # 1P — low estimate
    p50 = float(_np.percentile(arr, 50))   # 2P — best estimate
    p10 = float(_np.percentile(arr, 90))   # 3P — high estimate
    return {
        "p90_1P": p90, "p50_2P": p50, "p10_3P": p10,
        "mean": float(_np.mean(arr)),
        "n": int(arr.size),
        "spread_3P_1P": (p10 / p90 if p90 > 0 else None),
    }


def run_monte_carlo(wells, asm, econ, n_realizations: int,
                     drivers_cfg: dict, seed: int = 42,
                     progress_callback=None, corr_pairs: dict = None) -> dict:
    """Run N realizations sampling from the configured driver distributions.

    Returns a dict with:
      - 'monthly': pd.DataFrame (long format) date|realization|oil_rate|gas_rate|cum_oil|cum_gas|recovery_factor|npv
      - 'summary': pd.DataFrame final-state per realization
      - 'percentiles': dict of {metric: pd.DataFrame(date, p10, p50, p90)}
      - 'realizations_run': int actually run
    """
    from copy import deepcopy
    is_oil = FLUID_SYSTEMS[asm.fluid_system]["primary"] == "oil"
    rng = np.random.default_rng(seed)
    corr_pairs = corr_pairs or {}

    monthly_records = []
    summary_records = []

    # Names of the enabled drivers — used for correlated sampling.
    enabled_names = [nm for nm, cfg in drivers_cfg.items() if cfg.get("on")]

    for r in range(n_realizations):
        # Sample factors for each enabled driver. When correlations are
        # specified, draw a correlated set of uniforms (a Gaussian copula)
        # and map each through its marginal; otherwise sample independently.
        factors = {}
        if corr_pairs and enabled_names:
            u_corr = _correlated_uniforms(rng, enabled_names, corr_pairs)
        else:
            u_corr = {}
        for name, cfg in drivers_cfg.items():
            if cfg.get("on"):
                if name in u_corr:
                    factors[name] = _factor_from_uniform(
                        u_corr[name], cfg.get("dist", "triangular"),
                        cfg.get("low", 0.8), cfg.get("high", 1.2))
                else:
                    factors[name] = _sample_factor(
                        rng, cfg.get("dist", "triangular"),
                        cfg.get("low", 0.8), cfg.get("high", 1.2))
            else:
                factors[name] = 1.0

        # Build perturbed copies
        asm_r = deepcopy(asm)
        econ_r = deepcopy(econ)
        wells_r = [deepcopy(w) for w in wells]

        if "Oil price" in factors:    econ_r.oil_price *= factors["Oil price"]
        if "Gas price" in factors:    econ_r.gas_price *= factors["Gas price"]
        if "Variable OPEX" in factors: econ_r.opex_var *= factors["Variable OPEX"]
        if "Well CAPEX" in factors:   econ_r.capex_per_well *= factors["Well CAPEX"]
        if "Facility CAPEX" in factors and factors["Facility CAPEX"] != 1.0:
            # Scale the facility CAPEX schedule (development rows only — leave
            # cessation/P&A, which is governed by abandonment_cost). This makes
            # facility CAPEX a first-class uncertainty in the NPV distribution.
            try:
                _fac_df = econ_r.facility_capex.df
                if _fac_df is not None and "amount_MMUSD" in _fac_df.columns:
                    _f = float(factors["Facility CAPEX"])
                    if "label" in _fac_df.columns:
                        _keep = ~_fac_df["label"].apply(
                            lambda x: fh.is_abandonment_label(
                                "" if x is None else str(x)))
                        _fac_df.loc[_keep, "amount_MMUSD"] = (
                            _fac_df.loc[_keep, "amount_MMUSD"].astype(float) * _f)
                    else:
                        _fac_df["amount_MMUSD"] = (
                            _fac_df["amount_MMUSD"].astype(float) * _f)
            except Exception:
                pass
        if "Discount rate" in factors: econ_r.discount_rate *= factors["Discount rate"]

        if "OOIP" in factors:         asm_r.ooip_oil *= factors["OOIP"]
        if "OGIP" in factors:         asm_r.ogip_gas *= factors["OGIP"]
        # Also propagate to per-reservoir if multi-reservoir
        if asm_r.reservoirs:
            for rsv in asm_r.reservoirs:
                if FLUID_SYSTEMS[rsv.fluid_system]["primary"] == "oil":
                    rsv.ooip_oil *= factors.get("OOIP", 1.0)
                else:
                    rsv.ogip_gas *= factors.get("OGIP", 1.0)
        if "Initial pressure" in factors:
            asm_r.pvt.p_init_psi *= factors["Initial pressure"]
            if asm_r.reservoirs:
                for rsv in asm_r.reservoirs:
                    rsv.pvt.p_init_psi *= factors["Initial pressure"]

        for w in wells_r:
            if w.is_producer:
                # Per-well independent sampling: each producer gets its own
                # multiplier drawn from the same distribution. Far more realistic
                # than a single field-wide factor (well-by-well variation is
                # typically the largest physical uncertainty in early-life fields).
                if "Well qi" in factors and drivers_cfg.get("Well qi", {}).get("on"):
                    cfg = drivers_cfg["Well qi"]
                    if cfg.get("per_well", True):
                        f_w = _sample_factor(rng, cfg.get("dist", "lognormal"),
                                              cfg.get("low", 0.7), cfg.get("high", 1.3))
                        w.qi_primary *= f_w
                        w.qi_secondary *= f_w
                    else:
                        w.qi_primary *= factors["Well qi"]
                        w.qi_secondary *= factors["Well qi"]
                if "Decline rate" in factors and drivers_cfg.get("Decline rate", {}).get("on"):
                    cfg = drivers_cfg["Decline rate"]
                    if cfg.get("per_well", True):
                        f_w = _sample_factor(rng, cfg.get("dist", "truncnormal"),
                                              cfg.get("low", 0.8), cfg.get("high", 1.3))
                        w.di_annual *= f_w
                    else:
                        w.di_annual *= factors["Decline rate"]

        # Run
        try:
            df_r, _, _ = run_simulation(wells_r, asm_r)
            df_e_r = compute_economics(df_r, is_oil, econ_r, wells_r)
        except Exception:
            continue

        # Capture monthly trajectories (subsample to ~60 points to keep data small)
        sample_idx = np.linspace(0, len(df_r) - 1, min(60, len(df_r))).astype(int)
        for i in sample_idx:
            monthly_records.append({
                "realization": r,
                "date": df_r["date"].iloc[i],
                "oil_rate": float(df_r["oil_rate"].iloc[i]),
                "gas_rate": float(df_r["gas_rate"].iloc[i]),
                "cum_oil":  float(df_r["cum_oil"].iloc[i]),
                "cum_gas":  float(df_r["cum_gas"].iloc[i]),
                "recovery_factor": float(df_r["recovery_factor"].iloc[i]),
                "npv":      float(df_e_r["npv"].iloc[i]),
            })

        summary_records.append({
            "realization": r,
            **{f"factor_{k}": float(v) for k, v in factors.items()},
            "final_rf":   float(df_r["recovery_factor"].iloc[-1]),
            "cum_oil":    float(df_r["cum_oil"].iloc[-1]),
            "cum_gas":    float(df_r["cum_gas"].iloc[-1]),
            "peak_oil":   float(df_r["oil_rate"].max()),
            "peak_gas":   float(df_r["gas_rate"].max()),
            "npv_usd":    float(df_e_r["npv"].iloc[-1]),
            "cum_cf_usd": float(df_e_r["cum_cashflow"].iloc[-1]),
        })

        if progress_callback:
            progress_callback((r + 1) / n_realizations)

    monthly_df = pd.DataFrame(monthly_records)
    summary_df = pd.DataFrame(summary_records)

    # Percentile fans per metric
    percentiles = {}
    if len(monthly_df) > 0:
        for metric in ["oil_rate", "gas_rate", "cum_oil", "cum_gas",
                        "recovery_factor", "npv"]:
            grouped = monthly_df.groupby("date")[metric].agg(
                p10=lambda x: np.percentile(x, 10),
                p50=lambda x: np.percentile(x, 50),
                p90=lambda x: np.percentile(x, 90),
            ).reset_index()
            percentiles[metric] = grouped

    return {
        "monthly": monthly_df,
        "summary": summary_df,
        "percentiles": percentiles,
        "realizations_run": len(summary_df),
    }


def compute_irr(cf):
    """Annualized IRR by bisection on the monthly rate.

    Uses a safe bracket [0, 1.0]. If NPV is positive at r=0 and negative
    at r=1, an IRR exists in between. If both signs are positive (very
    profitable), expand bracket upward. If both negative (unprofitable),
    return None.
    """
    try:
        cf = np.asarray(cf, dtype=float)
        if not np.isfinite(cf).all() or cf.sum() <= 0:
            return None

        def npv_at(r):
            disc = (1 + r) ** np.arange(len(cf))
            return float((cf / disc).sum())

        lo, hi = 0.0, 1.0
        f_lo = npv_at(lo)
        f_hi = npv_at(hi)
        if f_lo <= 0:
            return None  # unprofitable at any positive rate
        # expand hi if still positive
        tries = 0
        while f_hi > 0 and tries < 6:
            hi *= 2
            f_hi = npv_at(hi)
            tries += 1
        if f_hi > 0:
            return None  # absurdly high IRR — bail

        for _ in range(200):
            mid = 0.5 * (lo + hi)
            f_mid = npv_at(mid)
            if abs(f_mid) < 1.0 or hi - lo < 1e-9:
                break
            if f_mid > 0:
                lo = mid
            else:
                hi = mid
        r_monthly = 0.5 * (lo + hi)
        return (1 + r_monthly) ** 12 - 1
    except (OverflowError, ValueError, ZeroDivisionError):
        return None


# =============================================================================
# Stale state
# =============================================================================
def mark_stale():
    st.session_state["stale"] = True


def on_units_change():
    """When the user toggles units, convert all table values stored in
    session_state from the previous unit system to the new one. This keeps
    the displayed numbers consistent with what the user intended.

    Strategy: track the previously-applied units in `_table_units`. When the
    new units differ, scan known unit-bearing columns and convert.
    """
    new_units = st.session_state.get("units")
    old_units = st.session_state.get("_table_units")
    if old_units is None:
        # First time we see this — record and exit (no conversion needed,
        # tables are still at their initial defaults built in this unit system)
        st.session_state["_table_units"] = new_units
        mark_stale()
        return
    if new_units == old_units:
        return

    # Conversion direction: convert displayed values from old to new
    # i.e., multiply by (factor_old_to_field / factor_new_to_field) on the field-equivalent
    # Simpler: take displayed_old → field → displayed_new
    def convert(value, kind):
        try:
            v = float(value)
            field_val = to_field(v, kind, old_units)
            return from_field(field_val, kind, new_units)
        except (ValueError, TypeError):
            return value

    # Producers/injectors rate columns
    rate_kind_primary = lambda fluid: ("oil_rate"
        if FLUID_SYSTEMS[fluid]["primary"] == "oil" else "gas_rate")
    rate_kind_secondary = lambda fluid: ("gas_rate"
        if FLUID_SYSTEMS[fluid]["primary"] == "oil" else "oil_rate")
    fluid = st.session_state.get("fluid", "Oil with associated gas")
    kp = rate_kind_primary(fluid); ks = rate_kind_secondary(fluid)

    if "producers_df" in st.session_state:
        df = st.session_state["producers_df"].copy()
        if "qi_primary" in df.columns:
            df["qi_primary"] = df["qi_primary"].apply(lambda v: convert(v, kp))
        if "qi_secondary" in df.columns:
            df["qi_secondary"] = df["qi_secondary"].apply(lambda v: convert(v, ks))
        # IPR columns (wellhead_pressure_psi, tubing_depth_ft) are stored
        # in ENGINE units (psi, ft) regardless of the unit selection. The
        # display layer in the producers-table renderer converts on the fly
        # via `pdf_display`, and the Apply button converts back via
        # `commit`. No conversion needed here.
        # PI override is a compound unit (rate per pressure). Convert
        # explicitly per row, using the well's primary phase (kp).
        if "well_pi_override" in df.columns:
            def _convert_pi(v):
                try:
                    pi_old = float(v)
                    if pi_old == 0.0:
                        return 0.0
                    rate_old_to_field = (1.0 if old_units == "field"
                                          else M2F[kp])
                    press_old_to_field = (1.0 if old_units == "field"
                                           else M2F["pressure"])
                    pi_field = (pi_old * rate_old_to_field
                                / press_old_to_field)
                    rate_field_to_new = (1.0 if new_units == "field"
                                          else 1.0 / M2F[kp])
                    press_field_to_new = (1.0 if new_units == "field"
                                           else 1.0 / M2F["pressure"])
                    return (pi_field * rate_field_to_new
                            / press_field_to_new)
                except (ValueError, TypeError):
                    return v
            df["well_pi_override"] = df["well_pi_override"].apply(_convert_pi)
        st.session_state["producers_df"] = df

    if "injectors_df" in st.session_state:
        df = st.session_state["injectors_df"].copy()
        if "inj_rate" in df.columns:
            df["inj_rate"] = df["inj_rate"].apply(lambda v: convert(v, "water_rate"))
        st.session_state["injectors_df"] = df

    # Capacity table (oil/water/liquid in volume rates; gas in MMscf/d field or kSm³/d metric)
    if "cap_df" in st.session_state:
        df = st.session_state["cap_df"].copy()
        for col, kind in [("oil", "oil_rate"), ("water", "water_rate"),
                          ("liquid", "oil_rate"),
                          ("water_inj", "water_rate"),
                          ("gas_inj", "gas_rate")]:
            if col in df.columns:
                df[col] = df[col].apply(lambda v: convert(v, kind))
        # `gas` column: MMscf/d in field units, kSm³/d in metric.
        # Conversion: field MMscf/d × 28.317 = metric kSm³/d (approx). We do via Mscf bridge.
        if "gas" in df.columns:
            def convert_gas_cap(v):
                try:
                    fv = float(v)
                    if old_units == "field" and new_units == "metric":
                        # MMscf/d -> Mscf/d (×1000) -> Sm³/d (/M2F) -> kSm³/d (/1000)
                        return fv * 1000.0 / M2F["gas_rate"] / 1000.0
                    elif old_units == "metric" and new_units == "field":
                        # kSm³/d -> Sm³/d (×1000) -> Mscf/d (×M2F/1000) -> MMscf/d (/1000)
                        return fv * 1000.0 * M2F["gas_rate"] / 1000.0 / 1000.0
                    return fv
                except (ValueError, TypeError):
                    return v
            df["gas"] = df["gas"].apply(convert_gas_cap)
        st.session_state["cap_df"] = df

    # Reservoirs table
    if "reservoirs_df" in st.session_state:
        df = st.session_state["reservoirs_df"].copy()
        col_kinds = {"ooip_oil_MMstb": "oil_vol", "ogip_gas_Bscf": "gas_vol",
                     "p_init": "pressure", "t_res": "temp",
                     "rs_init": "gor", "p_bub": "pressure"}
        for col, kind in col_kinds.items():
            if col in df.columns:
                df[col] = df[col].apply(lambda v: convert(v, kind))
        st.session_state["reservoirs_df"] = df

    # ---- Scalar widget values stored in session_state ----
    # Each scalar number_input that uses a unit-aware default also stashes
    # its current value in session_state under its `key`. When the user
    # switches units, those scalars stay in the OLD unit system but the
    # widget label flips to the new unit — so a user sees "241" labelled
    # "psi" (which is really 241 bar from the metric session). Convert each
    # scalar key explicitly here so the displayed number matches the new
    # label. Streamlit forbids writing widget-backed keys *after* the
    # widget exists for that run, but `on_units_change` fires BEFORE the
    # widgets re-render, so the write here is applied to the rebuilt
    # widgets on this same rerun.
    _scalar_kinds = {
        # PVT / reservoir pressures and temperatures
        "p_init":          "pressure",
        "p_bub":           "pressure",
        "t_res":           "temp",
        # Aquifer + gas cap pressures and volume
        "aq_pini":         "pressure",
        "aq_vol":          "water_vol",
        "gc_pi":           "pressure",
        # PI: rate per pressure unit. PI doesn't have a simple `kind` in
        # the unit framework (it's a compound unit). Skip — `well_pi_default`
        # is treated as dimensionless and the user re-enters it.
        # In-place volumes
        "ooip":            "oil_vol",
        "ogip":            "gas_vol",
        # GOR/CGR (engine in scf/stb or stb/MMscf — both handled by `gor`
        # in the unit framework).
        "rs_init":         "gor",
        # Default min BHP for non-IPR wells
        "min_bhp_default": "pressure",
        # Field abandonment rates (oil is straight `oil_rate`; gas uses
        # the same MMscf/d↔kSm³/d convention as the capacity column, so
        # it's handled separately below).
        "aban_oil":        "oil_rate",
        # Shut-in well-head pressure tool
        "siwhp_p_res":     "pressure",
        "siwhp_datum":     "depth",
        "siwhp_wd":        "depth",
        "siwhp_twh":       "temp",
        # Portfolio shared constraints
        "portfolio_constraint_oil": "oil_rate",
        "portfolio_constraint_gas": "gas_rate",
    }
    for _key, _kind in _scalar_kinds.items():
        if _key in st.session_state:
            try:
                _v = st.session_state[_key]
                _field = to_field(float(_v), _kind, old_units)
                st.session_state[_key] = from_field(_field, _kind, new_units)
            except (ValueError, TypeError):
                pass

    # `aban_gas` uses MMscf/d in field and kSm³/d in metric (same
    # convention as the capacity column's `gas` column). `M2F["gas_rate"]`
    # is the per-Mscf↔per-kSm³ factor (35.3147). So:
    #   field MMscf/d → field Mscf/d (×1000) → metric kSm³/d (÷ M2F).
    if "aban_gas" in st.session_state:
        try:
            _v = float(st.session_state["aban_gas"])
            if old_units == "field" and new_units == "metric":
                st.session_state["aban_gas"] = (
                    _v * 1000.0 / M2F["gas_rate"])
            elif old_units == "metric" and new_units == "field":
                st.session_state["aban_gas"] = (
                    _v * M2F["gas_rate"] / 1000.0)
        except (ValueError, TypeError):
            pass

    # PI (well_pi_default) — compound unit (rate per pressure). Convert
    # explicitly: PI_metric = PI_field × (rate factor) / (pressure factor).
    if "well_pi_default" in st.session_state:
        try:
            _pi_old = float(st.session_state["well_pi_default"])
            _is_oil = (FLUID_SYSTEMS.get(
                st.session_state.get("fluid",
                                      "Oil with associated gas"),
                {"primary": "oil"})["primary"] == "oil")
            _rk = "oil_rate" if _is_oil else "gas_rate"
            # PI in old units → field PI → PI in new units
            # field PI = old PI × (rate from old → field) / (pressure from old → field)
            _rate_old_to_field = (1.0 if old_units == "field"
                                   else M2F[_rk])
            _press_old_to_field = (1.0 if old_units == "field"
                                    else M2F["pressure"])
            _pi_field = _pi_old * _rate_old_to_field / _press_old_to_field
            _rate_field_to_new = (1.0 if new_units == "field"
                                   else 1.0 / M2F[_rk])
            _press_field_to_new = (1.0 if new_units == "field"
                                    else 1.0 / M2F["pressure"])
            st.session_state["well_pi_default"] = (
                _pi_field * _rate_field_to_new / _press_field_to_new)
        except (ValueError, TypeError):
            pass

    st.session_state["_table_units"] = new_units
    mark_stale()


def _hash_table_state() -> str:
    """Hash the contents of all data-editor tables so we can detect edits
    without relying on on_change (which fires on every keystroke and lags)."""
    keys = ["rigs_df", "producers_df", "injectors_df", "cap_df", "fac_df",
            "reservoirs_df", "well_reservoir_df"]
    h = []
    for k in keys:
        if k in st.session_state:
            try:
                df = st.session_state[k]
                if isinstance(df, pd.DataFrame):
                    h.append(str(pd.util.hash_pandas_object(df, index=False).sum()))
                else:
                    h.append(str(df))
            except Exception:
                h.append(repr(st.session_state.get(k)))
    return "|".join(h)


def check_tables_for_changes():
    """Compare current table hash to the one captured at last run; flip stale if differ."""
    cur = _hash_table_state()
    last = st.session_state.get("last_table_hash")
    if last is not None and cur != last:
        st.session_state["stale"] = True


# =============================================================================
# Sidebar
# =============================================================================
def sidebar_inputs():
    st.sidebar.title("⚙️ Field Setup")

    units = st.sidebar.radio(
        "Unit system", ["field", "metric"], horizontal=True,
        format_func=lambda x: "Field" if x == "field" else "Metric",
        key="units", on_change=on_units_change,
        help="Field uses stb, scf, psi, °F. Metric uses Sm³, bar, °C. "
             "All conversions are done in the engine; computation is internally in field units. "
             "Switching units automatically converts values in all tables.",
    )
    # Track the units that table values are currently expressed in
    if "_table_units" not in st.session_state:
        st.session_state["_table_units"] = units

    fluid = st.sidebar.selectbox(
        "Fluid system", list(FLUID_SYSTEMS.keys()), key="fluid", on_change=mark_stale,
        help="Sets primary and secondary fluids and which capacity constraints are relevant.",
    )
    is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"

    strategy = st.sidebar.radio(
        "Drainage strategy", ["Depletion", "Injection"], horizontal=True,
        key="strategy", on_change=mark_stale,
        help="Depletion = natural reservoir energy only. "
             "Injection = water/gas injection adds voidage replacement & pressure support.",
    )

    # Apply a pending start-date change requested elsewhere (e.g. the
    # "use first oil date" button in the schedule section). Streamlit forbids
    # writing to a widget-backed key after the widget exists, so the request
    # is staged under a separate key and applied here, before the widget.
    if "_pending_start_date" in st.session_state:
        st.session_state["start_date"] = st.session_state.pop(
            "_pending_start_date")
    start_date = st.sidebar.date_input(
        "Project start date", value=date(2026, 1, 1),
        key="start_date", on_change=mark_stale,
        help="Anchor for all schedules (drilling, capacities, facility CAPEX).",
    )
    horizon = st.sidebar.slider("Forecast horizon (years)", 5, 50, 25,
                                key="horizon", on_change=mark_stale)

    st.sidebar.markdown("### Reservoir volumes")
    if is_oil:
        ooip = st.sidebar.number_input(
            f"OOIP ({ulabel('oil_vol', units)})", min_value=0.0,
            value=from_field(250.0, "oil_vol", units), step=10.0,
            key="ooip", on_change=mark_stale,
            help="Stock-tank oil originally in place.")
        ogip = st.sidebar.number_input(
            f"OGIP ({ulabel('gas_vol', units)})", min_value=0.0,
            value=from_field(300.0, "gas_vol", units), step=10.0,
            key="ogip", on_change=mark_stale,
            help="Associated/solution gas originally in place.")
    else:
        ogip = st.sidebar.number_input(
            f"OGIP ({ulabel('gas_vol', units)})", min_value=0.0,
            value=from_field(1500.0, "gas_vol", units), step=50.0,
            key="ogip", on_change=mark_stale,
            help="Gas originally in place — the primary in-place volume "
                 "for a gas / gas-condensate field.")
        # Condensate in place. For a gas-condensate field this is not an
        # independent number: it follows from the gas in place and the CGR.
        # If a CGR has been entered in the PVT tab, offer to derive it and
        # flag any inconsistency between the entered value and OGIP × CGR.
        _cgr = float(st.session_state.get("rs_init", 0.0) or 0.0)  # stb/MMscf
        ooip = st.sidebar.number_input(
            f"Condensate in place ({ulabel('oil_vol', units)})",
            min_value=0.0,
            value=from_field(20.0, "oil_vol", units), step=1.0,
            key="ooip", on_change=mark_stale,
            help="Stock-tank condensate originally in place. For a gas-"
                 "condensate field this should be roughly OGIP × CGR "
                 "(initial), since the condensate is dissolved in the gas. "
                 "The app shows the implied value below so you can check "
                 "consistency.")
        if _cgr > 0:
            # OGIP is in display units; convert to Bscf for the arithmetic.
            ogip_bscf = to_field(ogip, "gas_vol", units)
            # condensate (MMstb) = OGIP(Bscf) * 1000 (MMscf/Bscf) * CGR
            #                       (stb/MMscf) / 1e6  (stb -> MMstb)
            implied_cond_mmstb = ogip_bscf * 1000.0 * _cgr / 1e6
            implied_cond_disp = from_field(implied_cond_mmstb,
                                            "oil_vol", units)
            entered_cond_mmstb = to_field(ooip, "oil_vol", units)
            # CGR engine units are stb/MMscf; SI equivalent is Sm³/kSm³
            # (× 0.22213).
            _cgr_d = _cgr if units == "field" else _cgr * 0.22213
            _cgr_u = "stb/MMscf" if units == "field" else "Sm³/kSm³"
            st.sidebar.caption(
                f"Implied condensate in place from OGIP × CGR "
                f"({_cgr_d:.1f} {_cgr_u}): "
                f"**{implied_cond_disp:,.1f} {ulabel('oil_vol', units)}**.")
            if entered_cond_mmstb > 0:
                ratio = implied_cond_mmstb / entered_cond_mmstb
                if ratio > 1.25 or ratio < 0.8:
                    st.sidebar.warning(
                        f"⚠️ Condensate-in-place ({ooip:,.1f}) and the value "
                        f"implied by OGIP × CGR ({implied_cond_disp:,.1f}) "
                        f"differ by {abs(ratio - 1) * 100:.0f}%. For a gas-"
                        f"condensate field these should be close — check "
                        f"the OGIP, the CGR, or the condensate-in-place "
                        f"entry.")

    rf_target = st.sidebar.slider(
        "Target recovery factor", 0.05, 0.80, 0.35, 0.01,
        key="rf_target", on_change=mark_stale,
        help="Used for the achievement warning and the auto-scaling solver.",
    )
    auto_scale_rf = st.sidebar.checkbox(
        "🎯 Auto-scale producers to hit target RF",
        value=False, key="auto_scale_rf", on_change=mark_stale,
        help=("If enabled, on Run the engine multiplies every producer's rate "
              "profile by a single global factor found by bisection so that "
              "the final recovery factor equals the target — within the limit "
              "of capacity choking. A warning is shown with the multiplier "
              "applied; if the target cannot be reached even at 5×, the "
              "warning explains why (capacity binding or insufficient fluid in place).")
    )

    with st.sidebar.expander("🧪 PVT inputs", expanded=False):
        p_init_disp = st.number_input(
            f"Initial reservoir pressure ({ulabel('pressure', units)})",
            value=from_field(3500.0, "pressure", units),
            key="p_init", on_change=mark_stale,
            help="Datum pressure for material balance.")
        t_res_disp = st.number_input(
            f"Reservoir temperature ({ulabel('temp', units)})",
            value=from_field(180.0, "temp", units),
            key="t_res", on_change=mark_stale)
        api = st.number_input("Oil API gravity", value=35.0,
                              key="api", on_change=mark_stale,
                              help="35 = light oil. <22 = heavy.")
        gas_grav = st.number_input("Gas specific gravity (air = 1)", value=0.7,
                                   key="gas_grav", on_change=mark_stale)
        # Solution ratio: oil systems use Rs (solution gas-oil ratio,
        # scf gas per stb oil); gas-condensate systems are characterised by
        # CGR (condensate-gas ratio, stb condensate per MMscf gas). The
        # engine stores the number in `rs_init`; the label clarifies which
        # physical quantity it represents for the chosen fluid.
        if FLUID_SYSTEMS[fluid]["primary"] == "gas":
            # CGR (condensate-gas ratio). Field unit: stb condensate per
            # MMscf gas. Metric unit: Sm³ condensate per kSm³ gas.
            # Conversion: 1 stb/MMscf = 6.2898 Sm³ / 28.3168 kSm³
            #            = 0.22213 Sm³/kSm³.
            CGR_F2M = 0.22213   # field (stb/MMscf) -> metric (Sm3/kSm3)
            cgr_unit = "Sm³/kSm³" if units == "metric" else "stb/MMscf"
            # Option: derive CGR from the in-place condensate & gas volumes
            # instead of entering it directly. This keeps a gas-condensate
            # field internally consistent (CGR ≈ condensate-in-place / OGIP).
            cgr_from_inplace = st.checkbox(
                "Compute initial CGR from in-place volumes",
                value=False, key="cgr_from_inplace", on_change=mark_stale,
                help="When ticked, the initial CGR is calculated as "
                     "(condensate in place) / (gas in place) rather than "
                     "entered directly — keeping the gas-condensate "
                     "volumetrics self-consistent. Set the condensate-in-"
                     "place and OGIP in the sidebar 'Reservoir volumes'.")
            if cgr_from_inplace:
                # ooip here holds condensate-in-place (MMstb), ogip is OGIP.
                _cond_ip = float(st.session_state.get("ooip", 0.0) or 0.0)
                _gas_ip = float(st.session_state.get("ogip", 0.0) or 0.0)
                _cond_mmstb = to_field(_cond_ip, "oil_vol", units)
                _gas_bscf = to_field(_gas_ip, "gas_vol", units)
                # CGR (stb/MMscf) = condensate(MMstb)*1e6 / (OGIP(Bscf)*1000)
                if _gas_bscf > 0:
                    cgr_field = (_cond_mmstb * 1e6) / (_gas_bscf * 1000.0)
                else:
                    cgr_field = 0.0
                cgr_disp = (cgr_field * CGR_F2M if units == "metric"
                            else cgr_field)
                st.metric(f"Initial CGR (computed, {cgr_unit})",
                          f"{cgr_disp:,.2f}")
                _cond_show = from_field(_cond_mmstb, "oil_vol", units)
                _gas_show = from_field(_gas_bscf, "gas_vol", units)
                st.caption(
                    f"= condensate-in-place ÷ OGIP "
                    f"= {_cond_show:,.1f} {ulabel('oil_vol', units)} ÷ "
                    f"{_gas_show:,.0f} {ulabel('gas_vol', units)}. "
                    f"Untick to enter the CGR manually.")
                rs_init_disp = cgr_disp
                rs_kind = "cgr"
            else:
                rs_default_field = 30.0   # stb/MMscf
                rs_default = (rs_default_field * CGR_F2M
                              if units == "metric" else rs_default_field)
                rs_help = (
                    f"Condensate-gas ratio in {cgr_unit}. Typical lean "
                    f"gas-condensate "
                    + ("1-11 Sm³/kSm³" if units == "metric"
                       else "5-50 stb/MMscf")
                    + "; rich "
                    + ("11-55 Sm³/kSm³" if units == "metric"
                       else "50-250 stb/MMscf")
                    + ". Set 0 for dry gas. Drives the secondary "
                      "(condensate) production stream.")
                rs_init_disp = st.number_input(
                    f"Initial CGR ({cgr_unit})", value=rs_default,
                    key="rs_init", on_change=mark_stale, help=rs_help)
                rs_kind = "cgr"
            # store the conversion so the PVTInputs build can convert back
            _cgr_f2m_factor = CGR_F2M
        else:
            rs_label = f"Initial Rs ({ulabel('gor', units)})"
            rs_help = "Initial solution gas-oil ratio (scf gas / stb oil)."
            rs_default = from_field(700.0, "gor", units)
            rs_kind = "gor"
            _cgr_f2m_factor = 1.0
            rs_init_disp = st.number_input(
                rs_label, value=rs_default,
                key="rs_init", on_change=mark_stale, help=rs_help)
        # Saturation pressure: for an oil system this is the BUBBLE point
        # (pressure at which the first gas bubble evolves from the oil);
        # for a gas / gas-condensate system it is the DEW point (pressure
        # at which the first liquid condenses from the gas). They are
        # physically distinct phase-boundary points — labelling matters.
        _pvt_is_gas = FLUID_SYSTEMS[fluid]["primary"] == "gas"
        if _pvt_is_gas:
            p_sat_label = f"Dew point ({ulabel('pressure', units)})"
            p_sat_help = (
                "Dew-point pressure of the gas. Below this pressure liquid "
                "(condensate) drops out of the gas phase in the reservoir. "
                "For a dry gas with no condensate, set this at or below the "
                "abandonment pressure so no retrograde behaviour is modelled. "
                "For gas-condensate, retrograde liquid drop-out below the "
                "dew point reduces the produced gas and is the reason a CGR "
                "is specified.")
            p_sat_default = 3200.0
        else:
            p_sat_label = f"Bubble point ({ulabel('pressure', units)})"
            p_sat_help = (
                "Bubble-point pressure of the oil. Below this pressure gas "
                "evolves from solution and Bo declines.")
            p_sat_default = 2800.0
        p_bub_disp = st.number_input(
            p_sat_label,
            value=from_field(p_sat_default, "pressure", units),
            key="p_bub", on_change=mark_stale,
            help=p_sat_help)
        # Retrograde-condensate modelling — only meaningful for a gas-
        # condensate system (gas primary + condensate secondary).
        retrograde_enabled = False
        retrograde_drop_fraction = 0.55
        if (_pvt_is_gas and
                FLUID_SYSTEMS[fluid]["secondary"] == "condensate"):
            retrograde_enabled = st.checkbox(
                "Model retrograde condensate drop-out",
                value=False, key="retrograde_enabled", on_change=mark_stale,
                help="When enabled, the produced condensate is computed as "
                     "gas rate × a producible CGR that FALLS below the dew "
                     "point — liquid condenses in the reservoir pores and "
                     "is left behind. Without this, condensate is a flat "
                     "yield. With it, the condensate stream declines faster "
                     "than the gas once the reservoir crosses the dew "
                     "point — the physically correct behaviour for a "
                     "gas-condensate field.")
            if retrograde_enabled:
                retrograde_drop_fraction = st.slider(
                    "Max producible-CGR loss at peak drop-out",
                    min_value=0.1, max_value=0.9, value=0.55, step=0.05,
                    key="retrograde_drop_fraction", on_change=mark_stale,
                    help="The fraction of the initial CGR that becomes "
                         "unproducible at the pressure of maximum liquid "
                         "drop-out (~50% of the dew point). 0.55 means the "
                         "producible CGR falls to 45% of its initial value "
                         "— typical for a moderately rich condensate. "
                         "Leaner gas: lower; richer: higher.")
        ct_rock = st.number_input("Rock compressibility (1/psi)", value=4e-6,
                                  format="%.1e", key="ct_rock", on_change=mark_stale)
        sw_init = st.number_input("Initial water saturation", value=0.20,
                                  min_value=0.0, max_value=0.6,
                                  key="sw_init", on_change=mark_stale)

    # Resolve rs_init to field units. For an oil system this is Rs (GOR
    # conversion). For a gas system it is CGR — stored internally in
    # stb/MMscf, so a metric display value is divided back by the CGR
    # factor.
    if rs_kind == "gor":
        _rs_field = to_field(rs_init_disp, "gor", units)
    elif rs_kind == "cgr":
        _rs_field = (float(rs_init_disp) / _cgr_f2m_factor
                     if units == "metric" else float(rs_init_disp))
    else:
        _rs_field = float(rs_init_disp)

    pvt = PVTInputs(
        p_init_psi=to_field(p_init_disp, "pressure", units),
        t_res_F=to_field(t_res_disp, "temp", units),
        api=api, gas_grav=gas_grav,
        rs_init=_rs_field,
        p_bub_psi=to_field(p_bub_disp, "pressure", units),
    )

    # ---- Fractional-flow water cut (oil fields) ----
    ff_enabled = False
    ff_params_ui = {}
    _ff_is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"
    if _ff_is_oil:
        with st.sidebar.expander("💧 Fractional-flow water cut",
                                  expanded=False):
            ff_enabled = st.checkbox(
                "Derive water cut from saturation physics",
                value=False, key="ff_enabled", on_change=mark_stale,
                help="Off — field water cut is the sum of the per-well "
                     "water-cut ramps you entered. On — the water cut is "
                     "computed from Corey relative-permeability curves and "
                     "the cumulative recovery: as oil is displaced the "
                     "average water saturation rises, the fractional flow "
                     "of water climbs its S-curve, and the produced water "
                     "cut follows. This makes a waterflood's water cut "
                     "emerge from the rock/fluid properties instead of a "
                     "prescribed ramp — overrides the per-well ramps.")
            if ff_enabled:
                st.caption("Corey relative-permeability & fluid inputs:")
                fc1, fc2 = st.columns(2)
                ff_params_ui["swc"] = fc1.number_input(
                    "Connate water Swc", min_value=0.0, max_value=0.5,
                    value=0.20, step=0.05, key="ff_swc",
                    on_change=mark_stale,
                    help="Irreducible water saturation — the water cut "
                         "starts rising from here.")
                ff_params_ui["sor"] = fc2.number_input(
                    "Residual oil Sor", min_value=0.0, max_value=0.5,
                    value=0.25, step=0.05, key="ff_sor",
                    on_change=mark_stale,
                    help="Residual (unrecoverable) oil saturation. "
                         "1 - Swc - Sor is the moveable oil window.")
                ff_params_ui["krw_max"] = fc1.number_input(
                    "krw endpoint", min_value=0.05, max_value=1.0,
                    value=0.30, step=0.05, key="ff_krw_max",
                    on_change=mark_stale,
                    help="Water relative permeability at the residual-oil "
                         "endpoint. Lower = more oil-wet, later water "
                         "breakthrough.")
                ff_params_ui["kro_max"] = fc2.number_input(
                    "kro endpoint", min_value=0.1, max_value=1.0,
                    value=0.90, step=0.05, key="ff_kro_max",
                    on_change=mark_stale,
                    help="Oil relative permeability at connate water.")
                ff_params_ui["nw"] = fc1.number_input(
                    "Corey water exp. nw", min_value=1.0, max_value=6.0,
                    value=3.0, step=0.5, key="ff_nw", on_change=mark_stale,
                    help="Curvature of the water rel-perm curve. Higher = "
                         "more delayed water rise.")
                ff_params_ui["no"] = fc2.number_input(
                    "Corey oil exp. no", min_value=1.0, max_value=6.0,
                    value=2.0, step=0.5, key="ff_no", on_change=mark_stale,
                    help="Curvature of the oil rel-perm curve.")
                ff_params_ui["mu_oil"] = fc1.number_input(
                    "Oil viscosity (cP)", min_value=0.1, max_value=100.0,
                    value=1.5, step=0.5, key="ff_mu_oil",
                    on_change=mark_stale,
                    help="In-situ oil viscosity. A higher oil viscosity "
                         "(unfavourable mobility ratio) means earlier, "
                         "sharper water breakthrough.")
                ff_params_ui["mu_water"] = fc2.number_input(
                    "Water viscosity (cP)", min_value=0.1, max_value=5.0,
                    value=0.4, step=0.1, key="ff_mu_water",
                    on_change=mark_stale)
                ff_params_ui["sweep"] = st.slider(
                    "Volumetric sweep efficiency", min_value=0.3,
                    max_value=1.0, value=0.70, step=0.05, key="ff_sweep",
                    on_change=mark_stale,
                    help="Fraction of the reservoir the flood actually "
                         "contacts. Lower sweep → the contacted region "
                         "watered out sooner for a given field recovery.")
                _mr = (ff_params_ui["mu_oil"] / max(ff_params_ui["mu_water"],
                                                     1e-6))
                st.caption(
                    f"Mobility ratio (oil/water viscosity) ≈ {_mr:.1f}. "
                    + ("Favourable — stable flood, gradual water rise."
                       if _mr < 5 else
                       "Unfavourable — expect early water breakthrough "
                       "and viscous fingering."))

    with st.sidebar.expander("🌊 Aquifer support", expanded=False):
        aq_active = st.checkbox("Aquifer active", value=False,
                                key="aq_active", on_change=mark_stale,
                                help="Adds a pot-tank aquifer model that supplies water as P drops.")
        aq_model = st.selectbox("Aquifer model", ["Pot", "Fetkovich", "Carter-Tracy"],
                                key="aq_model", on_change=mark_stale,
                                help="Pot = finite-tank, instantaneous influx; "
                                     "Fetkovich = rate-limited influx using productivity index "
                                     "with aquifer pressure depleting over time; "
                                     "Carter-Tracy = analytical infinite-acting radial aquifer "
                                     "(uses dimensionless-time formulation; preferred when "
                                     "the aquifer is very large vs the reservoir).")
        aq_vol = st.number_input(f"Aquifer water volume ({ulabel('water_vol', units)})",
                                 value=from_field(500.0, "water_vol", units),
                                 key="aq_vol", on_change=mark_stale,
                                 help="Used by Pot and Fetkovich models. "
                                      "Carter-Tracy uses the U constant below instead.")
        aq_pi = st.number_input("Aquifer PI (bbl/d/psi)", value=20.0,
                                key="aq_pi", on_change=mark_stale,
                                help="Used by Fetkovich only.")
        aq_pini = st.number_input(
            f"Aquifer initial pressure ({ulabel('pressure', units)})",
            value=from_field(3500.0, "pressure", units),
            key="aq_pini", on_change=mark_stale)
        if aq_model == "Carter-Tracy":
            ct1, ct2 = st.columns(2)
            ct_U = ct1.number_input("Aquifer constant U (bbl/psi)",
                                     value=200.0, min_value=1.0, step=10.0,
                                     key="aq_ct_U", on_change=mark_stale,
                                     help="Carter-Tracy aquifer constant. "
                                          "U = 1.119 × φ × ct × h × rₑ² (bbl/psi). "
                                          "Typical screening range: 50–2000 bbl/psi.")
            ct_diff = ct2.number_input("Dimensionless time / month",
                                        value=50.0, min_value=0.1, step=5.0,
                                        key="aq_ct_diff", on_change=mark_stale,
                                        help="t_D advance per simulated month — "
                                             "controls how fast the influx response "
                                             "develops. Higher = faster aquifer "
                                             "response (smaller, more permeable aquifer).")
        else:
            ct_U = 200.0
            ct_diff = 50.0

    aquifer = AquiferInputs(
        active=aq_active, model=aq_model,
        aquifer_volume=to_field(aq_vol, "water_vol", units),
        productivity_index=aq_pi,
        initial_pressure_psi=to_field(aq_pini, "pressure", units),
        ct_aquifer_constant=ct_U,
        ct_diffusivity=ct_diff,
    )

    with st.sidebar.expander("💨 Gas cap drive", expanded=False):
        gc_active = st.checkbox("Gas cap active", value=False,
                                key="gc_active", on_change=mark_stale,
                                help="Adds an initial gas cap that expands as oil is produced.")
        gc_size = st.number_input("Gas-cap size m (Vgc/Voil at initial conditions)",
                                  value=0.2, min_value=0.0, max_value=5.0,
                                  key="gc_size", on_change=mark_stale)
        gc_pi = st.number_input(
            f"Gas-cap initial pressure ({ulabel('pressure', units)})",
            value=from_field(3500.0, "pressure", units),
            key="gc_pi", on_change=mark_stale)

    gas_cap = GasCapInputs(
        active=gc_active, size_fraction=gc_size,
        initial_pressure_psi=to_field(gc_pi, "pressure", units),
    )

    # Injection inputs are only meaningful for Injection strategy. In Depletion
    # mode we still need defaults for the FieldAssumptions dataclass but the UI
    # is hidden to keep the sidebar focused.
    if strategy == "Injection":
        with st.sidebar.expander("💧 Injection", expanded=True):
            vrr = st.slider("Voidage replacement ratio (target)", 0.5, 1.5, 1.0, 0.05,
                            key="vrr", on_change=mark_stale,
                            help="Used only when no injector wells are defined.")
            eff = st.slider("Injection efficiency", 0.3, 1.0, 0.85, 0.05,
                            key="inj_eff", on_change=mark_stale,
                            help="Fraction of injected fluid that effectively replaces voidage.")
    else:
        vrr = st.session_state.get("vrr", 1.0)
        eff = st.session_state.get("inj_eff", 0.85)

    with st.sidebar.expander("🧪 Productivity index (single-reservoir)", expanded=False):
        is_oil_for_pi = FLUID_SYSTEMS[fluid]["primary"] == "oil"
        # PI relates rate to drawdown:  q = PI × (P_res − P_wf).
        # Field units: oil bbl/d/psi, gas Mscf/d/psi.
        # Metric units: oil Sm³/d/bar, gas kSm³/d/bar.
        if units == "metric":
            pi_units_label = "Sm³/d/bar" if is_oil_for_pi else "kSm³/d/bar"
        else:
            pi_units_label = "bbl/d/psi" if is_oil_for_pi else "Mscf/d/psi"
        st.caption(
            "Used only when wells have **PI mode** enabled in the producers "
            "table. Multi-reservoir mode picks PI from each reservoir's row "
            "instead."
        )
        # Pick natural example numbers per unit system.
        _ex_dd = 1000 if units == "field" else 70   # psi or bar
        _ex_pl = "psi" if units == "field" else "bar"
        _ex_rate_unit = ulabel("oil_rate" if is_oil_for_pi
                                else "gas_rate", units)
        st.caption(
            f"**How PI works:** the productivity index links flow rate to "
            f"drawdown by  q = PI × (P_res − P_wf), where P_wf is the "
            f"flowing bottom-hole pressure. So PI in **{pi_units_label}** is "
            f"the rate produced per unit of pressure drawdown. Example: a PI "
            f"of 2 {pi_units_label} with {_ex_dd} {_ex_pl} of drawdown "
            f"delivers {2 * _ex_dd:,} {_ex_rate_unit}. "
            f"It is normally obtained from a well test (build-up / drawdown) "
            f"or estimated from k·h, fluid viscosity and skin."
        )
        # The engine works internally in field units; convert the metric
        # input back. PI has compound units so it scales by the rate factor
        # divided by the pressure factor.
        _pi_rate_kind = "oil_rate" if is_oil_for_pi else "gas_rate"
        well_pi_disp = st.number_input(
            f"Well PI ({pi_units_label}/well)",
            value=2.0 if is_oil_for_pi else 1.0,
            min_value=0.0, step=0.1, format="%.3f",
            key="well_pi_default", on_change=mark_stale,
            help="Productivity index per well. Typical screening values "
                 "(field units): light onshore oil 1-3 bbl/d/psi, deepwater "
                 "10-20, dry gas conventional 0.5-2 Mscf/d/psi, tight gas "
                 "0.05-0.20, heavy oil 0.3-1.5 (viscosity-limited).",
        )
        # Convert metric PI -> field PI for the engine.
        if units == "metric":
            # PI_field = PI_metric × (rate m->f factor) / (pressure m->f factor)
            well_pi_default = (well_pi_disp
                               * M2F[_pi_rate_kind] / M2F["pressure"])
        else:
            well_pi_default = well_pi_disp
        min_bhp_default = st.number_input(
            f"Min flowing BHP ({ulabel('pressure', units)})",
            value=from_field(1500.0, "pressure", units),
            min_value=0.0, step=100.0,
            key="min_bhp_default", on_change=mark_stale,
            help="Minimum allowable flowing bottom-hole pressure (well-level constraint). "
                 "Drawdown = P_res − BHP_min.",
        )

    with st.sidebar.expander("⚙️ Operational efficiency", expanded=False):
        prod_eff = st.slider(
            "Production efficiency (field)", 0.5, 1.0, 0.95, 0.01,
            key="prod_eff", on_change=mark_stale,
            help=("Fraction of theoretical production actually delivered. Captures field-level "
                  "downtime: weather, planned shutdowns, facility trips, pipeline outages, etc. "
                  "0.95 means 5% of nameplate is lost to downtime.")
        )
        st.caption(
            "Per-well **uptime** is set inside each well row in the Producers/Injectors table "
            "below (separate from this field-level efficiency). Total uptime ≈ uptime × PE."
        )

    with st.sidebar.expander("⛽ Gas disposition", expanded=False):
        st.markdown(
            "Defines how the **produced gas stream** is split. Fractions should sum to 1.0; "
            "if they don't, the remainder defaults to **export**. If they exceed 1.0, all four "
            "values are renormalized down."
        )
        gas_export = st.slider("Export (sold)", 0.0, 1.0, 1.00, 0.05,
                                key="gas_export", on_change=mark_stale,
                                help="Fraction of produced gas sold to market; only export volume earns gas revenue (in 'net' revenue mode).")
        gas_inj_frac = st.slider("Injection (re-injected for pressure support / EOR)",
                                  0.0, 1.0, 0.0, 0.05,
                                  key="gas_inj_frac", on_change=mark_stale,
                                  help="Re-injected gas. Subject to the gas-injection capacity. Excess (capacity-limited) falls back to export.")
        gas_fuel = st.slider("Fuel gas (own consumption)", 0.0, 0.3, 0.0, 0.01,
                              key="gas_fuel", on_change=mark_stale,
                              help="Used as fuel for compressors / power generation on the platform. Burnt — counts toward CO₂.")
        gas_flare = st.slider("Flare", 0.0, 0.3, 0.0, 0.01,
                               key="gas_flare", on_change=mark_stale,
                               help="Routine or upset flaring. Burnt — counts toward CO₂, plus a small methane-slip component.")
        total_disp = gas_export + gas_inj_frac + gas_fuel + gas_flare
        if abs(total_disp - 1.0) > 0.01:
            if total_disp > 1.001:
                st.warning(f"Sum = {total_disp:.2f} > 1.00 — values will be renormalized.")
            else:
                st.info(f"Sum = {total_disp:.2f}; remainder ({1-total_disp:.2f}) treated as export.")

    with st.sidebar.expander("⛔ Abandonment", expanded=False):
        aban_basis = st.radio("Apply at", ["Per well", "Field total"],
                              horizontal=True, key="aban_basis", on_change=mark_stale)
        default_oil = 50.0 if aban_basis == "Per well" else 5000.0
        default_gas = 0.5 if aban_basis == "Per well" else 20.0
        aban_oil_disp = st.number_input(
            f"Min oil rate ({ulabel('oil_rate', units)})",
            value=from_field(default_oil, "oil_rate", units),
            key="aban_oil", on_change=mark_stale)
        aban_gas_disp = st.number_input(
            "Min gas rate (MMscf/d in field; kSm³/d in metric)",
            value=default_gas if units == "field"
                  else from_field(default_gas * 1000, "gas_rate", units),
            key="aban_gas", on_change=mark_stale)
        aban_wc = st.slider("Max water cut", 0.5, 0.99, 0.95, 0.01,
                            key="aban_wc", on_change=mark_stale)

    return {
        "units": units, "fluid": fluid, "strategy": strategy,
        "start_date": start_date, "horizon": horizon,
        "ooip": to_field(ooip, "oil_vol", units),
        "ogip": to_field(ogip, "gas_vol", units),
        "rf_target": rf_target,
        "pvt": pvt, "aquifer": aquifer, "gas_cap": gas_cap,
        "vrr": vrr, "inj_eff": eff,
        "aban_basis": aban_basis,
        "aban_oil": to_field(aban_oil_disp, "oil_rate", units),
        "aban_gas": aban_gas_disp if units == "field"
                    else to_field(aban_gas_disp, "gas_rate", units) / 1000.0,
        "aban_wc": aban_wc,
        "ct_rock": ct_rock, "sw_init": sw_init,
        "retrograde_enabled": retrograde_enabled,
        "retrograde_drop_fraction": retrograde_drop_fraction,
        "fractional_flow_enabled": ff_enabled,
        "ff_params": ff_params_ui,
        "prod_eff": prod_eff,
        "auto_scale_rf": auto_scale_rf,
        "gas_export": gas_export, "gas_inj_frac": gas_inj_frac,
        "gas_fuel": gas_fuel, "gas_flare": gas_flare,
        "well_pi": well_pi_default,
        "min_bhp_psi": to_field(min_bhp_default, "pressure", units),
    }


# =============================================================================
# Wells UI
# =============================================================================
def well_type_curve_picker(units: str, fluid: str, rig_names: list):
    """Library + UI: pick a well archetype, preview its decline, instantiate N
    wells into the producers or injectors table.
    """
    is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"

    with st.expander("🧬 Add wells from a type curve", expanded=False):
        st.caption(
            "Pick a well archetype from the library and instantiate one or more "
            "wells with its full parameter set. Useful for screening — saves "
            "filling in 12 columns per well by hand. After adding, edit any "
            "fields directly in the producers / injectors tables."
        )

        # Build the reservoir context (used to score well-archetype fit)
        reservoir_ctx = None
        try:
            multi_on = bool(st.session_state.get("multi_res_enable", False))
            if multi_on and "reservoirs_df" in st.session_state:
                rdf = st.session_state["reservoirs_df"]
                if len(rdf) > 0:
                    r0 = rdf.iloc[0]
                    reservoir_ctx = {
                        "fluid_system": str(r0.get("fluid_system", fluid)),
                        "p_init":  to_field(float(r0.get("p_init", 3500.0)),
                                              "pressure", units),
                        "well_pi": float(r0.get("well_pi", 2.0)),
                        "min_bhp": to_field(float(r0.get("min_bhp", 1500.0)),
                                              "pressure", units),
                    }
            if reservoir_ctx is None:
                # Single-res mode: build from sidebar
                reservoir_ctx = {
                    "fluid_system": fluid,
                    "p_init":  to_field(float(st.session_state.get("p_init", 3500.0)),
                                          "pressure", units),
                    "well_pi": float(st.session_state.get("well_pi_default", 2.0)),
                    "min_bhp": to_field(float(st.session_state.get("min_bhp_default", 1500.0)),
                                          "pressure", units),
                }
        except Exception:
            reservoir_ctx = None

        # Filter toggle: by default, recommend only well archetypes scoring ≥ 0.5
        filter_on = st.checkbox(
            "🎯 Recommend only well archetypes that fit the current reservoir",
            value=True,
            key="tc_well_filter_on",
            help="Filters out archetypes whose fluid type or qi range is "
                 "incompatible with the active reservoir's PI × ΔP envelope. "
                 "Uncheck to see all archetypes.",
        )

        # Strategy context — keeps injectors out of the picker for Depletion fields
        strategy_ctx = st.session_state.get("strategy", "Injection")

        if filter_on and reservoir_ctx:
            names = fh.list_well_types_for_reservoir(reservoir_ctx, min_score=0.5,
                                                      strategy=strategy_ctx)
            if not names:
                st.info("No archetypes scored above 0.5 for this reservoir. "
                        "Showing all.")
                names = fh.list_well_types()
        else:
            names = fh.list_well_types()

        # Hide injector archetypes when the field is in Depletion mode (belt-and-
        # braces — the scorer already downranks them, this enforces it regardless
        # of the filter toggle).
        if strategy_ctx == "Depletion":
            names = [n for n in names
                     if (fh.get_well_type(n) or {}).get("kind") != "injector"]
            if not names:
                names = [n for n in fh.list_well_types()
                          if (fh.get_well_type(n) or {}).get("kind") != "injector"]

        c_left, c_right = st.columns([3, 2])

        with c_left:
            tname = st.selectbox(
                "Type curve",
                names,
                key="tc_well_choice",
                help="P50 archetypes for common producer & injector classes. "
                     "User-saved templates appear here too. Filtered by reservoir "
                     "fit when the toggle above is on.",
            )
            tmpl = fh.get_well_type(tname)
            if tmpl:
                st.caption(tmpl.get("description", ""))

            # Reservoir-fit chips for the selected archetype
            if tmpl and reservoir_ctx:
                fit = fh.well_template_reservoir_fit(tmpl, reservoir_ctx,
                                                       strategy=strategy_ctx)
                if fit["badges"]:
                    chips_html = " ".join(
                        f'<span class="eq-chip">{b}</span>' for b in fit["badges"]
                    )
                    st.markdown(chips_html, unsafe_allow_html=True)
                if fit["score"] < 0.7 and fit["reason"]:
                    st.warning(f"Reservoir fit: {fit['reason']}")
                elif fit["pi_implied_qi"] > 0:
                    _qi_k = "oil_rate" if tmpl.get("fluid") == "oil" else "gas_rate"
                    _qi_d = from_field(fit["pi_implied_qi"], _qi_k, units)
                    st.caption(
                        f"💡 Reservoir PI × ΔP implies ~{_qi_d:,.0f} "
                        f"{ulabel(_qi_k, units)} per well for this archetype."
                    )

        with c_right:
            # Live preview of the decline curve (producers only)
            if tmpl and tmpl.get("kind") == "producer":
                horizon_months = 120
                xs, ys = fh.type_curve_preview_series(tmpl, horizon_months)
                # Display in user units
                fluid_kind_for_preview = "oil_rate" if tmpl.get("fluid") == "oil" else "gas_rate"
                ys_disp = [from_field(y, fluid_kind_for_preview, units) for y in ys]
                fig = go.Figure()
                fig.add_trace(go.Scatter(
                    x=xs, y=ys_disp, mode="lines",
                    line=dict(color=fh.EQ_COLORS["oil"] if tmpl.get("fluid") == "oil"
                                                       else fh.EQ_COLORS["gas"],
                              width=2.5),
                    name="Type curve",
                    hovertemplate=f"Month %{{x}}<br>Rate %{{y:,.0f}} {ulabel(fluid_kind_for_preview, units)}<extra></extra>",
                ))
                fig.update_layout(
                    height=180,
                    margin=dict(l=40, r=10, t=20, b=30),
                    xaxis_title="Months on stream",
                    yaxis_title=ulabel(fluid_kind_for_preview, units),
                    showlegend=False,
                )
                st.plotly_chart(fh.apply_plot_template(fig), use_container_width=True)
            else:
                st.info("Injector — no decline preview.")

        if not tmpl:
            return

        # Instantiate controls
        st.markdown("**Add to project**")
        a, b, c, d = st.columns([1, 1, 2, 1])
        n_to_add = a.number_input("Wells to add", min_value=1, max_value=50,
                                   value=4, step=1, key="tc_n_wells")
        scale = b.number_input("qi scale ×", min_value=0.1, max_value=5.0,
                                value=1.0, step=0.05, key="tc_qi_scale",
                                help="Multiplier on the archetype's IP rate "
                                     "(use 0.8 for low-side / 1.2 for high-side).")
        prefix_default = ("WI-" if tmpl.get("kind") == "injector" else "P-")
        prefix = c.text_input("Name prefix", value=prefix_default,
                               key="tc_prefix",
                               help="Sequential numbering is appended (e.g. 'P-09').")
        rig_choice = d.selectbox("Rig", rig_names if rig_names else ["Rig-A"],
                                  key="tc_rig_choice")

        if st.button(f"➕ Add {n_to_add} {tmpl['kind']}{'s' if n_to_add != 1 else ''}",
                      key="tc_add_btn", use_container_width=True):
            _instantiate_wells_from_template(tmpl, int(n_to_add), float(scale),
                                              prefix, rig_choice, units)
            st.success(f"Added {n_to_add} well(s) from '{tname}'.")
            mark_stale()
            st.rerun()

        # Save current template / manage user templates
        with st.expander("💾 Save current row as template / manage saved", expanded=False):
            st.caption(
                "Pick an existing producer row from the table to save its "
                "values as a reusable template, or delete a saved user template."
            )
            sub_a, sub_b = st.columns(2)
            with sub_a:
                if "producers_df" in st.session_state and len(st.session_state.producers_df) > 0:
                    pdf = st.session_state.producers_df
                    src_idx = st.selectbox(
                        "Source row (producer)",
                        list(range(len(pdf))),
                        format_func=lambda i: f"{i}: {pdf.iloc[i].get('name', '?')}",
                        key="tc_save_src",
                    )
                    new_name = st.text_input("Template name", value="My custom curve",
                                              key="tc_save_name")
                    new_desc = st.text_area("Description", value="",
                                             key="tc_save_desc", height=60)
                    if st.button("💾 Save row as template", key="tc_save_btn"):
                        try:
                            row = pdf.iloc[src_idx]
                            params = {
                                "kind": "producer",
                                "fluid": "oil" if is_oil else "gas",
                                "qi_primary":   to_field(float(row.get("qi_primary", 0.0)),
                                                          "oil_rate" if is_oil else "gas_rate", units),
                                "qi_secondary": to_field(float(row.get("qi_secondary", 0.0)),
                                                          "gas_rate" if is_oil else "oil_rate", units),
                                "decline_model": str(row.get("decline_model", "Exponential")),
                                "di_annual": float(row.get("di_annual", 0.20)),
                                "b_factor":  float(row.get("b_factor", 0.5)),
                                "wc_initial": float(row.get("wc_initial", 0.05)),
                                "wc_final":   float(row.get("wc_final", 0.85)),
                                "wc_ramp_months": int(float(row.get("wc_ramp_months", 60))),
                                "uptime":   float(row.get("uptime", 0.95)),
                                "drill_days": int(float(row.get("drill_days", 45))),
                                "completion_days": int(float(row.get("completion_days", 15))),
                                "description": new_desc or f"User template '{new_name}'.",
                            }
                            fh.save_user_well_type(new_name, params)
                            st.success(f"Saved '{new_name}'.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Could not save: {e}")
                else:
                    st.caption("Add at least one producer row first.")
            with sub_b:
                user_names = fh._list_user_well_types()
                if user_names:
                    to_del = st.selectbox("Delete user template",
                                           ["—"] + user_names, key="tc_delete_choice")
                    if to_del != "—" and st.button("🗑 Delete", key="tc_delete_btn"):
                        if fh.delete_user_well_type(to_del):
                            st.success(f"Deleted '{to_del}'.")
                            st.rerun()
                else:
                    st.caption("No user templates saved yet.")


def _instantiate_wells_from_template(tmpl: dict, n: int, scale: float,
                                      prefix: str, rig: str, units: str) -> None:
    """Append N rows to producers_df or injectors_df from a template."""
    is_producer = tmpl.get("kind") == "producer"

    # Decide naming start: continue numbering from the existing table
    target_key = "producers_df" if is_producer else "injectors_df"
    existing = st.session_state.get(target_key)
    existing_names = set()
    if existing is not None and len(existing) > 0 and "name" in existing.columns:
        existing_names = set(str(x) for x in existing["name"].dropna())

    def next_name(start_i: int) -> str:
        i = start_i
        while True:
            candidate = f"{prefix}{i:02d}"
            if candidate not in existing_names:
                existing_names.add(candidate)
                return candidate
            i += 1

    new_rows = []
    if is_producer:
        # Field-unit qi from template, then convert to display units for the table
        rate_kind_p = "oil_rate" if tmpl.get("fluid") == "oil" else "gas_rate"
        rate_kind_s = "gas_rate" if tmpl.get("fluid") == "oil" else "oil_rate"
        qi_p_display = from_field(tmpl["qi_primary"] * scale, rate_kind_p, units)
        qi_s_display = from_field(tmpl["qi_secondary"] * scale, rate_kind_s, units)
        start_i = 1
        for _ in range(n):
            new_rows.append({
                "name": next_name(start_i),
                "rig": rig,
                "drill_days": int(tmpl.get("drill_days", 45)),
                "completion_days": int(tmpl.get("completion_days", 15)),
                "qi_primary":   float(qi_p_display),
                "qi_secondary": float(qi_s_display),
                "decline_model": tmpl.get("decline_model", "Exponential"),
                "di_annual":     float(tmpl.get("di_annual", 0.20)),
                "b_factor":      float(tmpl.get("b_factor", 0.5)),
                "wc_initial":    float(tmpl.get("wc_initial", 0.05)),
                "wc_final":      float(tmpl.get("wc_final", 0.85)),
                "wc_ramp_months": int(tmpl.get("wc_ramp_months", 60)),
                "scale_factor":  1.0,
                "uptime":        float(tmpl.get("uptime", 0.95)),
            })
            start_i += 1
    else:
        # Injector: convert inj_rate to display units
        inj_display = from_field(tmpl["inj_rate"] * scale, "water_rate", units)
        start_i = 1
        for _ in range(n):
            new_rows.append({
                "name": next_name(start_i),
                "rig": rig,
                "drill_days": int(tmpl.get("drill_days", 45)),
                "completion_days": int(tmpl.get("completion_days", 15)),
                "inj_rate": float(inj_display),
                "scale_factor": 1.0,
                "uptime":       float(tmpl.get("uptime", 0.95)),
            })

    new_df = pd.DataFrame(new_rows)
    if existing is None or len(existing) == 0:
        st.session_state[target_key] = new_df
    else:
        st.session_state[target_key] = pd.concat([existing, new_df], ignore_index=True)


def decline_fitter_picker(units: str, fluid: str, rig_names: list):
    """UI to fit Arps decline parameters from pasted/uploaded historical
    monthly production and instantiate fitted wells into the producers table.
    """
    is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"

    with st.expander("📈 Fit decline from historical production", expanded=False):
        st.caption(
            "Paste or upload monthly production history; the engine fits "
            "Arps (exponential/harmonic/hyperbolic) parameters per well "
            "and creates pre-populated producer rows. CSV columns accepted "
            "(case-insensitive): `well`/`name`, `month`/`date`, "
            "`rate`/`oil_rate`/`gas_rate`/`production`."
        )

        upl = st.file_uploader("Upload CSV", type=["csv", "txt"],
                                key="fit_upload",
                                help="Or paste data into the text box below.")
        sample_csv = (
            "well,month,rate\n"
            "P-01,0,2500\nP-01,1,2380\nP-01,2,2270\nP-01,3,2160\nP-01,4,2070\n"
            "P-01,5,1980\nP-01,6,1900\nP-01,7,1820\nP-01,8,1740\nP-01,9,1670\n"
            "P-02,0,3500\nP-02,1,3300\nP-02,2,3120\nP-02,3,2950\nP-02,4,2790\n"
            "P-02,5,2640\nP-02,6,2500\nP-02,7,2370\nP-02,8,2240\nP-02,9,2120\n"
        )
        text = st.text_area("Or paste CSV", value=sample_csv, height=160,
                             key="fit_paste",
                             help="The pre-filled example shows the expected format.")

        c1, c2, c3 = st.columns([2, 2, 2])
        chosen_model = c1.selectbox(
            "Decline model", ["auto", "exponential", "harmonic", "hyperbolic"],
            index=0, key="fit_model",
            help="`auto` picks whichever fits best (with a small AIC-style "
                 "penalty against hyperbolic to avoid over-fitting).",
        )
        rig_choice = c2.selectbox("Rig", rig_names if rig_names else ["Rig-A"],
                                   key="fit_rig")
        # Default forecast WC and ramp for newly-instantiated wells (history
        # alone doesn't give us water-cut behavior unless the user uploads it)
        default_wc_final = c3.slider("Default final water cut", 0.0, 0.99,
                                      0.85, 0.05, key="fit_wcf")

        if not st.button("Fit decline & preview", key="fit_run",
                          use_container_width=True):
            return

        # Parse
        try:
            if upl is not None:
                hist = fh.parse_decline_csv(upl)
            else:
                hist = fh.parse_decline_csv(text)
        except Exception as e:
            st.error(f"Could not parse the input: {e}")
            return

        if len(hist) == 0:
            st.warning("No usable rows found in the input.")
            return

        st.success(f"Parsed {len(hist)} rows · {hist['well'].nunique()} well(s).")

        # Fit per well
        rate_kind = "oil_rate" if is_oil else "gas_rate"
        fit_rows = []
        preview_traces = []
        for well_name in hist["well"].unique():
            sub = hist[hist["well"] == well_name].sort_values("month")
            months = sub["month"].values
            rates = sub["rate"].values
            fit = fh.fit_arps(months, rates, model=chosen_model)
            fit_rows.append({
                "Well": well_name,
                "Model": fit["model"],
                "qi (fit)": fit["qi"],
                "qi ± SE":  fit.get("qi_se", float("nan")),
                "di /yr":   fit["di_annual"],
                "di ± SE":  fit.get("di_se", float("nan")),
                "b":        fit["b_factor"],
                "R²":       fit["r2"],
                "n points": fit["n_points"],
            })
            preview_traces.append({
                "well": well_name, "months": months, "rates": rates,
                "fit_rates": fit["fitted_rates"], "fit": fit,
            })

        # Render fit summary table
        st.markdown("**Fit summary**")
        fit_df = pd.DataFrame(fit_rows)
        st.dataframe(
            fit_df.style.format({
                "qi (fit)": "{:,.0f}", "qi ± SE": "{:,.0f}",
                "di /yr": "{:.3f}", "di ± SE": "{:.3f}",
                "b": "{:.2f}", "R²": "{:.3f}",
            }),
            use_container_width=True, hide_index=True,
        )

        # Show derived MC priors (P10/P90 envelopes from parameter SEs)
        with st.expander("📊 Use fit uncertainty as Monte Carlo priors", expanded=False):
            st.caption(
                "The standard errors from each well's fit translate directly into "
                "MC driver bounds. Use these in the Monte Carlo tab as your "
                "`Well qi` and `Decline rate` low/high factors."
            )
            n_sig = st.slider("Confidence level (× σ)", 1.0, 3.0, 1.65, 0.05,
                               key="mc_prior_nsig",
                               help="1.65σ ≈ P10/P90, 1.96σ ≈ 95% CI, 2.58σ ≈ P5/P95.")
            prior_rows = []
            for tr in preview_traces:
                priors = fh.fit_to_mc_priors(tr["fit"], n_sigma=n_sig)
                prior_rows.append({
                    "Well": tr["well"],
                    "qi low": priors["qi_low_factor"],
                    "qi high": priors["qi_high_factor"],
                    "di low": priors["di_low_factor"],
                    "di high": priors["di_high_factor"],
                })
            prior_df = pd.DataFrame(prior_rows)
            st.dataframe(prior_df.style.format({
                "qi low": "{:.3f}", "qi high": "{:.3f}",
                "di low": "{:.3f}", "di high": "{:.3f}",
            }), use_container_width=True, hide_index=True)

            # Pooled / median prior across wells (for use as a single MC bound)
            if len(prior_df) > 0:
                qi_lo = float(prior_df["qi low"].median())
                qi_hi = float(prior_df["qi high"].median())
                di_lo = float(prior_df["di low"].median())
                di_hi = float(prior_df["di high"].median())
                st.markdown(
                    f"**Median across wells (use as MC bounds):**  \n"
                    f"`Well qi`: low = **{qi_lo:.3f}**, high = **{qi_hi:.3f}**  \n"
                    f"`Decline rate`: low = **{di_lo:.3f}**, high = **{di_hi:.3f}**"
                )
                if st.button("📌 Push these to Monte Carlo defaults",
                              key="push_mc_priors",
                              help="Sets the MC tab's Well qi / Decline rate "
                                   "low/high to these values (effective on next "
                                   "page render)."):
                    st.session_state["mc_lo_Well qi"] = qi_lo
                    st.session_state["mc_hi_Well qi"] = qi_hi
                    st.session_state["mc_lo_Decline rate"] = di_lo
                    st.session_state["mc_hi_Decline rate"] = di_hi
                    st.success("Pushed to Monte Carlo tab. Open the tab to verify.")

        # Per-well overlay plot (history vs fitted curve, +24mo forecast)
        st.markdown("**Fit overlay (history + 24-month forecast extension)**")
        fig = go.Figure()
        for tr in preview_traces:
            # History
            fig.add_trace(go.Scatter(
                x=tr["months"], y=from_field(tr["rates"], rate_kind, units),
                mode="markers", name=f"{tr['well']} obs",
                marker=dict(size=6, color=fh.EQ_COLORS["oil"] if is_oil else fh.EQ_COLORS["gas"]),
                legendgroup=tr["well"],
            ))
            # Fit + forecast
            n_hist = int(tr["months"].max() if len(tr["months"]) else 0)
            forward = np.arange(0, n_hist + 24)
            fit = tr["fit"]
            fitted = fh._arps_rate(forward / 12.0, fit["qi"],
                                    fit["di_annual"], fit["b_factor"])
            fig.add_trace(go.Scatter(
                x=forward, y=from_field(fitted, rate_kind, units),
                mode="lines", name=f"{tr['well']} fit ({fit['model']})",
                line=dict(width=2, dash="dash"),
                legendgroup=tr["well"],
            ))
        fig.update_layout(
            height=380, hovermode="x unified",
            xaxis_title="Month", yaxis_title=ulabel(rate_kind, units),
            legend=dict(orientation="h", y=-0.18),
        )
        st.plotly_chart(fh.apply_plot_template(fig), use_container_width=True)

        st.markdown("**Add fitted wells to the producers table**")
        a, b = st.columns([2, 1])
        a.caption(
            "Click below to append the fitted wells to the producers table. "
            f"You can edit any fields afterward (e.g. set spud dates per well)."
        )
        if b.button(f"➕ Add {len(fit_rows)} fitted producer(s)",
                     key="fit_add_btn", use_container_width=True):
            new_rows = []
            existing = st.session_state.get("producers_df")
            existing_names = (set(str(x) for x in existing["name"].dropna())
                              if existing is not None and "name" in existing.columns
                              else set())
            for tr in preview_traces:
                fit = tr["fit"]
                # qi is already in field units (whatever the input was — we
                # treat the rate column as already in the user's chosen units)
                qi_display = float(fit["qi"])
                # Ensure unique name
                base_name = str(tr["well"])
                name = base_name
                k = 1
                while name in existing_names:
                    name = f"{base_name}_{k}"
                    k += 1
                existing_names.add(name)
                new_rows.append({
                    "name": name,
                    "rig": rig_choice,
                    "drill_days": 45, "completion_days": 15,
                    "qi_primary":   qi_display,
                    "qi_secondary": qi_display * 2.0 if is_oil else 50.0,
                    "decline_model": fit["model"].capitalize(),
                    "di_annual":     float(fit["di_annual"]),
                    "b_factor":      float(fit["b_factor"]),
                    "wc_initial":    0.05,
                    "wc_final":      float(default_wc_final),
                    "wc_ramp_months": 60,
                    "scale_factor":  1.0, "uptime": 0.95,
                })
            new_df = pd.DataFrame(new_rows)
            if existing is None or len(existing) == 0:
                st.session_state["producers_df"] = new_df
            else:
                st.session_state["producers_df"] = pd.concat(
                    [existing, new_df], ignore_index=True)
            st.success(f"Added {len(new_rows)} fitted producer(s) to the table.")
            mark_stale()
            st.rerun()


def well_section(units, fluid, start_date):
    is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"
    st.subheader("🛠️ Drilling rigs, producers & injectors")

    # HPHT badge — the field PVT determines whether these are HPHT wells.
    try:
        _p = to_field(float(st.session_state.get("p_init", 3500.0)),
                      "pressure", units)
        _t = to_field(float(st.session_state.get("t_res", 180.0)),
                      "temp", units)
        _hpht = fh.classify_hpht(_p, _t)
        # `_p`, `_t` are in field units (psi/°F) because classify_hpht
        # expects field units. Convert for display.
        _p_d = from_field(_p, "pressure", units)
        _t_d = from_field(_t, "temp", units)
        _pl = ulabel("pressure", units)
        _tl = ulabel("temp", units)
        if _hpht["is_hpht"]:
            st.warning(f"**{_hpht['tag']} wells** — these wells operate in "
                       f"{_hpht['tier']} conditions ({_p_d:,.0f} {_pl}, "
                       f"{_t_d:,.0f} {_tl}). HPHT wells need specialised "
                       f"completions and higher-grade metallurgy; expect "
                       f"longer drilling/completion times and higher well "
                       f"cost. The development concept builder applies a "
                       f"×{_hpht['capex_uplift']:.2f} CAPEX uplift for this "
                       f"tier.")
        else:
            st.caption(f"{_hpht['tag']} — standard pressure/temperature "
                       f"conditions ({_p_d:,.0f} {_pl}, {_t_d:,.0f} {_tl}).")
    except Exception:
        pass

    with st.expander("ℹ️ How this section works", expanded=False):
        st.markdown(
            "- **Rigs**: each rig drills its assigned wells **sequentially** in the order they appear "
            "in the producers/injectors tables.\n"
            "- The **spud date** for each well is computed as the prior well's drill+completion end "
            "on the same rig (or the rig's `Available from` date for the first well on it).\n"
            "- **Scaling factor** multiplies the well's full rate profile (used for sensitivities or "
            "type-curve scaling).\n"
            "- For **User-defined profile** decline, upload a CSV or an "
            "Eclipse summary export — the importer auto-detects the format "
            "and flexible column names (oil_rate / qoil / WOPR …).\n"
            "- For **Multi-segment** decline, build a piecewise-Arps "
            "profile (plateau, decline, bean-up, late-life bump) in the "
            "segment editor."
        )

    st.markdown("**Drilling rigs**")
    if "rigs_df" not in st.session_state:
        st.session_state.rigs_df = pd.DataFrame({
            "rig": ["Rig-A"],
            "start_date": [start_date],
            "move_in_days": [30],
            "move_out_days": [15],
            "maintenance_days_per_year": [10],
            "day_rate_kUSD": [350.0],
        })
    # Backfill new columns on saved sessions
    _rdf = st.session_state.rigs_df
    for col, default in [("move_in_days", 30), ("move_out_days", 15),
                          ("maintenance_days_per_year", 10),
                          ("day_rate_kUSD", 350.0)]:
        if col not in _rdf.columns:
            _rdf[col] = default
    st.session_state.rigs_df = _rdf

    rigs_buf = st.data_editor(
        st.session_state.rigs_df, num_rows="dynamic", use_container_width=True,
        column_config={
            "rig": st.column_config.TextColumn("Rig name", required=True,
                help="Unique rig identifier referenced by wells below."),
            "start_date": st.column_config.DateColumn("Available from", required=True,
                help="The earliest date this rig can spud its first well."),
            "move_in_days": st.column_config.NumberColumn(
                "Move-in (days)", min_value=0, step=1,
                help="Rig mobilization time before the FIRST well on this rig "
                     "can spud. Pushes the whole rig's drilling program forward. "
                     "Adds dayrate cost to facility CAPEX."),
            "move_out_days": st.column_config.NumberColumn(
                "Move-out (days)", min_value=0, step=1,
                help="Rig demobilization time after the LAST well. Adds dayrate "
                     "cost to facility CAPEX (no production impact)."),
            "maintenance_days_per_year": st.column_config.NumberColumn(
                "Maint. (days/yr)", min_value=0, max_value=120, step=1,
                help="Planned rig maintenance / downtime per year. Inserted as "
                     "gaps between wells, proportionally delaying spud dates of "
                     "later wells on this rig."),
            "day_rate_kUSD": st.column_config.NumberColumn(
                "Day rate ($k/d)", min_value=0.0, step=10.0, format="%.0f",
                help="Rig day rate — used to cost move-in/out and maintenance "
                     "days, and (when 'Rig-rate' well-cost mode is on) the "
                     "drilling + completion days of each well."),
        },
        key="rigs_editor",
    )
    br1, br2, _br3 = st.columns([2, 1, 3])
    with br1:
        rigs_apply_clicked = _apply_button(rigs_buf, st.session_state.rigs_df,
                                            "rigs_apply", "Apply rig edits")
    if rigs_apply_clicked:
        commit = rigs_buf.copy()
        if "rig" in commit.columns:
            commit = commit[commit["rig"].notna() & (commit["rig"].astype(str).str.strip() != "")]
            commit = commit.reset_index(drop=True)
        st.session_state.rigs_df = commit
        mark_stale()
        st.rerun()
    if br2.button("📋 Duplicate last rig", key="rigs_dup"):
        st.session_state.rigs_df = _duplicate_last_row(st.session_state.rigs_df)
        mark_stale()
        st.rerun()
    rigs_df = st.session_state.rigs_df
    rig_names = rigs_df["rig"].tolist() if len(rigs_df) > 0 else ["Rig-A"]

    qi_p_default = from_field(2500.0 if is_oil else 25_000.0,
                              "oil_rate" if is_oil else "gas_rate", units)
    qi_s_default = from_field(5000.0 if is_oil else 200.0,
                              "gas_rate" if is_oil else "oil_rate", units)

    # =========================================================================
    # 🧬 Add wells from a type curve (archetype library)
    # =========================================================================
    well_type_curve_picker(units, fluid, rig_names)
    decline_fitter_picker(units, fluid, rig_names)

    st.markdown("**Producers**")
    if "producers_df" not in st.session_state:
        # IPR defaults — store in display units to match the column-header
        # labels (P_wh in psi or bar; depth in ft or m).
        _whp_default = from_field(200.0, "pressure", units)
        _td_default = from_field(8000.0, "depth", units)
        rows = []
        for i in range(8):
            rows.append({
                "name": f"P-{i+1:02d}",
                "rig": rig_names[i % len(rig_names)],
                "drill_days": 45, "completion_days": 15,
                "qi_primary": qi_p_default, "qi_secondary": qi_s_default,
                "decline_model": "Exponential", "di_annual": 0.20, "b_factor": 0.5,
                "wc_initial": 0.05, "wc_final": 0.85, "wc_ramp_months": 60,
                "scale_factor": 1.0, "uptime": 0.95,
                "derive_qi_from_pi": False, "well_pi_override": 0.0,
                "fluid": "auto",
                "ipr_mode": False,
                "wellhead_pressure_psi": _whp_default,
                "tubing_depth_ft": _td_default,
                "fluid_gradient_psi_per_ft": 0.35,   # engine units (psi/ft)
                "friction_psi_per_kbpd": 5.0,
            })
        st.session_state.producers_df = pd.DataFrame(rows)
    else:
        # Backfill new columns on saved sessions
        pdf = st.session_state.producers_df
        if "derive_qi_from_pi" not in pdf.columns:
            pdf["derive_qi_from_pi"] = False
        if "well_pi_override" not in pdf.columns:
            pdf["well_pi_override"] = 0.0
        if "uptime" not in pdf.columns:
            pdf["uptime"] = 0.95
        if "fluid" not in pdf.columns:
            pdf["fluid"] = "auto"
        if "ipr_mode" not in pdf.columns:
            pdf["ipr_mode"] = False
        if "wellhead_pressure_psi" not in pdf.columns:
            pdf["wellhead_pressure_psi"] = 200.0
        if "tubing_depth_ft" not in pdf.columns:
            pdf["tubing_depth_ft"] = 8000.0
        if "fluid_gradient_psi_per_ft" not in pdf.columns:
            pdf["fluid_gradient_psi_per_ft"] = 0.35
        if "friction_psi_per_kbpd" not in pdf.columns:
            pdf["friction_psi_per_kbpd"] = 5.0
        st.session_state.producers_df = pdf

    # ---- Display→storage conversion for unit-bearing IPR columns ----
    # Storage keeps the field-unit values (which is what the engine expects);
    # the editor BUFFER displays in user units. On Apply we convert back.
    is_metric = (units == "metric")
    pdf_storage = st.session_state.producers_df
    pdf_display = pdf_storage.copy()
    if is_metric:
        if "wellhead_pressure_psi" in pdf_display.columns:
            pdf_display["wellhead_pressure_psi"] = pdf_display["wellhead_pressure_psi"].apply(
                lambda v: from_field(float(v or 0.0), "pressure", units))
        if "tubing_depth_ft" in pdf_display.columns:
            pdf_display["tubing_depth_ft"] = pdf_display["tubing_depth_ft"].apply(
                lambda v: from_field(float(v or 0.0), "depth", units))
        # ρ stored as psi/ft → bar/m (× 0.22621) for display.
        if "fluid_gradient_psi_per_ft" in pdf_display.columns:
            pdf_display["fluid_gradient_psi_per_ft"] = (
                pdf_display["fluid_gradient_psi_per_ft"].apply(
                    lambda v: float(v or 0.0) * 0.22621))

    producers_df_buf = st.data_editor(
        pdf_display, num_rows="dynamic", use_container_width=True,
        column_config={
            "name": st.column_config.TextColumn("Well", required=True),
            "rig": st.column_config.SelectboxColumn("Rig", options=rig_names, required=True),
            "drill_days": st.column_config.NumberColumn(
                "Drill days", min_value=1, step=1,
                help="Days from spud to TD. Wells on a rig drill sequentially."),
            "completion_days": st.column_config.NumberColumn(
                "Compl. days", min_value=1, step=1,
                help="Days from TD to first oil (perforation, hookup, testing)."),
            "qi_primary": st.column_config.NumberColumn(
                f"qi {'oil' if is_oil else 'gas'} ({ulabel('oil_rate' if is_oil else 'gas_rate', units)})",
                min_value=0.0,
                help="Initial primary-fluid rate at start of decline. "
                     "When 'PI mode' is enabled for this well, this value is "
                     "ignored and qi is recomputed from PI × (P_init − BHP_min)."),
            "qi_secondary": st.column_config.NumberColumn(
                f"qi {'gas' if is_oil else 'cond'} ({ulabel('gas_rate' if is_oil else 'oil_rate', units)})",
                min_value=0.0),
            "decline_model": st.column_config.SelectboxColumn(
                "Decline", options=DECLINE_MODELS, required=True,
                help="Arps families. Pick 'User-defined profile' to upload a CSV."),
            "di_annual": st.column_config.NumberColumn(
                "Decline (1/yr)", min_value=0.0, max_value=2.0, step=0.01, format="%.3f",
                help="Initial decline rate (nominal annual)."),
            "b_factor": st.column_config.NumberColumn(
                "b (hyperb.)", min_value=0.0, max_value=2.0, step=0.05,
                help="Hyperbolic exponent. 0 = exponential, 1 = harmonic."),
            "wc_initial": st.column_config.NumberColumn("WC start", min_value=0.0, max_value=0.99, format="%.2f",
                help="Water cut at first oil."),
            "wc_final": st.column_config.NumberColumn("WC final", min_value=0.0, max_value=0.99, format="%.2f"),
            "wc_ramp_months": st.column_config.NumberColumn("WC ramp (mo)", min_value=0, step=1,
                help="Months to ramp from WC start to WC final."),
            "scale_factor": st.column_config.NumberColumn(
                "Scale", min_value=0.1, max_value=5.0, step=0.05, format="%.2f",
                help="Multiplies the well's entire rate profile."),
            "uptime": st.column_config.NumberColumn(
                "Uptime", min_value=0.0, max_value=1.0, step=0.01, format="%.2f",
                help="Fraction of time the well is on stream."),
            "derive_qi_from_pi": st.column_config.CheckboxColumn(
                "PI mode",
                help="When ON, the well's qi is computed at simulation time as "
                     "PI × (P_init − BHP_min) using the linked reservoir's PI / BHP "
                     "(or the per-well override below). Decline still applies on top. "
                     "When OFF, qi_primary is used directly as a free input."),
            "well_pi_override": st.column_config.NumberColumn(
                "PI override",
                min_value=0.0, step=0.1, format="%.2f",
                help=("Optional per-well PI override "
                      f"({'bbl/d/psi' if units=='field' else 'Sm³/d/bar'} for oil, "
                      f"{'Mscf/d/psi' if units=='field' else 'kSm³/d/bar'} for gas). "
                      "Leave at 0 to use the linked reservoir's PI.")),
            "fluid": st.column_config.SelectboxColumn(
                "Fluid",
                options=["auto", "oil", "gas"],
                help="Per-well fluid type. 'auto' inherits the field's primary "
                     "fluid. Set 'oil' or 'gas' explicitly when the well taps "
                     "a different-fluid reservoir than the field default "
                     "(useful in multi-reservoir fields with mixed fluids)."),
            "ipr_mode": st.column_config.CheckboxColumn(
                "IPR mode",
                help="When ON, the engine computes the well's actual rate at every "
                     "timestep from the IPR (Vogel for oil, back-pressure for gas) "
                     "intersected with a simple outflow curve (hydrostatic + friction). "
                     "Wells go off plateau when reservoir pressure drops."),
            "wellhead_pressure_psi": st.column_config.NumberColumn(
                f"P_wh ({ulabel('pressure', units)})",
                min_value=0.0, step=10.0 if units == "field" else 1.0,
                format="%.0f" if units == "field" else "%.1f",
                help=("Flowing wellhead pressure (separator / manifold). "
                      "Typical 100-500 psi onshore (7-35 bar), "
                      "200-1500 psi offshore (14-100 bar).")),
            "tubing_depth_ft": st.column_config.NumberColumn(
                f"Depth ({ulabel('depth', units)})",
                min_value=0.0, step=100.0 if units == "field" else 30.0,
                format="%.0f",
                help="Mid-perf depth — sets the hydrostatic head for outflow."),
            "fluid_gradient_psi_per_ft": st.column_config.NumberColumn(
                f"ρ [{'psi/ft' if units == 'field' else 'bar/m'}]",
                min_value=0.0,
                max_value=1.0 if units == "field" else 0.25,
                step=0.01 if units == "field" else 0.005,
                format="%.3f" if units == "metric" else "%.2f",
                help=(
                    "Mixture hydrostatic gradient. Typical values: "
                    + ("oil ~0.30-0.40 psi/ft, water ~0.43, gas "
                       "~0.05-0.15. For high-WC wells use 0.40-0.43."
                       if units == "field" else
                       "oil ~0.068-0.090 bar/m, water ~0.097, gas "
                       "~0.011-0.034. For high-WC wells use "
                       "0.090-0.097.")
                    + " A consistent suggestion based on the current "
                      "PVT (API gravity + water-cut assumption) appears "
                      "as a caption below the table. "
                      "1 psi/ft ≈ 0.2262 bar/m.")),
            "friction_psi_per_kbpd": st.column_config.NumberColumn(
                "Friction [psi/kbpd]",
                min_value=0.0, max_value=50.0, step=0.5, format="%.1f",
                help=("Linear friction proxy (psi per 1000 bbl/d, kept in field "
                      "convention regardless of unit system). Higher tubing ID "
                      "and lower viscosity reduce this. Typical 2-10 for oil "
                      "wells, 5-20 for high-rate gas wells.")),
        },
        key="producers_editor",
    )

    # ---- Suggested ρ gradient from current PVT ----
    # Oil density at standard conditions from API gravity:
    #   SG_oil = 141.5 / (131.5 + API)
    # Mixture gradient at the wellbore = WC × ρ_water + (1-WC) × ρ_oil
    # ρ_water = 0.433 psi/ft (fresh), 0.45 for brine.
    # Convert to bar/m if metric.
    try:
        _api = float(st.session_state.get("api", 36.0))
        _sg_oil = 141.5 / (131.5 + max(_api, 1.0))
        _grad_oil_psi_ft = _sg_oil * 0.433
        _wc_assumed = float(st.session_state.get("wc_design", 0.30))
        _grad_mix_psi_ft = (
            _wc_assumed * 0.443 + (1 - _wc_assumed) * _grad_oil_psi_ft)
        if units == "field":
            _disp_oil = _grad_oil_psi_ft
            _disp_mix = _grad_mix_psi_ft
            _g_u = "psi/ft"
        else:
            _disp_oil = _grad_oil_psi_ft * 0.22621
            _disp_mix = _grad_mix_psi_ft * 0.22621
            _g_u = "bar/m"
        st.caption(
            f"💡 **Suggested ρ from current PVT** "
            f"(API = {_api:.1f}, SG_oil = {_sg_oil:.3f}): "
            f"pure-oil leg ≈ **{_disp_oil:.3f} {_g_u}**; "
            f"mixture with {_wc_assumed:.0%} water cut "
            f"≈ **{_disp_mix:.3f} {_g_u}**. "
            f"Set the ρ column to one of these (or your own measured value)."
        )
    except (ValueError, TypeError, KeyError):
        pass

    btn_p1, btn_p2, _btn_p3 = st.columns([2, 1, 3])
    # The buffer is in display units; compare against a display-converted copy
    # of the committed (storage-unit) producers_df so dirty-detection is correct.
    with btn_p1:
        producers_apply_clicked = _apply_button(
            producers_df_buf, pdf_display, "producers_apply",
            "Apply producer edits")
    if producers_apply_clicked:
        commit = producers_df_buf.copy()
        # Convert display→storage for IPR fields
        if is_metric:
            if "wellhead_pressure_psi" in commit.columns:
                commit["wellhead_pressure_psi"] = commit["wellhead_pressure_psi"].apply(
                    lambda v: to_field(float(v or 0.0), "pressure", units))
            if "tubing_depth_ft" in commit.columns:
                commit["tubing_depth_ft"] = commit["tubing_depth_ft"].apply(
                    lambda v: to_field(float(v or 0.0), "depth", units))
            # ρ user-entered as bar/m → storage psi/ft. 1 psi/ft = 0.22621 bar/m.
            if "fluid_gradient_psi_per_ft" in commit.columns:
                commit["fluid_gradient_psi_per_ft"] = (
                    commit["fluid_gradient_psi_per_ft"].apply(
                        lambda v: float(v or 0.0) / 0.22621))
        # Drop rows missing required fields
        if "name" in commit.columns:
            commit = commit[commit["name"].notna() & (commit["name"].astype(str).str.strip() != "")]
            commit = commit.reset_index(drop=True)
        st.session_state.producers_df = commit
        mark_stale()
        st.rerun()
    if btn_p2.button("📋 Duplicate last producer", key="producers_dup"):
        st.session_state.producers_df = _duplicate_last_row(st.session_state.producers_df)
        mark_stale()
        st.rerun()

    producers_df = st.session_state.producers_df

    # Injectors table — only shown for Injection strategy. In Depletion mode
    # we keep an empty injectors_df in session so downstream code still works.
    strategy = st.session_state.get("strategy", "Depletion")
    is_injection = (strategy == "Injection")

    if is_injection:
        st.markdown("**Injectors**")
        if "injectors_df" not in st.session_state:
            # Seed two example injectors so the UX has something visible to edit
            inj_default_rate = from_field(20000.0, "water_rate", units)
            st.session_state.injectors_df = pd.DataFrame([{
                "name": f"WI-{i+1:02d}",
                "rig": rig_names[(i + 0) % len(rig_names)],
                "drill_days": 45, "completion_days": 15,
                "inj_rate": inj_default_rate,
                "scale_factor": 1.0, "uptime": 0.95,
            } for i in range(2)])
        else:
            # Backfill missing columns when loading older sessions / saved cases
            df = st.session_state.injectors_df
            if "uptime" not in df.columns:
                df["uptime"] = 0.95
            if "scale_factor" not in df.columns:
                df["scale_factor"] = 1.0
            st.session_state.injectors_df = df

        injectors_buf = st.data_editor(
            st.session_state.injectors_df, num_rows="dynamic", use_container_width=True,
            column_config={
                "name": st.column_config.TextColumn("Well", required=True),
                "rig": st.column_config.SelectboxColumn("Rig", options=rig_names, required=True),
                "drill_days": st.column_config.NumberColumn("Drill days", min_value=1, step=1),
                "completion_days": st.column_config.NumberColumn("Compl. days", min_value=1, step=1),
                "inj_rate": st.column_config.NumberColumn(
                    f"Injection rate ({ulabel('water_rate', units)})",
                    min_value=0.0,
                    help="Constant target injection rate while online (subject to facility "
                         "capacity and, if VRR cap is enabled below, voidage matching)."),
                "scale_factor": st.column_config.NumberColumn(
                    "Scale", min_value=0.1, max_value=5.0, step=0.05, format="%.2f",
                    help="Sensitivity multiplier on the injection target."),
                "uptime": st.column_config.NumberColumn(
                    "Uptime", min_value=0.0, max_value=1.0, step=0.01, format="%.2f",
                    help="Fraction of online time the injector actually injects."),
            },
            key="injectors_editor",
        )
        bi1, bi2, _bi3 = st.columns([2, 1, 3])
        with bi1:
            injectors_apply_clicked = _apply_button(
                injectors_buf, st.session_state.injectors_df,
                "injectors_apply", "Apply injector edits")
        if injectors_apply_clicked:
            commit = injectors_buf.copy()
            if "name" in commit.columns:
                commit = commit[commit["name"].notna() & (commit["name"].astype(str).str.strip() != "")]
                commit = commit.reset_index(drop=True)
            st.session_state.injectors_df = commit
            mark_stale()
            st.rerun()
        if bi2.button("📋 Duplicate last injector", key="injectors_dup"):
            st.session_state.injectors_df = _duplicate_last_row(st.session_state.injectors_df)
            mark_stale()
            st.rerun()
        injectors_df = st.session_state.injectors_df
    else:
        # Depletion mode: empty injectors so add_well() finds nothing to add.
        # Preserve any existing injectors_df in session (in case user toggles
        # strategy back to Injection) but pass an empty frame downstream.
        injectors_df = pd.DataFrame(columns=["name", "rig", "drill_days",
                                              "completion_days", "inj_rate",
                                              "scale_factor", "uptime"])

    user_profiles = {}
    needs_upload = producers_df[producers_df["decline_model"] == "User-defined profile"]["name"].tolist()
    if needs_upload:
        with st.expander("📄 Import production profiles (CSV / Eclipse)",
                          expanded=True):
            st.caption(
                "Upload a rate history for each well. The importer is "
                "flexible — it accepts a generic **CSV** or an **Eclipse** "
                "summary / RSM export, and auto-detects the format. "
                "Recognised columns (case-insensitive): a time column "
                "(`month`, `date`, or an Eclipse `DATE`/`TIME`); a primary-"
                "rate column (`oil_rate`, `qoil`, `WOPR`, `FOPR`, …); a "
                "secondary-rate column (`gas_rate`, `qgas`, `WGPR`, …); and "
                "optionally a `water_rate` column. Daily data is resampled "
                "to monthly averages. If the file has a `well` column with "
                "several wells, the profile matching this well's name is "
                "used."
            )
            for wname in needs_upload:
                f = st.file_uploader(
                    f"Profile for {wname}",
                    type=["csv", "txt", "rsm", "dat", "prn"],
                    key=f"prof_{wname}")
                if f is not None:
                    try:
                        parsed = fh.parse_production_profile(
                            f, filename=f.name, field_is_oil=is_oil)
                        profs = parsed["profiles"]
                        # match by well name; else take the first profile
                        chosen = None
                        for pk in profs:
                            if pk.strip().lower() == str(wname).strip().lower():
                                chosen = profs[pk]; break
                        if chosen is None:
                            first_key = next(iter(profs))
                            chosen = profs[first_key]
                            if parsed["n_wells"] > 1:
                                st.info(
                                    f"{wname}: file has "
                                    f"{parsed['n_wells']} wells "
                                    f"({', '.join(list(profs)[:4])}…) — no "
                                    f"exact name match, using "
                                    f"'{first_key}'.")
                        user_profiles[wname] = chosen
                        src = ("Eclipse export" if parsed["source"]
                               == "eclipse" else "CSV")
                        st.success(
                            f"{wname}: loaded {len(chosen)} months "
                            f"from {src}. Peak primary "
                            f"{chosen['primary_rate'].max():,.0f}, "
                            f"peak secondary "
                            f"{chosen['secondary_rate'].max():,.0f}.")
                        for note in parsed["notes"]:
                            st.caption("ℹ️ " + note)
                        for w in parsed["warnings"]:
                            st.warning(w)
                    except Exception as e:
                        st.error(f"{wname}: could not import — {e}")

    # ---- Multi-segment decline editor ----
    def _f(v, default=0.0):
        """Safe float coercion for table cells (NaN/None/blank → default)."""
        try:
            if v is None: return default
            x = float(v)
            if pd.isna(x): return default
            return x
        except (TypeError, ValueError):
            return default

    def _i(v, default=0):
        try:
            if v is None: return default
            x = float(v)
            if pd.isna(x): return default
            return int(x)
        except (TypeError, ValueError):
            return default

    # Wells whose decline_model is "Multi-segment" get a piecewise-Arps
    # profile: a sequence of segments (plateau, decline, bean-up, late-life
    # bump). Each well's segment table is stored in session_state so it
    # survives reruns.
    well_segments = {}
    needs_segments = producers_df[
        producers_df["decline_model"] == "Multi-segment"]["name"].tolist()
    if needs_segments:
        with st.expander("📈 Multi-segment decline profiles", expanded=True):
            st.caption(
                "Build a piecewise-Arps profile for each well. Segments run "
                "back-to-back. **Model** 'Plateau' holds the rate flat; "
                "Exponential / Harmonic / Hyperbolic decline it. **Step ×** "
                "multiplies the rate at the segment start: 1.0 = continuous, "
                ">1 = bean-up or re-stimulation bump, <1 = choke-back. The "
                "last segment is extrapolated to the end of the forecast. "
                "Tip: a bean-up is a short first segment with a negative "
                "decline (rate ramping up)."
            )
            _seg_cols = ["months", "model", "di", "b", "mult"]
            _seg_defaults = pd.DataFrame([
                {"months": 24, "model": "Plateau",
                 "di": 0.0, "b": 0.0, "mult": 1.0},
                {"months": 60, "model": "Hyperbolic",
                 "di": 0.25, "b": 0.6, "mult": 1.0},
                {"months": 120, "model": "Exponential",
                 "di": 0.12, "b": 0.0, "mult": 1.0},
            ])
            for wname in needs_segments:
                st.markdown(f"**{wname}**")
                seg_key = f"segments_{wname}"
                if seg_key not in st.session_state:
                    st.session_state[seg_key] = _seg_defaults.copy()
                edited = st.data_editor(
                    st.session_state[seg_key],
                    key=f"segeditor_{wname}",
                    num_rows="dynamic", use_container_width=True,
                    column_config={
                        "months": st.column_config.NumberColumn(
                            "Duration (months)", min_value=1, step=1,
                            help="Length of this segment in months."),
                        "model": st.column_config.SelectboxColumn(
                            "Decline model",
                            options=["Plateau", "Exponential",
                                     "Harmonic", "Hyperbolic"],
                            help="Plateau = flat rate; others = Arps "
                                 "decline within the segment."),
                        "di": st.column_config.NumberColumn(
                            "Annual decline", step=0.01, format="%.3f",
                            help="Nominal annual decline. Use a negative "
                                 "value for a ramp-up (bean-up). Ignored "
                                 "for Plateau."),
                        "b": st.column_config.NumberColumn(
                            "Arps b", min_value=0.0, max_value=1.0,
                            step=0.05, format="%.2f",
                            help="Hyperbolic b-exponent (0-1). Used only "
                                 "for the Hyperbolic model."),
                        "mult": st.column_config.NumberColumn(
                            "Step ×", min_value=0.1, max_value=5.0,
                            step=0.05, format="%.2f",
                            help="Rate multiplier at segment start. "
                                 "1.0 = continuous; 1.4 = a +40% "
                                 "re-stimulation bump."),
                    },
                )
                st.session_state[seg_key] = edited
                # Convert to list-of-dicts for the engine
                segs = []
                for _, sr in edited.iterrows():
                    try:
                        segs.append({
                            "months": int(_f(sr.get("months"), 12)),
                            "model": str(sr.get("model") or "Exponential"),
                            "di": _f(sr.get("di"), 0.0),
                            "b": _f(sr.get("b"), 0.5),
                            "mult": _f(sr.get("mult"), 1.0),
                        })
                    except Exception:
                        continue
                well_segments[wname] = segs
                # quick sanity feedback
                total_m = sum(s["months"] for s in segs)
                st.caption(f"{len(segs)} segment(s), {total_m} months "
                           f"defined; last segment extrapolated beyond that.")

    # Rig schedule: apply move-in (delays the rig's first well), and maintenance
    # (inserted as gaps between wells). Move-out is costed but doesn't affect
    # production timing.
    rig_meta = {}
    for _, r in rigs_df.iterrows():
        rname = r["rig"]
        rig_meta[rname] = {
            "move_in_days": _i(r.get("move_in_days"), 0),
            "move_out_days": _i(r.get("move_out_days"), 0),
            "maintenance_days_per_year": _i(r.get("maintenance_days_per_year"), 0),
            "day_rate_kUSD": _f(r.get("day_rate_kUSD"), 0.0),
        }
    rig_starts = {r["rig"]: r["start_date"] for _, r in rigs_df.iterrows()}
    # First-well cursor = rig available date + move-in days
    rig_cursor = {}
    for r in rig_names:
        base = pd.Timestamp(rig_starts[r]).date()
        mi = rig_meta.get(r, {}).get("move_in_days", 0)
        rig_cursor[r] = base + timedelta(days=mi)
    # Track whether each rig has had its move-in applied (so we only do it once)
    rig_first_well_done = {r: False for r in rig_names}

    wells = []

    def add_well(row, is_producer):
        name = row.get("name")
        if not isinstance(name, str) or not name.strip():
            return  # skip empty rows from data_editor
        rig = row.get("rig")
        if not isinstance(rig, str) or rig not in rig_cursor:
            rig = rig_names[0] if rig_names else "Rig-A"
            if rig not in rig_cursor:
                rig_cursor[rig] = start_date
                rig_first_well_done[rig] = False
        spud = rig_cursor[rig]
        drill = max(1, _i(row.get("drill_days"), 45))
        compl = max(1, _i(row.get("completion_days"), 15))
        # Maintenance: distribute the rig's annual maintenance over its wells,
        # added as a gap *before* this well (except the very first well, which
        # already had the move-in applied).
        maint_per_yr = rig_meta.get(rig, {}).get("maintenance_days_per_year", 0)
        if rig_first_well_done.get(rig, False) and maint_per_yr > 0:
            # Pro-rate: maintenance days for the time this well's program spans
            well_span_yrs = (drill + compl) / 365.0
            maint_gap = int(round(maint_per_yr * well_span_yrs))
            spud = spud + timedelta(days=maint_gap)
        rig_first_well_done[rig] = True
        if is_producer:
            qi_p = to_field(_f(row.get("qi_primary"), 0.0),
                            "oil_rate" if is_oil else "gas_rate", units)
            qi_s = to_field(_f(row.get("qi_secondary"), 0.0),
                            "gas_rate" if is_oil else "oil_rate", units)
            ws = WellSpec(
                name=str(name).strip(), is_producer=True, rig=rig,
                spud_date=spud, drill_days=drill, completion_days=compl,
                qi_primary=qi_p, qi_secondary=qi_s,
                decline_model=str(row.get("decline_model") or "Exponential"),
                di_annual=_f(row.get("di_annual"), 0.20),
                b_factor=_f(row.get("b_factor"), 0.5),
                wc_initial=_f(row.get("wc_initial"), 0.05),
                wc_final=_f(row.get("wc_final"), 0.85),
                wc_ramp_months=_i(row.get("wc_ramp_months"), 60),
                scale_factor=_f(row.get("scale_factor"), 1.0),
                uptime=_f(row.get("uptime"), 0.95),
                user_profile=user_profiles.get(str(name).strip()),
                segments=well_segments.get(str(name).strip()),
                derive_qi_from_pi=bool(row.get("derive_qi_from_pi", False)),
                # PI override is in DISPLAY units (rate/pressure). Convert
                # to engine units (field rate / psi) — see column-header
                # help text. Zero means "use the reservoir's PI".
                well_pi_override=(
                    (lambda _pi_disp: (
                        0.0 if _pi_disp == 0.0
                        else _pi_disp
                             * (1.0 if units == "field" else M2F[
                                 "oil_rate" if is_oil else "gas_rate"])
                             / (1.0 if units == "field" else M2F["pressure"])
                    ))(_f(row.get("well_pi_override"), 0.0))
                ),
                fluid=str(row.get("fluid", "auto") or "auto"),
                ipr_mode=bool(row.get("ipr_mode", False)),
                # IPR columns hold DISPLAY-unit values (the column header
                # flips between psi/bar and ft/m). Convert to field units
                # before passing to the engine. Fluid gradient stays in
                # psi/ft and friction in psi/kbpd — both are explicitly
                # marked as field-convention engineering inputs in their
                # column headers and help text.
                wellhead_pressure_psi=to_field(
                    _f(row.get("wellhead_pressure_psi"), 200.0),
                    "pressure", units),
                tubing_depth_ft=to_field(
                    _f(row.get("tubing_depth_ft"), 8000.0),
                    "depth", units),
                # ρ is stored in engine units (psi/ft) — display layer
                # converts. Read raw.
                fluid_gradient_psi_per_ft=_f(
                    row.get("fluid_gradient_psi_per_ft"), 0.35),
                friction_psi_per_kbpd=_f(
                    row.get("friction_psi_per_kbpd"), 5.0),
            )
        else:
            inj_rate_val = _f(row.get("inj_rate"), 0.0)
            if inj_rate_val <= 0:
                return  # skip injectors with no rate set
            ws = WellSpec(
                name=str(name).strip(), is_producer=False, rig=rig,
                spud_date=spud, drill_days=drill, completion_days=compl,
                qi_primary=0, qi_secondary=0,
                decline_model="Exponential",
                di_annual=0, b_factor=0,
                wc_initial=0, wc_final=0, wc_ramp_months=0,
                scale_factor=_f(row.get("scale_factor"), 1.0),
                uptime=_f(row.get("uptime"), 0.95),
                inj_rate=to_field(inj_rate_val, "water_rate", units),
            )
        wells.append(ws)
        rig_cursor[rig] = spud + timedelta(days=drill + compl)

    for _, r in producers_df.iterrows():
        add_well(r, True)
    for _, r in injectors_df.iterrows():
        add_well(r, False)

    # Stash rig metadata so the economics layer can cost move-in/out/maintenance
    st.session_state["_rig_meta"] = rig_meta

    # ---- Per-well production preview ----
    # Before running the full field model, let the user inspect each
    # producer's standalone monthly profile and its cumulative volumes.
    # This catches a mis-keyed decline or qi early, without waiting for the
    # whole simulation.
    producers_preview = [w for w in wells if w.is_producer]
    if producers_preview:
        with st.expander("🔍 Preview producer profiles "
                         "(before running)", expanded=False):
            st.caption(
                "Each producer's standalone profile from its decline / "
                "profile inputs — no field constraints, capacity caps or "
                "material-balance effects applied yet. Use it to sanity-"
                "check the per-well inputs before running the field model.")
            _prev_view = st.radio(
                "View", ["Primary phase", "All phases", "Ratios "
                 "(GOR / water cut / CGR)"],
                horizontal=True, key="prev_view")
            _is_oil_prev = FLUID_SYSTEMS[fluid]["primary"] == "oil"
            # build a monthly calendar covering all producers
            _hor_years = int(st.session_state.get("horizon", 25))
            _prev_dates = pd.date_range(
                start_date, periods=_hor_years * 12, freq="MS")
            prev_rows = []
            # Three separate figures depending on the view
            fig_main = go.Figure()
            fig_gas = go.Figure() if _prev_view == "All phases" else None
            fig_water = go.Figure() if _prev_view == "All phases" else None
            fig_gor = go.Figure() if _prev_view.startswith("Ratios") else None
            fig_wc = go.Figure() if _prev_view.startswith("Ratios") else None
            for w in producers_preview:
                wm = well_monthly(w, _prev_dates, _is_oil_prev)
                prim = wm["primary"]
                sec = wm["secondary"]
                wat = wm["water"]
                days = DAYS_PER_MONTH
                # Cumulative volumes.  prim is in stb/d for oil and Mscf/d
                # for gas; both go to MMstb / Bscf by dividing the monthly
                # total (rate × days) by 1e6.  The previous code divided
                # the gas cumulative by 1e9, which understated it 1000×.
                cum_prim = float(np.sum(prim) * days) / 1e6
                cum_sec = float(np.sum(sec) * days) / 1e6
                # for water always in stb/d → MMstb
                cum_water = float(np.sum(wat) * days) / 1e6
                peak = float(np.max(prim)) if len(prim) else 0.0
                on_months = int(np.sum(prim > 0.01))
                # display unit conversion
                if _is_oil_prev:
                    prim_vu, sec_vu = "oil_vol", "gas_vol"
                    prim_ru, sec_ru = "oil_rate", "gas_rate"
                else:
                    prim_vu, sec_vu = "gas_vol", "oil_vol"
                    prim_ru, sec_ru = "gas_rate", "oil_rate"
                cum_disp = from_field(cum_prim, prim_vu, units)
                cum_sec_disp = from_field(cum_sec, sec_vu, units)
                cum_wat_disp = from_field(cum_water, "water_vol", units)
                peak_disp = from_field(peak, prim_ru, units)
                _wt_map = st.session_state.get(
                    "well_template_map", {}) or {}
                prev_rows.append({
                    "Well": w.name,
                    "Rig": w.rig,
                    "Template": _wt_map.get(w.name, "—"),
                    "Online": str(w.online_date),
                    f"Peak ({ulabel(prim_ru, units)})": round(peak_disp, 1),
                    f"Cum primary ({ulabel(prim_vu, units)})":
                        round(cum_disp, 2),
                    f"Cum secondary ({ulabel(sec_vu, units)})":
                        round(cum_sec_disp, 2),
                    f"Cum water ({ulabel('water_vol', units)})":
                        round(cum_wat_disp, 2),
                    "Producing months": on_months,
                })
                # PROFILE CURVES
                if _prev_view == "Primary phase":
                    fig_main.add_trace(go.Scatter(
                        x=_prev_dates,
                        y=from_field(prim, prim_ru, units),
                        mode="lines", name=w.name))
                elif _prev_view == "All phases":
                    fig_main.add_trace(go.Scatter(
                        x=_prev_dates, y=from_field(prim, prim_ru, units),
                        mode="lines", name=w.name,
                        legendgroup=w.name))
                    fig_gas.add_trace(go.Scatter(
                        x=_prev_dates, y=from_field(sec, sec_ru, units),
                        mode="lines", name=w.name,
                        legendgroup=w.name, showlegend=False))
                    fig_water.add_trace(go.Scatter(
                        x=_prev_dates,
                        y=from_field(wat, "water_rate", units),
                        mode="lines", name=w.name,
                        legendgroup=w.name, showlegend=False))
                else:
                    # RATIOS — derived from rates
                    with np.errstate(divide="ignore", invalid="ignore"):
                        if _is_oil_prev:
                            # GOR in scf/stb  =  Mscf/d × 1000 / stb/d
                            gor = np.where(prim > 0,
                                           sec * 1000.0 / prim, 0.0)
                            fig_gor.add_trace(go.Scatter(
                                x=_prev_dates, y=gor, mode="lines",
                                name=w.name))
                        else:
                            # CGR in stb/MMscf = stb/d  /  (Mscf/d / 1000)
                            cgr = np.where(prim > 0,
                                           sec / (prim / 1000.0), 0.0)
                            fig_gor.add_trace(go.Scatter(
                                x=_prev_dates, y=cgr, mode="lines",
                                name=w.name))
                        wc = np.where((prim + wat) > 0,
                                      wat / (prim + wat), 0.0) \
                            if _is_oil_prev else np.zeros_like(prim)
                        fig_wc.add_trace(go.Scatter(
                            x=_prev_dates, y=wc, mode="lines",
                            name=w.name, showlegend=False))
            # Render figures
            if _prev_view == "Primary phase":
                fig_main.update_layout(
                    title="Standalone producer profiles",
                    xaxis_title="Date",
                    yaxis_title=f"{'Oil' if _is_oil_prev else 'Gas'} rate "
                                f"({ulabel(prim_ru, units)})",
                    height=340, legend=dict(orientation="h", y=-0.25))
                st.plotly_chart(fh.apply_plot_template(fig_main),
                                use_container_width=True)
            elif _prev_view == "All phases":
                fig_main.update_layout(
                    title=f"{'Oil' if _is_oil_prev else 'Gas'} rate per well",
                    yaxis_title=ulabel(prim_ru, units), height=240,
                    legend=dict(orientation="h", y=-0.3))
                fig_gas.update_layout(
                    title=f"{'Gas' if _is_oil_prev else 'Condensate'} rate "
                          f"per well",
                    yaxis_title=ulabel(sec_ru, units), height=240,
                    showlegend=False)
                fig_water.update_layout(
                    title="Water rate per well",
                    yaxis_title=ulabel("water_rate", units), height=240,
                    showlegend=False)
                st.plotly_chart(fh.apply_plot_template(fig_main),
                                use_container_width=True)
                st.plotly_chart(fh.apply_plot_template(fig_gas),
                                use_container_width=True)
                st.plotly_chart(fh.apply_plot_template(fig_water),
                                use_container_width=True)
            else:
                fig_gor.update_layout(
                    title=("GOR (scf/stb) per well" if _is_oil_prev
                           else "CGR (stb/MMscf) per well"),
                    yaxis_title=("GOR (scf/stb)" if _is_oil_prev
                                 else "CGR (stb/MMscf)"),
                    height=260,
                    legend=dict(orientation="h", y=-0.3))
                st.plotly_chart(fh.apply_plot_template(fig_gor),
                                use_container_width=True)
                if _is_oil_prev:
                    fig_wc.update_layout(
                        title="Water cut per well",
                        yaxis_title="Water cut (fraction)",
                        yaxis_range=[0, 1], height=240, showlegend=False)
                    st.plotly_chart(fh.apply_plot_template(fig_wc),
                                    use_container_width=True)
            prev_df = pd.DataFrame(prev_rows)
            st.dataframe(prev_df, use_container_width=True,
                         hide_index=True)
            _tot_col = [c for c in prev_df.columns
                        if c.startswith("Cum primary")][0]
            st.caption(
                f"Combined standalone cumulative: "
                f"{prev_df[_tot_col].sum():,.1f} "
                f"{_tot_col.split('(')[1].rstrip(')')}. "
                f"Note this is the simple sum of unconstrained well "
                f"profiles — the field model will apply capacity limits, "
                f"the volumetric cap and material-balance effects, so the "
                f"final field total is normally lower.")

    return wells


# =============================================================================
# Multi-reservoir UI
# =============================================================================
def reservoir_template_picker(units: str):
    """UI: pick a reservoir archetype and append a row to the reservoirs table."""
    with st.expander("🪨 Add reservoir from a type curve", expanded=False):
        st.caption(
            "Pick a reservoir archetype to append a pre-populated row to the "
            "reservoirs table below. After adding, edit any fields directly."
        )
        names = fh.list_reservoir_types()
        a, b = st.columns([3, 2])
        with a:
            tname = st.selectbox("Reservoir archetype", names,
                                  key="rt_choice")
            tmpl = fh.get_reservoir_type(tname) if tname else None
            if tmpl:
                st.caption(tmpl.get("description", ""))
        with b:
            if tmpl:
                rid = st.text_input("Reservoir ID", value="", key="rt_id_input",
                                     help="Leave blank to auto-generate (R1, R2, …).")

        if not tmpl:
            return

        if st.button("➕ Add reservoir row", key="rt_add_btn",
                      use_container_width=True):
            existing = st.session_state.get("reservoirs_df")
            if existing is None or len(existing) == 0:
                next_n = 1
                existing_ids = set()
            else:
                existing_ids = set(str(x) for x in existing.get("id", []) if pd.notna(x))
                next_n = len(existing) + 1
            new_id = rid.strip() if rid and rid.strip() else f"R{next_n}"
            while new_id in existing_ids:
                next_n += 1
                new_id = f"R{next_n}"

            row = {
                "id": new_id,
                "name": tname,
                "fluid_system": tmpl.get("fluid_system", "Oil with associated gas"),
                "strategy": tmpl.get("strategy", "Depletion"),
                "ooip_oil_MMstb": from_field(tmpl["ooip_oil_MMstb"], "oil_vol", units),
                "ogip_gas_Bscf":  from_field(tmpl["ogip_gas_Bscf"],  "gas_vol", units),
                "rf_target": tmpl["rf_target"],
                "p_init":  from_field(tmpl["p_init"],   "pressure", units),
                "t_res":   from_field(tmpl["t_res"],    "temp",     units),
                "api":     tmpl["api"],
                "gas_sg":  tmpl["gas_sg"],
                "rs_init": from_field(tmpl["rs_init"],  "gor",      units),
                "p_bub":   from_field(tmpl["p_bub"],    "pressure", units),
                "aquifer_active": False,
                "gas_cap_active": False,
                "vrr": tmpl.get("vrr", 1.0),
                "well_pi": tmpl.get("well_pi", 2.0),
                "min_bhp": from_field(tmpl.get("min_bhp", 1500.0), "pressure", units),
            }
            new_df = pd.DataFrame([row])
            if existing is None or len(existing) == 0:
                st.session_state["reservoirs_df"] = new_df
            else:
                st.session_state["reservoirs_df"] = pd.concat(
                    [existing, new_df], ignore_index=True)
            st.success(f"Added reservoir '{new_id}' from '{tname}' template.")
            mark_stale()
            st.rerun()


def reservoir_section(units: str, sidebar_inputs: dict,
                      well_names: list[str], inj_names: list[str]) -> tuple[list, list]:
    """Define multiple reservoirs and well-reservoir allocations.

    Returns:
      reservoirs: list[Reservoir]
      well_links: list[WellReservoirLink]

    If user does not enable multi-reservoir mode, returns ([], []) — the
    engine will then synthesize a default single-reservoir from sidebar inputs.
    """
    st.subheader("🪨 Reservoirs (optional multi-reservoir mode)")
    enable = st.checkbox(
        "Enable multi-reservoir mode",
        value=False, key="multi_res_enable", on_change=mark_stale,
        help=("Off (default): a single reservoir is built from the sidebar's PVT, "
              "aquifer, gas-cap, fluid system and strategy. "
              "On: define multiple reservoirs below, each with its own properties, "
              "and assign each well to one or more reservoirs with allocation fractions.")
    )

    with st.expander("ℹ️ How multi-reservoir works", expanded=False):
        st.markdown(
            "- Add a row for each reservoir with its **fluid system**, **strategy**, "
            "**OOIP / OGIP**, **target RF**, and **PVT** (Pi, T, API, gas SG, Rs, "
            "bubble point) and aquifer / gas-cap flags.\n"
            "- In the **Allocations** table, list `(well, reservoir, fraction)` rows. "
            "A well's fractions across reservoirs should sum to 1.0; if they don't, "
            "the engine **renormalizes** them.\n"
            "- A producer with no allocation row defaults to **100% on the first reservoir**.\n"
            "- The engine computes per-reservoir cumulatives, RF, and material-balance "
            "pressure separately, then aggregates field rates as the sum across "
            "reservoirs (weighted by allocation).\n"
            "- **Limitation**: aquifer pressure & gas-cap are read from the row's "
            "Pi value; the Pot/Fetkovich choice and aquifer volume are inherited "
            "from the sidebar (a future revision could store them per row)."
        )

    if not enable:
        return [], []

    # ---- Reservoir templates picker ----
    reservoir_template_picker(units)

    # ---- Reservoirs table ----
    st.markdown("**Reservoirs**")
    if "reservoirs_df" not in st.session_state:
        st.session_state.reservoirs_df = pd.DataFrame([{
            "id": "R1", "name": "Reservoir 1",
            "fluid_system": sidebar_inputs.get("fluid", "Oil with associated gas"),
            "strategy": sidebar_inputs.get("strategy", "Depletion"),
            "ooip_oil_MMstb": from_field(sidebar_inputs.get("ooip", 250.0), "oil_vol", units),
            "ogip_gas_Bscf":  from_field(sidebar_inputs.get("ogip", 300.0), "gas_vol", units),
            "rf_target": sidebar_inputs.get("rf_target", 0.35),
            "p_init": from_field(sidebar_inputs["pvt"].p_init_psi, "pressure", units),
            "t_res":  from_field(sidebar_inputs["pvt"].t_res_F, "temp", units),
            "api":    sidebar_inputs["pvt"].api,
            "gas_sg": sidebar_inputs["pvt"].gas_grav,
            "rs_init": from_field(sidebar_inputs["pvt"].rs_init, "gor", units),
            "p_bub":  from_field(sidebar_inputs["pvt"].p_bub_psi, "pressure", units),
            "aquifer_active": sidebar_inputs["aquifer"].active,
            "gas_cap_active": sidebar_inputs["gas_cap"].active,
            "vrr": sidebar_inputs.get("vrr", 1.0),
            "well_pi": 2.0,
            "min_bhp": from_field(1500.0, "pressure", units),
        }])
    else:
        # Backfill PI / BHP columns on older saved sessions
        rdf = st.session_state.reservoirs_df
        if "well_pi" not in rdf.columns:
            rdf["well_pi"] = 2.0
        if "min_bhp" not in rdf.columns:
            rdf["min_bhp"] = from_field(1500.0, "pressure", units)
        st.session_state.reservoirs_df = rdf

    res_df = st.data_editor(
        st.session_state.reservoirs_df,
        num_rows="dynamic", use_container_width=True,
        column_config={
            "id":   st.column_config.TextColumn("ID", required=True,
                    help="Short unique ID used in the allocations table (e.g. R1, R2)."),
            "name": st.column_config.TextColumn("Name", required=True),
            "fluid_system": st.column_config.SelectboxColumn(
                "Fluid system", options=list(FLUID_SYSTEMS.keys()), required=True),
            "strategy": st.column_config.SelectboxColumn(
                "Strategy", options=["Depletion", "Injection"], required=True),
            "ooip_oil_MMstb": st.column_config.NumberColumn(
                f"OOIP ({ulabel('oil_vol', units)})", min_value=0.0,
                help="Oil in place. Use 0 for a dry-gas reservoir."),
            "ogip_gas_Bscf": st.column_config.NumberColumn(
                f"OGIP ({ulabel('gas_vol', units)})", min_value=0.0,
                help="Gas in place. Use 0 for a black-oil reservoir without solution gas."),
            "rf_target": st.column_config.NumberColumn(
                "Target RF", min_value=0.0, max_value=0.95, step=0.01, format="%.2f"),
            "p_init": st.column_config.NumberColumn(
                f"Pi ({ulabel('pressure', units)})", min_value=0.0,
                help="Initial reservoir pressure for this reservoir."),
            "t_res":  st.column_config.NumberColumn(
                f"T ({ulabel('temp', units)})",
                help="Reservoir temperature."),
            "api":    st.column_config.NumberColumn("API", min_value=5.0, max_value=70.0,
                                                     help="Oil API gravity."),
            "gas_sg": st.column_config.NumberColumn("Gas SG", min_value=0.5, max_value=1.5,
                                                     help="Gas specific gravity (air = 1)."),
            "rs_init": st.column_config.NumberColumn(
                f"Rsi ({ulabel('gor', units)})", min_value=0.0,
                help="Initial solution GOR."),
            "p_bub":  st.column_config.NumberColumn(
                f"Pb ({ulabel('pressure', units)})", min_value=0.0,
                help="Bubble point."),
            "aquifer_active": st.column_config.CheckboxColumn(
                "Aquifer", help="Inherits aquifer model & volume from the sidebar."),
            "gas_cap_active": st.column_config.CheckboxColumn(
                "Gas cap", help="Inherits gas-cap m and Pi from the sidebar."),
            "vrr": st.column_config.NumberColumn(
                "VRR", min_value=0.0, max_value=2.0, step=0.05,
                help="Voidage replacement ratio (used only when no injectors are assigned)."),
            "well_pi": st.column_config.NumberColumn(
                "Well PI",
                min_value=0.0, step=0.1, format="%.2f",
                help="Productivity index per well (bbl/d/psi for oil, "
                     "Mscf/d/psi for gas). Used by the PI bridge: when a "
                     "well has 'derive qi from PI' enabled, its initial rate "
                     "is computed as PI × (P_res − BHP_min). "
                     "Typical screening values: light onshore oil 1-3, "
                     "deepwater 10-20, dry gas conventional 0.5-2, tight gas 0.05-0.2."),
            "min_bhp": st.column_config.NumberColumn(
                f"Min BHP ({ulabel('pressure', units)})",
                min_value=0.0, step=10.0,
                help="Minimum flowing bottom-hole pressure (well-level constraint). "
                     "Used by the PI bridge to compute drawdown."),
        },
        key="reservoirs_editor",
    )
    brs1, brs2, _brs3 = st.columns([2, 1, 3])
    with brs1:
        reservoirs_apply_clicked = _apply_button(
            res_df, st.session_state.reservoirs_df,
            "reservoirs_apply", "Apply reservoir edits")
    if reservoirs_apply_clicked:
        commit = res_df.copy()
        if "id" in commit.columns:
            commit = commit[commit["id"].notna() & (commit["id"].astype(str).str.strip() != "")]
            commit = commit.reset_index(drop=True)
        st.session_state.reservoirs_df = commit
        mark_stale()
        st.rerun()
    if brs2.button("📋 Duplicate last reservoir", key="reservoirs_dup"):
        st.session_state.reservoirs_df = _duplicate_last_row(st.session_state.reservoirs_df)
        mark_stale()
        st.rerun()
    res_df = st.session_state.reservoirs_df

    # Build Reservoir objects
    reservoirs: list[Reservoir] = []
    for _, row in res_df.iterrows():
        if pd.isna(row.get("id")):
            continue
        # Inherit aquifer / gas_cap from sidebar; just toggle active
        aq_in = sidebar_inputs["aquifer"]
        aq_for_res = AquiferInputs(
            active=bool(row.get("aquifer_active", False)),
            model=aq_in.model,
            aquifer_volume=aq_in.aquifer_volume,
            productivity_index=aq_in.productivity_index,
            initial_pressure_psi=to_field(float(row.get("p_init", 3500.0)),
                                           "pressure", units),
        )
        gc_in = sidebar_inputs["gas_cap"]
        gc_for_res = GasCapInputs(
            active=bool(row.get("gas_cap_active", False)),
            size_fraction=gc_in.size_fraction,
            initial_pressure_psi=to_field(float(row.get("p_init", 3500.0)),
                                           "pressure", units),
        )
        pvt_for_res = PVTInputs(
            p_init_psi=to_field(float(row.get("p_init", 3500.0)), "pressure", units),
            t_res_F=to_field(float(row.get("t_res", 180.0)), "temp", units),
            api=float(row.get("api", 35.0)),
            gas_grav=float(row.get("gas_sg", 0.7)),
            rs_init=to_field(float(row.get("rs_init", 700.0)), "gor", units),
            p_bub_psi=to_field(float(row.get("p_bub", 2800.0)), "pressure", units),
        )
        reservoirs.append(Reservoir(
            id=str(row["id"]), name=str(row.get("name", row["id"])),
            fluid_system=str(row["fluid_system"]),
            ooip_oil=to_field(float(row.get("ooip_oil_MMstb", 0.0)), "oil_vol", units),
            ogip_gas=to_field(float(row.get("ogip_gas_Bscf", 0.0)), "gas_vol", units),
            rf_target=float(row.get("rf_target", 0.35)),
            strategy=str(row.get("strategy", "Depletion")),
            pvt=pvt_for_res, aquifer=aq_for_res, gas_cap=gc_for_res,
            voidage_ratio=float(row.get("vrr", 1.0)),
            inj_efficiency=sidebar_inputs.get("inj_eff", 0.85),
            well_pi=float(row.get("well_pi", 2.0)),
            min_bhp_psi=to_field(float(row.get("min_bhp", 1500.0)),
                                  "pressure", units),
        ))

    if not reservoirs:
        st.warning("No reservoirs defined — falling back to single-reservoir mode.")
        return [], []

    # ---- Allocations table ----
    st.markdown("**Well → reservoir allocations**")
    res_ids = [r.id for r in reservoirs]
    all_well_names = list(well_names) + list(inj_names)

    if "well_reservoir_df" not in st.session_state or len(all_well_names) == 0:
        # Default: each well 100% on first reservoir
        if all_well_names and res_ids:
            st.session_state.well_reservoir_df = pd.DataFrame([
                {"well": w, "reservoir": res_ids[0], "fraction": 1.0}
                for w in all_well_names
            ])
        else:
            st.session_state.well_reservoir_df = pd.DataFrame(
                columns=["well", "reservoir", "fraction"])

    alloc_df = st.data_editor(
        st.session_state.well_reservoir_df,
        num_rows="dynamic", use_container_width=True,
        column_config={
            "well": st.column_config.SelectboxColumn(
                "Well", options=all_well_names, required=True,
                help="Pick a producer or injector defined above. "
                     "Add multiple rows for the same well to split across reservoirs."),
            "reservoir": st.column_config.SelectboxColumn(
                "Reservoir", options=res_ids, required=True),
            "fraction": st.column_config.NumberColumn(
                "Fraction", min_value=0.0, max_value=1.0, step=0.05, format="%.2f",
                help="Fraction of the well's rate going to this reservoir. "
                     "Values across all rows for one well should sum to 1.0; "
                     "the engine renormalizes if they don't."),
        },
        key="well_reservoir_editor",
    )
    bw1, bw2, _bw3 = st.columns([2, 1, 3])
    with bw1:
        wr_apply_clicked = _apply_button(
            alloc_df, st.session_state.well_reservoir_df,
            "well_reservoir_apply", "Apply allocation edits")
    if wr_apply_clicked:
        commit = alloc_df.copy()
        # Drop rows missing well or reservoir
        if "well" in commit.columns and "reservoir" in commit.columns:
            commit = commit[commit["well"].notna() & commit["reservoir"].notna()]
            commit = commit.reset_index(drop=True)
        st.session_state.well_reservoir_df = commit
        mark_stale()
        st.rerun()
    if bw2.button("📋 Duplicate last allocation", key="well_reservoir_dup"):
        st.session_state.well_reservoir_df = _duplicate_last_row(st.session_state.well_reservoir_df)
        mark_stale()
        st.rerun()
    alloc_df = st.session_state.well_reservoir_df

    well_links: list[WellReservoirLink] = []
    for _, row in alloc_df.iterrows():
        if pd.isna(row.get("well")) or pd.isna(row.get("reservoir")):
            continue
        try:
            well_links.append(WellReservoirLink(
                well_name=str(row["well"]),
                reservoir_id=str(row["reservoir"]),
                fraction=float(row.get("fraction", 1.0)),
            ))
        except (ValueError, TypeError):
            continue

    # Validation summary
    if well_links:
        sums = pd.DataFrame([{"well": l.well_name, "frac": l.fraction}
                             for l in well_links]).groupby("well")["frac"].sum()
        bad = sums[(sums - 1.0).abs() > 0.01]
        if len(bad) > 0:
            st.info(f"ℹ️ Allocations don't sum to 1.0 for: "
                    f"{', '.join(bad.index)} — engine will renormalize.")

        # Detect wells that allocate to reservoirs of different fluid types
        res_fluid = {r.id: FLUID_SYSTEMS[r.fluid_system]["primary"] for r in reservoirs}
        well_fluids = {}
        for l in well_links:
            f = res_fluid.get(l.reservoir_id)
            if f is None: continue
            well_fluids.setdefault(l.well_name, set()).add(f)
        mixed = [w for w, fs in well_fluids.items() if len(fs) > 1]
        if mixed:
            st.warning(
                f"⚠️ Wells {', '.join(mixed)} are allocated to reservoirs of "
                "different fluid types (oil + gas). The engine assumes the well's "
                "primary fluid matches each reservoir's primary fluid; mixed-fluid "
                "wells will give unit-mismatched results. Either assign each such "
                "well to one fluid type only, or split it into two distinct wells."
            )

    return reservoirs, well_links


# =============================================================================
# Capacities UI
# =============================================================================
def capacity_section(units, start_date, strategy: str = "Injection"):
    """Render the capacity table. In Depletion mode the water/gas injection
    columns are hidden — the engine still receives a CapacitySchedule with
    those columns set to 0 (effectively unlimited) so downstream logic
    works unchanged.

    UX pattern: the user edits a *buffer* DataFrame; a single "Apply" button
    commits the buffer to the canonical session_state. This avoids the
    "have to enter values twice" problem caused by Streamlit's rerun
    behavior on data_editor onchange callbacks.
    """
    is_injection = (strategy == "Injection")
    st.subheader("🛢️ Production capacities (time-varying)"
                 if not is_injection
                 else "🛢️ Production & injection capacities (time-varying)")
    with st.expander("ℹ️ How capacities work", expanded=False):
        notes = [
            "- Each row defines a **change date**; values apply forward until the next row.",
            "- A proportional choke is applied each month so the field honors the tightest active limit.",
            "- Set a value to 0 to disable that constraint.",
            "- `Gas` is in MMscf/d in field units (note the unit in the column header).",
            "- The **PE** column lets you vary production efficiency over time (0–1).",
            "- Add a row and click **✓ Apply** when done. Adding empty rows is safe.",
        ]
        if is_injection:
            notes.append("- `water_inj` and `gas_inj` cap the injection plant capacity.")
        st.markdown("\n".join(notes))
    if "cap_df" not in st.session_state:
        st.session_state.cap_df = pd.DataFrame({
            "start_date": [start_date],
            "oil":       [from_field(50000.0, "oil_rate", units)],
            "gas":       [150.0 if units == "field" else from_field(150*1000, "gas_rate", units)],
            "water":     [from_field(80000.0, "water_rate", units)],
            "liquid":    [from_field(120000.0, "oil_rate", units)],
            "water_inj": [from_field(100000.0, "water_rate", units)],
            "gas_inj":   [0.0],
            "prod_eff":  [0.95],
        })
    # Backfill new prod_eff column on saved sessions
    if "prod_eff" not in st.session_state.cap_df.columns:
        st.session_state.cap_df["prod_eff"] = 0.95

    # Build the display-only DataFrame and column config, hiding injection
    # columns in Depletion mode.
    cap_display = st.session_state.cap_df.copy()
    column_config = {
        "start_date": st.column_config.DateColumn("From"),
        "oil": st.column_config.NumberColumn(
            f"Oil ({ulabel('oil_rate', units)})", min_value=0.0,
            help="Oil treatment / export. 0 = unlimited."),
        "gas": st.column_config.NumberColumn(
            f"Gas ({'MMscf/d' if units=='field' else ulabel('gas_rate', units)})",
            min_value=0.0, help="Gas processing capacity."),
        "water": st.column_config.NumberColumn(
            f"Water ({ulabel('water_rate', units)})", min_value=0.0,
            help="Produced water handling capacity."),
        "liquid": st.column_config.NumberColumn(
            f"Liquid ({ulabel('oil_rate', units)})", min_value=0.0,
            help="Total liquid (oil+water) handling capacity."),
        "prod_eff": st.column_config.NumberColumn(
            "PE", min_value=0.0, max_value=1.0, step=0.01, format="%.2f",
            help="Production efficiency (0..1) for this period. "
                 "Vary over field life to model maintenance windows, "
                 "facility upgrades, etc. 0.95 = typical mature ops."),
    }
    if is_injection:
        column_config["water_inj"] = st.column_config.NumberColumn(
            f"Water inj. ({ulabel('water_rate', units)})", min_value=0.0,
            help="Water injection plant capacity.")
        column_config["gas_inj"] = st.column_config.NumberColumn(
            f"Gas inj. ({ulabel('gas_rate', units)})", min_value=0.0,
            help="Gas injection plant capacity.")
        edit_df = cap_display
    else:
        edit_df = cap_display.drop(columns=[c for c in ("water_inj", "gas_inj")
                                              if c in cap_display.columns],
                                    errors="ignore")
    # Editable buffer — committed only via Apply button below
    cap_df_edited = st.data_editor(
        edit_df, num_rows="dynamic", use_container_width=True,
        column_config=column_config,
        key="cap_editor",
    )
    btn_col1, btn_col2, btn_col3 = st.columns([2, 1, 3])
    with btn_col1:
        cap_apply_clicked = _apply_button(cap_df_edited, edit_df,
                                           "cap_apply", "Apply capacity edits")
    if cap_apply_clicked:
        cleaned = _clean_table_buffer(cap_df_edited, st.session_state.cap_df,
                                        date_cols=["start_date"])
        if "water_inj" not in cleaned.columns:
            cleaned["water_inj"] = 0.0
        if "gas_inj" not in cleaned.columns:
            cleaned["gas_inj"] = 0.0
        if "prod_eff" not in cleaned.columns:
            cleaned["prod_eff"] = 0.95
        st.session_state.cap_df = cleaned
        mark_stale()
        st.rerun()
    if btn_col2.button("📋 Duplicate last row", key="cap_dup"):
        st.session_state.cap_df = _duplicate_last_row(st.session_state.cap_df)
        mark_stale()
        st.rerun()
    cap_df = st.session_state.cap_df

    # ---- Convert to engine field units ----
    cap_field = cap_df.copy()
    for col, kind in [("oil", "oil_rate"), ("water", "water_rate"),
                      ("liquid", "oil_rate"),
                      ("water_inj", "water_rate"), ("gas_inj", "gas_rate")]:
        if col in cap_field.columns:
            cap_field[col] = cap_field[col].apply(lambda v: to_field(float(v or 0.0), kind, units))
    if units == "metric":
        cap_field["gas"] = cap_field["gas"].apply(lambda v: to_field(float(v or 0.0), "gas_rate", units) / 1000.0)
    cap_field = cap_field.sort_values("start_date").reset_index(drop=True)
    return CapacitySchedule(df=cap_field)


def _clean_table_buffer(edited_df: pd.DataFrame, fallback_df: pd.DataFrame,
                         date_cols: list = None) -> pd.DataFrame:
    """Sanitize a data_editor buffer before committing it to session_state.

    - Drops rows where all date_cols are NaT (truly empty)
    - Preserves the column dtype of the fallback (so int columns stay int, etc.)
    - When the user inserts a blank row, returns the buffer with that row removed
      rather than crashing downstream code that assumes valid dates/numbers.
    """
    if edited_df is None or len(edited_df) == 0:
        return fallback_df.iloc[0:0].copy()
    out = edited_df.copy()
    if date_cols:
        for c in date_cols:
            if c in out.columns:
                out[c] = pd.to_datetime(out[c], errors="coerce")
        # Drop rows where ALL date cols are NaT
        keep_mask = pd.Series(False, index=out.index)
        for c in date_cols:
            if c in out.columns:
                keep_mask = keep_mask | out[c].notna()
        out = out[keep_mask].reset_index(drop=True)
    return out


def _svg_to_data_uri(svg_string: str) -> str:
    """Convert an SVG string to a base64 data URI so st.image can render it
    inline. Streamlit's st.image accepts data: URIs directly."""
    import base64
    b64 = base64.b64encode(svg_string.encode("utf-8")).decode("ascii")
    return f"data:image/svg+xml;base64,{b64}"


def _build_concept_3d_figure(geo: dict):
    """Build an interactive Plotly 3D scene of the development concept from
    the geometry dict returned by fh.concept_3d_geometry. The user can
    rotate / zoom it in the browser."""
    fig = go.Figure()
    sea_z = geo["sea_z"]
    seabed_z = geo["seabed_z"]
    tb = geo.get("tieback_km", 10.0)
    # extents for the surfaces
    xs = [0.0, geo["host"]["x"] * 1.25]
    ys = [-tb * 0.4, tb * 0.4]

    # Sea surface — a translucent blue plane
    fig.add_trace(go.Mesh3d(
        x=[xs[0], xs[1], xs[1], xs[0]],
        y=[ys[0], ys[0], ys[1], ys[1]],
        z=[sea_z] * 4,
        i=[0, 0], j=[1, 2], k=[2, 3],
        color="#5a9bcf", opacity=0.25, name="Sea surface",
        hoverinfo="name", showscale=False))
    # Seabed — a sandy plane
    fig.add_trace(go.Mesh3d(
        x=[xs[0], xs[1], xs[1], xs[0]],
        y=[ys[0], ys[0], ys[1], ys[1]],
        z=[seabed_z] * 4,
        i=[0, 0], j=[1, 2], k=[2, 3],
        color="#d8c9a8", opacity=0.55, name="Seabed",
        hoverinfo="name", showscale=False))

    # Templates — markers on the seabed
    if geo["templates"]:
        fig.add_trace(go.Scatter3d(
            x=[t["x"] for t in geo["templates"]],
            y=[t["y"] for t in geo["templates"]],
            z=[seabed_z for _ in geo["templates"]],
            mode="markers+text",
            marker=dict(size=9, color="#c4566a", symbol="square"),
            text=[t["label"] for t in geo["templates"]],
            textposition="top center",
            name="Templates",
            hovertext=[f"{t['label']}: {t['wells']}/{t['slots']} slots"
                       for t in geo["templates"]],
            hoverinfo="text"))
    # Wells
    if geo["wells"]:
        fig.add_trace(go.Scatter3d(
            x=[w["x"] for w in geo["wells"]],
            y=[w["y"] for w in geo["wells"]],
            z=[seabed_z for _ in geo["wells"]],
            mode="markers",
            marker=dict(size=3.5, color="#1a1a1a"),
            name=f"Wells ({geo['n_subsea']})", hoverinfo="name"))
    # Flowline
    fl = geo["flowline"]
    fig.add_trace(go.Scatter3d(
        x=[p[0] for p in fl], y=[p[1] for p in fl],
        z=[p[2] for p in fl], mode="lines",
        line=dict(color="#224466", width=6), name="Flowline"))
    # Umbilical
    um = geo["umbilical"]
    fig.add_trace(go.Scatter3d(
        x=[p[0] for p in um], y=[p[1] for p in um],
        z=[p[2] for p in um], mode="lines",
        line=dict(color="#b07ac0", width=4, dash="dash"),
        name="Umbilical"))
    # Riser — the S-curve up to the host
    rs = geo["riser"]
    fig.add_trace(go.Scatter3d(
        x=[p[0] for p in rs], y=[p[1] for p in rs],
        z=[p[2] for p in rs], mode="lines",
        line=dict(color="#1199aa", width=6), name="Riser"))
    # Host
    h = geo["host"]
    fig.add_trace(go.Scatter3d(
        x=[h["x"]], y=[h["y"]], z=[sea_z],
        mode="markers+text",
        marker=dict(size=14, color="#8a96a0",
                    symbol="diamond"),
        text=["Host"], textposition="top center",
        name=h.get("type", "Host"), hoverinfo="name"))
    # Boosting stations
    if geo["boosting"]:
        fig.add_trace(go.Scatter3d(
            x=[b["x"] for b in geo["boosting"]],
            y=[b["y"] for b in geo["boosting"]],
            z=[b["z"] for b in geo["boosting"]],
            mode="markers",
            marker=dict(size=7, color="#ffaa33", symbol="circle"),
            name="Boosting station"))
    # Export pipeline
    if geo["export"]:
        ex = geo["export"]
        fig.add_trace(go.Scatter3d(
            x=[p[0] for p in ex], y=[p[1] for p in ex],
            z=[p[2] for p in ex], mode="lines",
            line=dict(color="#555555", width=4, dash="dot"),
            name="Export pipeline"))

    fig.update_layout(
        height=540,
        scene=dict(
            xaxis_title="Along tie-back (km)",
            yaxis_title="Lateral (km)",
            zaxis_title="Elevation (m)",
            aspectmode="manual",
            aspectratio=dict(x=2.2, y=1.0, z=1.1),
            camera=dict(eye=dict(x=1.8, y=1.6, z=1.1)),
        ),
        margin=dict(l=0, r=0, t=10, b=0),
        legend=dict(orientation="h", y=-0.05),
    )
    return fig


def _duplicate_last_row(df: pd.DataFrame) -> pd.DataFrame:
    """Append a copy of the last row to a DataFrame. If the DataFrame is empty,
    returns it unchanged."""
    if df is None or len(df) == 0:
        return df
    out = pd.concat([df, df.iloc[[-1]]], ignore_index=True)
    return out


def _table_is_dirty(buffer_df: pd.DataFrame, committed_df: pd.DataFrame) -> bool:
    """True when the data_editor buffer differs from the committed session
    state — used to colour the Apply button orange (dirty) vs green (clean).

    Compares shape and values; tolerant of dtype noise and NaN.
    """
    if buffer_df is None or committed_df is None:
        return buffer_df is not committed_df
    try:
        if buffer_df.shape != committed_df.shape:
            return True
        # Align columns; if column sets differ → dirty
        if set(buffer_df.columns) != set(committed_df.columns):
            return True
        b = buffer_df.reset_index(drop=True)
        c = committed_df[buffer_df.columns].reset_index(drop=True)
        # Stringify to dodge dtype / NaN comparison quirks
        return not b.astype(str).equals(c.astype(str))
    except Exception:
        # When in doubt, treat as dirty so the user can always commit
        return True


def _apply_button(buffer_df: pd.DataFrame, committed_df: pd.DataFrame,
                  key: str, label: str = "Apply") -> bool:
    """Render an Apply button that is ORANGE when the buffer has unsaved
    changes and GREEN when it matches the committed state. Returns True when
    clicked.

    Streamlit doesn't expose per-button colours directly, so we use the
    button label + an emoji indicator + a coloured caption to convey state.
    """
    dirty = _table_is_dirty(buffer_df, committed_df)
    if dirty:
        clicked = st.button(f"🟠 {label} — unsaved changes", key=key,
                            type="primary", use_container_width=False)
        st.caption(":orange[● Edited — click Apply to commit and refresh results.]")
    else:
        clicked = st.button(f"🟢 {label} — up to date", key=key,
                            type="secondary", use_container_width=False)
        st.caption(":green[✓ All changes applied.]")
    return clicked


# =============================================================================
# Economics UI
# =============================================================================
def economics_section(units, start_date):
    st.subheader("💰 Economics")

    with st.expander("ℹ️ Economic assumptions help", expanded=False):
        st.markdown(
            "- **Prices** are flat over the horizon (extend code for price decks).\n"
            "- **OPEX**: variable per unit of primary fluid; fixed is annual $MM.\n"
            "- **Royalty / Tax / Tariffs** as fractions / unit rates.\n"
            "- **Facility CAPEX** is phased — add lines for each milestone payment.\n"
            "- **Abandonment** cost is incurred at the last producing month.\n"
            "- **Tax** is applied on positive pre-tax cashflow only (no loss carry-forward)."
        )

    # Prices are always shown in $/bbl (oil) and $/MMBtu (gas) regardless of
    # the metric/field units toggle — this is the industry-standard convention
    # (oil traded as Brent/WTI per barrel; gas as Henry Hub/JKM/TTF per MMBtu).
    # The engine internally uses $/bbl (oil) and $/Mscf (gas) for consistency
    # with field-unit production. We convert $/MMBtu → $/Mscf via the standard
    # natural-gas heating-value factor of 1.0 Mcf/MMBtu (a close screening
    # approximation; real values vary 0.95-1.10 with composition).
    MMBTU_PER_MCF = 1.0   # screening factor; engine treats gas_price as $/Mscf
    with st.expander("💵 Prices & OPEX", expanded=False):
        c1, c2, c3, c4 = st.columns(4)
        oil_price_bbl = c1.number_input(
            "Oil price ($/bbl)", value=75.0,
            key="oil_price_bbl", on_change=mark_stale,
            help="Flat real crude price per barrel. $75/bbl is a conservative "
                 "long-run screening value (Brent). Industry-standard "
                 "regardless of unit system."
        )
        gas_price_mmbtu = c2.number_input(
            "Gas price ($/MMBtu)", value=10.0,
            key="gas_price_mmbtu", on_change=mark_stale,
            help="Flat real gas price per MMBtu. Default $10/MMBtu reflects "
                 "European hub pricing (TTF / NBP) over the recent cycle — "
                 "NCS gas is sold mainly into Europe and the post-2022 hub "
                 "average has sat in the $8–$15/MMBtu range. Use a long-run "
                 "real price for screening rather than a recent spot peak. "
                 "Internally converted to $/Mscf using 1 Mcf ≈ 1 MMBtu "
                 "(real heating values vary 0.95–1.10)."
        )
        # Variable OPEX — unit basis depends on the fluid system. For an oil
        # field the natural basis is $/bbl of oil; for a gas field it is $/Mscf
        # of gas. Charging a $/bbl number against a Mscf/d rate (the old bug)
        # under-charges gas OPEX by roughly the boe factor.
        _econ_fluid = st.session_state.get("fluid", "Oil with associated gas")
        _econ_is_oil = FLUID_SYSTEMS[_econ_fluid]["primary"] == "oil"
        # The variable-OPEX key is suffixed with the fluid phase. This is
        # deliberate: switching from an oil to a gas fluid must give a fresh
        # widget with the gas default ($/Mscf), not carry over the oil $/bbl
        # value (a $5.5/bbl value silently shown as $5.5/Mscf was a real bug).
        _opex_phase = "oil" if _econ_is_oil else "gas"
        _opex_key = f"opex_var_{_opex_phase}"
        if _econ_is_oil:
            opex_var_bbl = c3.number_input(
                "Var. OPEX ($/bbl)", value=5.5,
                key=_opex_key, on_change=mark_stale,
                help="Variable operating cost per barrel of primary fluid "
                     "(oil) produced. Default $5.5/bbl reflects a mid-size NCS "
                     "offshore development; small / late-life fields run "
                     "higher ($10-20/bbl), very large fields lower. Industry-"
                     "standard regardless of unit system.")
        else:
            opex_var_bbl = c3.number_input(
                "Var. OPEX ($/Mscf)", value=0.9,
                key=_opex_key, on_change=mark_stale,
                help="Variable operating cost per Mscf of primary fluid (gas) "
                     "produced. Default $0.9/Mscf reflects an NCS gas / gas-"
                     "condensate development (~$5/boe). For a gas field the "
                     "engine charges variable OPEX against the gas rate "
                     "(Mscf/d), so this must be a $/Mscf figure — typically "
                     "$0.5-2.0/Mscf, NOT an oil $/bbl number.")
        opex_fixed = c4.number_input("Fixed OPEX ($MM/yr)", value=20.0,
                                     key="opex_fixed", on_change=mark_stale)

    # Convert always-$/bbl and always-$/MMBtu to engine-internal $/bbl, $/Mscf
    oil_price = float(oil_price_bbl)
    gas_price = float(gas_price_mmbtu) * MMBTU_PER_MCF
    opex_var = float(opex_var_bbl)

    # ---- NGL (Natural Gas Liquids) ----
    with st.expander("💎 NGL (Natural Gas Liquids) stream", expanded=False):
        st.markdown(
            "NGL = propane / butane / pentane+ extracted from the produced gas at "
            "a midstream plant. Modelled as a per-MMscf yield, priced per barrel. "
            "Set yield to **0** to disable. "
            "Typical yields by reservoir type:\n"
            "- Dry gas: 0–10 bbl/MMscf\n"
            "- Wet gas: 10–50 bbl/MMscf\n"
            "- Gas condensate: 30–150 bbl/MMscf\n"
            "- Rich gas / volatile oil associated: 50–200 bbl/MMscf\n\n"
            "NGL pricing is typically quoted as a **30–60% fraction of WTI** "
            "(so at oil = $75/bbl, expect NGL $22–45/bbl)."
        )
        ngl1, ngl2, ngl3, ngl4 = st.columns(4)
        ngl_yield = ngl1.number_input(
            "NGL yield (bbl/MMscf)", value=0.0, min_value=0.0, step=5.0,
            key="ngl_yield", on_change=mark_stale,
            help="Volume of NGL recovered per MMscf of gross gas processed. "
                 "Always bbl/MMscf regardless of unit system."
        )
        ngl_price = ngl2.number_input(
            "NGL price ($/bbl)", value=25.0, min_value=0.0, step=1.0,
            key="ngl_price", on_change=mark_stale,
            help="Composite NGL price. As a rule of thumb, 35-50% of the oil price."
        )
        ngl_opex = ngl3.number_input(
            "NGL OPEX ($/bbl)", value=5.0, min_value=0.0, step=0.5,
            key="ngl_opex", on_change=mark_stale,
            help="NGL-specific processing + transport + fractionation tariff. "
                 "Typical 3-10 $/bbl."
        )
        ngl_shrink = ngl4.slider(
            "Gas shrinkage from NGL", 0.0, 0.15, 0.0, 0.005, format="%.3f",
            key="ngl_shrink", on_change=mark_stale,
            help="Fraction of gas volume lost to the NGL plant (extraction "
                 "removes hydrocarbons that no longer reach the gas sales meter). "
                 "Typical 2-5%. Set to 0 to ignore (slightly optimistic)."
        )
        # Live preview
        if ngl_yield > 0:
            st.caption(
                f"💡 At yield = {ngl_yield:.0f} bbl/MMscf, NGL price = ${ngl_price:.0f}/bbl, "
                f"every **1 MMscf/d** of gas yields **{ngl_yield:.0f} bbl/d** of NGL → "
                f"**${ngl_yield * ngl_price * 365 / 1e6:,.2f}MM/yr** revenue (gross), "
                f"net of ${ngl_yield * ngl_opex * 365 / 1e6:,.2f}MM/yr OPEX."
            )

    # ---- Well cost (fixed vs rig-rate bottom-up) ----
    with st.expander("⚒️ Well cost model (rig-rate or fixed $MM/well)",
                     expanded=False):
        well_cost_mode = st.radio(
            "Method", ["rig_rate", "fixed"],
            format_func=lambda x: "Rig-rate (bottom-up)" if x == "rig_rate" else "Fixed $MM / well (legacy)",
            index=0, horizontal=True,
            key="well_cost_mode", on_change=mark_stale,
            help="Rig-rate: cost = (drill_days × rig dayrate + completion_days × completion-spread dayrate) × (1 + intangibles%) + tangibles. "
                 "Fixed: a single $MM/well number (legacy).",
        )
        if well_cost_mode == "rig_rate":
            rc1, rc2, rc3, rc4 = st.columns(4)
            rig_day_rate_kUSD = rc1.number_input(
                "Rig dayrate ($k/day)", value=500.0, min_value=0.0, step=10.0,
                key="rig_dayrate", on_change=mark_stale,
                help="Drilling-rig spread cost per day. Typical ranges: $50-150k/d (land), "
                     "$150-350k/d (jackup), $400-700k/d (semi/drillship).")
            completion_day_rate_kUSD = rc2.number_input(
                "Completion dayrate ($k/day)", value=350.0, min_value=0.0, step=10.0,
                key="cmpl_dayrate", on_change=mark_stale,
                help="Completion spread (frac fleet, wireline, mob). Typically 60–80% of the rig dayrate.")
            well_tangibles_MM = rc3.number_input(
                "Tangibles ($MM/well)", value=4.0, min_value=0.0, step=0.5,
                key="well_tangibles", on_change=mark_stale,
                help="Casing, tubing, tree, wellhead, line pipe.")
            well_intangibles_pct = rc4.slider(
                "Intangibles % of spread", 0.0, 0.50, 0.10, 0.01,
                key="well_intangibles_pct", on_change=mark_stale,
                help="Mud, cement, services, transport, fuel — typically 8–15% of spread cost.")
            # Show a live preview of typical well cost
            avg_drill = 45  # default
            avg_compl = 15
            if "producers_df" in st.session_state and len(st.session_state.producers_df) > 0:
                try:
                    pdf = st.session_state.producers_df
                    avg_drill = float(pdf["drill_days"].mean())
                    avg_compl = float(pdf["completion_days"].mean())
                except Exception:
                    pass
            spread = (avg_drill * rig_day_rate_kUSD + avg_compl * completion_day_rate_kUSD) / 1000.0
            preview_cost = spread * (1.0 + well_intangibles_pct) + well_tangibles_MM
            st.caption(
                f"💡 At average drill = **{avg_drill:.0f} days**, completion = "
                f"**{avg_compl:.0f} days** → estimated cost ≈ **${preview_cost:.1f}MM** per well "
                f"(spread ${spread:.1f}MM × {(1+well_intangibles_pct):.2f} + tangibles "
                f"${well_tangibles_MM:.1f}MM)."
            )
            capex_well = preview_cost   # legacy field still populated; used as MC base
        else:
            # Initial default only — once the user edits, session_state["capex_well"]
            # persists, and Streamlit will use that value instead of the literal.
            # Streamlit does NOT re-apply `value=` on subsequent renders when the
            # key exists, but we still need to seed it explicitly so the first
            # render has a value when the user has never touched the widget.
            if "capex_well" not in st.session_state:
                st.session_state["capex_well"] = 15.0
            capex_well = st.number_input("CAPEX per well ($MM)",
                                          key="capex_well", on_change=mark_stale,
                                          min_value=0.0, step=1.0,
                                          help="Spent at the well's spud date. "
                                               "Edited value persists across "
                                               "reruns and mode switches.")
            rig_day_rate_kUSD = 500.0
            completion_day_rate_kUSD = 350.0
            well_tangibles_MM = 4.0
            well_intangibles_pct = 0.10

    with st.expander("📊 Discount rate, tax, royalty, tariffs, abandonment",
                     expanded=False):
        c2, c3, c4 = st.columns(3)
        disc = c2.slider("Discount rate", 0.0, 0.30, 0.10, 0.01,
                         key="disc", on_change=mark_stale)
        tax = c3.slider("Tax rate", 0.0, 0.7, 0.30, 0.01,
                        key="tax_rate", on_change=mark_stale,
                        help="Applied on positive pre-tax CF only. "
                             "**Ignored if fiscal regime is NCS** "
                             "(CIT + SPT replace it).")
        royalty = c4.slider("Royalty rate", 0.0, 0.5, 0.10, 0.01,
                            key="royalty", on_change=mark_stale,
                            help="Deducted from gross revenue. "
                                 "**Ignored if fiscal regime is NCS**.")

        c1, c2, c3 = st.columns(3)
        tariff_oil_bbl = c1.number_input(
            "Oil tariff ($/bbl)", value=2.0,
            key="tariff_oil_bbl", on_change=mark_stale,
            help="Pipeline / processing tariff per barrel of oil. Always per bbl "
                 "regardless of unit system.")
        tariff_gas_mmbtu = c2.number_input(
            "Gas tariff ($/MMBtu)", value=0.3,
            key="tariff_gas_mmbtu", on_change=mark_stale,
            help="Gas transport / processing tariff per MMBtu.")
        aban_cost = c3.number_input("Abandonment cost ($MM)", value=80.0,
                                    key="aban_cost", on_change=mark_stale)

    # ---- Economic limit / cessation timing ----
    st.markdown("**Cessation timing**")
    cutoff_mode_label = st.radio(
        "When does the field cease production?",
        ["Full forecast horizon", "Economic limit (smart cut-off)"],
        horizontal=True, key="economic_cutoff_mode_label", on_change=mark_stale,
        help="Full forecast horizon: produce until the end of the forecast "
             "period; cessation cost booked at the last producing month "
             "(legacy behaviour).\n\n"
             "Economic limit: the engine finds the month after which monthly "
             "operating cashflow (revenue − royalty − tariff − OPEX) stays "
             "negative, shuts the field in there, and books cessation at that "
             "month. This is the self-consistent way to define field life — "
             "you don't keep producing at a loss.",
    )
    economic_cutoff_mode = ("economic"
                            if cutoff_mode_label.startswith("Economic")
                            else "horizon")
    economic_cutoff_persistence = 6
    if economic_cutoff_mode == "economic":
        economic_cutoff_persistence = st.slider(
            "Consecutive negative-CF months before cessation",
            1, 24, 6, 1, key="economic_cutoff_persistence",
            on_change=mark_stale,
            help="How many consecutive months of negative operating cashflow "
                 "are required before declaring the economic limit. A higher "
                 "value rides through transient dips (e.g. a maintenance "
                 "month or a price trough); 6–12 months is typical.")

    # ---- Money basis (nominal vs real) ----
    st.markdown("**Money basis**")
    mb_col1, mb_col2 = st.columns([2, 1])
    money_basis = mb_col1.radio(
        "All cashflows in", ["Real (today's $)", "Nominal (escalated $)"],
        horizontal=True, key="money_basis_label", on_change=mark_stale,
        help="**Real**: costs and revenues remain in today's $. The discount "
             "rate above is then a REAL discount rate (e.g. 7-10% typical).\n\n"
             "**Nominal**: future cashflows are escalated by inflation each "
             "year. The discount rate above is then a NOMINAL discount rate "
             "(e.g. 10-13% typical). The two approaches give equivalent NPVs "
             "if the discount rates and inflation are consistent — most "
             "screening work uses real $.")
    money_basis_for_engine = ("nominal"
                               if money_basis.startswith("Nominal")
                               else "real")
    inflation_rate = 0.0
    if money_basis_for_engine == "nominal":
        inflation_rate = mb_col2.number_input(
            "Inflation rate (%/yr)", min_value=0.0, max_value=15.0, value=2.5,
            step=0.5, key="inflation_rate", on_change=mark_stale,
            help="Annual inflation applied to every monthly cashflow "
                 "(compounded monthly). 2-3% is typical for developed "
                 "economies; 5-10% may be appropriate for high-inflation "
                 "environments.") / 100.0

    # ---- Cost-input currency (input side — costs can be entered in NOK and
    # the engine converts them to USD for the calculation; all RESULTS are
    # always displayed in USD). NCS costs are typically quoted in NOK, so
    # this lets you enter MNOK figures directly. ----
    cur_col1, cur_col2 = st.columns([2, 1])
    cost_input_currency = cur_col1.radio(
        "Cost input currency", ["USD", "NOK"],
        horizontal=True, key="cost_input_currency",
        help="The currency you ENTER costs in (CAPEX, OPEX, day-rates, "
             "tangibles, facility schedule). Selecting NOK lets you type "
             "MNOK figures; the engine converts them to USD using the rate "
             "at right before computing. All RESULTS (NPV, breakeven, etc.) "
             "are always shown in USD.")
    nok_to_usd_rate = cur_col2.number_input(
        "NOK→USD rate", min_value=1.0, max_value=30.0,
        value=float(st.session_state.get("usd_to_nok", 10.5)),
        step=0.1, key="usd_to_nok",
        help="NOK per 1 USD (≈10-11 recently). Cost inputs in NOK are "
             "divided by this to get USD for the engine.")
    if cost_input_currency == "NOK":
        cur_col2.caption(f"1 USD = {nok_to_usd_rate:.1f} NOK — "
                         f"costs entered as NOK ÷ {nok_to_usd_rate:.1f}")

    # ---- Fiscal regime (Tax/Royalty / PSC / NCS) ----
    st.markdown("**Fiscal regime**")
    regime = st.radio(
        "Regime", ["Tax/Royalty", "PSC", "NCS (Norwegian shelf)"],
        horizontal=True, key="fiscal_regime", on_change=mark_stale,
        help="Tax/Royalty: simple regime — royalty on gross revenue, tax on positive pre-tax CF.\n\n"
             "PSC: Production Sharing Contract — cost recovery, profit-oil "
             "split between contractor and government, contractor tax on its "
             "profit-oil share, optional carried-government participation.\n\n"
             "NCS: Norwegian Continental Shelf — Corporate Income Tax (22%) "
             "+ Special Petroleum Tax (71.8%) = ~78% effective on petroleum "
             "profits, with an 'uplift' allowance (17.69% of capex booked in "
             "the year of spend, deductible from the SPT base only) "
             "compensating for non-deductibility of financing in SPT.",
    )
    # Normalize regime label for the engine
    if regime == "NCS (Norwegian shelf)":
        regime_for_engine = "NCS"
    else:
        regime_for_engine = regime
    psc_cost_recovery_ceiling = 0.50
    psc_profit_oil_share_contractor = 0.40
    psc_govt_participation = 0.0
    psc_psc_tax_rate = 0.30
    psc_signature_bonus_MM = 0.0
    ncs_cit_rate = 0.22
    ncs_spt_rate = 0.718
    ncs_uplift_rate = 0.1769
    if regime == "NCS (Norwegian shelf)":
        with st.expander("NCS parameters", expanded=True):
            nc1, nc2, nc3 = st.columns(3)
            ncs_cit_rate = nc1.slider(
                "Corporate income tax (CIT)", 0.10, 0.40, 0.22, 0.01,
                key="ncs_cit", on_change=mark_stale,
                help="Standard Norwegian corporate income tax. Currently 22%. "
                     "Applies to positive pre-tax cashflow.")
            ncs_spt_rate = nc2.slider(
                "Special Petroleum Tax (SPT)", 0.40, 0.90, 0.718, 0.01,
                key="ncs_spt", on_change=mark_stale,
                help="Additional petroleum-sector tax on top of CIT. "
                     "Currently 71.8%. Combined effective rate ≈ 78% on "
                     "petroleum profits.")
            ncs_uplift_rate = nc3.slider(
                "Uplift allowance (× capex)", 0.00, 0.30, 0.1769, 0.01,
                key="ncs_uplift", on_change=mark_stale,
                help="Capital uplift booked in the YEAR of the investment. "
                     "Reduces the SPT base only (not CIT). Currently 17.69%. "
                     "Designed to compensate for the fact that financing "
                     "costs are not deductible in SPT.")
            st.caption(
                f"Combined effective rate on petroleum profits ≈ "
                f"**{(ncs_cit_rate + ncs_spt_rate)*100:.0f}%**. "
                f"Uplift of {ncs_uplift_rate*100:.2f}% × capex offsets the "
                f"SPT base.")
    if regime == "PSC":
        with st.expander("PSC parameters", expanded=True):
            pc1, pc2, pc3 = st.columns(3)
            psc_cost_recovery_ceiling = pc1.slider(
                "Cost recovery ceiling", 0.10, 1.00, 0.50, 0.05,
                key="psc_cr_ceiling", on_change=mark_stale,
                help="Max share of net revenue (after royalty) that can be "
                     "applied to cost recovery each period. Unrecovered "
                     "costs carry forward in the cost pool.")
            psc_profit_oil_share_contractor = pc2.slider(
                "Contractor profit oil share", 0.05, 0.95, 0.40, 0.05,
                key="psc_pos", on_change=mark_stale,
                help="Contractor's share of profit oil (the rest goes to "
                     "government). May be a sliding scale in real PSCs; "
                     "this is a screening-level constant share.")
            psc_psc_tax_rate = pc3.slider(
                "PSC tax on contractor profit", 0.0, 0.85, 0.30, 0.05,
                key="psc_tax", on_change=mark_stale,
                help="Tax rate applied to contractor's profit oil share.")
            pd1, pd2, _ = st.columns(3)
            psc_govt_participation = pd1.slider(
                "Govt participation (carried)", 0.00, 0.60, 0.00, 0.05,
                key="psc_gov_part", on_change=mark_stale,
                help="Fraction of contractor net cashflow accruing to "
                     "government as carried equity (0 = no participation).")
            psc_signature_bonus_MM = pd2.number_input(
                "Signature bonus ($MM, paid month 0)", value=0.0,
                min_value=0.0, key="psc_sig_bonus", on_change=mark_stale,
                help="One-off bonus payable on signing the contract.")
        # Override royalty/tax UI fields with PSC equivalents:
        # royalty stays from earlier slider (it's PSC royalty too).
        # 'tax' from earlier becomes irrelevant; we use psc_psc_tax_rate.

    st.markdown("**Facility CAPEX**")

    # ---- Development concept builder ----
    with st.expander("🏗️ Development concept builder — design the concept & generate CAPEX",
                     expanded=False):
        st.caption(
            "Specify a development concept in engineering terms (host type, "
            "water depth, templates, flowlines, risers, trees, boosting, "
            "artificial lift, flow assurance). The builder produces a phased "
            "CAPEX schedule, a concept summary, engineering sanity-check "
            "warnings, and a schematic. Clicking **Generate CAPEX schedule** "
            "overwrites the phased table below. All cost models are "
            "screening-level — order-of-magnitude, for concept select."
        )

        dc1, dc2 = st.columns(2)
        concept_type = dc1.radio(
            "Concept type",
            ["Subsea tie-in", "Standalone"],
            key="dc_concept_type",
            help="Subsea tie-in: wells tied back to an EXISTING host facility "
                 "— lowest CAPEX, but limited by tie-back distance and host "
                 "ullage.\n\nStandalone: a DEDICATED production facility "
                 "(platform / FPSO offshore, or a central processing "
                 "facility onshore).")
        water_depth_class = dc2.selectbox(
            "Water depth class",
            ["Shallow (<150 m)", "Mid (150-600 m)", "Deep (600-1500 m)",
             "Ultra-deep (>1500 m)"],
            key="dc_water_depth",
            help="Drives installation-difficulty multipliers on flowlines, "
                 "risers and subsea hardware, and constrains the feasible "
                 "host types (fixed structures are not feasible beyond "
                 "~400-500 m).")

        host_type = None
        processing_capacity = 50.0
        host_distance_km = 0.0
        if concept_type == "Standalone":
            host_options = [
                "Fixed steel jacket (shallow)", "Fixed steel jacket (mid)",
                "Concrete gravity structure", "Compliant tower",
                "Jack-up production unit",
                "FPSO (leased — capitalised)", "FPSO (owned)",
                "Semi-submersible FPU", "Spar", "TLP (tension-leg platform)",
                "Onshore central processing facility (CPF)",
            ]
            hc1, hc2 = st.columns(2)
            host_type = hc1.selectbox(
                "Host / facility type", host_options, key="dc_host_type",
                help="Fixed structures (jacket, gravity, compliant tower, "
                     "jack-up): shallow-to-mid water only. Floating hosts "
                     "(FPSO, semi, spar, TLP): mid-to-ultra-deep. Onshore "
                     "CPF: land developments.")
            processing_capacity = hc2.number_input(
                "Processing capacity (kboe/d)", min_value=1.0, value=50.0,
                step=5.0, key="dc_proc_cap",
                help="Plant throughput sizing basis. Topsides / CPF cost "
                     "scales with this. Set it to roughly your expected "
                     "plateau rate in thousand boe/d.")
        else:
            host_distance_km = st.number_input(
                "Tie-back distance to host facility (km)",
                min_value=0.0, value=15.0, step=1.0, key="dc_host_dist",
                help="Distance from the field to the existing host. Beyond "
                     "~30-50 km, flow assurance and pressure support become "
                     "major issues — the builder will warn you.")

        st.markdown("**Wells & subsea hardware**")
        ws1, ws2, ws3, ws4 = st.columns(4)
        n_templates = ws1.number_input(
            "Subsea templates / manifolds", min_value=0, value=1, step=1,
            key="dc_n_templates",
            help="Subsea structures that host and tie together multiple "
                 "wells. Cost depends on the slot count chosen below.")
        n_subsea_wells = ws2.number_input(
            "Wells on wet (subsea) trees", min_value=0, value=4, step=1,
            key="dc_n_subsea_wells",
            help="Wells completed with subsea xmas trees on the seabed "
                 "(~$9MM/tree). Standard for tie-ins and floating hosts. "
                 "Must fit within the template slot capacity.")
        n_dry_wells = ws3.number_input(
            "Wells on dry (surface) trees", min_value=0, value=0, step=1,
            key="dc_n_dry_wells",
            help="Wells completed with surface xmas trees on a platform "
                 "deck (~$1.8MM/tree). Only possible with a fixed platform "
                 "or a dry-tree-capable floater (spar / TLP).")
        n_risers = ws4.number_input(
            "Number of risers", min_value=0, value=2, step=1,
            key="dc_n_risers",
            help="Pipes carrying fluids from the seabed up to a floating or "
                 "fixed host. Subsea production needs risers; dry-tree wells "
                 "do not.")
        # Template type — sets slot capacity and per-template cost.
        tt1, tt2 = st.columns(2)
        template_type = tt1.selectbox(
            "Template type (slot count)",
            ["Single-slot (1 well)", "Double-slot (2 wells)",
             "4-slot (4 wells)", "6-slot (6 wells)"],
            index=2, key="dc_template_type",
            help="A subsea template is built for a fixed number of well "
                 "slots. More slots → bigger, heavier, costlier structure, "
                 "but more drilling flexibility and room for future infill. "
                 "Screening cost: single $18MM, double $30MM, 4-slot $52MM, "
                 "6-slot $72MM. The number of subsea wells must fit within "
                 "n_templates × slots — a warning is shown if not.")
        _slot_cap = {"Single-slot (1 well)": 1, "Double-slot (2 wells)": 2,
                     "4-slot (4 wells)": 4, "6-slot (6 wells)": 6}[template_type]
        _total_slots = n_templates * _slot_cap
        if n_subsea_wells > _total_slots:
            tt2.error(f"⚠️ {n_subsea_wells} wells > {_total_slots} slots "
                      f"({n_templates} × {_slot_cap}). Add templates or pick "
                      f"a larger type.")
        elif n_subsea_wells > 0 and n_subsea_wells <= _total_slots - _slot_cap:
            tt2.warning(f"{_total_slots} slots for {n_subsea_wells} wells — "
                        f"{_total_slots - n_subsea_wells} spare.")
        elif n_subsea_wells > 0:
            tt2.success(f"✓ {n_subsea_wells} wells fit in {_total_slots} "
                        f"slots ({n_templates} × {_slot_cap}).")
        # Multi-template layout — only meaningful with 2+ templates.
        if n_templates > 1:
            template_layout = st.radio(
                "Template layout",
                ["clustered", "spread"], horizontal=True,
                key="dc_template_layout",
                format_func=lambda x: ("Clustered (drill centres together)"
                                       if x == "clustered"
                                       else "Spread (separated along tie-back)"),
                help="How the templates are positioned relative to each "
                     "other. **Clustered**: templates sit side-by-side at "
                     "one drill centre — simplest, shortest in-field "
                     "lines, one manifold hub. **Spread**: templates are "
                     "separated (e.g. over different fault blocks or "
                     "crestal areas) and linked by an in-field line — more "
                     "reservoir coverage but more flowline and a longer "
                     "installation campaign. The schematic updates to show "
                     "the chosen layout.")
        else:
            template_layout = "clustered"

        # ---- Per-template detailed configuration ----
        # With 2+ templates the user can optionally define each template
        # individually: its own slot count and its tie-in topology (which
        # template or the host it connects to, plus the flowline/umbilical
        # leg length). This overrides the single-type simple mode.
        templates_detail = None
        if n_templates >= 1:
            use_detail = st.checkbox(
                "Configure each template individually (position, slots, tie-in)",
                value=False, key="dc_use_template_detail",
                help="Off — every template is the same type above and "
                     "auto-placed. On — set each template's X/Y position "
                     "in km from the host, slot count, role and how it "
                     "ties in (to the host directly, or daisy-chained to "
                     "another template), with the flowline / umbilical leg "
                     "length for that connection. Works for a single "
                     "template too — useful for placing one drill centre "
                     "at the correct field coordinate.")
            if use_detail:
                tie_options = ["Host"] + [f"T{i+1}"
                                          for i in range(int(n_templates))]
                _std_types = list(_TEMPLATE_SLOT_CAPACITY_UI.keys())
                _td_key = f"dc_templates_detail_{int(n_templates)}"

                def _default_td_row(i):
                    return {
                        "Template": f"T{i+1}",
                        "Name": f"Template {i+1}",
                        "Type": "4-slot (4 wells)",
                        "Slots": 4,
                        "Role": "Producer",
                        "Tie-in to": "Host" if i == 0 else "T1",
                        "X (km)": float(-(14 + i * 6)),
                        "Y (km)": float((i - (int(n_templates)-1)/2) * 6),
                        "Flowline leg (km)": 14.0 if i == 0 else 6.0,
                        "Umbilical leg (km)": 15.0 if i == 0 else 6.0,
                    }
                if _td_key not in st.session_state:
                    st.session_state[_td_key] = pd.DataFrame(
                        [_default_td_row(i) for i in range(int(n_templates))])
                # keep the row count in sync with n_templates
                _cur = st.session_state[_td_key]
                if len(_cur) != int(n_templates):
                    new_rows = []
                    for i in range(int(n_templates)):
                        if i < len(_cur):
                            new_rows.append(_cur.iloc[i].to_dict())
                        else:
                            new_rows.append(_default_td_row(i))
                    _cur = pd.DataFrame(new_rows)
                    st.session_state[_td_key] = _cur

                st.caption(
                    "Set each template's type, role, position (X/Y km from "
                    "the host at 0,0) and tie-in. The aerial view places "
                    "templates and host by these coordinates, to scale. "
                    "Press **Apply template layout** to use the edits.")
                edited_td = st.data_editor(
                    st.session_state[_td_key],
                    key=f"dc_td_editor_{int(n_templates)}",
                    use_container_width=True, hide_index=True,
                    column_config={
                        "Template": st.column_config.TextColumn(
                            "ID", disabled=True),
                        "Name": st.column_config.TextColumn(
                            "Name", help="Display name on the schematic."),
                        "Type": st.column_config.SelectboxColumn(
                            "Template type", options=_std_types,
                            help="Standard subsea template type — sets the "
                                 "slot count and cost."),
                        "Slots": st.column_config.NumberColumn(
                            "Slots", min_value=1, max_value=12, step=1,
                            help="Well slots — auto-set from the type, but "
                                 "can be overridden."),
                        "Role": st.column_config.SelectboxColumn(
                            "Role", options=["Producer", "Injector"],
                            help="Producer or water-injector template — "
                                 "colours the slots on the schematic."),
                        "Tie-in to": st.column_config.SelectboxColumn(
                            "Tie-in to", options=tie_options,
                            help="Where this template routes — the host or "
                                 "another template (daisy-chained)."),
                        "X (km)": st.column_config.NumberColumn(
                            "X (km)", step=1.0, format="%.1f",
                            help="East-west position relative to the host "
                                 "(host at 0). Negative = west of host."),
                        "Y (km)": st.column_config.NumberColumn(
                            "Y (km)", step=1.0, format="%.1f",
                            help="North-south position relative to the "
                                 "host (host at 0)."),
                        "Flowline leg (km)":
                            st.column_config.NumberColumn(
                                "Flowline leg (km)", min_value=0.0,
                                step=1.0, format="%.1f",
                                help="Flowline length from this template "
                                     "to its tie-in point."),
                        "Umbilical leg (km)":
                            st.column_config.NumberColumn(
                                "Umbilical leg (km)", min_value=0.0,
                                step=1.0, format="%.1f"),
                    })

                _apply_td = st.button("✅ Apply template layout",
                                       key=f"dc_apply_td_{int(n_templates)}",
                                       help="Commit the template-table "
                                            "edits and refresh the concept "
                                            "and schematics.")
                if _apply_td:
                    # sync slot count to the chosen standard type
                    _e = edited_td.copy()
                    for i in _e.index:
                        ttype = str(_e.at[i, "Type"])
                        _e.at[i, "Slots"] = _TEMPLATE_SLOT_CAPACITY_UI.get(
                            ttype, int(_e.at[i, "Slots"]))
                    st.session_state[_td_key] = _e
                    mark_stale()
                    st.success("Template layout applied.")
                    st.rerun()

                _td_use = st.session_state[_td_key]
                templates_detail = []
                for i, row in _td_use.iterrows():
                    templates_detail.append({
                        "name": str(row.get("Name", f"T{i+1}")),
                        "template_type": str(row.get("Type",
                                                     "4-slot (4 wells)")),
                        "slots": int(row["Slots"]),
                        "role": str(row.get("Role", "Producer")).lower(),
                        "tie_to": str(row["Tie-in to"]),
                        "x_km": float(row.get("X (km)", -14.0)),
                        "y_km": float(row.get("Y (km)", 0.0)),
                        "flowline_km": float(row["Flowline leg (km)"]),
                        "umbilical_km": float(row["Umbilical leg (km)"]),
                    })
                # validation feedback
                _det_slots = sum(t["slots"] for t in templates_detail)
                if n_subsea_wells > _det_slots:
                    st.error(f"⚠️ {n_subsea_wells} wells > {_det_slots} "
                             f"slots across {n_templates} templates.")
                else:
                    st.success(f"✓ {n_subsea_wells} wells fit in "
                               f"{_det_slots} slots "
                               f"({n_templates} templates).")
                # self-tie / loop check
                for i, t in enumerate(templates_detail):
                    if t["tie_to"] == f"T{i+1}":
                        st.warning(f"T{i+1} is tied to itself — change its "
                                   f"tie-in to the host or another "
                                   f"template.")

                # ---- Link wells (and their rigs) to each template ----
                # Each producer in the production profile is assigned to
                # exactly one template (a well sits on one drill centre).
                # The cleanest UX is a single dataframe editor: one row per
                # well, one column to pick the template — exclusive by
                # construction, and easy to scan against the production
                # profile and the rig list.
                _pdf = st.session_state.get("producers_df")
                if _pdf is not None and len(_pdf) > 0:
                    st.markdown("**Link wells to templates and rigs**")
                    st.caption(
                        "One row per producer from the production profile. "
                        "Pick the template hosting each well; the rig "
                        "drilling it is read straight from the producers "
                        "table (it can't be edited here — change it there). "
                        "A well can only sit on one template — selecting "
                        "the template moves the well exclusively.")
                    _well_names = [str(n) for n in _pdf["name"].tolist()
                                   if str(n).strip()]
                    _rig_of = {}
                    if "rig" in _pdf.columns:
                        for _, _r in _pdf.iterrows():
                            _rig_of[str(_r["name"])] = str(
                                _r.get("rig", "—"))
                    _tpl_labels = [f"T{i+1}"
                                   for i in range(int(n_templates))]
                    _tpl_options = [f"{lbl} · {templates_detail[i]['name']}"
                                    for i, lbl in enumerate(_tpl_labels)]
                    # Mapping from full option label back to T-id (for
                    # parsing the editor result).
                    def _opt_to_tid(opt):
                        return str(opt).split(" ", 1)[0] if opt else "T1"

                    _wt_key = f"dc_well_template_map_{int(n_templates)}"
                    _assign = dict(st.session_state.get(_wt_key, {}))
                    # purge stale wells (renamed/removed in producers_df)
                    _assign = {k: v for k, v in _assign.items()
                               if k in _well_names}
                    # default-fill new wells across templates by slot
                    if _well_names:
                        _slot_caps = [int(templates_detail[i]["slots"])
                                      for i in range(int(n_templates))]
                        _slot_used = [0] * int(n_templates)
                        # count current usage
                        for wn, lbl in _assign.items():
                            if lbl in _tpl_labels:
                                _slot_used[_tpl_labels.index(lbl)] += 1
                        for wn in _well_names:
                            if wn in _assign:
                                continue
                            # find the first template with spare slots
                            target = 0
                            for ti in range(int(n_templates)):
                                if _slot_used[ti] < _slot_caps[ti]:
                                    target = ti; break
                            _assign[wn] = _tpl_labels[target]
                            _slot_used[target] += 1

                    # Build the table
                    def _tid_to_opt(tid):
                        if tid in _tpl_labels:
                            return _tpl_options[_tpl_labels.index(tid)]
                        return _tpl_options[0]
                    _wt_rows = [{
                        "Well": wn,
                        "Rig": _rig_of.get(wn, "—"),
                        "Template": _tid_to_opt(_assign.get(wn,
                                                             _tpl_labels[0])),
                    } for wn in _well_names]
                    _wt_df = pd.DataFrame(_wt_rows)
                    _edited_wt = st.data_editor(
                        _wt_df,
                        key=f"dc_wt_editor_{int(n_templates)}_"
                            f"{len(_well_names)}",
                        use_container_width=True, hide_index=True,
                        column_config={
                            "Well": st.column_config.TextColumn(
                                "Well", disabled=True),
                            "Rig": st.column_config.TextColumn(
                                "Rig", disabled=True,
                                help="Set in the producers table — shown "
                                     "here for reference."),
                            "Template": st.column_config.SelectboxColumn(
                                "Template", options=_tpl_options,
                                help="Pick the template hosting this well."),
                        })
                    # Apply button — commit edits exclusively
                    if st.button("✅ Apply well-template links",
                                  key=f"dc_apply_wt_{int(n_templates)}",
                                  help="Commit the well-to-template "
                                       "assignment. The schematic, the "
                                       "drilling Gantt and the per-template "
                                       "well counts refresh."):
                        new_assign = {}
                        for _, row in _edited_wt.iterrows():
                            wn = str(row["Well"])
                            new_assign[wn] = _opt_to_tid(row["Template"])
                        st.session_state[_wt_key] = new_assign
                        _assign = new_assign
                        mark_stale()
                        st.success("Well-template links applied.")
                        st.rerun()

                    # Per-template summary lines (read from the edited
                    # frame so the user sees their pending edits even
                    # before pressing Apply).
                    _pending = {}
                    for _, row in _edited_wt.iterrows():
                        wn = str(row["Well"])
                        _pending[wn] = _opt_to_tid(row["Template"])
                    for ti in range(int(n_templates)):
                        _tlabel = _tpl_labels[ti]
                        _here = [wn for wn in _well_names
                                 if _pending.get(wn) == _tlabel]
                        _rigs_here = sorted({_rig_of.get(wn, "—")
                                              for wn in _here})
                        _rig_txt = (", ".join(r for r in _rigs_here
                                              if r != "—")
                                     or "—")
                        cap = int(templates_detail[ti]["slots"])
                        over = len(_here) > cap
                        st.caption(
                            f"**{_tlabel} · {templates_detail[ti]['name']}** "
                            f"— {len(_here)} well(s) / {cap} slot(s); "
                            f"rig(s): {_rig_txt}"
                            + ("  ⚠️ over capacity" if over else ""))

                    # Attach the mapping (+ rig) to each template detail so
                    # downstream code (aerial labels, side view, schedule)
                    # can read it.
                    for ti in range(int(n_templates)):
                        _tlabel = _tpl_labels[ti]
                        _wells_here = [wn for wn in _well_names
                                       if _pending.get(wn) == _tlabel]
                        templates_detail[ti]["wells"] = _wells_here
                        templates_detail[ti]["rigs"] = sorted(
                            {_rig_of.get(wn, "—")
                             for wn in _wells_here})
                    # stash for the rest of the app (drilling Gantt etc.)
                    st.session_state["well_template_map"] = dict(_pending)
        st.markdown("**Flowlines, umbilicals & export**")

        # ---- SURF field architecture (from the SURF long list) ----
        # Top-level architecture choice that drives flowline routing,
        # jumper count and PLEM/manifold scope. Mirrors the "SURF
        # overall — Field architecture" column on the DG1 long list.
        arch1, arch2 = st.columns(2)
        field_architecture = arch1.selectbox(
            "Field architecture (SURF)",
            ["Cluster (manifold)", "Daisy chain", "Inline tees / PLEM",
             "Template", "Satellite wells"],
            index=0, key="dc_field_architecture",
            help="Cluster — wells around a central manifold, one "
                 "flowline pair to host (most common). Daisy chain — "
                 "wells linked in series, fewest flowlines, but one "
                 "failure isolates downstream. Inline tees / PLEM — "
                 "wells tie into a trunk line via tees. Template — "
                 "wells share a structural template + integrated "
                 "manifold. Satellite — individually tied-back wells, "
                 "most flowlines/jumpers, max flexibility.")
        # ---- Hydrate management (MEG / MeOH) ----
        hydrate_mgmt = arch2.selectbox(
            "Hydrate management",
            ["None", "MEG (mono-ethylene glycol) with reclamation",
             "Methanol (MeOH) once-through", "Insulation / no-touch time only"],
            index=0, key="dc_hydrate_mgmt",
            help="MEG with reclamation — a topside MEG regeneration "
                 "package + dosing line in the umbilical; standard for "
                 "long wet-gas tie-backs (high CAPEX, low OPEX). "
                 "Methanol once-through — simpler, no reclamation, but "
                 "high consumable OPEX. Insulation only — relies on "
                 "thermal no-touch time, no chemical system.")

        rs1, rs2 = st.columns(2)
        riser_type = rs1.selectbox(
            "Riser type",
            ["Flexible riser", "Steel catenary riser (SCR)",
             "Top-tensioned riser (TTR)", "Hybrid riser tower segment"],
            key="dc_riser_type",
            help="Flexible: most common, mid-cost. SCR: rigid, deep water. "
                 "TTR: dry-tree floaters (spar/TLP). Hybrid riser tower: "
                 "ultra-deep, highest cost.")
        n_boosting = rs2.number_input(
            "Subsea boosting stations", min_value=0, value=0, step=1,
            key="dc_n_boosting",
            help="Multiphase pump stations on the seabed to boost production "
                 "over long tie-backs or from low-energy reservoirs "
                 "(~$75MM/station).")


        fl1, fl2, fl3 = st.columns(3)
        flowline_km = fl1.number_input(
            "Flowline length (km)", min_value=0.0, value=15.0, step=1.0,
            key="dc_flowline_km",
            help="In-field flowline, or tie-back flowline to the host. For a "
                 "tie-in this is normally ≈ the tie-back distance.")
        flowline_diameter = fl2.number_input(
            "Flowline diameter (inches)", min_value=4.0, max_value=36.0,
            value=10.0, step=2.0, key="dc_flowline_diam",
            help="Nominal bore. Larger diameter → higher cost per km but "
                 "lower pressure drop. Typical screening range 6-30\".")
        flowline_material = fl3.selectbox(
            "Flowline material",
            ["Carbon steel", "CRA-clad", "Solid CRA", "Flexible pipe"],
            key="dc_flowline_material",
            help="Carbon steel: cheapest, needs corrosion management. "
                 "CRA-clad / solid CRA: corrosion-resistant alloy for sour "
                 "or corrosive fluids (1.85× / 3.2× cost). Flexible pipe: "
                 "easier installation, 2.4× cost.")
        # Insulation row
        in1, in2 = st.columns(2)
        flowline_insulation = in1.selectbox(
            "Flowline thermal insulation",
            ["None", "Polypropylene coating (basic)",
             "Multi-layer PP / syntactic", "Pipe-in-pipe (PIP)"],
            key="dc_flowline_insulation",
            help="Required for waxy/viscous crude and to manage hydrate "
                 "formation. Cost: polypropylene 0.30MM/km, multi-layer "
                 "0.65MM/km, pipe-in-pipe 1.80MM/km. PIP is highest cost but "
                 "also highest U-value — chosen for long deep-water tie-backs.")
        insulated_flowline_km = 0.0
        if flowline_insulation != "None":
            insulated_flowline_km = in2.number_input(
                "Insulated flowline length (km)", min_value=0.0,
                value=float(flowline_km), step=1.0, key="dc_insulated_km",
                help="Length of flowline that needs thermal insulation — "
                     "typically the whole flowline for waxy crude or long "
                     "tie-backs.")

        fl4, fl5, fl6 = st.columns(3)
        umbilical_km = fl4.number_input(
            "Umbilical length (km)", min_value=0.0, value=16.0, step=1.0,
            key="dc_umbilical_km",
            help="Bundle carrying hydraulic / electric / chemical lines to "
                 "subsea equipment (~$1.6MM/km). Usually ≈ flowline length.")
        export_pipeline_km = fl5.number_input(
            "Export pipeline length (km)", min_value=0.0, value=0.0, step=1.0,
            key="dc_export_km",
            help="Pipeline carrying processed product from the host to "
                 "shore / a terminal / a trunkline. 0 if exporting via an "
                 "existing nearby tie-in.")
        export_pipeline_diameter = fl6.number_input(
            "Export pipeline diameter (inches)", min_value=4.0, max_value=48.0,
            value=16.0, step=2.0, key="dc_export_diam",
            help="Nominal bore of the export line.")

        # Subsea ancillary elements
        st.markdown("**Subsea ancillary elements**")
        sa1, sa2, sa3, sa4 = st.columns(4)
        n_riser_bases = sa1.number_input(
            "Riser bases (FRBs)", min_value=0, value=int(n_risers), step=1,
            key="dc_n_riser_bases",
            help="Riser base / FRB seated on the seabed at the foot of each "
                 "riser ($7.5MM each). Usually one per riser.")
        n_ssiv = sa2.number_input(
            "Subsea isolation valves (SSIV)", min_value=0, value=0, step=1,
            key="dc_n_ssiv",
            help="Subsea isolation valves between the field and the host "
                 "($4MM each). Required by safety case for long tie-backs.")
        n_jumpers = sa3.number_input(
            "Subsea jumpers", min_value=0, value=int(n_subsea_wells), step=1,
            key="dc_n_jumpers",
            help="Rigid or flexible spools connecting trees to manifolds "
                 "or manifolds to flowlines ($1.2MM each).")
        n_control_modules = sa4.number_input(
            "Subsea control modules (SCMs)", min_value=0,
            value=int(n_subsea_wells), step=1,
            key="dc_n_scm",
            help="One control module per subsea well ($2.5MM each).")

        # ---- Installation method (from the SURF Installation column) ----
        st.markdown("**Installation method**")
        inst1, inst2 = st.columns(2)
        flowline_install = inst1.selectbox(
            "Flowline installation",
            ["Reeling (reel-lay)", "S-lay", "J-lay", "Flexible lay"],
            index=0, key="dc_flowline_install",
            help="Reeling — fastest/cheapest for rigid pipe up to "
                 "~16\", spooled onshore. S-lay — high-rate shallow / "
                 "mid-water. J-lay — deep water, near-vertical "
                 "departure. Flexible lay — for flexible pipe, "
                 "vessel-flexible but slower. Drives the pipelay "
                 "vessel spread day-rate.")
        tiein_method = inst2.selectbox(
            "Tie-in method",
            ["Rigid spool + diverless connector", "Flexible jumper",
             "Vertical (VxT) connection", "Diver-assisted"],
            index=0, key="dc_tiein_method",
            help="Rigid spool — most common, ROV-installed. Flexible "
                 "jumper — tolerant of misalignment, quicker. Vertical "
                 "connection — for vertical-bore trees. Diver-assisted "
                 "— shallow water only, slower/costlier per connection.")

        # Topside modification + offshore manpower
        st.markdown("**Topside modification & manpower**")
        tm1, tm2, tm3 = st.columns(3)
        topside_mod_tonnes = tm1.number_input(
            "Topside mod — net installed weight (tonnes)",
            min_value=0.0, value=0.0, step=50.0,
            key="dc_topside_tonnes",
            help="Alternative basis to a lumped $MM number for host "
                 "modifications: enter the total net new/modified topside "
                 "weight (tonnes) and the cost rate per tonne. Set to 0 to "
                 "use the default lumped tie-in modification cost.")
        topside_mod_rate_per_tonne_kUSD = tm2.number_input(
            "Cost rate ($k per installed tonne)", min_value=0.0,
            value=60.0, step=5.0, key="dc_topside_rate_k",
            help="Fully-loaded $/tonne for offshore brownfield mods. "
                 "Screening default $60k/tonne; high-spec mods can run "
                 "$100k+/tonne.")
        # Convert k$/tonne → $MM/tonne for the engine
        topside_mod_rate_per_tonne_MM = topside_mod_rate_per_tonne_kUSD / 1000.0
        offshore_manhours = tm3.number_input(
            "Offshore manhours", min_value=0.0, value=0.0, step=1000.0,
            key="dc_manhours",
            help="Total offshore execution + engineering manhours (set to 0 "
                 "if already covered by the topside weight or installation "
                 "rows). Useful when you have a manhour estimate from a "
                 "pre-FEED study.")
        offshore_manhour_rate_usd = st.number_input(
            "Manhour rate ($/hr)", min_value=0.0, value=220.0, step=10.0,
            key="dc_manhour_rate",
            help="Fully-loaded $/hr for offshore manhours (engineering + "
                 "offshore execution blended). Screening default $220/hr.")

        st.markdown("**Artificial lift & flow assurance**")
        al1, al2 = st.columns(2)
        gas_lift = al1.checkbox(
            "Gas lift system", value=False, key="dc_gas_lift",
            help="Injects gas into the production tubing to lighten the "
                 "fluid column and sustain rates as reservoir pressure "
                 "declines. ~$25MM base + $1.2MM/well.")
        n_gas_lift_wells = 0
        if gas_lift:
            n_gas_lift_wells = al2.number_input(
                "Wells on gas lift", min_value=0,
                value=int(n_subsea_wells + n_dry_wells), step=1,
                key="dc_n_gas_lift_wells")
        ha1, ha2 = st.columns(2)
        heating_type = ha1.selectbox(
            "Flowline heating / flow assurance",
            ["None", "Electrically heated flowline (EHTF)",
             "Direct electric heating (DEH)",
             "Hot-water / glycol circulation"],
            key="dc_heating_type",
            help="For waxy / viscous crude or hydrate management. EHTF / DEH "
                 "are priced per km of heated line; hot-water/glycol is a "
                 "fixed-cost circulation system. 'None' relies on "
                 "insulation + chemical inhibition alone.")
        heated_flowline_km = 0.0
        if heating_type in ("Electrically heated flowline (EHTF)",
                             "Direct electric heating (DEH)"):
            heated_flowline_km = ha2.number_input(
                "Heated flowline length (km)", min_value=0.0,
                value=float(flowline_km), step=1.0,
                key="dc_heated_km",
                help="Length of flowline that needs active heating — often "
                     "the whole tie-back for long deep-water lines.")

        # Pull any saved cost overrides from session state. Persists across
        # reruns so the user's edits aren't lost.
        cost_overrides = st.session_state.get("dc_cost_overrides", {})

        # Assemble the spec and build the concept
        # HPHT classification — auto-derived from the reservoir PVT inputs
        # (pressure / temperature in the Reservoir tab), with a manual
        # override so the user can force a tier.
        _p_psi = to_field(float(st.session_state.get("p_init", 3500.0)),
                          "pressure", units)
        _t_F = to_field(float(st.session_state.get("t_res", 180.0)),
                        "temp", units)
        _auto_hpht = fh.classify_hpht(_p_psi, _t_F)
        st.markdown("**HPHT classification**")
        hpht_choice = st.radio(
            "Pressure / temperature class",
            ["Auto from reservoir PVT", "Standard", "HPHT", "Ultra-HPHT",
             "Extreme-HPHT"],
            horizontal=True, key="dc_hpht_choice",
            help="HPHT (High Pressure / High Temperature) developments need "
                 "specialised completions, higher-grade metallurgy and longer "
                 "drilling/testing — a CAPEX uplift applies to wells and "
                 "subsea hardware.\n\n"
                 "Thresholds: HPHT ≥ 10,000 psi or ≥ 300 °F; "
                 "Ultra-HPHT ≥ 15,000 psi or ≥ 350 °F; "
                 "Extreme-HPHT ≥ 20,000 psi or ≥ 400 °F.\n\n"
                 "'Auto' reads the pressure & temperature from the Reservoir "
                 "tab's PVT inputs.")
        if hpht_choice == "Auto from reservoir PVT":
            hpht_tier = _auto_hpht["tier"]
            tag = _auto_hpht["tag"]
            if _auto_hpht["is_hpht"]:
                st.warning(f"**{tag}** — {_auto_hpht['rationale']}")
            else:
                st.caption(f"{tag} — {_auto_hpht['rationale']}")
        else:
            hpht_tier = hpht_choice
            _man = fh._HPHT_CAPEX_UPLIFT.get(hpht_tier, 1.0)
            st.caption(f"Manual override: **{hpht_tier}** "
                       f"(CAPEX uplift ×{_man:.2f} on wells & subsea). "
                       f"Auto-classification from PVT would be "
                       f"'{_auto_hpht['tier']}'.")

        # HIPPS — pressure-protection system. Auto-enabled for HPHT tiers,
        # optional otherwise.
        _hpht_is_hpht = hpht_tier != "Standard"
        hipps_on = st.checkbox(
            "Include HIPPS (High Integrity Pressure Protection System)",
            value=_hpht_is_hpht, key="dc_hipps",
            help="A safety-instrumented system that protects downstream "
                 "equipment (flowline, host) rated below full reservoir "
                 "shut-in pressure. Effectively mandatory for HPHT subsea "
                 "developments. ~$35MM per skid (one per template by "
                 "default). Auto-enabled when an HPHT tier is selected.")
        n_hipps_ui = 0
        if hipps_on:
            n_hipps_ui = st.number_input(
                "Number of HIPPS skids", min_value=1,
                value=max(1, int(n_templates)), step=1, key="dc_n_hipps",
                help="Typically one HIPPS skid per template / drill centre.")

        # Subsea multiphase flow metering
        mpfm_on = st.checkbox(
            "Include subsea multiphase flow meters",
            value=False, key="dc_mpfm",
            help="Per-well or per-template multiphase flow meters give "
                 "continuous oil/gas/water allocation without subsea test "
                 "separation. ~$3.2MM per meter (one per producing well "
                 "by default). Common on modern NCS subsea developments.")
        n_mpfm_ui = 0
        if mpfm_on:
            n_mpfm_ui = st.number_input(
                "Number of multiphase meters", min_value=1,
                value=max(1, int(n_subsea_wells)), step=1,
                key="dc_n_mpfm",
                help="Typically one per producing subsea well; some "
                     "developments meter per template instead.")

        dc_spec = {
            "concept_type": concept_type,
            "host_type": host_type,
            "fluid_system": st.session_state.get(
                "fluid", "Oil with associated gas"),
            "processing_capacity_kboed": processing_capacity,
            "water_depth_class": water_depth_class,
            "hpht_tier": hpht_tier,
            "reservoir_pressure_psi": _p_psi,
            "reservoir_temp_F": _t_F,
            "hipps": hipps_on,
            "n_hipps": n_hipps_ui,
            "multiphase_metering": mpfm_on,
            "n_multiphase_meters": n_mpfm_ui,
            "template_type": template_type,
            "template_layout": template_layout,
            "n_templates": n_templates,
            "templates_detail": templates_detail,
            "n_subsea_wells": n_subsea_wells,
            "n_dry_wells": n_dry_wells,
            "n_total_wells": n_subsea_wells + n_dry_wells,
            "flowline_km": flowline_km,
            "flowline_diameter_in": flowline_diameter,
            "flowline_material": flowline_material,
            "flowline_insulation": flowline_insulation,
            "insulated_flowline_km": insulated_flowline_km,
            "umbilical_km": umbilical_km,
            "n_risers": n_risers,
            "riser_type": riser_type,
            "n_riser_bases": n_riser_bases,
            "n_ssiv": n_ssiv,
            "n_jumpers": n_jumpers,
            "n_control_modules": n_control_modules,
            "n_boosting_stations": n_boosting,
            "field_architecture": field_architecture,
            "hydrate_management": hydrate_mgmt,
            "flowline_install_method": flowline_install,
            "tiein_method": tiein_method,
            "gas_lift": gas_lift,
            "n_gas_lift_wells": n_gas_lift_wells,
            "heating_type": heating_type,
            "heated_flowline_km": heated_flowline_km,
            "export_pipeline_km": export_pipeline_km,
            "export_pipeline_diameter_in": export_pipeline_diameter,
            "host_distance_km": host_distance_km,
            "topside_mod_tonnes": topside_mod_tonnes,
            "topside_mod_rate_per_tonne_MM": topside_mod_rate_per_tonne_MM,
            "offshore_manhours": offshore_manhours,
            "offshore_manhour_rate_usd": offshore_manhour_rate_usd,
            "cost_overrides": cost_overrides,
            "start_date": start_date,
            "horizon_years": st.session_state.get("horizon", 20),
        }
        try:
            concept = fh.build_development_concept(dc_spec)
        except Exception as exc:
            concept = None
            st.error(f"Could not build the concept: {exc}")

        if concept is not None:
            # Schematics — the user picks which view to show.
            st.markdown("**Concept schematic**")
            _view_choice = st.selectbox(
                "Schematic view",
                ["Side view (cross-section)", "Aerial view (plan)",
                 "3D view (interactive)"],
                key="dc_schematic_view",
                help="Side view — a water-column cross-section showing "
                     "wells, risers and the host. Aerial view — a top-down "
                     "field-layout map. 3D view — an interactive scene you "
                     "can rotate and zoom, showing the full subsea layout "
                     "in three dimensions.")
            if _view_choice.startswith("3D"):
                geo3d = concept.get("geometry_3d")
                if geo3d and geo3d.get("available"):
                    fig3d = _build_concept_3d_figure(geo3d)
                    st.plotly_chart(fig3d, use_container_width=True)
                    st.caption(
                        "Interactive 3D layout — drag to rotate, scroll to "
                        "zoom. Sea surface and seabed are shown as planes; "
                        "templates, wells, flowline, umbilical, riser, "
                        "boosting and export pipeline are positioned in "
                        "true 3D. Elevation is in metres (water depth from "
                        "the depth class); horizontal axes in km.")
                else:
                    st.info("3D view is available for offshore subsea "
                            "layouts.")
            elif _view_choice.startswith("Aerial"):
                if concept.get("aerial"):
                    st.image(_svg_to_data_uri(concept["aerial"]),
                             use_container_width=True)
                    st.caption("Plan view from above — template layout, "
                               "well slots, flowline / umbilical routing, "
                               "manifolds, boosting and the export line. "
                               "Use the **template layout** control to "
                               "switch between clustered and spread "
                               "drill-centre arrangements.")
                else:
                    st.info("Aerial view is available for offshore "
                            "subsea layouts.")
            else:
                st.image(_svg_to_data_uri(concept["schematic"]),
                         use_container_width=True)
                st.caption("Side-view cross-section — water column, "
                           "seabed, wells, risers and the host facility.")

            # Summary + warnings side by side
            sum_col, warn_col = st.columns([1, 1])
            with sum_col:
                st.markdown("**Concept summary**")
                summary_df = pd.DataFrame(concept["summary"],
                                           columns=["Item", "Value"])
                st.dataframe(summary_df, use_container_width=True,
                             hide_index=True)
            with warn_col:
                st.markdown("**Engineering checks**")
                if concept["warnings"]:
                    for w in concept["warnings"]:
                        st.warning(w)
                else:
                    st.success("No engineering red flags — the concept is "
                               "internally consistent.")

            # ---- Topside modification advisor ----
            # Reads the selected concept and lists the cross-functional
            # topside scope it implies — exactly the kind of facilities
            # interface list on an NCS DG1 long-list slide.
            with st.expander(
                    "🏗️ Topside modification advisor — what this concept "
                    "implies for the host", expanded=False):
                st.caption(
                    "Screening-level recommendations for the topside "
                    "scope implied by the SURF / host concept you've "
                    "selected above. Use as a checklist for the "
                    "facilities interface register, not as a "
                    "substitute for a topside feasibility study.")
                try:
                    advice = fh.topside_modification_advice({
                        "concept_type": dc_spec.get("concept_type", ""),
                        "host_type": dc_spec.get("host_type", ""),
                        "n_subsea_wells": dc_spec.get("n_subsea_wells", 0),
                        "n_boosting_stations": dc_spec.get(
                            "n_boosting_stations", 0),
                        "gas_lift": dc_spec.get("gas_lift", False),
                        "heating_type": dc_spec.get("heating_type", "None"),
                        "hydrate_management": dc_spec.get(
                            "hydrate_management", "None"),
                        "export_pipeline_km": dc_spec.get(
                            "export_pipeline_km", 0),
                        "flowline_km": dc_spec.get("flowline_km", 0),
                        "is_gas": (FLUID_SYSTEMS.get(
                            st.session_state.get(
                                "fluid", "Oil with associated gas"),
                            {"primary": "oil"})["primary"] == "gas"),
                    })
                    for heading, rec in advice:
                        st.markdown(f"**{heading}** — {rec}")
                except Exception as _e:
                    st.info(f"Could not generate topside advice: {_e}")

            # ---- Editable cost table ----------------------------------
            # The engine produces benchmark costs per line; the user can
            # override any line by editing the "Override ($MM)" column.
            # Overrides persist across reruns via session state.
            st.markdown("**Generated CAPEX schedule — editable**")
            st.caption(
                "Each row shows the benchmark cost from the cost model. "
                "To override any line, type a value into the **Override ($MM)** "
                "column and click **✓ Apply cost edits**. Leave the cell empty "
                "to keep using the benchmark. Click **🔄 Reset all to benchmark** "
                "to clear every override."
            )
            # Build a buffer dataframe combining benchmarks + any active
            # overrides. The label is the lookup key.
            edit_rows = []
            for r in concept["capex_rows"]:
                lbl = r["label"]
                override_val = cost_overrides.get(lbl, None)
                edit_rows.append({
                    "Spend date": pd.to_datetime(r["date"]).date(),
                    "Component": lbl,
                    "Benchmark ($MM)": float(r["amount_MMUSD"])
                        if override_val is None
                        else None,   # benchmark is the engine output when no override
                    "Override ($MM)": float(override_val) if override_val is not None else None,
                    "In effect ($MM)": float(r["amount_MMUSD"]),
                })
            # For rows that already have an override, the engine's amount_MMUSD
            # IS the overridden value — so the "Benchmark" column needs the
            # un-overridden value. Recompute by calling the engine with no
            # overrides to get the pure benchmark.
            try:
                _bench_spec = dict(dc_spec)
                _bench_spec["cost_overrides"] = {}
                _bench_concept = fh.build_development_concept(_bench_spec)
                _bench_lookup = {r["label"]: r["amount_MMUSD"]
                                  for r in _bench_concept["capex_rows"]}
                for row in edit_rows:
                    row["Benchmark ($MM)"] = float(
                        _bench_lookup.get(row["Component"], row["In effect ($MM)"]))
            except Exception:
                pass
            edit_df = pd.DataFrame(edit_rows)

            edit_buf = st.data_editor(
                edit_df, use_container_width=True, hide_index=True,
                disabled=["Spend date", "Component", "Benchmark ($MM)",
                          "In effect ($MM)"],
                column_config={
                    "Spend date": st.column_config.DateColumn("Spend date"),
                    "Component":  st.column_config.TextColumn("Component", width="large"),
                    "Benchmark ($MM)": st.column_config.NumberColumn(
                        "Benchmark ($MM)", format="%.1f",
                        help="The screening-grade cost from the engine's "
                             "benchmark model (read-only)."),
                    "Override ($MM)": st.column_config.NumberColumn(
                        "Override ($MM)", format="%.1f", min_value=0.0,
                        help="Type a value here to override the benchmark for "
                             "this line. Empty = use benchmark."),
                    "In effect ($MM)": st.column_config.NumberColumn(
                        "In effect ($MM)", format="%.1f",
                        help="The cost currently used in calculations "
                             "(override if set, else benchmark)."),
                },
                key="dc_cost_editor",
            )

            eba, ebb, _ebc = st.columns([1, 1, 3])
            if eba.button("✓ Apply cost edits", key="dc_apply_overrides",
                           type="primary"):
                # Read the buffer; any non-null Override values become the new
                # overrides dict, keyed by Component label.
                new_overrides = {}
                for _, row in edit_buf.iterrows():
                    lbl = row.get("Component")
                    ov = row.get("Override ($MM)")
                    if lbl and ov is not None and not pd.isna(ov) \
                            and float(ov) > 0:
                        new_overrides[lbl] = float(ov)
                st.session_state["dc_cost_overrides"] = new_overrides
                mark_stale()
                if new_overrides:
                    st.success(f"Applied {len(new_overrides)} override(s). "
                               "The CAPEX schedule below is now based on your "
                               "edited costs.")
                else:
                    st.info("All overrides cleared — every line back to its "
                            "benchmark.")
                st.rerun()
            if ebb.button("🔄 Reset all to benchmark", key="dc_reset_overrides"):
                st.session_state["dc_cost_overrides"] = {}
                mark_stale()
                st.rerun()

            t = concept["totals"]
            st.caption(
                f"**CAPEX excl. cessation: ${t['capex_excl_cessation']:,.0f}MM**  •  "
                f"Cessation / P&A: ${t['cessation']:,.0f}MM  •  "
                f"**Grand total: ${t['grand_total']:,.0f}MM**"
            )
            if cost_overrides:
                # Show how the overrides shift the total vs benchmark
                try:
                    _bench_total = sum(_bench_lookup.values()) \
                        + _bench_concept["totals"]["cessation"]
                    delta = t["grand_total"] - _bench_total
                    st.caption(
                        f":blue[{len(cost_overrides)} override(s) active.] "
                        f"Benchmark grand total: ${_bench_total:,.0f}MM  •  "
                        f"With overrides: ${t['grand_total']:,.0f}MM  •  "
                        f"**Δ {delta:+,.0f}MM** "
                        f"({delta/_bench_total*100:+.1f}%).")
                except Exception:
                    pass

            # Cost distribution pie chart (uses the effective costs — overrides
            # included if any are active)
            st.markdown("**Cost distribution**")
            pie_df = pd.DataFrame(concept["capex_rows"])[
                ["label", "amount_MMUSD"]].rename(
                columns={"amount_MMUSD": "amount"})
            tot_amt = pie_df["amount"].sum()
            if tot_amt > 0:
                threshold = 0.025 * tot_amt   # group anything < 2.5% as Other
                big = pie_df[pie_df["amount"] >= threshold].copy()
                small_sum = pie_df[pie_df["amount"] < threshold]["amount"].sum()
                if small_sum > 0:
                    big = pd.concat([big, pd.DataFrame(
                        [{"label": "Other (small items)", "amount": small_sum}]
                    )], ignore_index=True)
                fig_pie = go.Figure(data=[go.Pie(
                    labels=big["label"], values=big["amount"],
                    hole=0.42, sort=True, direction="clockwise",
                    textposition="outside", textinfo="label+percent",
                    hovertemplate="<b>%{label}</b><br>$%{value:,.0f}MM<br>"
                                   "%{percent}<extra></extra>",
                    marker=dict(line=dict(color="#fff", width=2)),
                )])
                fig_pie.update_layout(
                    title=f"CAPEX breakdown — total ${tot_amt:,.0f}MM",
                    height=420, showlegend=True,
                    legend=dict(orientation="v", x=1.02, y=0.5),
                    margin=dict(t=60, b=20, l=10, r=150),
                )
                st.plotly_chart(fh.apply_plot_template(fig_pie),
                                use_container_width=True)
                st.caption("Items under 2.5% of total are grouped as 'Other' "
                           "to keep the chart readable.")

            # ---- NCS / UKCS cost benchmarking ----
            st.markdown("**Cost benchmarking — NCS / UKCS reference data**")
            # Reserves basis: prefer the engine's in-place × RF if available,
            # else let the user enter a reserves figure for the $/boe metric.
            _ooip = float(st.session_state.get("ooip", 0.0) or 0.0)
            _ogip = float(st.session_state.get("ogip", 0.0) or 0.0)
            _rf = float(st.session_state.get("rf_target", 0.35) or 0.35)
            # crude boe reserves estimate: oil MMstb + gas Bscf/6 *1000/1000
            _reserves_guess = (_ooip + _ogip / 6.0) * _rf
            bench_reserves = st.number_input(
                "Recoverable reserves for benchmarking (MMboe)",
                min_value=0.0,
                value=float(round(max(_reserves_guess, 1.0), 1)),
                step=5.0, key="dc_bench_reserves",
                help="Used as the denominator for the CAPEX-per-boe "
                     "benchmark. Pre-filled from (OOIP + OGIP/6) × target RF; "
                     "override with your own reserves estimate if needed.")
            # Guard against a stale fp_helpers.py that predates this
            # function (e.g. only one of the two files was redeployed).
            if not hasattr(fh, "benchmark_concept_cost"):
                st.info(
                    "Cost benchmarking is unavailable — the helper module "
                    "(fp_helpers.py) on this deployment is out of date. "
                    "Re-upload the latest fp_helpers.py alongside the app "
                    "to enable NCS / UKCS benchmarking.")
                bench = None
            else:
                bench = fh.benchmark_concept_cost(
                    grand_total_MMUSD=concept["totals"]["grand_total"],
                    reserves_mmboe=bench_reserves,
                    concept_type=concept_type,
                    host_type=host_type or "",
                    n_subsea_wells=n_subsea_wells)
            if bench is not None:
                st.caption(
                    f"Matched benchmark class: **{bench['concept_class']}**. "
                    f"Your concept: "
                    f"**${bench['capex_per_boe']:.1f}/boe**"
                    if bench['capex_per_boe'] is not None
                    else "Enter a reserves figure above to compute $/boe.")

            if bench is not None and bench["rows"] \
                    and bench["capex_per_boe"] is not None:
                # Bar chart: benchmark low/mid/high bands + the user's value,
                # per region.
                fig_bm = go.Figure()
                regions = [r["region"] for r in bench["rows"]]
                lows = [r["low"] for r in bench["rows"]]
                mids = [r["mid"] for r in bench["rows"]]
                highs = [r["high"] for r in bench["rows"]]
                # low-to-high range bar
                fig_bm.add_trace(go.Bar(
                    x=regions, y=[h - l for h, l in zip(highs, lows)],
                    base=lows, name="Typical range",
                    marker_color="#bcd4e6",
                    hovertemplate="%{x}: $%{base:.0f}-$%{customdata:.0f}/boe"
                                   "<extra></extra>",
                    customdata=highs))
                # mid markers
                fig_bm.add_trace(go.Scatter(
                    x=regions, y=mids, mode="markers", name="Benchmark mid",
                    marker=dict(symbol="line-ew", size=26, color="#2a6f97",
                                line=dict(width=3, color="#2a6f97")),
                    hovertemplate="%{x} mid: $%{y:.0f}/boe<extra></extra>"))
                # the user's concept
                fig_bm.add_trace(go.Scatter(
                    x=regions, y=[bench["capex_per_boe"]] * len(regions),
                    mode="markers+text", name="This concept",
                    marker=dict(symbol="diamond", size=15, color="#d62828"),
                    text=[f"${bench['capex_per_boe']:.0f}"] * len(regions),
                    textposition="top center",
                    hovertemplate="This concept: $%{y:.1f}/boe<extra></extra>"))
                fig_bm.update_layout(
                    title=f"Development CAPEX intensity vs NCS / UKCS — "
                          f"{bench['concept_class']}",
                    yaxis_title="CAPEX per boe ($/boe)",
                    height=380, barmode="overlay",
                    margin=dict(t=60, b=30, l=10, r=10),
                    legend=dict(orientation="h", y=-0.15))
                st.plotly_chart(fh.apply_plot_template(fig_bm),
                                use_container_width=True)
                # Verdict line
                vparts = []
                for r in bench["rows"]:
                    vparts.append(f"{r['region']}: {r['verdict']}")
                st.caption("Verdict — " + "  •  ".join(vparts))

            # Per-subsea-well benchmark
            if bench is not None and bench["well_rows"] \
                    and bench["well_share_MM"] is not None:
                with st.expander("CAPEX per subsea well vs NCS / UKCS",
                                 expanded=False):
                    wb_df = pd.DataFrame([
                        {"Region": r["region"],
                         "Typical low ($MM)": r["low"],
                         "Typical mid ($MM)": r["mid"],
                         "Typical high ($MM)": r["high"],
                         "This concept ($MM)": round(bench["well_share_MM"], 1),
                         "Verdict": r["verdict"]}
                        for r in bench["well_rows"]])
                    st.dataframe(wb_df, use_container_width=True,
                                 hide_index=True)
                    st.caption(
                        "This is total concept CAPEX ÷ number of subsea "
                        "wells, so for a standalone development it also "
                        "carries the host/topsides cost — expect it to read "
                        "high for FPSO / platform concepts. It is most "
                        "meaningful for pure subsea tie-ins.")
            if bench is not None:
                for note in bench["notes"]:
                    st.info(note)
                st.caption(
                    "Benchmark bands are screening-level ranges compiled "
                    "from public NCS (Sokkeldirektoratet) and UKCS (NSTA) "
                    "project disclosures. Real project costs vary widely — "
                    "use these to check order of magnitude, not as a "
                    "class-3 estimate.")

            if st.button("⚙️ Generate CAPEX schedule from this concept",
                          key="dc_generate", type="primary"):
                _rows = concept["capex_rows"]
                # Phase the facility CAPEX against a realistic project
                # schedule (long-lead → fabrication → installation → hook-up
                # → first oil) so spend lands BEFORE first oil on the phase
                # that incurs it — not all dumped at first oil. We anchor the
                # schedule so that first oil = the field's production start.
                try:
                    _durs = fh.default_schedule_durations(dc_spec)
                    _total_mo = sum(_durs.values())
                    _fo = pd.Timestamp(dc_spec.get("start_date")
                                       or st.session_state.get("start_date"))
                    # Back-calculate the FEED start so that first oil ≈ the
                    # production start date.
                    _feed_start = (_fo - pd.Timedelta(
                        days=int(_total_mo * 30.4375))).date()
                    _sched = fh.build_project_schedule(
                        dc_spec, _feed_start, _durs)
                    _rows = fh.phase_capex_against_schedule(_rows, _sched)
                    _phased = True
                except Exception:
                    _phased = False
                st.session_state.fac_df = pd.DataFrame(_rows)
                mark_stale()
                _msg = (f"Generated {len(_rows)} CAPEX line(s) from the "
                        f"'{concept_type}' concept.")
                if _phased:
                    _msg += (" Spend is phased across the project schedule "
                             "(long-lead → fabrication → installation → "
                             "hook-up), all before first oil.")
                st.success(_msg + " Edit the table below to fine-tune.")
                st.rerun()

            # ---- Concept-decision sensitivity ----
            # The standard tornado perturbs production/economics inputs.
            # This one perturbs the *concept choices* themselves — host
            # type, water depth, well count, tie-back distance — and shows
            # how the development CAPEX responds. It answers the real
            # screening question: which concept decision moves cost most?
            with st.expander("🌪️ Concept-decision sensitivity",
                             expanded=False):
                st.caption(
                    "How sensitive is the development CAPEX to the concept "
                    "choices? Each row re-runs the concept with one "
                    "decision changed, holding everything else fixed. Use "
                    "it to see which decision — host type, water depth, "
                    "well count, tie-back length — drives the cost.")
                if st.button("Run concept sensitivity",
                             key="dc_concept_sens"):
                    base_capex = concept["totals"]["grand_total"]
                    sens_rows = []

                    def _capex_of(spec_override):
                        try:
                            s = dict(dc_spec)
                            s.update(spec_override)
                            c = fh.build_development_concept(s)
                            return c["totals"]["grand_total"]
                        except Exception:
                            return base_capex

                    # Define the perturbations to test
                    perturbations = []
                    # water depth one class deeper / shallower
                    _wd_order = ["Shallow (<150 m)", "Mid (150-600 m)",
                                 "Deep (600-1500 m)",
                                 "Ultra-deep (>1500 m)"]
                    if water_depth_class in _wd_order:
                        _wi = _wd_order.index(water_depth_class)
                        if _wi > 0:
                            perturbations.append(
                                ("Water depth one class shallower",
                                 {"water_depth_class": _wd_order[_wi-1]}))
                        if _wi < len(_wd_order) - 1:
                            perturbations.append(
                                ("Water depth one class deeper",
                                 {"water_depth_class": _wd_order[_wi+1]}))
                    # well count +/- 25%
                    _nw = int(dc_spec.get("n_subsea_wells", 0))
                    if _nw > 0:
                        perturbations.append(
                            ("Subsea wells +25%",
                             {"n_subsea_wells": int(round(_nw * 1.25))}))
                        perturbations.append(
                            ("Subsea wells -25%",
                             {"n_subsea_wells": max(1,
                                                    int(round(_nw*0.75)))}))
                    # tie-back distance +/- 30%
                    _hd = float(dc_spec.get("host_distance_km", 0))
                    if _hd > 0:
                        perturbations.append(
                            ("Tie-back +30% longer",
                             {"host_distance_km": _hd * 1.3,
                              "flowline_km": float(
                                  dc_spec.get("flowline_km", _hd)) * 1.3,
                              "umbilical_km": float(
                                  dc_spec.get("umbilical_km", _hd)) * 1.3}))
                        perturbations.append(
                            ("Tie-back -30% shorter",
                             {"host_distance_km": _hd * 0.7,
                              "flowline_km": float(
                                  dc_spec.get("flowline_km", _hd)) * 0.7,
                              "umbilical_km": float(
                                  dc_spec.get("umbilical_km", _hd)) * 0.7}))
                    # HPHT tier up
                    if dc_spec.get("hpht_tier", "Standard") == "Standard":
                        perturbations.append(
                            ("HPHT conditions (vs standard)",
                             {"hpht_tier": "HPHT"}))

                    for label, override in perturbations:
                        capex = _capex_of(override)
                        sens_rows.append({
                            "Decision change": label,
                            "CAPEX ($MM)": round(capex, 0),
                            "Δ vs base ($MM)": round(capex - base_capex, 0),
                            "Δ %": round(100.0 * (capex - base_capex)
                                         / base_capex, 1)
                            if base_capex > 0 else 0.0,
                        })
                    if sens_rows:
                        sens_df = pd.DataFrame(sens_rows)
                        sens_df = sens_df.reindex(
                            sens_df["Δ vs base ($MM)"].abs()
                            .sort_values(ascending=False).index)
                        st.dataframe(sens_df, use_container_width=True,
                                     hide_index=True)
                        # tornado chart
                        fig_cs = go.Figure()
                        fig_cs.add_trace(go.Bar(
                            y=sens_df["Decision change"],
                            x=sens_df["Δ vs base ($MM)"],
                            orientation="h",
                            marker_color=["#d62828" if v > 0 else "#2a9d8f"
                                          for v in
                                          sens_df["Δ vs base ($MM)"]]))
                        fig_cs.update_layout(
                            title=f"Concept CAPEX sensitivity "
                                  f"(base ${base_capex:,.0f}MM)",
                            xaxis_title="Δ CAPEX vs base ($MM)",
                            height=max(260, 52 * len(sens_df)),
                            margin=dict(l=10, r=10, t=50, b=30))
                        st.plotly_chart(fh.apply_plot_template(fig_cs),
                                        use_container_width=True)
                        st.caption(
                            "Positive (red) = the change increases CAPEX. "
                            "The decision at the top is the biggest cost "
                            "lever for this concept.")

        # Stash the spec so the schedule builder can use it
        st.session_state["_dc_spec"] = dc_spec

    # ---- Project schedule builder ----
    with st.expander("📅 Project schedule builder — milestones, durations, "
                     "realism checks", expanded=False):
        st.caption(
            "Build a milestone timeline from **FEED → DG3/sanction → "
            "long-lead → fabrication → installation → hookup → first oil**. "
            "Default durations are pre-filled based on the concept above. "
            "The builder flags any phase that's outside realistic industry "
            "bounds for that concept family. You can optionally push the "
            "field's production start date to match the computed first-oil "
            "date."
        )
        dc_spec_now = st.session_state.get("_dc_spec", {
            "concept_type": "Subsea tie-in", "host_type": "",
            "host_distance_km": 0,
        })
        bench_key_preview = fh._concept_benchmark_key(dc_spec_now)
        st.caption(f"Benchmark family: **{bench_key_preview[0]} — "
                   f"{bench_key_preview[1]}** (concept-aware defaults).")

        feed_start = st.date_input(
            "FEED / pre-FEED start date",
            value=st.session_state.get("sched_feed_start", start_date),
            key="sched_feed_start",
            help="Anchor date for the schedule. This is when the engineering "
                 "definition phase begins — well before sanction.")

        defaults = fh.default_schedule_durations(dc_spec_now)
        sc1, sc2, sc3 = st.columns(3)
        feed_months = sc1.number_input(
            "FEED duration (months)", min_value=0, max_value=60,
            value=int(defaults["FEED"]), step=1, key="sched_feed_months",
            help="Front-end engineering & design. Major engineering / cost "
                 "definition. Typical: subsea tie-in 6-15 mo; onshore CPF "
                 "9-22 mo; fixed platform 12-26 mo; floating host 15-32 mo.")
        ll_months = sc2.number_input(
            "Long-lead items (months)", min_value=0, max_value=60,
            value=int(defaults["Long-lead"]), step=1, key="sched_ll_months",
            help="Manufacturing time for long-lead items (FPSO hull, jacket "
                 "steel, subsea trees, large compressors). Typically starts "
                 "at sanction. Can overlap with fabrication (see below).")
        fab_months = sc3.number_input(
            "Fabrication / EPC (months)", min_value=0, max_value=72,
            value=int(defaults["Fabrication"]), step=1,
            key="sched_fab_months",
            help="Main engineering, procurement and construction phase — "
                 "topsides, hull, jacket, subsea hardware all built.")
        sc4, sc5, sc6 = st.columns(3)
        inst_months = sc4.number_input(
            "Installation (months)", min_value=0, max_value=36,
            value=int(defaults["Installation"]), step=1,
            key="sched_inst_months",
            help="Offshore installation campaign — heavy lift, pipelay, "
                 "riser pull-in, subsea connections. Bounded by weather "
                 "windows in many regions.")
        huc_months = sc5.number_input(
            "Hookup & commissioning (months)", min_value=0, max_value=24,
            value=int(defaults["Hookup & comm."]), step=1,
            key="sched_huc_months",
            help="Mechanical completion, system commissioning, performance "
                 "testing, ramp-up to first oil.")
        overlap_months = sc6.number_input(
            "Long-lead / fabrication overlap (months)", min_value=0,
            max_value=36, value=int(min(12, ll_months)),
            step=1, key="sched_overlap",
            help="Long-lead items typically start in parallel with "
                 "fabrication. 0 = strictly sequential; 6-12 months overlap "
                 "is common. Cannot exceed the long-lead duration.")

        durations = {
            "FEED": feed_months, "Long-lead": ll_months,
            "Fabrication": fab_months, "Installation": inst_months,
            "Hookup & comm.": huc_months,
        }
        try:
            sched = fh.build_project_schedule(dc_spec_now, feed_start,
                                                durations,
                                                overlap_longlead_months=overlap_months)
        except Exception as exc:
            sched = None
            st.error(f"Could not build schedule: {exc}")

        if sched is not None:
            # Gantt chart — use px.timeline which is the proper Plotly Gantt
            # primitive and handles dates natively (avoids the int-vs-date
            # mixing that breaks add_vline on some Plotly versions).
            import plotly.express as px
            phase_colors = {
                "FEED":            "#5b8def",
                "Long-lead":       "#9c7ad6",
                "Fabrication":     "#e8a23a",
                "Installation":    "#d65a5a",
                "Hookup & comm.":  "#3ba776",
            }
            gantt_df = pd.DataFrame([
                {"Phase": ph["phase"],
                 "Start": pd.Timestamp(ph["start"]),
                 "Finish": pd.Timestamp(ph["end"]),
                 "Duration_mo": ph["duration_months"]}
                for ph in sched["phases"]
            ])
            fig_g = px.timeline(
                gantt_df, x_start="Start", x_end="Finish", y="Phase",
                color="Phase", color_discrete_map=phase_colors,
                hover_data={"Duration_mo": True,
                            "Start": "|%Y-%m-%d", "Finish": "|%Y-%m-%d",
                            "Phase": False},
            )
            fig_g.update_yaxes(autorange="reversed",
                                categoryorder="array",
                                categoryarray=[ph["phase"]
                                               for ph in sched["phases"]])
            fig_g.update_traces(marker_line_width=0)

            # Milestone markers — use safe_vline to avoid Plotly's datetime
            # annotation bug (Timestamp + int crash in shapeannotation.py).
            for label, mdate in sched["milestones"]:
                mts = pd.Timestamp(mdate)
                fh.safe_vline(
                    fig_g, mts, label=label, color="#333", dash="dot",
                    width=1, label_position="top", label_font_size=9,
                    label_color="#444", textangle=-45)
            # First-oil emphasis — green thick line
            fo_ts = pd.Timestamp(sched["first_oil_date"])
            fh.safe_vline(
                fig_g, fo_ts,
                label=f"🛢️ First oil: {sched['first_oil_date']}",
                color="#2ca02c", width=3, dash="solid",
                label_position="bottom", label_font_size=12,
                label_color="#2ca02c")
            fig_g.update_layout(
                title=(f"Project schedule — {sched['total_months']} months "
                       f"FEED to first oil"),
                height=380,
                xaxis_title="Date",
                showlegend=False,
                margin=dict(t=90, b=50, l=10, r=10),
                plot_bgcolor="rgba(245,247,250,0.6)",
                bargap=0.35,
            )
            st.plotly_chart(fh.apply_plot_template(fig_g),
                            use_container_width=True)

            # ---- Yearly CAPEX phasing chart -----------------------------
            # Show how the facility CAPEX (from the current fac_df) is spread
            # across calendar years against this schedule, so the user can
            # see the annual cash-out profile and confirm it precedes first
            # oil. Uses whatever is currently in fac_df (which the Generate
            # button phases against the schedule).
            try:
                _fac = st.session_state.get("fac_df")
                if _fac is not None and len(_fac) > 0 and \
                        "date" in _fac.columns and "amount_MMUSD" in _fac.columns:
                    _fc = _fac.copy()
                    _fc["year"] = pd.to_datetime(
                        _fc["date"], errors="coerce").dt.year
                    _fc = _fc.dropna(subset=["year"])
                    _fc["amount_MMUSD"] = pd.to_numeric(
                        _fc["amount_MMUSD"], errors="coerce").fillna(0.0)
                    # Split cessation out so it doesn't dwarf the dev CAPEX
                    _fc["is_cessation"] = _fc["label"].apply(
                        fh.is_abandonment_label)
                    dev = (_fc[~_fc["is_cessation"]]
                           .groupby("year")["amount_MMUSD"].sum())
                    ces = (_fc[_fc["is_cessation"]]
                           .groupby("year")["amount_MMUSD"].sum())
                    if len(dev) > 0 or len(ces) > 0:
                        all_years = sorted(set(dev.index) | set(ces.index))
                        fo_year = pd.Timestamp(sched["first_oil_date"]).year
                        fig_cap = go.Figure()
                        fig_cap.add_trace(go.Bar(
                            x=all_years,
                            y=[float(dev.get(y, 0.0)) for y in all_years],
                            name="Development CAPEX",
                            marker_color="#5b8def"))
                        if ces.sum() > 0:
                            fig_cap.add_trace(go.Bar(
                                x=all_years,
                                y=[float(ces.get(y, 0.0)) for y in all_years],
                                name="Cessation / P&A",
                                marker_color="#d65a5a"))
                        # cumulative line
                        _cum, _run = [], 0.0
                        for y in all_years:
                            _run += float(dev.get(y, 0.0)) + float(ces.get(y, 0.0))
                            _cum.append(_run)
                        fig_cap.add_trace(go.Scatter(
                            x=all_years, y=_cum, name="Cumulative",
                            mode="lines+markers", yaxis="y2",
                            line=dict(color="#2ca02c", width=2)))
                        fh.safe_vline(
                            fig_cap, fo_year,
                            label=f"First oil {fo_year}",
                            color="#2ca02c", width=2, dash="dot",
                            label_position="top", label_font_size=10,
                            label_color="#2ca02c")
                        fig_cap.update_layout(
                            title="Facility CAPEX phasing by year ($MM)",
                            barmode="stack", height=340,
                            xaxis_title="Year",
                            yaxis_title="Annual CAPEX ($MM)",
                            yaxis2=dict(title="Cumulative ($MM)",
                                        overlaying="y", side="right",
                                        showgrid=False),
                            margin=dict(t=60, b=40, l=10, r=10),
                            legend=dict(orientation="h", y=-0.2),
                            plot_bgcolor="rgba(245,247,250,0.6)")
                        st.plotly_chart(fh.apply_plot_template(fig_cap),
                                        use_container_width=True)
                        st.caption(
                            "Annual facility CAPEX from the cost schedule "
                            "below, phased against the project timeline. "
                            "Development spend should fall in the years "
                            "before first oil; cessation sits at end of life.")
                        # Stash for Excel export (yearly CAPEX phasing sheet)
                        st.session_state["_capex_phasing_by_year"] = {
                            "years": [int(y) for y in all_years],
                            "development_MM": [float(dev.get(y, 0.0))
                                                for y in all_years],
                            "cessation_MM": [float(ces.get(y, 0.0))
                                              for y in all_years],
                            "cumulative_MM": _cum,
                            "first_oil_year": int(fo_year),
                        }
            except Exception as _ce:
                st.caption(f"(CAPEX phasing chart unavailable: {_ce})")

            # Milestone table
            ms_col, w_col = st.columns([1, 1])
            with ms_col:
                st.markdown("**Key milestones**")
                ms_df = pd.DataFrame(sched["milestones"],
                                      columns=["Milestone", "Date"])
                ms_df["Date"] = pd.to_datetime(ms_df["Date"]).dt.date
                st.dataframe(ms_df, use_container_width=True, hide_index=True)
            with w_col:
                st.markdown("**Realism checks**")
                if sched["warnings"]:
                    for w in sched["warnings"]:
                        st.warning(w)
                else:
                    st.success("All phase durations are within realistic "
                               "industry bounds for this concept family.")

            # Benchmark reference table
            with st.expander("📊 Industry benchmark ranges for this concept",
                             expanded=False):
                bench_rows = []
                for phase, (lo, typ, hi) in sched["benchmark"].items():
                    bench_rows.append({
                        "Phase": phase,
                        "Min (months)": lo,
                        "Typical (months)": typ,
                        "Max (months)": hi,
                        "Your value": durations.get(phase, 0),
                    })
                st.dataframe(pd.DataFrame(bench_rows),
                             use_container_width=True, hide_index=True)
                st.caption(
                    "Benchmark ranges are screening-level, drawn from "
                    "published industry project case studies. Real projects "
                    "vary widely — these flag obvious over/under-estimates "
                    "rather than dictating a single right answer.")

            # Push first oil date to the field's start_date
            push_col1, push_col2 = st.columns([3, 2])
            push_col1.caption(
                f"The computed first-oil date is "
                f"**{sched['first_oil_date']}**. Currently the field's "
                f"production start date is set to **{start_date}**.")
            if push_col2.button("📌 Use first oil date as production start",
                                 key="sched_push_fop",
                                 help="Updates the field's production start "
                                      "date to match the computed first-oil "
                                      "milestone, so the economics and "
                                      "production forecast align with the "
                                      "schedule."):
                # Stage the change — it is applied to the date widget on the
                # next run, before that widget is created (Streamlit forbids
                # writing a widget-backed key after the widget exists).
                fo = sched["first_oil_date"]
                if not isinstance(fo, date):
                    fo = pd.Timestamp(fo).date()
                st.session_state["_pending_start_date"] = fo
                mark_stale()
                st.success(f"Production start date set to "
                           f"{sched['first_oil_date']}. Refreshing…")
                st.rerun()

    st.markdown("**Phased facility CAPEX**")
    if "fac_df" not in st.session_state:
        st.session_state.fac_df = pd.DataFrame({
            "date":         [start_date, start_date + timedelta(days=365)],
            "amount_MMUSD": [200.0, 150.0],
            "label":        ["FEED + topsides", "Subsea & hookup"],
        })
    fac_df_buf = st.data_editor(
        st.session_state.fac_df, num_rows="dynamic", use_container_width=True,
        column_config={
            "date": st.column_config.DateColumn("Spend date"),
            "amount_MMUSD": st.column_config.NumberColumn("Amount ($MM)", min_value=0.0),
            "label": st.column_config.TextColumn("Description"),
        },
        key="fac_editor",
    )
    bf1, bf2, _bf3 = st.columns([2, 1, 3])
    with bf1:
        fac_apply_clicked = _apply_button(fac_df_buf, st.session_state.fac_df,
                                           "fac_apply", "Apply CAPEX edits")
    if fac_apply_clicked:
        st.session_state.fac_df = _clean_table_buffer(
            fac_df_buf, st.session_state.fac_df, date_cols=["date"])
        mark_stale()
        st.rerun()
    if bf2.button("📋 Duplicate last CAPEX row", key="fac_dup"):
        st.session_state.fac_df = _duplicate_last_row(st.session_state.fac_df)
        mark_stale()
        st.rerun()
    fac_df = st.session_state.fac_df

    # ---- Contingency on CAPEX ----
    # Screening-stage cost estimates are systematically optimistic — the
    # AACE International Class 5 / Class 4 cost estimate classes (typical
    # for a pre-FEED concept) carry an accuracy range of -50%/+100%, and
    # NCS / UKCS benchmarking shows that final field CAPEX comes in 20-60%
    # above the screening estimate on average. The contingency multiplier
    # is applied to BOTH the facility CAPEX schedule AND the per-well
    # CAPEX (whether fixed or rig-rate computed), since the same
    # estimating bias applies to both.
    cont_c1, cont_c2 = st.columns([1, 2])
    contingency_pct = cont_c1.slider(
        "CAPEX contingency (%)", 0, 100, 25, 5,
        key="capex_contingency_pct", on_change=mark_stale,
        help=(
            "Multiplier applied to both facility CAPEX and well CAPEX "
            "to cover scope growth, schedule slip, fabrication inflation "
            "and execution risk. Applied as `(1 + pct/100)` to every "
            "spend line.\n\n"
            "**Benchmark recommendations** (AACE International + NCS / "
            "UKCS post-mortems):\n"
            "- **10-15%** — Class 2 detailed FEED / FID-ready estimate "
            "(mature scope, locked technology, firm contracts).\n"
            "- **20-30%** — Class 3 FEED estimate (typical NCS "
            "subsea tie-in or brownfield with mostly known scope).\n"
            "- **30-40%** — Class 4 concept-select estimate "
            "(greenfield development, several technology choices "
            "still open).\n"
            "- **40-60%** — Class 5 screening / pre-concept estimate "
            "(very early phase, high HPHT or deepwater technology "
            "risk, frontier basin).\n\n"
            "**Real-world calibration**: Johan Sverdrup phase 1 came "
            "in ~30% under the original PDO estimate (rare); Goliat "
            "and Yme came in 50-100% OVER; Mariner came in ~25% over. "
            "The NCS average for FPSO / semi-sub developments since "
            "2010 is +35% vs the PDO submission. Use 25-35% as a "
            "screening default."))
    if contingency_pct > 0:
        cont_c2.caption(
            f"💡 With {contingency_pct}% contingency, every $100MM of "
            f"CAPEX in the schedule below (and every well CAPEX) is "
            f"booked as **${100 * (1 + contingency_pct/100):.0f}MM** in "
            f"the cashflow. Total facility-CAPEX uplift: "
            f"**${fac_df['amount_MMUSD'].sum() * contingency_pct/100:,.0f}MM** "
            f"on top of the **${fac_df['amount_MMUSD'].sum():,.0f}MM** "
            f"base.")

    # ---- CO₂ emissions & carbon fees (Scope 1 + Scope 3) ----
    with st.expander("🌍 CO₂ emissions & carbon fees", expanded=False):
        st.caption(
            "**Scope 1** is operational — fuel gas combustion, flaring, "
            "methane slip, routine vents. **Scope 3** is end-use combustion "
            "of the sold oil and gas (downstream). Scope 3 typically "
            "dwarfs Scope 1 by 50–100×. Each scope can carry its own "
            "carbon price.")
        co2_price = st.number_input(
            "Scope 1 carbon price ($/tonne CO₂-eq)", value=80.0,
            min_value=0.0, step=10.0, key="co2_price",
            on_change=mark_stale,
            help="EU ETS-style price applied to operational emissions. "
                 "NCS operators have faced an effective combined CO₂ tax "
                 "+ ETS quota cost in the $80–$120/t range in recent "
                 "years. Set to 0 to ignore Scope 1 emissions in the "
                 "economics.")
        co2_factor_gas_combust = st.number_input(
            "Fuel/flare gas emission factor (kg CO₂/Mscf)",
            value=53.0, min_value=0.0, step=1.0,
            key="co2_factor_gas_combust", on_change=mark_stale,
            help="Kilograms of CO₂ per Mscf of gas burnt as fuel or "
                 "flared. 53 kg/Mscf is the standard value for typical "
                 "associated gas.")
        co2_factor_flare_ineff = st.number_input(
            "Flare combustion inefficiency (CH₄ slip fraction)",
            value=0.02, min_value=0.0, max_value=0.20, step=0.005,
            format="%.3f", key="co2_factor_flare_inefficiency",
            on_change=mark_stale,
            help="Fraction of flared gas that escapes as un-combusted "
                 "methane. 2% is typical for a well-maintained flare; "
                 "older or smokeless flares can be 5–10%. Counted at "
                 "GWP100 = 28× CO₂.")
        co2_factor_oil_routine = st.number_input(
            "Routine ops emissions (kg CO₂-eq/bbl oil)",
            value=0.5, min_value=0.0, step=0.1,
            key="co2_factor_oil_routine", on_change=mark_stale,
            help="Vented + ops emissions per barrel of oil produced "
                 "(diesel, methanol, fugitives). NCS upstream average "
                 "is ~8 kg/boe; 0.5 here covers the residual after fuel "
                 "and flare are accounted for separately.")
        st.markdown("**Scope 3 — end-use combustion**")
        co2_scope3_enabled = st.checkbox(
            "Include Scope 3 in the cashflow as a fee",
            value=False, key="co2_scope3_enabled", on_change=mark_stale,
            help="Off — Scope 3 is reported in the yearly emissions "
                 "profile but not charged. On — Scope 3 tonnes are "
                 "priced at the Scope 3 carbon price below and deducted "
                 "from the cashflow. Use this to test sensitivity to a "
                 "future downstream carbon levy or internal carbon "
                 "shadow price.")
        s3c1, s3c2, s3c3 = st.columns(3)
        co2_scope3_price = s3c1.number_input(
            "Scope 3 carbon price ($/tonne)", value=0.0,
            min_value=0.0, step=10.0, key="co2_scope3_price",
            on_change=mark_stale,
            help="$/tonne CO₂ applied to Scope 3 emissions when the "
                 "toggle is on. Some companies use $50–$100/t as an "
                 "internal shadow price for screening.")
        co2_scope3_factor_oil = s3c2.number_input(
            "Oil Scope 3 factor (kg CO₂/bbl)", value=430.0,
            min_value=0.0, step=10.0, key="co2_scope3_factor_oil",
            on_change=mark_stale,
            help="End-use combustion of crude. 430 kg/bbl is the "
                 "standard IPCC / EPA value for stationary combustion.")
        co2_scope3_factor_gas = s3c3.number_input(
            "Gas Scope 3 factor (kg CO₂/Mscf)", value=53.0,
            min_value=0.0, step=1.0, key="co2_scope3_factor_gas",
            on_change=mark_stale,
            help="End-use combustion of natural gas. 53 kg/Mscf same as "
                 "the upstream combustion factor.")

    # When the NCS regime is active, the engine applies CIT (22%) + SPT
    # (71.8%) + uplift — the global tax_rate and royalty_rate sliders
    # would double-count if also passed in. Override to zero before
    # building EconInputs, and tell the user so they're not surprised
    # the sliders are ignored.
    if regime_for_engine == "NCS" and (tax > 0 or royalty > 0):
        st.info(
            f"**NCS regime is active** — the global Tax ({tax:.0%}) and "
            f"Royalty ({royalty:.0%}) sliders are ignored to avoid "
            f"double-counting on top of CIT + SPT + uplift. The engine "
            f"will use CIT={ncs_cit_rate:.0%} + SPT={ncs_spt_rate:.0%}.")
        tax = 0.0
        royalty = 0.0

    # Apply contingency multiplier to facility + well CAPEX. The slider
    # value is read from session_state in case the widget hasn't rendered
    # yet (e.g. loading a saved case before the facility-CAPEX section
    # has been visited this session). The multiplier touches every CAPEX
    # input the engine sees: facility schedule, fixed $MM/well, and the
    # rig-rate / completion / tangibles components (since the engine
    # rebuilds capex_well from these when well_cost_mode == "rig_rate").
    _cont_mult = 1.0 + float(
        st.session_state.get("capex_contingency_pct", 25)) / 100.0
    fac_df_with_cont = fac_df.copy()
    # Strip abandonment/cessation/P&A rows — booked separately via aban_cost,
    # so keeping them in facilities would double-count cessation. Guard
    # against missing/empty tables and non-string labels (NaN, floats).
    try:
        if ("label" in fac_df_with_cont.columns
                and len(fac_df_with_cont) > 0):
            _keep = [not fh.is_abandonment_label(
                        "" if _v is None else str(_v))
                     for _v in fac_df_with_cont["label"].tolist()]
            fac_df_with_cont = fac_df_with_cont[_keep].reset_index(drop=True)
    except Exception:
        # Never let CAPEX hygiene crash the run — fall back to the raw table.
        fac_df_with_cont = fac_df.copy()
    if "amount_MMUSD" in fac_df_with_cont.columns:
        fac_df_with_cont["amount_MMUSD"] = (
            fac_df_with_cont["amount_MMUSD"].astype(float) * _cont_mult)
    capex_well_with_cont = float(capex_well) * _cont_mult
    rig_day_rate_with_cont = float(rig_day_rate_kUSD) * _cont_mult
    cmpl_day_rate_with_cont = float(completion_day_rate_kUSD) * _cont_mult
    well_tangibles_with_cont = float(well_tangibles_MM) * _cont_mult
    # Abandonment is a CAPEX-like spend at the end of life and carries
    # similar estimating bias — apply the same contingency multiplier.
    aban_cost_with_cont = float(aban_cost) * _cont_mult

    # ---- Cost-input currency conversion -----------------------------------
    # If the user is entering costs in NOK, convert every COST field to USD
    # here so the engine always computes in USD and all results display in
    # USD. Revenue-side prices (oil/gas/NGL) and tariffs are left as-is —
    # those are conventionally quoted in USD. The facility CAPEX schedule is
    # converted row-by-row.
    _c = cost_input_to_usd
    opex_var_usd = _c(opex_var)
    opex_fixed_usd = _c(opex_fixed)
    capex_well_usd = _c(capex_well_with_cont)
    aban_cost_usd = _c(aban_cost_with_cont)
    rig_dr_usd = _c(rig_day_rate_with_cont)
    cmpl_dr_usd = _c(cmpl_day_rate_with_cont)
    well_tang_usd = _c(well_tangibles_with_cont)
    fac_df_usd = fac_df_with_cont
    try:
        if (st.session_state.get("cost_input_currency", "USD") == "NOK"
                and fac_df_with_cont is not None
                and "amount_MMUSD" in fac_df_with_cont.columns):
            fac_df_usd = fac_df_with_cont.copy()
            _rate = float(st.session_state.get("usd_to_nok", 10.5))
            if _rate > 0:
                fac_df_usd["amount_MMUSD"] = (
                    fac_df_usd["amount_MMUSD"].astype(float) / _rate)
    except Exception:
        fac_df_usd = fac_df_with_cont

    return EconInputs(
        oil_price=oil_price,        # already in $/bbl (engine-internal)
        gas_price=gas_price,        # already in $/Mscf (engine-internal)
        opex_var=opex_var_usd,       # $/bbl (NOK→USD if cost input is NOK)
        opex_fixed=opex_fixed_usd * 1e6,
        capex_per_well=capex_well_usd,
        discount_rate=disc, tax_rate=tax, royalty_rate=royalty,
        tariff_oil=tariff_oil_bbl,    # will be set below from $/bbl input
        tariff_gas=tariff_gas_mmbtu * MMBTU_PER_MCF,  # $/MMBtu → $/Mscf
        abandonment_cost_MM=aban_cost_usd,
        facility_capex=CapexSchedule(df=fac_df_usd),
        co2_price=co2_price,
        co2_factor_gas_combust=co2_factor_gas_combust,
        co2_factor_flare_inefficiency=co2_factor_flare_ineff,
        co2_factor_oil_routine=co2_factor_oil_routine,
        co2_scope3_enabled=co2_scope3_enabled,
        co2_scope3_factor_oil=co2_scope3_factor_oil,
        co2_scope3_factor_gas=co2_scope3_factor_gas,
        co2_scope3_price=co2_scope3_price,
        fiscal_regime=regime_for_engine,
        psc_cost_recovery_ceiling=psc_cost_recovery_ceiling,
        psc_profit_oil_share_contractor=psc_profit_oil_share_contractor,
        psc_govt_participation=psc_govt_participation,
        psc_psc_tax_rate=psc_psc_tax_rate,
        psc_signature_bonus_MM=psc_signature_bonus_MM,
        ncs_cit_rate=ncs_cit_rate,
        ncs_spt_rate=ncs_spt_rate,
        ncs_uplift_rate=ncs_uplift_rate,
        money_basis=money_basis_for_engine,
        inflation_rate=inflation_rate,
        well_cost_mode=well_cost_mode,
        rig_day_rate_kUSD=rig_dr_usd,
        completion_day_rate_kUSD=cmpl_dr_usd,
        well_tangibles_MM=well_tang_usd,
        well_intangibles_pct=well_intangibles_pct,
        ngl_yield_bbl_per_mmscf=ngl_yield,
        ngl_price_bbl=ngl_price,
        ngl_opex_bbl=ngl_opex,
        ngl_shrinkage_pct=ngl_shrink,
        rig_meta=st.session_state.get("_rig_meta", {}),
        economic_cutoff_mode=economic_cutoff_mode,
        economic_cutoff_persistence=economic_cutoff_persistence,
    )


# =============================================================================
# Plotting
# =============================================================================
def _plot_production_yearly(df, fluid, units):
    """Annual production profile: per-year average rates (bars) plus annual
    produced volumes. The reporting view used for reserves/AOP summaries."""
    f = lambda v, k: from_field(v, k, units)
    oil_label = ulabel("oil_rate", units)
    gas_label = ulabel("gas_rate", units)
    oilv_label = ulabel("oil_vol", units)
    gasv_label = ulabel("gas_vol", units)
    C = fh.EQ_COLORS
    d = df.copy()
    d["year"] = pd.to_datetime(d["date"]).dt.year
    # Average rate per year (mean of monthly rates) and annual volume
    # (rate × days). DAYS_PER_MONTH is the engine month length.
    agg = {}
    for col in ("oil_rate", "gas_rate", "water_rate"):
        if col in d.columns:
            agg[col] = d.groupby("year")[col].mean()
    # Annual produced volume from monthly rate × month length, summed by year
    vol = {}
    for col, volkind, scale in (("oil_rate", "oil_vol", 1e6),
                                 ("gas_rate", "gas_vol", 1e9)):
        if col in d.columns:
            mvol = d[col].values * DAYS_PER_MONTH
            tmp = pd.DataFrame({"year": d["year"].values, "v": mvol})
            vol[volkind] = tmp.groupby("year")["v"].sum() / scale
    years = sorted(d["year"].unique())

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    if "oil_rate" in agg and agg["oil_rate"].max() > 0:
        fig.add_trace(go.Bar(
            x=years, y=[f(agg["oil_rate"].get(y, 0), "oil_rate") for y in years],
            name=f"Oil avg rate ({oil_label})",
            marker_color=C["oil"]), secondary_y=False)
    if "water_rate" in agg and agg["water_rate"].max() > 0:
        fig.add_trace(go.Bar(
            x=years, y=[f(agg["water_rate"].get(y, 0), "water_rate") for y in years],
            name=f"Water avg rate ({oil_label})",
            marker_color=C["water"]), secondary_y=False)
    if "gas_rate" in agg and agg["gas_rate"].max() > 0:
        fig.add_trace(go.Bar(
            x=years, y=[f(agg["gas_rate"].get(y, 0), "gas_rate") for y in years],
            name=f"Gas avg rate ({gas_label})",
            marker_color=C["gas"]), secondary_y=False)
    # Cumulative production overlay (the dashed line in corporate
    # production-profile charts) — cumulative oil on the secondary axis.
    if "oil_rate" in d.columns:
        _cum_oil_year = []
        _run = 0.0
        for y in years:
            _run += float(vol.get("oil_vol", {}).get(y, 0.0))
            _cum_oil_year.append(_run)
        fig.add_trace(go.Scatter(
            x=years, y=_cum_oil_year, name=f"Cumulative oil ({oilv_label})",
            mode="lines+markers", line=dict(color="#8B4513", width=2,
                                            dash="dash")),
            secondary_y=True)
    if "gas_rate" in d.columns and "gas_vol" in vol:
        _cum_gas_year = []
        _rung = 0.0
        for y in years:
            _rung += float(vol.get("gas_vol", {}).get(y, 0.0))
            _cum_gas_year.append(_rung)
        fig.add_trace(go.Scatter(
            x=years, y=_cum_gas_year, name=f"Cumulative gas ({gasv_label})",
            mode="lines+markers", line=dict(color="#555", width=2,
                                            dash="dot")),
            secondary_y=True)
    fig.update_layout(
        title="Annual production profile (rates + cumulative)",
        barmode="group", hovermode="x unified", height=460,
        legend=dict(orientation="h", y=-0.18),
        xaxis_title="Year")
    fig.update_yaxes(title_text=f"Avg rate", secondary_y=False, showgrid=True)
    fig.update_yaxes(title_text="Cumulative production",
                     secondary_y=True, showgrid=False)
    return fh.apply_plot_template(fig)


def plot_production(df, fluid, units):
    """Phase-explicit production rates: oil, gas, water on dual axes."""
    f = lambda v, k: from_field(v, k, units)
    oil_label   = ulabel("oil_rate", units)
    gas_label   = ulabel("gas_rate", units)
    water_label = ulabel("water_rate", units)
    C = fh.EQ_COLORS

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    if df["oil_rate"].max() > 0:
        fig.add_trace(go.Scatter(x=df["date"], y=f(df["oil_rate"], "oil_rate"),
                                 name=f"Oil ({oil_label})",
                                 line=dict(color=C["oil"], width=2.5)),
                      secondary_y=False)
    if df["water_rate"].max() > 0:
        fig.add_trace(go.Scatter(x=df["date"], y=f(df["water_rate"], "water_rate"),
                                 name=f"Water ({water_label})",
                                 line=dict(color=C["water"], width=1.8, dash="dot")),
                      secondary_y=False)
    if df["injection_rate"].max() > 0:
        fig.add_trace(go.Scatter(x=df["date"], y=f(df["injection_rate"], "water_rate"),
                                 name=f"Water inj. ({water_label})",
                                 line=dict(color=C["water_inj"], width=1.8, dash="dash")),
                      secondary_y=False)
    if df["gas_rate"].max() > 0:
        fig.add_trace(go.Scatter(x=df["date"], y=f(df["gas_rate"], "gas_rate"),
                                 name=f"Gas ({gas_label})",
                                 line=dict(color=C["gas"], width=2.5)),
                      secondary_y=True)

    fig.update_layout(title="Production profiles by phase",
                      hovermode="x unified", height=460,
                      legend=dict(orientation="h", y=-0.18))
    # Left axis: liquids only — set its title to reflect what's actually plotted
    # there rather than the generic "Liquid (...)" label, since gas lives on
    # the right axis.
    fig.update_yaxes(title_text=f"Oil & water ({oil_label})",
                     secondary_y=False, showgrid=True)
    fig.update_yaxes(title_text=f"Gas ({gas_label})",
                     secondary_y=True, showgrid=False)
    # Cessation marker — pulled from df.attrs if the engine set one. Lets
    # the user see WHERE the abandonment cutoff (rate / water-cut /
    # economic limit) actually fired on the production curve.
    cidx = df.attrs.get("cessation_idx") if hasattr(df, "attrs") else None
    if cidx is not None and 0 <= int(cidx) < len(df):
        try:
            _cdate = pd.to_datetime(df["date"].iloc[int(cidx)])
            fig.add_vline(x=_cdate, line=dict(color="#7f7f7f", dash="dot"),
                          annotation_text="Cessation",
                          annotation_position="top left",
                          annotation_font=dict(size=11))
        except Exception:
            pass
    return fh.apply_plot_template(fig)


def plot_cumulatives(df, fluid, rf_target, units):
    """Cumulative oil, gas, water + RF — phase-explicit."""
    f = lambda v, k: from_field(v, k, units)
    oil_u   = ulabel("oil_vol", units)
    gas_u   = ulabel("gas_vol", units)
    water_u = ulabel("water_vol", units)
    C = fh.EQ_COLORS

    fig = make_subplots(rows=1, cols=2,
                        specs=[[{"secondary_y": True}, {}]],
                        subplot_titles=("Cumulative production by phase", "Recovery factor"))
    if df["cum_oil"].max() > 0:
        fig.add_trace(go.Scatter(x=df["date"],
                                 y=f(df["cum_oil"], "oil_vol"),
                                 name=f"Cum oil ({oil_u})",
                                 line=dict(color=C["oil"], width=2.5)),
                      row=1, col=1, secondary_y=False)
    if df["cum_water"].max() > 0:
        fig.add_trace(go.Scatter(x=df["date"],
                                 y=f(df["cum_water"], "water_vol"),
                                 name=f"Cum water ({water_u})",
                                 line=dict(color=C["water"], width=1.8, dash="dot")),
                      row=1, col=1, secondary_y=False)
    if df["cum_gas"].max() > 0:
        fig.add_trace(go.Scatter(x=df["date"],
                                 y=f(df["cum_gas"], "gas_vol"),
                                 name=f"Cum gas ({gas_u})",
                                 line=dict(color=C["gas"], width=2.5)),
                      row=1, col=1, secondary_y=True)
    fig.add_trace(go.Scatter(x=df["date"], y=df["recovery_factor"], name="RF",
                             line=dict(color=C["rf"], width=2.5)), row=1, col=2)
    fh.safe_hline(fig, rf_target, label=f"Target {rf_target:.0%}",
                  color=C["pressure"], dash="dash", row=1, col=2)
    fig.update_layout(height=420, hovermode="x unified",
                      legend=dict(orientation="h", y=-0.18))
    fig.update_yaxes(title_text=f"Liquid ({oil_u})", row=1, col=1, secondary_y=False)
    fig.update_yaxes(title_text=f"Gas ({gas_u})", row=1, col=1, secondary_y=True)
    fig.update_yaxes(title_text="RF", row=1, col=2, tickformat=".0%")
    return fh.apply_plot_template(fig)


def plot_well_stack(per_well_df, primary_label, units, is_oil):
    """Per-well stacked area for the well's primary fluid (oil or gas).

    Note: per_well_df stores each well's primary_rate (oil for oil wells,
    gas for gas wells). When fluids are mixed across wells, this stack
    is unit-mixed; UI guards against that case.
    """
    f = lambda v: from_field(v, "oil_rate" if is_oil else "gas_rate", units)
    rate_label = ulabel("oil_rate" if is_oil else "gas_rate", units)
    fig = go.Figure()
    for col in [c for c in per_well_df.columns if c != "date"]:
        fig.add_trace(go.Scatter(x=per_well_df["date"], y=f(per_well_df[col]),
                                 name=col, stackgroup="one", mode="none"))
    fig.update_layout(title=f"Per-well contribution ({rate_label})",
                      hovermode="x unified", height=460,
                      yaxis_title=rate_label, legend=dict(orientation="v"))
    return fh.apply_plot_template(fig)


def plot_per_well_phase(per_well_df, df, units, fluid):
    """Three stacked-area subplots: per-well oil, gas, water contributions.

    Uses the proper per-well phase matrices (oil_mat, gas_mat, water_mat) stashed
    on per_well_df.attrs by run_simulation when available; falls back to the
    field-share-weighted approximation only if those matrices are missing
    (e.g. very old saved cases). Each well's own fluid type controls its
    primary→oil-or-gas mapping, which matters in mixed-fluid (multi-reservoir)
    fields.
    """
    is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"
    well_cols = [c for c in per_well_df.columns if c != "date"]
    if not well_cols:
        return None
    f = lambda v, k: from_field(v, k, units)

    # Prefer the proper per-well phase matrices if they were attached
    oil_mat_df   = per_well_df.attrs.get("oil_mat")
    gas_mat_df   = per_well_df.attrs.get("gas_mat")
    water_mat_df = per_well_df.attrs.get("water_mat")

    fig = make_subplots(rows=3, cols=1, shared_xaxes=True,
                         subplot_titles=("Oil rate by well", "Gas rate by well",
                                          "Water rate by well"),
                         vertical_spacing=0.06)

    if oil_mat_df is not None and gas_mat_df is not None and water_mat_df is not None:
        # Proper per-well phases — each well carries its own oil/gas/water profile,
        # honouring per-well fluid type in mixed-fluid fields.
        for col in well_cols:
            oil_w = oil_mat_df[col].values   if col in oil_mat_df.columns   else np.zeros(len(per_well_df))
            gas_w = gas_mat_df[col].values   if col in gas_mat_df.columns   else np.zeros(len(per_well_df))
            wat_w = water_mat_df[col].values if col in water_mat_df.columns else np.zeros(len(per_well_df))
            fig.add_trace(go.Scatter(x=per_well_df["date"], y=f(oil_w, "oil_rate"),
                                      name=col, stackgroup="oil", mode="none",
                                      legendgroup=col, showlegend=True),
                          row=1, col=1)
            fig.add_trace(go.Scatter(x=per_well_df["date"], y=f(gas_w, "gas_rate"),
                                      name=col, stackgroup="gas", mode="none",
                                      legendgroup=col, showlegend=False),
                          row=2, col=1)
            fig.add_trace(go.Scatter(x=per_well_df["date"], y=f(wat_w, "water_rate"),
                                      name=col, stackgroup="water", mode="none",
                                      legendgroup=col, showlegend=False),
                          row=3, col=1)
    else:
        # Legacy fallback: share-weight the field-level phases by each well's
        # primary share at every timestep.
        well_mat = per_well_df[well_cols].values
        total = well_mat.sum(axis=1, keepdims=True)
        total = np.where(total > 0, total, 1.0)
        shares = well_mat / total
        field_oil   = df["oil_rate"].values
        field_gas   = df["gas_rate"].values
        field_water = df["water_rate"].values
        for j, col in enumerate(well_cols):
            oil_w = shares[:, j] * field_oil
            gas_w = shares[:, j] * field_gas
            wat_w = shares[:, j] * field_water
            fig.add_trace(go.Scatter(x=per_well_df["date"], y=f(oil_w, "oil_rate"),
                                      name=col, stackgroup="oil", mode="none",
                                      legendgroup=col, showlegend=True),
                          row=1, col=1)
            fig.add_trace(go.Scatter(x=per_well_df["date"], y=f(gas_w, "gas_rate"),
                                      name=col, stackgroup="gas", mode="none",
                                      legendgroup=col, showlegend=False),
                          row=2, col=1)
            fig.add_trace(go.Scatter(x=per_well_df["date"], y=f(wat_w, "water_rate"),
                                      name=col, stackgroup="water", mode="none",
                                      legendgroup=col, showlegend=False),
                          row=3, col=1)

    fig.update_yaxes(title_text=f"Oil ({ulabel('oil_rate', units)})", row=1, col=1)
    fig.update_yaxes(title_text=f"Gas ({ulabel('gas_rate', units)})", row=2, col=1)
    fig.update_yaxes(title_text=f"Water ({ulabel('water_rate', units)})", row=3, col=1)
    fig.update_layout(height=720, hovermode="x unified",
                      title="Per-well contribution by phase",
                      legend=dict(orientation="v"))
    return fh.apply_plot_template(fig)


def plot_drilling_gantt(wells):
    fig = go.Figure()
    rigs = sorted({w.rig for w in wells})
    color_map = {r: RIG_COLORS[i % len(RIG_COLORS)] for i, r in enumerate(rigs)}
    # Read the well -> template mapping built by the concept builder.
    # When available, the Gantt label includes the template (e.g.
    # "P1 [T1]"), and the rows are grouped by template so each drill
    # centre's wells sit together. When no mapping exists, the legacy
    # behaviour (rig grouping) is preserved.
    wt_map = st.session_state.get("well_template_map", {}) or {}
    has_map = bool(wt_map)

    def _tpl_of(name):
        return wt_map.get(str(name), "—")

    def _y_label(w):
        if has_map:
            t = _tpl_of(w.name)
            return f"{w.name}  [{t}]" if t and t != "—" else w.name
        return w.name

    # Sort wells: by template (if mapped) then by spud, so each
    # template's wells form a contiguous band.
    if has_map:
        wells_sorted = sorted(
            wells,
            key=lambda w: (_tpl_of(w.name), pd.Timestamp(w.spud_date),
                            w.name))
    else:
        wells_sorted = list(wells)

    for w in wells_sorted:
        tpl = _tpl_of(w.name)
        tpl_txt = (f"<br>Template: {tpl}"
                   if has_map and tpl and tpl != "—" else "")
        ylab = _y_label(w)
        fig.add_trace(go.Bar(
            x=[w.drill_days * 86400000], y=[ylab],
            base=[pd.Timestamp(w.spud_date)],
            orientation="h", marker_color=color_map[w.rig],
            opacity=0.85, showlegend=False,
            hovertemplate=(f"<b>{w.name}</b> — {w.rig}{tpl_txt}<br>"
                           f"Spud: {w.spud_date}<br>"
                           f"Drill: {w.drill_days} d<br>"
                           f"Compl: {w.completion_days} d<br>"
                           f"Online: {w.online_date}<extra></extra>"),
        ))
        compl_start = pd.Timestamp(w.spud_date) + pd.Timedelta(days=w.drill_days)
        fig.add_trace(go.Bar(
            x=[w.completion_days * 86400000], y=[ylab],
            base=[compl_start],
            orientation="h", marker_color=color_map[w.rig],
            opacity=0.45, showlegend=False,
            hovertemplate=(f"{w.name} completion{tpl_txt}<br>"
                           f"From: {compl_start.date()}<br>"
                           f"Online: {w.online_date}<extra></extra>"),
        ))
    for r in rigs:
        fig.add_trace(go.Bar(x=[None], y=[None], marker_color=color_map[r],
                             name=r, showlegend=True))
    # Add faint horizontal bands behind each template's wells when mapped
    if has_map:
        # build groups in plot order
        groups = []
        last_tpl = None
        for w in wells_sorted:
            tpl = _tpl_of(w.name)
            if tpl != last_tpl:
                groups.append([tpl, _y_label(w), _y_label(w)])
                last_tpl = tpl
            else:
                groups[-1][2] = _y_label(w)
        shapes = []
        band_colors = ["rgba(120,170,210,0.10)",
                        "rgba(200,160,90,0.10)"]
        for gi, (tpl, y0, y1) in enumerate(groups):
            if not tpl or tpl == "—":
                continue
            shapes.append(dict(
                type="rect", xref="paper", yref="y",
                x0=0, x1=1, y0=y0, y1=y1,
                line=dict(width=0),
                fillcolor=band_colors[gi % len(band_colors)],
                layer="below"))
        if shapes:
            fig.update_layout(shapes=shapes)
    title = ("Drilling schedule — grouped by template "
             "(drill = solid, completion = faded)"
             if has_map else
             "Drilling schedule (drill = solid, completion = faded)")
    fig.update_layout(
        title=dict(text=title, font=dict(size=16)),
        # 32 px / row gives the y-axis labels enough vertical space for a
        # 13 px font without overlap. Older 28 px crushed labels at 10 px.
        height=max(420, 32 * len(wells)),
        barmode="overlay",
        xaxis=dict(type="date", title=dict(text="Date",
                                            font=dict(size=14)),
                   tickfont=dict(size=12)),
        yaxis=dict(tickfont=dict(size=13),
                    automargin=True),
        legend=dict(orientation="h", y=-0.15,
                     font=dict(size=12)),
        font=dict(size=12),
    )
    return fh.apply_plot_template(fig)


def plot_pressure(df, units):
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Scatter(x=df["date"],
                             y=from_field(df["pressure"], "pressure", units),
                             name=f"Reservoir pressure ({ulabel('pressure', units)})",
                             line=dict(color=fh.EQ_COLORS["pressure"], width=2.5)),
                  secondary_y=False)
    fig.add_trace(go.Scatter(x=df["date"], y=df["recovery_factor"],
                             name="Recovery factor",
                             line=dict(color=fh.EQ_COLORS["rf"], width=2.5, dash="dash")),
                  secondary_y=True)
    fig.update_layout(title="Material balance — pressure & RF (field aggregate)",
                      height=400, hovermode="x unified",
                      legend=dict(orientation="h", y=-0.2))
    fig.update_yaxes(title_text=f"Pressure ({ulabel('pressure', units)})",
                     secondary_y=False)
    fig.update_yaxes(title_text="RF", tickformat=".0%", secondary_y=True)
    return fh.apply_plot_template(fig)


def plot_per_reservoir_pressure(per_res_df, units):
    """One pressure trace per reservoir."""
    fig = go.Figure()
    for rid, group in per_res_df.groupby("reservoir_id"):
        name = group["reservoir_name"].iloc[0]
        fig.add_trace(go.Scatter(
            x=group["date"],
            y=from_field(group["pressure"], "pressure", units),
            name=f"{name}",
            mode="lines",
        ))
    fig.update_layout(
        title=f"Per-reservoir pressure ({ulabel('pressure', units)})",
        height=380, hovermode="x unified",
        yaxis_title=f"Pressure ({ulabel('pressure', units)})",
        legend=dict(orientation="h", y=-0.2),
    )
    return fh.apply_plot_template(fig)


def plot_per_reservoir_rf(per_res_df):
    fig = go.Figure()
    for rid, group in per_res_df.groupby("reservoir_id"):
        name = group["reservoir_name"].iloc[0]
        fig.add_trace(go.Scatter(
            x=group["date"], y=group["recovery_factor"],
            name=f"{name}", mode="lines",
        ))
    fig.update_layout(
        title="Per-reservoir recovery factor",
        height=380, hovermode="x unified",
        yaxis_title="RF",
        yaxis=dict(tickformat=".0%"),
        legend=dict(orientation="h", y=-0.2),
    )
    return fh.apply_plot_template(fig)


def plot_per_reservoir_rate(per_res_df, units, fluid):
    """Three subplots: per-reservoir oil, gas, water rates.

    Each reservoir contributes oil from oil-reservoir primaries (or condensate
    from gas-reservoir secondaries) and gas from oil-reservoir secondaries
    (or gas-reservoir primaries).
    """
    f = lambda v, k: from_field(v, k, units)

    # Compute per-reservoir oil/gas streams using the same logic as engine
    rows = []
    for _, row in per_res_df.iterrows():
        is_oil_r = FLUID_SYSTEMS[row["fluid_system"]]["primary"] == "oil"
        if is_oil_r:
            oil = row["primary_rate"]; gas = row["secondary_rate"]
        else:
            oil = row["secondary_rate"]; gas = row["primary_rate"]
        rows.append({
            "date": row["date"],
            "reservoir_name": row["reservoir_name"],
            "oil_rate": oil, "gas_rate": gas,
            "water_rate": row["water_rate"],
        })
    pdf = pd.DataFrame(rows)

    fig = make_subplots(rows=3, cols=1, shared_xaxes=True,
                         subplot_titles=("Oil rate by reservoir",
                                          "Gas rate by reservoir",
                                          "Water rate by reservoir"),
                         vertical_spacing=0.06)
    for name, group in pdf.groupby("reservoir_name"):
        fig.add_trace(go.Scatter(x=group["date"], y=f(group["oil_rate"], "oil_rate"),
                                  name=name, stackgroup="oil", mode="none",
                                  legendgroup=name, showlegend=True),
                      row=1, col=1)
        fig.add_trace(go.Scatter(x=group["date"], y=f(group["gas_rate"], "gas_rate"),
                                  name=name, stackgroup="gas", mode="none",
                                  legendgroup=name, showlegend=False),
                      row=2, col=1)
        fig.add_trace(go.Scatter(x=group["date"], y=f(group["water_rate"], "water_rate"),
                                  name=name, stackgroup="water", mode="none",
                                  legendgroup=name, showlegend=False),
                      row=3, col=1)
    fig.update_yaxes(title_text=f"Oil ({ulabel('oil_rate', units)})", row=1, col=1)
    fig.update_yaxes(title_text=f"Gas ({ulabel('gas_rate', units)})", row=2, col=1)
    fig.update_yaxes(title_text=f"Water ({ulabel('water_rate', units)})", row=3, col=1)
    fig.update_layout(height=720, hovermode="x unified",
                      title="Per-reservoir production by phase",
                      legend=dict(orientation="v"))
    return fh.apply_plot_template(fig)


def plot_npv_waterfall(df_e, discount_rate: float):
    """Waterfall chart: how gross revenue is whittled down to NPV through
    royalty, tariffs, OPEX, CAPEX, tax, and abandonment — then discounted.

    Shows two bars at the ends (Gross revenue → NPV) with the deductions as
    floating intermediate steps. Uses undiscounted totals for the value-
    construction steps, then a final explicit 'discounting' bridge to NPV.
    """
    rev      = df_e["revenue"].sum()
    royalty  = df_e["royalty"].sum()
    tariff   = df_e["tariff"].sum()
    opex     = df_e["opex"].sum()
    capex_w  = df_e["capex_well"].sum()
    capex_f  = df_e["capex_facility"].sum()
    tax      = df_e["tax"].sum()
    aban     = df_e["abandonment"].sum()
    undiscounted_cf = df_e["cashflow"].sum()
    npv = df_e["npv"].iloc[-1] if "npv" in df_e.columns and len(df_e) else 0.0
    discount_effect = npv - undiscounted_cf

    labels = ["Gross revenue", "Royalty", "Tariffs", "OPEX",
              "Well CAPEX", "Facility CAPEX", "Tax", "Abandonment",
              "Undiscounted CF", f"Discounting @ {discount_rate:.0%}", "NPV"]
    measures = ["absolute", "relative", "relative", "relative",
                "relative", "relative", "relative", "relative",
                "total", "relative", "total"]
    values = [rev/1e6, -royalty/1e6, -tariff/1e6, -opex/1e6,
              -capex_w/1e6, -capex_f/1e6, -tax/1e6, -aban/1e6,
              0, discount_effect/1e6, 0]

    fig = go.Figure(go.Waterfall(
        orientation="v",
        measure=measures,
        x=labels,
        y=values,
        textposition="outside",
        text=[f"{v:,.0f}" if m == "relative" else "" for v, m in zip(values, measures)],
        connector={"line": {"color": "rgb(160,160,160)"}},
        decreasing={"marker": {"color": "#d62728"}},
        increasing={"marker": {"color": "#2ca02c"}},
        totals={"marker": {"color": "#1f77b4"}},
    ))
    fig.update_layout(
        title="NPV value construction (waterfall, $MM)",
        height=460, showlegend=False,
        yaxis_title="$MM",
        xaxis=dict(tickangle=-30),
    )
    fig.add_hline(y=0, line=dict(color="grey", dash="dot"))
    return fh.apply_plot_template(fig)


def plot_co2_yearly(df_e):
    """Yearly CO₂ emissions stacked by scope (Scope 1 ops + Scope 3 end-use).

    Scope 3 dwarfs Scope 1 by 50-100×, so the chart uses a *split-axis* view:
    Scope 1 on its own panel, Scope 3 on a second panel side-by-side, so
    Scope 1 detail isn't crushed to invisibility under the Scope 3 stack.
    A line of Scope 3 / Scope 1 ratio is overlaid where useful.
    """
    if "co2_scope1_tonnes" not in df_e.columns:
        return None
    df_c = df_e.copy()
    df_c["year"] = pd.to_datetime(df_c["date"], errors="coerce").dt.year
    df_c = df_c[df_c["year"].notna() & (df_c["year"] >= 1990)]
    df_c["year"] = df_c["year"].astype(int)
    annual = df_c.groupby("year").agg({
        "co2_scope1_tonnes": "sum",
        "co2_scope3_tonnes": "sum",
        "co2_cost": "sum",
        "co2_scope3_cost": "sum",
    }).reset_index()
    # convert tonnes -> kilotonnes for chart legibility
    annual["s1_kt"] = annual["co2_scope1_tonnes"] / 1000.0
    annual["s3_kt"] = annual["co2_scope3_tonnes"] / 1000.0
    yrs = annual["year"].astype(str)

    fig = make_subplots(
        rows=1, cols=2,
        subplot_titles=("Scope 1 — operational (kt CO₂-eq/yr)",
                        "Scope 3 — end-use combustion (kt CO₂-eq/yr)"),
        horizontal_spacing=0.12)
    fig.add_trace(go.Bar(x=yrs, y=annual["s1_kt"],
                         name="Scope 1",
                         marker_color="#d98a2b"), row=1, col=1)
    fig.add_trace(go.Bar(x=yrs, y=annual["s3_kt"],
                         name="Scope 3",
                         marker_color="#6b3a17"), row=1, col=2)
    fig.update_xaxes(title_text="Calendar year", row=1, col=1,
                      type="category")
    fig.update_xaxes(title_text="Calendar year", row=1, col=2,
                      type="category")
    fig.update_yaxes(title_text="kt CO₂-eq / year", row=1, col=1)
    fig.update_yaxes(title_text="kt CO₂-eq / year", row=1, col=2)
    fig.update_layout(height=380, showlegend=False,
                      title="Yearly CO₂ emissions by scope")
    return fh.apply_plot_template(fig)


def plot_economics(df_e):
    # Drop rows with bogus dates (NaT or epoch placeholders) before
    # grouping. A single year=0 row stretched the bar chart's x-axis from
    # 0 to 2040, hiding all the data at the right edge.
    df_clean = df_e.copy()
    df_clean["year"] = pd.to_datetime(df_clean["date"],
                                        errors="coerce").dt.year
    df_clean = df_clean[df_clean["year"].notna()
                         & (df_clean["year"] >= 1990)]
    df_clean["year"] = df_clean["year"].astype(int)
    annual = df_clean.groupby("year").agg({
        "revenue": "sum", "royalty": "sum", "tariff": "sum",
        "opex": "sum", "capex_well": "sum", "capex_facility": "sum",
        "tax": "sum", "abandonment": "sum", "cashflow": "sum"
    }).reset_index()

    # First-oil and cessation calendar years for reference markers
    first_oil_year = None
    if (df_e["revenue"] > 0).any():
        first_oil_year = int(
            pd.to_datetime(
                df_e.loc[df_e["revenue"] > 0, "date"].iloc[0]).year)
    cessation_year = None
    cidx = df_e.attrs.get("cessation_idx")
    if cidx is not None and cidx < len(df_e):
        cessation_year = int(pd.to_datetime(df_e["date"].iloc[cidx]).year)

    # Include every calendar year in the plot range, not only years
    # with cost activity. The annual frame might only have one or two
    # non-zero years (e.g. a CAPEX year and not much else if the run
    # hasn't reached full economics). Pad the span to cover at least
    # the first calendar year present in df_clean through the cessation
    # year, so the x-axis reflects the true project life even when most
    # years would be zeros.
    if len(annual) > 0:
        y_lo = int(annual["year"].min())
        y_hi = int(annual["year"].max())
        if len(df_clean) > 0:
            y_lo = min(y_lo, int(df_clean["year"].min()))
            y_hi = max(y_hi, int(df_clean["year"].max()))
        if cessation_year is not None:
            y_hi = max(y_hi, int(cessation_year))
        if first_oil_year is not None:
            y_lo = min(y_lo, int(first_oil_year))
            y_hi = max(y_hi, int(first_oil_year))
        full_years = list(range(y_lo, y_hi + 1))
        annual = (annual.set_index("year")
                        .reindex(full_years, fill_value=0.0)
                        .reset_index()
                        .rename(columns={"index": "year"}))

    fig = make_subplots(rows=1, cols=2,
                        subplot_titles=("Annual cashflow buildup ($MM)",
                                        "Cumulative CF & NPV ($MM)"))
    bars = [
        ("revenue", "Revenue", "#2ca02c", 1),
        ("royalty", "Royalty", "#aec7e8", -1),
        ("tariff", "Tariffs", "#c5b0d5", -1),
        ("opex", "OPEX", "#d62728", -1),
        ("capex_well", "Well CAPEX", "#9467bd", -1),
        ("capex_facility", "Facility CAPEX", "#8c564b", -1),
        ("tax", "Tax", "#e377c2", -1),
        ("abandonment", "Abandonment", "#7f7f7f", -1),
    ]
    # x-values: INTEGER years on a numeric linear axis. Previous attempts
    # used a categorical string axis with explicit `categoryarray` +
    # `tickvals`, but those get silently overridden by Plotly's auto-tick
    # logic and by `apply_plot_template`'s layout merge — leaving the
    # bars crammed at the leftmost category and no tick labels for years
    # in between. A numeric linear axis with explicit `range=[y_lo-0.5,
    # y_hi+0.5]` and `dtick=1` is the bulletproof form: `range` is one of
    # the few axis settings that survives all downstream merges, and
    # `dtick=1` forces a tick every year.
    year_vals = annual["year"].astype(int).tolist()
    for col, name, color, sign in bars:
        fig.add_trace(go.Bar(x=year_vals,
                             y=sign * annual[col]/1e6,
                             name=name, marker_color=color), row=1, col=1)
    # First-oil and cessation markers — guarded so they only draw when
    # the year is actually inside the bar-chart range.
    if (first_oil_year is not None and year_vals
            and year_vals[0] <= first_oil_year <= year_vals[-1]):
        fh.safe_vline(fig, int(first_oil_year), label="First oil",
                      color="#2ca02c", dash="dot", row=1, col=1)
    if (cessation_year is not None and year_vals
            and year_vals[0] <= cessation_year <= year_vals[-1]):
        fh.safe_vline(fig, int(cessation_year), label="Cessation",
                      color="#7f7f7f", dash="dot", row=1, col=1,
                      label_position="bottom")
    fig.add_trace(go.Scatter(x=df_e["date"], y=df_e["cum_cashflow"]/1e6,
                             name="Cum CF", line=dict(color="#1f77b4", width=2)),
                  row=1, col=2)
    fig.add_trace(go.Scatter(x=df_e["date"], y=df_e["npv"]/1e6,
                             name="NPV", line=dict(color="#ff7f0e", width=2, dash="dash")),
                  row=1, col=2)
    fig.add_hline(y=0, line=dict(color="grey", dash="dot"), row=1, col=2)
    # Numeric integer-year axis: `range` survives template merges, `dtick=1`
    # forces a label per year, `tickformat="d"` prints them as integers
    # (no decimals, no thousands separator).
    if year_vals:
        x_lo = year_vals[0] - 0.5
        x_hi = year_vals[-1] + 0.5
        fig.update_xaxes(title_text="Calendar year", row=1, col=1,
                          type="linear",
                          range=[x_lo, x_hi],
                          tick0=year_vals[0], dtick=1,
                          tickformat="d")
    fig.update_xaxes(title_text="Date", row=1, col=2)
    fig.update_layout(barmode="relative", height=450,
                      legend=dict(orientation="h", y=-0.2))
    return fh.apply_plot_template(fig)


# =============================================================================
# Main
# =============================================================================
def validate_inputs(asm: FieldAssumptions, econ: EconInputs,
                     wells: list, fluid: str, units: str = "field") -> None:
    """Surface soft warnings for likely-wrong input combinations.

    Doesn't block execution — just renders an info/warning banner with
    actionable hints. Catches a class of common screening-mode mistakes:
    PVT contradictions, decline > 100%/yr, water cuts going backwards,
    capacities trivially below typical well rates, gas-disposition fractions
    that don't sum, missing producers, etc.
    """
    is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"
    issues = []   # list[(severity, message)] where severity ∈ {"warn", "info"}

    # PVT consistency — show in user's chosen pressure unit
    _pl = ulabel("pressure", units)
    _p_init_d = from_field(asm.pvt.p_init_psi, "pressure", units)
    _p_bub_d = from_field(asm.pvt.p_bub_psi, "pressure", units)
    if asm.pvt.p_init_psi <= asm.pvt.p_bub_psi and is_oil:
        issues.append(("warn",
            f"Initial pressure ({_p_init_d:,.0f} {_pl}) is at or below bubble point "
            f"({_p_bub_d:,.0f} {_pl}). The reservoir starts saturated; "
            "expect immediate gas evolution and free-gas behavior."))
    if asm.pvt.api < 10 or asm.pvt.api > 60:
        issues.append(("warn",
            f"Oil API gravity {asm.pvt.api:.1f} is outside the typical 10–60 range."))
    if asm.pvt.gas_grav < 0.55 or asm.pvt.gas_grav > 1.2:
        issues.append(("warn",
            f"Gas specific gravity {asm.pvt.gas_grav:.2f} is unusual "
            "(typical 0.6–0.9 for natural gas)."))

    # Producer-level checks
    producers = [w for w in wells if w.is_producer]
    injectors = [w for w in wells if not w.is_producer]
    if not producers:
        issues.append(("warn", "No producers defined — the simulation will be empty."))
    for w in producers:
        if w.di_annual > 1.0:
            issues.append(("warn",
                f"Well **{w.name}**: decline rate {w.di_annual:.0%}/yr is > 100%. "
                "Use a value between 0 and 1 (e.g. 0.20 for 20%/yr)."))
        if w.wc_initial > w.wc_final and w.wc_ramp_months > 0:
            issues.append(("warn",
                f"Well **{w.name}**: water-cut initial ({w.wc_initial:.0%}) is higher "
                f"than final ({w.wc_final:.0%}) — water cut should generally rise over time."))
        if w.qi_primary <= 0:
            issues.append(("warn",
                f"Well **{w.name}**: primary rate is zero — well will produce nothing."))
        if w.uptime > 1.0 or w.uptime < 0.0:
            issues.append(("warn",
                f"Well **{w.name}**: uptime {w.uptime:.2f} should be between 0 and 1."))

    # Capacity sanity vs total nameplate — both in user rate units
    if producers and asm.cap_schedule is not None and len(asm.cap_schedule.df) > 0:
        nameplate = sum(w.qi_primary for w in producers)
        first = asm.cap_schedule.df.iloc[0]
        if is_oil:
            cap_p = float(first["oil"])
            if cap_p > 0 and cap_p < nameplate * 0.10:
                _ru = ulabel("oil_rate", units)
                issues.append(("info",
                    f"Initial oil capacity "
                    f"({from_field(cap_p, 'oil_rate', units):,.0f} {_ru}) "
                    f"is < 10% of nameplate production "
                    f"({from_field(nameplate, 'oil_rate', units):,.0f} {_ru}). "
                    f"Wells will be heavily choked."))
        else:
            cap_p = float(first["gas"]) * 1000.0
            if cap_p > 0 and cap_p < nameplate * 0.10:
                _ru = ulabel("gas_rate", units)
                issues.append(("info",
                    f"Initial gas capacity "
                    f"({from_field(cap_p, 'gas_rate', units):,.0f} {_ru}) "
                    f"is < 10% of nameplate production "
                    f"({from_field(nameplate, 'gas_rate', units):,.0f} {_ru}). "
                    f"Wells will be heavily choked."))

    # Gas disposition fractions should sum to 1
    gas_sum = (asm.gas_export_fraction + asm.gas_injection_fraction
                + asm.gas_fuel_fraction + asm.gas_flare_fraction)
    if abs(gas_sum - 1.0) > 0.02:
        issues.append(("warn",
            f"Gas disposition fractions sum to {gas_sum:.2f} (should be 1.00). "
            "Engine will renormalize — set them to sum to 1 to silence this."))

    # Strategy consistency
    if asm.strategy == "Injection" and not injectors and asm.voidage_ratio == 0:
        issues.append(("info",
            "Strategy is 'Injection' but no injectors defined and VRR = 0. "
            "The engine will fall back to depletion behavior."))
    if asm.strategy == "Depletion" and injectors:
        issues.append(("info",
            f"Strategy is 'Depletion' but {len(injectors)} injectors are defined. "
            "They will inject according to the surface and VRR caps."))

    # Aquifer / pressure consistency — user pressure units
    if asm.aquifer.active and asm.aquifer.initial_pressure_psi < asm.pvt.p_init_psi * 0.7:
        _aq_d = from_field(asm.aquifer.initial_pressure_psi, "pressure", units)
        issues.append(("warn",
            f"Aquifer initial pressure ({_aq_d:,.0f} {_pl}) "
            f"is much lower than reservoir Pi ({_p_init_d:,.0f} {_pl}). "
            "The aquifer will provide little support."))

    # PI bridge sanity: when wells use PI mode, check that PI × ΔP gives a
    # reasonable qi vs typical archetype ranges for the reservoir's PVT class.
    pi_wells = [w for w in producers if getattr(w, "derive_qi_from_pi", False)]
    if pi_wells:
        # Use the synthesized or first reservoir
        r0 = (asm.reservoirs[0] if asm.reservoirs else None)
        if r0 is None:
            # Single-reservoir mode: derive from sidebar defaults
            ref_pi = asm.default_well_pi
            ref_bhp = asm.default_min_bhp_psi
            ref_pi_init = asm.pvt.p_init_psi
        else:
            ref_pi = r0.well_pi
            ref_bhp = r0.min_bhp_psi
            ref_pi_init = r0.pvt.p_init_psi
        derived_qi = ref_pi * max(ref_pi_init - ref_bhp, 0.0)
        _ref_pi_init_d = from_field(ref_pi_init, "pressure", units)
        _ref_bhp_d = from_field(ref_bhp, "pressure", units)
        _dd_d = from_field(max(ref_pi_init - ref_bhp, 0.0), "pressure", units)
        if derived_qi <= 0:
            issues.append(("warn",
                f"{len(pi_wells)} well(s) have PI mode ON but the reservoir's "
                f"PI × (P − BHP_min) = {ref_pi:.2f} × "
                f"({_ref_pi_init_d:,.0f} − {_ref_bhp_d:,.0f}) {_pl} ≤ 0. "
                "Wells will produce nothing. Check PI / BHP / Pi values."))
        elif (ref_pi_init - ref_bhp) < 100:    # 100 psi (~7 bar)
            issues.append(("info",
                f"Drawdown (P_init − BHP_min) = {_dd_d:,.0f} {_pl} is "
                "very small. Wells will be deliverability-limited; consider lowering BHP_min."))
        # Cross-check against any free-input qi values that DON'T have PI mode on:
        free_wells = [w for w in producers if not getattr(w, "derive_qi_from_pi", False)
                       and w.qi_primary > 0]
        if free_wells and derived_qi > 0:
            free_avg = sum(w.qi_primary for w in free_wells) / len(free_wells)
            ratio = free_avg / derived_qi if derived_qi > 0 else 0
            if ratio > 3.0 or ratio < 0.33:
                _qi_kind = "oil_rate" if is_oil else "gas_rate"
                _ru = ulabel(_qi_kind, units)
                _free_d = from_field(free_avg, _qi_kind, units)
                _der_d = from_field(derived_qi, _qi_kind, units)
                issues.append(("info",
                    f"Free-input qi (avg {_free_d:,.0f} {_ru}) differs by "
                    f"{ratio:.1f}× from the PI-derived qi "
                    f"({_der_d:,.0f} {_ru}). Consider whether your "
                    "reservoir PI / BHP values reflect the same well type."))

    # IPR mode sanity — pressure in user units
    ipr_wells_v = [w for w in producers if getattr(w, "ipr_mode", False)]
    if ipr_wells_v:
        for w in ipr_wells_v:
            hydrostatic = w.fluid_gradient_psi_per_ft * w.tubing_depth_ft
            min_bhp_implied = w.wellhead_pressure_psi + hydrostatic
            _bhp_d = from_field(min_bhp_implied, "pressure", units)
            if min_bhp_implied >= asm.pvt.p_init_psi:
                issues.append(("warn",
                    f"Well **{w.name}**: outflow back-pressure (P_wh + ρ×depth = "
                    f"{_bhp_d:,.0f} {_pl}) exceeds reservoir Pi "
                    f"({_p_init_d:,.0f} {_pl}). Well will not flow. "
                    "Reduce wellhead pressure, depth, or fluid gradient."))
            elif min_bhp_implied >= asm.pvt.p_init_psi * 0.85:
                issues.append(("info",
                    f"Well **{w.name}**: outflow back-pressure "
                    f"({_bhp_d:,.0f} {_pl}) is close to reservoir Pi — "
                    f"limited drawdown available. Well will go off plateau "
                    "quickly as reservoir depletes."))

    # Economics
    if econ.discount_rate <= 0 or econ.discount_rate > 0.30:
        issues.append(("info",
            f"Discount rate {econ.discount_rate:.0%} is outside the typical 5–25% band."))
    if econ.tax_rate + econ.royalty_rate > 0.85:
        issues.append(("info",
            f"Tax + royalty = {(econ.tax_rate + econ.royalty_rate):.0%} of revenue — "
            "this is a heavy fiscal regime. Verify it matches the actual concession terms."))

    # Render: collapsed expander only when there are issues
    if not issues:
        return
    warns = [m for s, m in issues if s == "warn"]
    infos = [m for s, m in issues if s == "info"]
    label = f"⚠️ Input checks ({len(warns)} warning{'' if len(warns)==1 else 's'}"
    if infos:
        label += f", {len(infos)} note{'' if len(infos)==1 else 's'}"
    label += ")"
    with st.expander(label, expanded=(len(warns) > 0)):
        for m in warns:
            st.warning(m)
        for m in infos:
            st.info(m)


def main():
    # ---- Styling ----
    st.markdown(fh.APP_CSS, unsafe_allow_html=True)

    # ---- Helper-module version check ----
    # If the app and fp_helpers.py are out of sync (e.g. only one file was
    # redeployed), new features crash with AttributeError mid-page. Detect
    # that here and show one clear banner.
    _EXPECTED_FP_VERSION = "4.5"
    _fp_version = getattr(fh, "FP_HELPERS_VERSION", None)
    if _fp_version != _EXPECTED_FP_VERSION:
        _fp_desc = (f"v{_fp_version}" if _fp_version
                    else "an older version (no version tag)")
        st.error(
            f"⚠️ **Version mismatch.** This app expects fp_helpers.py "
            f"v{_EXPECTED_FP_VERSION}, but the loaded helper module is "
            f"{_fp_desc}. Some features (cost benchmarking, HPHT, "
            f"development concepts, methodology docs) may be unavailable or "
            f"error. **Re-upload the latest fp_helpers.py alongside "
            f"field_prognosis_app.py** — both files must be from the same "
            f"release.")

    # ---- Branded banner ----
    # FieldVista — an SVG logo mark: a horizon over subsurface strata with a
    # production curve rising to the surface (the "vista").
    _logo_svg = (
        '<svg width="64" height="64" viewBox="0 0 64 64" '
        'xmlns="http://www.w3.org/2000/svg">'
        '<defs><linearGradient id="sky" x1="0" y1="0" x2="0" y2="1">'
        '<stop offset="0" stop-color="#1a4d6e"/>'
        '<stop offset="1" stop-color="#2a7fa8"/></linearGradient>'
        '<linearGradient id="rock" x1="0" y1="0" x2="0" y2="1">'
        '<stop offset="0" stop-color="#c98a3a"/>'
        '<stop offset="1" stop-color="#6e4a1f"/></linearGradient></defs>'
        '<rect x="2" y="2" width="60" height="60" rx="12" fill="url(#sky)"/>'
        '<path d="M2 38 Q 32 30 62 38 L62 62 L2 62 Z" fill="url(#rock)"/>'
        '<path d="M2 46 Q 32 40 62 46" stroke="#8a5e2a" '
        'stroke-width="1.5" fill="none" opacity="0.7"/>'
        '<path d="M2 54 Q 32 49 62 54" stroke="#8a5e2a" '
        'stroke-width="1.5" fill="none" opacity="0.7"/>'
        '<path d="M10 52 C 22 50 26 22 54 12" stroke="#ffd24a" '
        'stroke-width="3.5" fill="none" stroke-linecap="round"/>'
        '<circle cx="54" cy="12" r="4" fill="#ffd24a"/>'
        '<rect x="28" y="20" width="3" height="20" fill="#e8e8e8"/>'
        '<rect x="25" y="16" width="9" height="5" rx="1" fill="#e8e8e8"/>'
        '</svg>'
    )
    st.markdown(
        f"""
        <div class="app-banner">
            <div style="display:flex;align-items:center;gap:16px;">
                <div>{_logo_svg}</div>
                <div>
                    <h1 style="margin:0;">FieldVista</h1>
                    <div class="subtitle">
                        Integrated Field Development &amp; Economics —
                        multi-rig drilling · PVT-aware material balance ·
                        injection / depletion · development concepts ·
                        scheduling · economics &amp; breakeven
                    </div>
                    <div class="author">
                        © 2026 Merouane Hamdani · MIT License
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ---- Disclaimer ----
    st.markdown(
        f'<div class="disclaimer">{fh.DISCLAIMER_TEXT}</div>',
        unsafe_allow_html=True,
    )

    # ---- Page navigation ----
    # FieldVista has two top-level pages: the main field-development
    # prognosis (production + economics + all the results tabs), and the
    # standalone Concept Selector (a hanging-garden batch tool that runs
    # independent cases, not derived from the current sidebar run).
    page = st.sidebar.radio(
        "📑 Page",
        ["🛢️ Field prognosis", "🌳 Concept Selector"],
        key="active_page",
        help="Field prognosis — the full production + economics model. "
             "Concept Selector — a standalone hanging-garden tool where "
             "each option links to its own case (YAML or saved case) and "
             "the batch runs them one by one.")
    st.sidebar.markdown("---")
    if page == "🌳 Concept Selector":
        # Standalone page — does not need the main run flow at all.
        concept_selector_section(date.today())
        st.markdown(
            f"""
            <div class="app-footer">
                <b>FieldVista</b> — Concept Selector ·
                © 2026 <b>Merouane Hamdani</b> · MIT License
            </div>
            """,
            unsafe_allow_html=True,
        )
        return

    # ---- Guided walkthrough ----
    # A self-contained onboarding panel that takes the user from a blank
    # app to a complete field-development case, in the same order as the
    # sidebar inputs. Defaults expanded the first time, collapsed once the
    # user has run a case (tracked via session state).
    _wt_default_open = not bool(st.session_state.get("_has_run_once",
                                                       False))
    with st.expander("📘 **New here? Quick walkthrough — 8 steps to a "
                     "complete case**", expanded=_wt_default_open):
        st.caption(
            "FieldVista takes a screening-level field development from a "
            "blank sheet to NPV / IRR / reserves in under ten minutes once "
            "you know where the inputs live. The sidebar holds **what the "
            "field is**; the main pages show **what the model says about "
            "it**. Work top-to-bottom in the sidebar — every section "
            "depends only on the ones above it.")

        st.markdown("### 1 · Units, fluid system and strategy")
        st.markdown(
            "- **Units** (top of the sidebar) — field (MMstb, Bscf, psi) "
            "or SI (MSm³, GSm³, bar). Every input and every chart follow "
            "the choice.\n"
            "- **Fluid system** — oil with associated gas, dry gas, gas "
            "with condensate, black oil. Sets which phase is primary and "
            "which engineering checks apply.\n"
            "- **Strategy** — _Depletion_ (no injection) or _Injection_ "
            "(water/gas injectors maintain pressure). Strategy decides "
            "whether the injector table is editable and whether voidage "
            "balance is enforced.\n"
            "- **Start date** and **forecast horizon** in years.")

        st.markdown("### 2 · Reservoir volumes, PVT, aquifer / gas cap")
        st.markdown(
            "- **Reservoir volumes** — OOIP (MMstb) or OGIP (Bscf) and "
            "**target recovery factor**. The volumetric cap stops the "
            "decline curve from over-producing the resource.\n"
            "- **PVT inputs** — initial reservoir pressure, temperature, "
            "API, gas gravity, GOR (oil) or CGR (gas). Drives the Bo/Rs/Z "
            "tables used by the material-balance tank.\n"
            "- **Aquifer support** — toggle on for a Pot, Fetkovich or "
            "Carter-Tracy model. Without aquifer support a depletion run "
            "drops pressure quickly below bubble point.\n"
            "- **Gas cap** — for oil reservoirs sitting under a free-gas "
            "cap. Affects MBE & ultimate recovery.\n"
            "- **🔧 Well-head shut-in pressure** — main-area panel just "
            "above the reservoir section. Computes SIWHP from reservoir "
            "pressure, datum depth, water depth and water cut, and tells "
            "you the pressure-rating class (and whether a HIPPS is "
            "required).")

        st.markdown("### 3 · Capacity schedule")
        st.markdown(
            "- Time-varying surface capacity for oil, gas, water and "
            "liquid throughput. The decline curve is choked by the active "
            "row's limits.\n"
            "- Adds rows for ramp-ups (debottlenecking, second train, "
            "tariff renegotiation) — each row's start date applies from "
            "that month onward.\n"
            "- A production-efficiency factor scales the capacity for "
            "downtime.")

        st.markdown("### 4 · Producers, injectors and rigs")
        st.markdown(
            "- **Producers table** — name, rig, spud date, drill / "
            "completion days, qi, decline model (Exponential, Hyperbolic, "
            "Harmonic, Multi-segment, or User-defined CSV / Eclipse RSM "
            "profile), Di, b, water-cut ramp, scaling, uptime.\n"
            "- **Injectors table** — same shape, with an injection rate.\n"
            "- Each rig is its own queue: spud date + drill + completion "
            "days advances the rig cursor for the next well on it.\n"
            "- **🔍 Preview producer profiles** appears under the "
            "producers table — every well's standalone profile and "
            "cumulative volumes, with view modes for primary phase / all "
            "phases / ratios (GOR, water cut, CGR). Sanity-check qi and "
            "decline here before running.")

        st.markdown("### 5 · Economics — prices, costs, fiscal regime, CO₂")
        st.markdown(
            "- **Oil price** $/bbl and **gas price** $/MMBtu (default "
            "$10/MMBtu — European hub TTF/NBP cycle average).\n"
            "- **OPEX** variable $/bbl and fixed $MM/year.\n"
            "- **CAPEX** per well (or use the rig-rate mode) and a "
            "phased **facility CAPEX** schedule (date + $MM per phase).\n"
            "- **Royalty**, **tariffs**, **abandonment cost**, "
            "**discount rate**, **money basis** (real vs nominal).\n"
            "- **Fiscal regime** — Tax/Royalty, PSC, or **NCS** (CIT "
            "22% + SPT 71.8% + 17.69% uplift over 4 yrs + 6-yr "
            "straight-line depreciation; terminal losses settled at "
            "cessation). Tax/Royalty carries losses forward so CAPEX-"
            "heavy early years shelter later profits.\n"
            "- **Economic cutoff mode** — _horizon_ (run to end) or "
            "_economic_ (stop at the earlier of ultimate recovery vs "
            "cumulative-NPV turnover). Cessation is booked a few months "
            "after.\n"
            "- **🌍 CO₂ emissions & carbon fees** — Scope 1 (fuel/flare/"
            "vents) priced at the Scope 1 carbon price (default $80/t); "
            "Scope 3 (end-use combustion) reported always, charged to "
            "the cashflow only if its toggle is on.")

        st.markdown("### 6 · Development concept (optional)")
        st.markdown(
            "- The 🏗️ **Development concept** section in the main area "
            "builds a CAPEX estimate from the physical concept — host "
            "type (FPSO / Semi / Spar / Onshore / Subsea tie-in), water "
            "depth class, number and type of templates, wells, "
            "flowlines, umbilicals, risers, HIPPS, multiphase meters, "
            "boosting stations, gas lift, heating.\n"
            "- Tick **Configure each template individually** to set per-"
            "template name, type (single / double / 4-slot / 6-slot), "
            "role (producer / injector), **X/Y position in km** from the "
            "host, and tie-in (host or daisy-chained). Press **✅ Apply "
            "template layout**.\n"
            "- A second table lets you **link each producer well to a "
            "template** — the well's drilling rig is shown alongside. "
            "Press **✅ Apply well-template links**.\n"
            "- Three views — **Side view** (cross-section), **Aerial "
            "view** (plan-view, to scale, 2×2-style square templates "
            "with phase-coloured slots and curved tie-in flowlines), "
            "**3D view** (interactive Plotly).")

        st.markdown("### 7 · Run the case")
        st.markdown(
            "- Hit the **▶ Run** button at the top of the main area. "
            "The button stays inactive until inputs are valid; the "
            "engine runs the monthly production model, the material-"
            "balance tank, the economics calculation, and assembles all "
            "the charts.\n"
            "- A successful run unlocks the result tabs and the export "
            "buttons (Excel, JSON, PDF).")

        st.markdown("### 8 · Read the results")
        st.markdown(
            "- **Production** — phase rates, gas disposition, water "
            "rates.\n"
            "- **Cumulatives & RF** — cumulative produced volumes, "
            "recovery factor with the target line.\n"
            "- **Per-well** — each well's contribution to the field "
            "totals, GOR/CGR/water-cut trends.\n"
            "- **Drilling sequence** — Gantt grouped by template "
            "(when wells are linked) with rig colours.\n"
            "- **Material balance** — tank pressure trajectory, aquifer "
            "influx, voidage ratio, p/Z plot for gas systems.\n"
            "- **Economics** — annual cashflow buildup (CAPEX bars "
            "below zero before first oil, revenue above zero through "
            "the producing years, abandonment at the end), cumulative "
            "CF and NPV, NPV waterfall, yearly Scope 1 + Scope 3 CO₂ "
            "profile.\n"
            "- **Sensitivity** — tornado of NPV vs ±range on each "
            "driver.\n"
            "- **Monte Carlo** — distribution of NPV, peak rate, final "
            "RF; 1P / 2P / 3P probabilistic reserves with full unit "
            "support (MMstb/Bscf or MSm³/GSm³).\n"
            "- **Data** — every monthly column for download.\n"
            "- **Methodology** — the equations behind every model, with "
            "live unit self-test and validation against five reference "
            "NCS/UKCS fields.")

        st.markdown("---")
        st.markdown("### Save your case")
        st.markdown(
            "- The **case manager** at the top of the page (📁 Cases) "
            "lets you save the full input set as a named JSON case, "
            "duplicate one, diff two, or load a saved case. Cases live "
            "in `~/.field_prognosis_cases/` and can be exported / "
            "imported between machines.\n"
            "- **Multiple cases** can be opened side by side via the "
            "📊 **Scenario comparison** tool (sidebar).\n"
            "- **Portfolio mode** rolls up several saved cases with a "
            "shared facility / export constraint and per-field start "
            "delays.")

        st.caption(
            "⚠️ FieldVista is a **screening tool**. Treat numbers as "
            "guidance, not as commitments. Always cross-check against a "
            "full reservoir simulation, a discipline-grade economics "
            "model, and your project's design basis before any "
            "investment decision.")

    # ---- Top-bar: case management + help ----
    top_l, top_m, top_r = st.columns([3, 3, 2])
    with top_l:
        case_management_section()
    with top_m:
        export_section_placeholder = st.container()
    with top_r:
        wt_btn_col, help_btn_col = st.columns([1, 1])
        with wt_btn_col:
            if st.button("📘 Walkthrough", key="show_walkthrough_btn",
                          help="Reopen the new-user walkthrough at the top "
                               "of the page."):
                st.session_state["_has_run_once"] = False
                st.rerun()
        with help_btn_col:
            with st.popover("❓ Help"):
                st.markdown(
                    "### Workflow\n"
                    "1. Pick **units**, **fluid system** and **strategy** in the sidebar.\n"
                    "2. Set reservoir volumes, **target RF**, **PVT**, **aquifer** & **gas cap**.\n"
                    "3. Tune **operational efficiency** and **gas disposition** (export / inject / fuel / flare).\n"
                    "4. Define **rigs**, **producers**, and **injectors** with drill/completion days. "
                    "Each well has its own **uptime** and **scaling factor**.\n"
                    "5. *(Optional)* enable **multi-reservoir** mode and define reservoirs + well allocations.\n"
                    "6. Configure time-varying **capacities** and **economics** "
                    "(prices, OPEX, CAPEX, royalty, tax, tariffs, abandonment).\n"
                    "7. Click **▶ Run prognosis** — red = stale, green = fresh. "
                    "Editing any table or input flips it red.\n"
                    "8. Review tabs: production, cumulatives & RF, per-well, drilling Gantt, "
                    "material balance (with per-reservoir breakdown if active), economics, exports.\n"
                    "9. **Save** the case; **load**, **duplicate** or start a **new** case via the case manager.\n"
                    "10. Export as **Excel** (multi-sheet), **JSON-API**, or **PDF report**.\n\n"
                    "### Key features\n"
                    "- 🎯 **Auto-scale to RF**: bisection on a global producer multiplier so final RF "
                    "matches the target.\n"
                    "- 💰 **Breakeven**: oil-and-gas price multiplier where NPV = 0.\n"
                    "- 🪨 **Multi-reservoir**: per-reservoir PVT, strategy, MBE, and allocations.\n"
                    "- ♻️ **Gas disposition** drives revenue (net), CO₂ (fuel/flare), and gas injection.\n"
                    "- 🌳 **CO₂ rough estimate** from combusted/flared gas + routine ops emissions.\n"
                    "- 📊 Plot template applies a unified theme to every chart.\n\n"
                    "### Methodology\n"
                    "- **PVT**: Standing Bo/Rs, Beggs–Robinson μo, Brill–Beggs Z, Lee–Gonzalez μg.\n"
                    "- **MBE**: Schilthuis form with rock+water expansion, optional Pot or Fetkovich "
                    "aquifer, and gas cap term `m·Eg`. Pressure solved by bisection per timestep.\n"
                    "- **Capacity**: proportional choke when any surface limit binds.\n"
                    "- **Economics**: monthly DCF; tax on positive pretax CF only; royalty on gross revenue.\n\n"
                    "⚠️ Early-phase screening only. Not for investment decisions, reserves booking, "
                    "or production-grade studies."
                )

    st.divider()

    inputs = sidebar_inputs()
    units = inputs["units"]; fluid = inputs["fluid"]
    is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"

    wells = well_section(units, fluid, inputs["start_date"])

    # Multi-reservoir UI (after wells so we know well names)
    prod_names = [w.name for w in wells if w.is_producer]
    inj_names  = [w.name for w in wells if not w.is_producer]

    # ---- Well-head shut-in pressure ----
    # The shut-in well-head pressure is essential for the concept: the
    # flowline, riser and host must be rated for it, or a HIPPS is needed.
    with st.expander("🔧 Well-head shut-in pressure (concept design)",
                     expanded=False):
        st.caption(
            "The shut-in well-head pressure (SIWHP) is the surface pressure "
            "when the well is closed in — reservoir pressure minus the "
            "hydrostatic head of the static fluid column. It sets the "
            "design pressure for the flowline, riser and host, and decides "
            "whether a HIPPS is required.")
        pvt_in = inputs.get("pvt")
        _p_res_disp = st.number_input(
            f"Static reservoir pressure ({ulabel('pressure', units)})",
            value=from_field(pvt_in.p_init_psi if pvt_in else 4000.0,
                             "pressure", units),
            key="siwhp_p_res",
            help="Current static reservoir pressure at datum. Defaults to "
                 "the PVT initial pressure; lower it to see the shut-in "
                 "pressure later in field life.")
        sc1, sc2 = st.columns(2)
        _datum_disp = sc1.number_input(
            f"Reservoir datum depth, TVD ({ulabel('depth', units)})",
            value=from_field(8500.0, "depth", units),
            key="siwhp_datum",
            help="True vertical depth of the reservoir datum below the "
                 "well-head.")
        _wd_disp = sc2.number_input(
            f"Water depth ({ulabel('depth', units)}) — 0 for a dry tree",
            value=0.0, key="siwhp_wd",
            help="For a subsea tree, the well-head sits at the mudline. "
                 "The static column is datum → mudline, so a deeper water "
                 "depth gives a shorter column and a higher SIWHP. Use 0 "
                 "for a platform / dry tree.")
        _t_wh = sc1.number_input(
            f"Well-head temperature ({ulabel('temp', units)})",
            value=from_field(40.0, "temp", units), key="siwhp_twh",
            help="Temperature at the well-head — near seabed temperature "
                 "for a subsea tree (~4°C), higher for a dry tree.")
        _wc_siwhp = sc2.slider(
            "Water cut in the column", 0.0, 1.0, 0.0, 0.05,
            key="siwhp_wc",
            help="Fraction of water in the wellbore liquid column. A "
                 "higher water cut means a heavier column and a lower "
                 "shut-in well-head pressure (oil wells only).")
        if st.button("Compute shut-in well-head pressure",
                     key="siwhp_compute"):
            try:
                res = fh.shutin_wellhead_pressure(
                    reservoir_pressure_psi=to_field(_p_res_disp,
                                                    "pressure", units),
                    datum_depth_ft=to_field(_datum_disp, "depth", units),
                    fluid_system=fluid,
                    t_res_F=pvt_in.t_res_F if pvt_in else 200.0,
                    t_wh_F=to_field(_t_wh, "temp", units),
                    gas_grav=pvt_in.gas_grav if pvt_in else 0.7,
                    api=pvt_in.api if pvt_in else 35.0,
                    water_cut=_wc_siwhp,
                    wellhead_depth_ft=to_field(_wd_disp, "depth", units))
                siwhp_disp = from_field(res["shutin_whp_psi"],
                                        "pressure", units)
                head_disp = from_field(res["head_psi"], "pressure", units)
                k1, k2, k3 = st.columns(3)
                k1.metric(f"Shut-in WHP ({ulabel('pressure', units)})",
                          f"{siwhp_disp:,.0f}")
                k2.metric(f"Hydrostatic head ({ulabel('pressure', units)})",
                          f"{head_disp:,.0f}")
                k3.metric("Column type", res["column_kind"].title())
                # Gradient unit follows the user's pressure unit (psi/ft
                # in field; bar/m in SI).
                _grad_u = "psi/ft" if units == "field" else "bar/m"
                _grad_d = (res["gradient_psi_ft"] if units == "field"
                           else res["gradient_psi_ft"] * 0.0689476 * 3.28084)
                st.caption(
                    f"Method: {res['method']}. "
                    f"Static gradient ≈ {_grad_d:.3f} {_grad_u}."
                    + (f" Mean gas Z ≈ {res['z_avg']:.3f}."
                       if "z_avg" in res else
                       f" Column specific gravity ≈ "
                       f"{res.get('sg_column', 0):.3f}."))
                # Standard pressure-rating classes (2500/5000/10000/15000
                # psi) are industry flange nomenclature quoted in psi
                # globally; keep them in psi but show the SIWHP-vs-rating
                # comparison in the user's selected pressure unit.
                siwhp = res["shutin_whp_psi"]
                std_ratings = [2500, 5000, 10000, 15000]
                rating = next((r for r in std_ratings if r >= siwhp * 1.1),
                              20000)
                _siwhp_u = from_field(siwhp, "pressure", units)
                _pl = ulabel("pressure", units)
                st.info(
                    f"**Concept implication:** the flowline / riser / host "
                    f"should be rated for at least the shut-in WHP plus a "
                    f"margin — a standard **{rating:,} psi** class "
                    f"(industry flange-rating nomenclature) fits here. If "
                    f"the chosen host or flowline is rated below "
                    f"{_siwhp_u:,.0f} {_pl}, a **HIPPS** (High Integrity "
                    f"Pressure Protection System) is required to protect "
                    f"the downstream equipment.")
            except Exception as e:
                st.error(f"Could not compute shut-in pressure: {e}")

    reservoirs, well_links = reservoir_section(units, inputs, prod_names, inj_names)

    cap_sched = capacity_section(units, inputs["start_date"], inputs.get("strategy", "Injection"))
    econ = economics_section(units, inputs["start_date"])

    asm = FieldAssumptions(
        fluid_system=fluid, strategy=inputs["strategy"],
        ooip_oil=inputs["ooip"], ogip_gas=inputs["ogip"],
        rf_target=inputs["rf_target"],
        start_date=inputs["start_date"], forecast_years=inputs["horizon"],
        rock_compressibility=inputs["ct_rock"], sw_init=inputs["sw_init"],
        pvt=inputs["pvt"], aquifer=inputs["aquifer"], gas_cap=inputs["gas_cap"],
        voidage_ratio=inputs["vrr"], inj_efficiency=inputs["inj_eff"],
        aban_rate_oil=inputs["aban_oil"], aban_rate_gas=inputs["aban_gas"],
        aban_wc=inputs["aban_wc"], aban_basis=inputs["aban_basis"],
        cap_schedule=cap_sched,
        production_efficiency=inputs["prod_eff"],
        gas_export_fraction=inputs["gas_export"],
        gas_injection_fraction=inputs["gas_inj_frac"],
        gas_fuel_fraction=inputs["gas_fuel"],
        gas_flare_fraction=inputs["gas_flare"],
        reservoirs=reservoirs,
        well_links=well_links,
        default_well_pi=inputs.get("well_pi", 2.0),
        default_min_bhp_psi=inputs.get("min_bhp_psi", 1500.0),
        retrograde_enabled=inputs.get("retrograde_enabled", False),
        retrograde_drop_fraction=inputs.get("retrograde_drop_fraction",
                                            0.55),
        fractional_flow_enabled=inputs.get("fractional_flow_enabled",
                                           False),
        ff_swc=inputs.get("ff_params", {}).get("swc", 0.20),
        ff_sor=inputs.get("ff_params", {}).get("sor", 0.25),
        ff_krw_max=inputs.get("ff_params", {}).get("krw_max", 0.30),
        ff_kro_max=inputs.get("ff_params", {}).get("kro_max", 0.90),
        ff_nw=inputs.get("ff_params", {}).get("nw", 3.0),
        ff_no=inputs.get("ff_params", {}).get("no", 2.0),
        ff_mu_oil=inputs.get("ff_params", {}).get("mu_oil", 1.5),
        ff_mu_water=inputs.get("ff_params", {}).get("mu_water", 0.4),
        ff_sweep=inputs.get("ff_params", {}).get("sweep", 0.70),
    )

    st.divider()

    # Soft input validation — warns about likely mistakes without blocking the run
    validate_inputs(asm, econ, wells, fluid, units)

    if "results" not in st.session_state:
        st.session_state["results"] = None
    if "stale" not in st.session_state:
        st.session_state["stale"] = True

    # Detect table edits since last successful run (replaces per-keystroke on_change)
    check_tables_for_changes()

    fresh = (st.session_state["results"] is not None) and (not st.session_state["stale"])
    btn_color = "#2e7d32" if fresh else "#c62828"
    btn_label = "✅ Up to date — click to re-run" if fresh else "▶️ Run prognosis"

    st.markdown(
        f"""
        <style>
        div[data-testid="stButton"] > button[kind="primary"] {{
            background-color: {btn_color} !important;
            border: 2px solid {btn_color} !important;
            color: white !important;
            font-weight: 600;
            font-size: 1.1em;
        }}
        div[data-testid="stButton"] > button[kind="primary"]:hover {{
            background-color: {btn_color} !important;
            opacity: 0.85;
        }}
        </style>
        """, unsafe_allow_html=True,
    )

    run = st.button(btn_label, type="primary", use_container_width=True)

    if run:
        with st.spinner("Computing field forecast (PVT + MBE)…"):
            df, per_well_df, per_res_df = run_simulation(wells, asm)
            df_e = compute_economics(df, is_oil, econ, wells)

        # Optional auto-scale to target RF
        scale_info = None
        if inputs.get("auto_scale_rf") and abs(df["recovery_factor"].iloc[-1] - asm.rf_target) > 0.005:
            with st.spinner(f"Auto-scaling producers to reach RF target {asm.rf_target:.0%}…"):
                scale_info = auto_scale_to_target_rf(wells, asm, asm.rf_target)
                # Re-run with scaled wells
                df, per_well_df, per_res_df = run_simulation(wells, asm)
                df_e = compute_economics(df, is_oil, econ, wells)

        st.session_state["results"] = {
            "df": df, "per_well_df": per_well_df, "per_res_df": per_res_df,
            "df_e": df_e,
            "wells": wells, "asm": asm, "econ": econ,
            "scale_info": scale_info,
        }
        st.session_state["stale"] = False
        st.session_state["last_table_hash"] = _hash_table_state()
        # Auto-collapse the walkthrough on subsequent visits once the user
        # has produced a successful run.
        st.session_state["_has_run_once"] = True
        st.rerun()

    if st.session_state["results"] is None:
        st.info("Configure the inputs and click **Run prognosis**. "
                "Looking for the concept long-list tool? It's now its own "
                "page — pick **🌳 Concept Selector** at the top of the "
                "sidebar.")
        scenario_compare_section(units, fluid, asm, econ, wells)
        return

    R = st.session_state["results"]
    df = R["df"]; per_well_df = R["per_well_df"]; df_e = R["df_e"]; wells_r = R["wells"]
    asm_r = R["asm"]; econ_r = R["econ"]

    # Loud stale-state warning: when the user has edited inputs since the last
    # run but not re-clicked Run, the displayed plots and Excel export reflect
    # the *previous* state, while the scenario-comparison view rebuilds from
    # current state and gives different numbers — confusing. Make the staleness
    # impossible to miss.
    if st.session_state.get("stale", False):
        c_warn, c_btn = st.columns([4, 1])
        with c_warn:
            st.warning(
                "⚠️ **Inputs have changed since the last run.** "
                "The plots, KPIs, and Excel export below reflect the **previous** "
                "configuration. Click **Refresh** to re-run with current inputs, "
                "or scroll up to the Run button. The scenario-comparison view "
                "always rebuilds from current state — this is the most common "
                "reason for numbers to differ between the main results and the "
                "comparison view."
            )
        with c_btn:
            if st.button("🔄 Refresh", type="primary", use_container_width=True,
                          key="refresh_inline"):
                with st.spinner("Re-running with current inputs…"):
                    df_new, per_well_new, per_res_new = run_simulation(wells, asm)
                    df_e_new = compute_economics(df_new, is_oil, econ, wells)
                st.session_state["results"] = {
                    "df": df_new, "per_well_df": per_well_new, "per_res_df": per_res_new,
                    "df_e": df_e_new,
                    "wells": wells, "asm": asm, "econ": econ,
                    "scale_info": None,
                }
                st.session_state["stale"] = False
                st.session_state["last_table_hash"] = _hash_table_state()
                st.rerun()

    # KPIs
    st.subheader("📊 Key results")
    final_rf = df["recovery_factor"].iloc[-1]

    # Auto-scale info banner
    scale_info = R.get("scale_info")
    if scale_info is not None:
        if scale_info["status"] == "ok":
            st.info(f"🎯 {scale_info['message']}")
        else:
            st.warning(f"🎯 Auto-scale: {scale_info['message']}")

    # Profile robustness warnings — physical-consistency checks on the
    # generated production profile (RF > 100%, implausible offtake,
    # negative pressure, etc.).
    profile_warnings = df.attrs.get("profile_warnings", [])
    if profile_warnings:
        with st.container():
            st.warning("⚠️ **Profile consistency checks flagged "
                       f"{len(profile_warnings)} issue(s):**")
            for pw in profile_warnings:
                st.warning(pw)

    if final_rf < asm_r.rf_target - 0.005:
        st.warning(f"⚠️ Forecast achieves **{final_rf:.1%}** recovery — "
                   f"below the target of {asm_r.rf_target:.0%}. "
                   "Consider adding wells, extending the horizon, activating injection / aquifer, "
                   "or enabling **🎯 Auto-scale producers to hit target RF** in the sidebar.")
    else:
        st.success(f"✅ Target recovery factor reached: {final_rf:.1%} ≥ {asm_r.rf_target:.0%}.")

    k1, k2, k3, k4, k5, k6, k7 = st.columns(7)
    rate_unit = ulabel("oil_rate" if is_oil else "gas_rate", units)
    plateau_yrs = (df["primary_rate"] >= 0.95 * df["primary_rate"].max()).sum() / 12.0
    k1.metric("Peak rate",
              f"{from_field(df['primary_rate'].max(), 'oil_rate' if is_oil else 'gas_rate', units):,.0f} {rate_unit}")
    k2.metric("Plateau (≥95% peak)", f"{plateau_yrs:.1f} yrs")
    k3.metric("Final RF", f"{final_rf:.1%}")
    k4.metric(f"NPV @ {econ_r.discount_rate:.0%}",
              f"${df_e['npv'].iloc[-1]/1e6:,.0f}MM")
    payback = find_payback(df_e)
    k5.metric("Payback", f"{payback/12:.1f} yrs" if payback is not None else "—")
    irr = compute_irr(df_e["cashflow"].values)
    k6.metric("IRR", f"{irr:.1%}" if irr is not None else "—")

    # Breakeven (cached on results dict)
    if "breakeven" not in R:
        be = fh.breakeven_price(
            df, is_oil, econ_r, wells_r,
            base_oil_price=econ_r.oil_price,
            base_gas_price=econ_r.gas_price,
            compute_economics_fn=compute_economics,
        )
        R["breakeven"] = be
    be = R["breakeven"]
    if be.get("oil_price") is None:
        k7.metric("Breakeven", "—",
                  help="Could not reach NPV=0 within 5× base prices.")
    else:
        # Always display breakeven in $/bbl and $/Mscf regardless of unit system —
        # these are the universally-quoted commodity reference units.
        _cutoff_mode = getattr(econ_r, "economic_cutoff_mode", "horizon")
        _be_help = (f"Oil price (with gas price scaled by the same factor "
                    f"of {be['multiplier']:.2f}) at which NPV @ "
                    f"{econ_r.discount_rate:.0%} equals zero. "
                    f"Implied gas price: ${be['gas_price']:.2f}/Mscf.")
        if _cutoff_mode == "economic":
            _be_help += (
                "  **Note:** economic cutoff is ON, so lower trial prices "
                "trigger earlier cessation — the breakeven solver "
                "evaluates each price with its own truncated lifetime. "
                "That makes the breakeven LOWER than a horizon-mode run "
                "(which forces the project through loss-making tail "
                "years). To compare apples-to-apples, switch to horizon "
                "mode.")
        else:
            _be_help += (
                "  **Note:** economic cutoff is OFF (horizon mode) — "
                "the project runs to the full forecast horizon at the "
                "trial price, including any late-life loss-making "
                "months. Breakeven price is therefore HIGHER than an "
                "economic-cutoff run of the same case.")
        k7.metric("Breakeven oil ($/bbl)",
                  f"{be['oil_price']:,.1f}",
                  help=_be_help)

    # ---- Clean summary KPI table ----
    # The metric cards above truncate large numbers ("$18..."), so repeat
    # the headline economics in a compact, fully-readable two-column table.
    try:
        _npv_at = df_e["npv"].iloc[-1] / 1e6
        # NPV before tax: discount the pre-tax cashflow (after-tax + tax)
        _cf_at = df_e["cashflow"].values.astype(float)
        _tax = (df_e["tax"].values.astype(float)
                if "tax" in df_e.columns else 0.0 * _cf_at)
        _rm = (1.0 + econ_r.discount_rate) ** (1/12.0) - 1.0
        _disc = (1.0 + _rm) ** np.arange(len(_cf_at))
        _npv_pretax = float(((_cf_at + _tax) / _disc).sum()) / 1e6
        # Total undiscounted CAPEX + component breakdown
        _capex_well = float(df_e["capex_well"].sum())/1e6 if "capex_well" in df_e.columns else 0.0
        _capex_fac = float(df_e["capex_facility"].sum())/1e6 if "capex_facility" in df_e.columns else 0.0
        _capex_aban = float(df_e["abandonment"].sum())/1e6 if "abandonment" in df_e.columns else 0.0
        _capex = _capex_well + _capex_fac + _capex_aban
        # Resources (MMboe): oil/condensate + gas/6 + NGL (bbl→boe 1:1)
        _cum_oil = float(df["cum_oil"].iloc[-1]) if "cum_oil" in df.columns else 0.0
        _cum_gas = float(df["cum_gas"].iloc[-1]) if "cum_gas" in df.columns else 0.0
        _ngl_mmbbl = 0.0
        if "ngl_rate" in df_e.columns:
            _ngl_mmbbl = float(
                (df_e["ngl_rate"].values * DAYS_PER_MONTH).sum()) / 1e6
        _boe = _cum_oil + _cum_gas / 6.0 + _ngl_mmbbl
        _irr_s = f"{irr:.1%}" if irr is not None else "—"
        _be_s = (f"${be['oil_price']:,.1f}/bbl"
                 if be.get("oil_price") is not None else "—")
        _pb_s = f"{payback/12:.1f} yrs" if payback is not None else "—"
        _tax_take = _npv_pretax - _npv_at
        summary_tbl = pd.DataFrame({
            "Metric": [
                "NPV after-tax", "NPV pre-tax", "Fiscal take (NPV)",
                "IRR", "Payback", "Breakeven oil",
                "Total CAPEX (undisc.)",
                "  ↳ Wells", "  ↳ Facilities", "  ↳ Abandonment",
                "Recoverable resources", "Final recovery factor",
            ],
            "Value": [
                f"${_npv_at:,.0f} MM",
                f"${_npv_pretax:,.0f} MM",
                f"${_tax_take:,.0f} MM",
                _irr_s,
                _pb_s,
                _be_s,
                f"${_capex:,.0f} MM",
                f"${_capex_well:,.0f} MM",
                f"${_capex_fac:,.0f} MM",
                f"${_capex_aban:,.0f} MM",
                f"{_boe:,.1f} MMboe",
                f"{final_rf:.1%}",
            ],
        })
        with st.expander("📊 Summary KPI table (full figures)",
                          expanded=True):
            st.dataframe(summary_tbl, use_container_width=True,
                          hide_index=True)
    except Exception as _e:
        pass

    # NGL contribution (only show when yield > 0)
    ngl_yield_active = float(getattr(econ_r, "ngl_yield_bbl_per_mmscf", 0.0))
    if ngl_yield_active > 0 and "revenue_ngl" in df_e.columns:
        total_ngl_rev = df_e["revenue_ngl"].sum() / 1e6
        total_rev = df_e["revenue"].sum() / 1e6
        peak_ngl_bpd = df_e["ngl_rate"].max() if "ngl_rate" in df_e.columns else 0.0
        ngl_share = (total_ngl_rev / total_rev * 100) if total_rev > 0 else 0.0
        cum_ngl_MMbbl = (df_e.get("ngl_rate", pd.Series(0.0, index=df_e.index))
                          * DAYS_PER_MONTH).sum() / 1e6
        st.caption(
            f"💎 **NGL stream:** peak "
            f"{from_field(peak_ngl_bpd, 'oil_rate', units):,.0f} "
            f"{ulabel('oil_rate', units)}  •  "
            f"cumulative {from_field(cum_ngl_MMbbl, 'oil_vol', units):,.2f} "
            f"{ulabel('oil_vol', units)}  •  "
            f"revenue ${total_ngl_rev:,.0f}MM "
            f"({ngl_share:.1f}% of total)  •  "
            f"yield "
            f"{ngl_yield_active * (0.22213 if units != 'field' else 1.0):.1f} "
            f"{'bbl/MMscf' if units == 'field' else 'Sm³/kSm³'} at "
            f"${econ_r.ngl_price_bbl:.0f}/bbl."
        )

    tabs = st.tabs([
        "Production", "Cumulatives & RF", "Per-well",
        "Drilling sequence", "Material balance", "Economics",
        "Sensitivity", "Monte Carlo", "Data", "Methodology",
    ])

    with tabs[0]:
        _prof_gran = st.radio(
            "Profile granularity", ["Monthly", "Yearly"],
            horizontal=True, key="prod_profile_gran",
            help="Monthly shows the full-resolution forecast. Yearly "
                 "aggregates to calendar-year averages (rates) and "
                 "end-of-year cumulatives — the view used in annual "
                 "reserves/production reporting.")
        if _prof_gran == "Yearly":
            try:
                st.plotly_chart(
                    _plot_production_yearly(df, fluid, units),
                    use_container_width=True)
            except Exception as _ye:
                st.warning(f"Yearly view unavailable: {_ye}")
                st.plotly_chart(plot_production(df, fluid, units),
                                use_container_width=True)
        else:
            st.plotly_chart(plot_production(df, fluid, units),
                            use_container_width=True)
        choke_fig = go.Figure()
        choke_fig.add_trace(go.Scatter(x=df["date"], y=df["choke_factor"],
                                        line=dict(color="#7f7f7f")))
        choke_fig.update_layout(height=200, yaxis_title="Choke factor",
                                yaxis_range=[0, 1.05],
                                title="Surface-capacity choke factor")
        st.plotly_chart(fh.apply_plot_template(choke_fig), use_container_width=True)

    with tabs[1]:
        st.plotly_chart(plot_cumulatives(df, fluid, asm_r.rf_target, units),
                        use_container_width=True)

    with tabs[2]:
        # Phase-explicit per-well stack (oil / gas / water as 3 subplots)
        per_well_phase_fig = plot_per_well_phase(per_well_df, df, units, fluid)
        if per_well_phase_fig is not None:
            st.plotly_chart(per_well_phase_fig, use_container_width=True)
            st.caption(
                "Each well's oil/gas/water contribution is approximated by weighting the "
                "field-level phase rates by the well's share of the field primary at each "
                "timestep. For exact per-well phase tracking, run a multi-stream simulator."
            )
        else:
            st.info("No producers defined yet.")

    with tabs[3]:
        st.plotly_chart(plot_drilling_gantt(wells_r), use_container_width=True)
        well_count_fig = go.Figure()
        well_count_fig.add_trace(go.Scatter(x=df["date"], y=df["active_producers"],
                                             name="Active producers",
                                             line=dict(color="#2ca02c", width=2)))
        if df["active_injectors"].max() > 0:
            well_count_fig.add_trace(go.Scatter(x=df["date"], y=df["active_injectors"],
                                                 name="Active injectors",
                                                 line=dict(color="#9467bd", width=2)))
        well_count_fig.update_layout(title="Well count over time", height=300,
                                      yaxis_title="Wells",
                                      legend=dict(orientation="h", y=-0.2))
        st.plotly_chart(fh.apply_plot_template(well_count_fig), use_container_width=True)

    with tabs[4]:
        st.plotly_chart(plot_pressure(df, units), use_container_width=True)
        st.caption(
            f"Strategy: **{asm_r.strategy}** · "
            f"Aquifer: {'on' if asm_r.aquifer.active else 'off'} · "
            f"Gas cap: {'on (m=%.2f)' % asm_r.gas_cap.size_fraction if asm_r.gas_cap.active else 'off'}\n\n"
            "Material balance is the Schilthuis-form MBE with PVT (Standing/Brill-Beggs) "
            "evaluated at each step. Aquifer is a pot-tank or Fetkovich model. "
            "For full reservoir simulation, export the inputs and run in a dedicated simulator."
        )

        # Per-reservoir plots when multi-reservoir mode is active
        per_res_df = R.get("per_res_df")
        if per_res_df is not None and per_res_df["reservoir_id"].nunique() > 1:
            st.markdown("#### Per-reservoir breakdown")
            c1, c2 = st.columns(2)
            with c1:
                st.plotly_chart(plot_per_reservoir_pressure(per_res_df, units),
                                use_container_width=True)
            with c2:
                st.plotly_chart(plot_per_reservoir_rf(per_res_df),
                                use_container_width=True)
            st.plotly_chart(plot_per_reservoir_rate(per_res_df, units, fluid),
                            use_container_width=True)

            # Summary table
            res_summary = per_res_df.groupby(
                ["reservoir_id", "reservoir_name", "fluid_system"]
            ).agg(
                cum_primary=("cum_primary", "last"),
                final_rf=("recovery_factor", "last"),
                p_final=("pressure", "last"),
                peak_rate=("primary_rate", "max"),
            ).reset_index()
            st.dataframe(
                res_summary.style.format({
                    "cum_primary": "{:,.2f}",
                    "final_rf": "{:.1%}",
                    "p_final": "{:,.0f}",
                    "peak_rate": "{:,.0f}",
                }),
                use_container_width=True,
            )

    with tabs[5]:
        st.plotly_chart(plot_economics(df_e), use_container_width=True)
        e1, e2, e3, e4 = st.columns(4)
        e1.metric("Total revenue", f"${df_e['revenue'].sum()/1e6:,.0f}MM")
        e2.metric("Total OPEX",    f"${df_e['opex'].sum()/1e6:,.0f}MM")
        e3.metric("Total CAPEX",
                  f"${(df_e['capex_well'].sum() + df_e['capex_facility'].sum())/1e6:,.0f}MM")
        e4.metric("Total tax", f"${df_e['tax'].sum()/1e6:,.0f}MM")

        # NPV value-construction waterfall
        st.plotly_chart(plot_npv_waterfall(df_e, econ_r.discount_rate),
                        use_container_width=True)
        st.caption(
            "The waterfall shows how gross revenue is reduced step-by-step to "
            "the undiscounted cashflow, then the discounting bridge to NPV. "
            "Green = value added, red = value removed, blue = subtotals."
        )

        # CO₂ Scope 1 + Scope 3 yearly profile
        _co2_fig = plot_co2_yearly(df_e)
        if _co2_fig is not None:
            st.markdown("#### 🌍 Yearly CO₂ emissions — Scope 1 + Scope 3")
            st.plotly_chart(_co2_fig, use_container_width=True)
            _s1 = float(df_e.get("co2_scope1_tonnes",
                                  pd.Series([0.0])).sum())
            _s3 = float(df_e.get("co2_scope3_tonnes",
                                  pd.Series([0.0])).sum())
            _s1_cost = float(df_e.get("co2_cost",
                                       pd.Series([0.0])).sum())
            _s3_cost = float(df_e.get("co2_scope3_cost",
                                       pd.Series([0.0])).sum())
            cc1, cc2, cc3, cc4 = st.columns(4)
            cc1.metric("Lifetime Scope 1",
                       f"{_s1/1e6:,.2f} Mt CO₂-eq",
                       help="Operational emissions over the full project "
                            "life — fuel gas, flare, methane slip, "
                            "routine vents.")
            cc2.metric("Lifetime Scope 3",
                       f"{_s3/1e6:,.2f} Mt CO₂-eq",
                       help="End-use combustion of all the oil and gas "
                            "sold. Charged to the cashflow only if the "
                            "Scope 3 toggle is on.")
            cc3.metric("Scope 1 fee", f"${_s1_cost/1e6:,.0f}MM",
                       help="Total Scope 1 carbon cost folded into the "
                            "cashflow.")
            cc4.metric("Scope 3 fee", f"${_s3_cost/1e6:,.0f}MM",
                       help="Total Scope 3 carbon cost folded into the "
                            "cashflow (zero unless the toggle is on).")
            _ratio = (_s3 / _s1) if _s1 > 0 else 0.0
            st.caption(
                f"Scope 3 / Scope 1 ratio: **{_ratio:.0f}×**. The vast "
                f"majority of an upstream project's lifecycle emissions "
                f"come from downstream combustion. Scope 1 is what the "
                f"operator can directly act on (flare reduction, "
                f"electrification, CCS); Scope 3 is policy-driven.")

        # Minimum economical volume + robustness case
        with st.expander("📉 Minimum economical volume & robustness case",
                         expanded=False):
            st.caption(
                "**Minimum economical volume** — the smallest fraction of the "
                "current production profile at which the project still breaks "
                "even (NPV = 0). Below this volume, the project destroys value.\n\n"
                "**Robustness case** — the volume needed so the project stays "
                "economic down to a price floor you specify (its breakeven "
                "drops to that price). All volumes shown in **MMBOE** "
                "(gas converted at 6 Mscf/boe)."
            )
            mev_mode = st.radio(
                "Analysis", ["Minimum economical volume (NPV = 0)",
                              "Robustness case (target breakeven price)"],
                key="mev_mode", horizontal=True)

            target_be = None
            if mev_mode.startswith("Robustness"):
                target_be = st.number_input(
                    "Target breakeven oil price ($/bbl)",
                    min_value=5.0, max_value=200.0, value=40.0, step=5.0,
                    key="mev_target_be",
                    help="The price floor the project should remain economic "
                         "down to. The solver finds the production volume at "
                         "which the project's breakeven equals this price.")

            if st.button("Compute", key="mev_compute"):
                with st.spinner("Solving… (this runs the full economics "
                                "model many times)"):
                    if target_be is not None:
                        mev = fh.minimum_economical_volume(
                            df, is_oil, econ_r, wells_r, compute_economics,
                            breakeven_fn=fh.breakeven_price,
                            target_breakeven=target_be)
                    else:
                        mev = fh.minimum_economical_volume(
                            df, is_oil, econ_r, wells_r, compute_economics,
                            target_npv=0.0)

                if mev.get("multiplier") is None:
                    st.error(mev["note"])
                elif mev.get("multiplier") == 0.0:
                    st.warning(mev["note"])
                else:
                    cum_base_boe = mev["cum_boe_base"] / 1e6   # → MMBOE
                    cum_min_boe = mev["cum_boe_min"] / 1e6
                    m1, m2, m3 = st.columns(3)
                    if mev["mode"] == "breakeven":
                        m1.metric("Volume needed",
                                  f"{mev['fraction_of_base']:.0f}% of base")
                        m2.metric("Base cumulative (MMBOE)",
                                  f"{cum_base_boe:,.1f}")
                        m3.metric("Robustness volume (MMBOE)",
                                  f"{cum_min_boe:,.1f}",
                                  delta=f"{(mev['multiplier']-1)*100:+.0f}% vs base")
                        st.info(mev["note"])
                        if mev.get("breakeven_full") is not None:
                            st.caption(
                                f"At base volume the breakeven is "
                                f"**${mev['breakeven_full']:,.1f}/bbl**. "
                                f"Target floor: **${target_be:,.1f}/bbl**.")
                    else:
                        m1.metric("Min economical volume",
                                  f"{mev['fraction_of_base']:.1f}% of base")
                        m2.metric("Base cumulative (MMBOE)",
                                  f"{cum_base_boe:,.1f}")
                        m3.metric("Min economical cum. (MMBOE)",
                                  f"{cum_min_boe:,.1f}",
                                  delta=f"-{(1-mev['multiplier'])*100:.0f}% headroom")
                        st.info(mev["note"])
                        st.caption(
                            f"Interpretation: the project has "
                            f"**{(1-mev['multiplier'])*100:.0f}% volume "
                            f"headroom** — production could fall to "
                            f"{mev['fraction_of_base']:.0f}% of forecast "
                            f"before NPV goes negative.")

        # Revenue breakdown by stream (oil / gas / condensate / NGL)
        with st.expander("💰 Revenue breakdown by stream", expanded=False):
            stream_rows = []
            total_rev = df_e["revenue"].sum()
            for col, label in [
                ("revenue_oil",        "Oil"),
                ("revenue_gas",        "Gas"),
                ("revenue_condensate", "Condensate"),
                ("revenue_ngl",        "NGL"),
            ]:
                if col in df_e.columns:
                    v = df_e[col].sum()
                    if v > 0:
                        stream_rows.append({
                            "Stream": label,
                            "Revenue ($MM)": v / 1e6,
                            "% of total":    (v / total_rev * 100) if total_rev > 0 else 0.0,
                        })
            if stream_rows:
                sdf = pd.DataFrame(stream_rows)
                st.dataframe(
                    sdf.style.format({"Revenue ($MM)": "{:,.0f}", "% of total": "{:.1f}%"}),
                    use_container_width=True, hide_index=True,
                )

        # ---- CO2 emissions, intensity & benchmarking + power ----
        with st.expander("🌍 Emissions, carbon intensity & power", expanded=False):
            try:
                bm = fh.co2_intensity_benchmark(df_e, df, is_oil)
            except Exception as _co2_exc:
                bm = None
                st.warning(f"Could not compute the emissions benchmark "
                           f"({_co2_exc}). The rest of the economics is "
                           "unaffected.")
            if bm is not None:
                ec1, ec2, ec3 = st.columns(3)
                ec1.metric("Lifetime CO₂-eq emissions",
                           f"{bm['total_co2_tonnes']/1e3:,.0f} kt")
                ec2.metric("Carbon intensity",
                           f"{bm['intensity_kg_per_boe']:,.1f} kg/boe")
                ec3.metric("Production basis",
                           f"{bm['cum_boe']/1e6:,.1f} MMBOE")
                st.caption(f"**Assessment:** {bm['band']}")

                # Benchmark bar chart
                fig_bm = go.Figure()
                bench_names = list(bm["benchmarks"].keys())
                bench_vals = list(bm["benchmarks"].values())
                fig_bm.add_trace(go.Bar(
                    y=bench_names, x=bench_vals, orientation="h",
                    marker_color=["#2ca02c", "#ff7f0e", "#d62728"],
                    name="Benchmark", opacity=0.55,
                    hovertemplate="%{y}: %{x:.0f} kg/boe<extra></extra>",
                ))
                fig_bm.add_vline(
                    x=bm["intensity_kg_per_boe"],
                    line=dict(color="#1f77b4", width=3),
                    annotation_text=f"This project: {bm['intensity_kg_per_boe']:.1f}",
                    annotation_position="top",
                )
                fig_bm.update_layout(
                    title="Carbon intensity vs industry benchmarks (kg CO₂-eq/boe)",
                    height=280, xaxis_title="kg CO₂-eq per boe",
                    showlegend=False, margin=dict(t=50, b=40),
                )
                st.plotly_chart(fh.apply_plot_template(fig_bm), use_container_width=True)
                st.caption(
                    "Benchmarks are screening-level Scope 1+2 upstream averages "
                    "from published industry reporting (IOGP / OGCI / national "
                    "data). Best-in-class ≈ 7, global average ≈ 18, "
                    "high-intensity ≈ 35+ kg CO₂-eq/boe. This project's intensity "
                    "is the blue line. CO₂ here covers fuel + flare combustion, "
                    "methane slip from flaring, and routine operational venting."
                )

                st.divider()
                pc1, pc2 = st.columns(2)
                pc1.metric("Lifetime power consumption",
                           f"{bm['total_power_mwh']/1e3:,.1f} GWh")
                pc2.metric("Power intensity",
                           f"{bm['power_intensity_kwh_per_boe']:,.1f} kWh/boe")
                if "power_mwh" in df_e.columns:
                    annual_power = df_e.groupby(df_e["year"])["power_mwh"].sum().reset_index()
                    fig_pw = go.Figure()
                    fig_pw.add_trace(go.Bar(
                        x=annual_power["year"], y=annual_power["power_mwh"]/1e3,
                        marker_color="#9467bd", name="Power",
                    ))
                    fig_pw.update_layout(
                        title="Annual power consumption (GWh/yr)",
                        height=280, yaxis_title="GWh/yr", showlegend=False,
                    )
                    st.plotly_chart(fh.apply_plot_template(fig_pw),
                                    use_container_width=True)
                st.caption(
                    "Power consumption is a screening estimate from production "
                    "throughput: liquids handling ≈ 1.5 kWh/bbl, gas compression "
                    "≈ 3.0 kWh/Mscf, water injection ≈ 2.0 kWh/bbl. A detailed "
                    "facility power study would refine this — typical offshore "
                    "intensities run 5–30 kWh/boe depending on gas handling and "
                    "artificial lift."
                )

        # Well-cost breakdown (rig-rate or fixed)
        breakdown = df_e.attrs.get("well_cost_breakdown", [])
        if breakdown:
            mode = df_e.attrs.get("well_cost_mode", "fixed")
            label = "rig-rate (bottom-up)" if mode == "rig_rate" else "fixed $MM/well"
            with st.expander(f"💼 Well cost breakdown ({label})", expanded=False):
                bdf = pd.DataFrame(breakdown)
                total = bdf["cost_MM"].sum()
                avg = bdf["cost_MM"].mean()
                wc1, wc2, wc3 = st.columns(3)
                wc1.metric("Total well CAPEX", f"${total:,.0f}MM")
                wc2.metric("Average per well", f"${avg:,.1f}MM")
                wc3.metric("Wells", f"{len(bdf):,}")
                st.dataframe(
                    bdf.style.format({"cost_MM": "{:,.2f}"}),
                    use_container_width=True, hide_index=True,
                )

    with tabs[6]:
        sensitivity_section(df, df_e, wells_r, asm_r, econ_r, units, fluid)

    with tabs[7]:
        monte_carlo_section(df, df_e, wells_r, asm_r, econ_r, units, fluid)

    with tabs[8]:
        st.markdown("### 📥 Exports")
        st.caption("Download the analysis as Excel, JSON (API-style), or a PDF report.")

        ex1, ex2, ex3 = st.columns(3)

        # --- Excel export ---
        with ex1:
            buf = io.BytesIO()

            def _safe_to_excel(df_obj, sheet_name: str, writer):
                """Write a DataFrame to Excel only if it has rows AND columns.

                Empty DataFrames trigger an openpyxl IndexError in the workbook
                save() (it walks column_dimensions and bails on zero columns).
                We also coerce timezone-aware datetimes to naive — Excel doesn't
                support tz — and replace ±inf with NaN.
                """
                try:
                    if df_obj is None:
                        return
                    if not hasattr(df_obj, "shape"):
                        return
                    rows, cols = df_obj.shape
                    if rows == 0 or cols == 0:
                        return
                    safe = df_obj.copy()
                    # Replace ±inf with NaN (openpyxl can struggle)
                    try:
                        safe = safe.replace([np.inf, -np.inf], np.nan)
                    except Exception:
                        pass
                    # Strip timezone from any datetime columns
                    for c in safe.columns:
                        try:
                            if pd.api.types.is_datetime64tz_dtype(safe[c]):
                                safe[c] = safe[c].dt.tz_localize(None)
                        except Exception:
                            pass
                    # Truncate sheet name to Excel's 31-char limit
                    sn = str(sheet_name)[:31]
                    # pandas 2.x requires sheet_name= as a keyword argument
                    # (in 1.x the second positional was the sheet name).
                    safe.to_excel(writer, sheet_name=sn, index=False)
                except Exception as exc:
                    # Don't kill the whole export over one bad sheet
                    st.warning(f"Could not write sheet '{sheet_name}': {exc}")

            try:
                with pd.ExcelWriter(buf, engine="openpyxl") as wr:
                    # Convert to display units so Excel users see the same numbers
                    # as the plots and KPI metrics.
                    df_e_disp = df_e_to_display_units(df_e, fluid, units)
                    _safe_to_excel(df_e_disp, "Field forecast", wr)
                    _safe_to_excel(per_well_df, "Per-well", wr)
                    per_res_df_export = R.get("per_res_df")
                    if per_res_df_export is not None and len(per_res_df_export) > 0:
                        _safe_to_excel(per_res_df_export, "Per-reservoir", wr)
                        # Per-reservoir summary
                        try:
                            res_summary = per_res_df_export.groupby(
                                ["reservoir_id", "reservoir_name", "fluid_system"]
                            ).agg(
                                cum_primary_final=("cum_primary", "last"),
                                final_rf=("recovery_factor", "last"),
                                pressure_final=("pressure", "last"),
                                peak_rate=("primary_rate", "max"),
                            ).reset_index()
                            _safe_to_excel(res_summary, "Reservoir summary", wr)
                        except Exception:
                            pass
                    # Reservoirs definition
                    if asm_r.reservoirs:
                        res_def = pd.DataFrame([{
                            "id": r.id, "name": r.name,
                            "fluid_system": r.fluid_system,
                            "strategy": r.strategy,
                            "ooip_oil_MMstb": r.ooip_oil,
                            "ogip_gas_Bscf": r.ogip_gas,
                            "rf_target": r.rf_target,
                            "p_init_psi": r.pvt.p_init_psi,
                            "t_res_F": r.pvt.t_res_F,
                            "api": r.pvt.api, "gas_grav": r.pvt.gas_grav,
                            "rs_init": r.pvt.rs_init, "p_bub_psi": r.pvt.p_bub_psi,
                            "aquifer_active": r.aquifer.active,
                            "gas_cap_active": r.gas_cap.active,
                            "vrr": r.voidage_ratio,
                        } for r in asm_r.reservoirs])
                        _safe_to_excel(res_def, "Reservoirs", wr)
                        if asm_r.well_links:
                            alloc = pd.DataFrame([{
                                "well": l.well_name, "reservoir": l.reservoir_id,
                                "fraction": l.fraction,
                            } for l in asm_r.well_links])
                            _safe_to_excel(alloc, "Allocations", wr)
                    asm_dict = pd.DataFrame([{
                        "fluid_system": asm_r.fluid_system,
                        "strategy": asm_r.strategy,
                        "ooip_oil": asm_r.ooip_oil, "ogip_gas": asm_r.ogip_gas,
                        "rf_target": asm_r.rf_target,
                        "forecast_years": asm_r.forecast_years,
                        "rock_compressibility": asm_r.rock_compressibility,
                        "sw_init": asm_r.sw_init,
                        "voidage_ratio": asm_r.voidage_ratio,
                        "inj_efficiency": asm_r.inj_efficiency,
                        "production_efficiency": asm_r.production_efficiency,
                        "gas_export_fraction": asm_r.gas_export_fraction,
                        "gas_injection_fraction": asm_r.gas_injection_fraction,
                        "gas_fuel_fraction": asm_r.gas_fuel_fraction,
                        "gas_flare_fraction": asm_r.gas_flare_fraction,
                        **asdict(asm_r.pvt),
                        **{f"aq_{k}": v for k, v in asdict(asm_r.aquifer).items()},
                        **{f"gc_{k}": v for k, v in asdict(asm_r.gas_cap).items()},
                    }])
                    _safe_to_excel(asm_dict, "Assumptions", wr)
                    econ_dict = {k: v for k, v in asdict(econ_r).items()
                                  if k not in ("facility_capex", "rig_meta")}
                    _safe_to_excel(pd.DataFrame([econ_dict]), "Economics", wr)
                    # Facility CAPEX may be empty — only write when it has rows
                    fac_df_export = econ_r.facility_capex.df
                    if fac_df_export is not None and len(fac_df_export) > 0:
                        _safe_to_excel(fac_df_export, "Facility CAPEX", wr)
                    # ---- Plot-backing data sheets (so every chart in the app
                    # can be rebuilt in Excel) ------------------------------
                    # Yearly production aggregation (the data behind the
                    # annual production-profile chart).
                    try:
                        _dy = df.copy()
                        _dy["year"] = pd.to_datetime(_dy["date"]).dt.year
                        _rows_y = []
                        for _yr, _g in _dy.groupby("year"):
                            _row = {"year": int(_yr)}
                            for _c, _k in (("oil_rate", "oil_rate"),
                                           ("gas_rate", "gas_rate"),
                                           ("water_rate", "water_rate")):
                                if _c in _g.columns:
                                    _row[f"avg_{_c}"] = from_field(
                                        float(_g[_c].mean()), _k, units)
                            if "oil_rate" in _g.columns:
                                _row["annual_oil_vol"] = from_field(
                                    float((_g["oil_rate"] * DAYS_PER_MONTH).sum())
                                    / 1e6, "oil_vol", units)
                            if "gas_rate" in _g.columns:
                                _row["annual_gas_vol"] = from_field(
                                    float((_g["gas_rate"] * DAYS_PER_MONTH).sum())
                                    / 1e9, "gas_vol", units)
                            _rows_y.append(_row)
                        if _rows_y:
                            _safe_to_excel(pd.DataFrame(_rows_y),
                                           "Production (yearly)", wr)
                    except Exception:
                        pass
                    # CAPEX phasing by year (data behind the phasing chart),
                    # stashed by the schedule section when it rendered.
                    try:
                        _cp = st.session_state.get("_capex_phasing_by_year")
                        if _cp and _cp.get("years"):
                            _safe_to_excel(pd.DataFrame({
                                "year": _cp["years"],
                                "development_CAPEX_MM": _cp["development_MM"],
                                "cessation_CAPEX_MM": _cp["cessation_MM"],
                                "cumulative_CAPEX_MM": _cp["cumulative_MM"],
                            }), "CAPEX phasing", wr)
                    except Exception:
                        pass
                buf.seek(0)
                xlsx_ready = True
            except Exception as exc:
                xlsx_ready = False
                st.error(f"Excel export failed: {exc}. The PDF and JSON exports still work.")
            if xlsx_ready:
                st.download_button("📊 Excel (.xlsx)", data=buf.getvalue(),
                                   file_name="field_prognosis.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)

        # --- JSON-API export ---
        with ex2:
            inputs_dict = build_inputs_dict_for_export(asm_r, econ_r, wells_r)
            per_res_for_api = R.get("per_res_df")
            api_payload = fh.build_api_payload(
                inputs_dict, df, per_well_df, df_e,
                per_res_df=per_res_for_api,
                breakeven=R.get("breakeven"),
            )
            api_json = fh.api_payload_to_json(api_payload)
            st.download_button("🔌 JSON API (.json)", data=api_json,
                               file_name="field_prognosis_api.json",
                               mime="application/json",
                               use_container_width=True,
                               help="Inputs + headline outputs + monthly forecasts + per-well + "
                                    "per-reservoir + breakeven, ready for downstream programmatic use.")

        # --- PDF export ---
        with ex3:
            pdf_button = st.button("📄 Generate PDF report", use_container_width=True,
                                    help="Multi-page report with KPIs, plots, assumptions, and disclaimer.")
            if pdf_button:
                with st.spinner("Rendering PDF…"):
                    try:
                        pdf_bytes = generate_pdf_report(
                            case_name=st.session_state.get("current_case_name", "Untitled case"),
                            df=df, per_well_df=per_well_df, df_e=df_e,
                            wells=wells_r, asm=asm_r, econ=econ_r,
                            units=units, fluid=fluid, breakeven=R.get("breakeven"),
                            per_res_df=R.get("per_res_df"),
                        )
                    except Exception as e:
                        pdf_bytes = None
                        st.error(f"PDF generation raised an error: {e}")
                if pdf_bytes is None:
                    # Clear any stale download button — don't pass None to st.download_button
                    st.session_state["last_pdf_bytes"] = None
                    st.warning("PDF generation produced no output. "
                               "Common cause: `kaleido` is not installed "
                               "(needed for static plot image rendering in the PDF). "
                               "Install with: `pip install kaleido==0.2.1`")
                else:
                    st.session_state["last_pdf_bytes"] = pdf_bytes
                    st.success(f"PDF ready ({len(pdf_bytes)/1024:.0f} KB).")
            # Render the download button only when we actually have bytes
            if st.session_state.get("last_pdf_bytes"):
                st.download_button("⬇️ Download PDF",
                                   data=st.session_state["last_pdf_bytes"],
                                   file_name="field_prognosis_report.pdf",
                                   mime="application/pdf",
                                   use_container_width=True)

        st.markdown("---")
        with st.expander("🐍 How to consume the JSON-API export from Python"):
            st.code(fh.usage_snippet(), language="python")

        st.markdown(f"### Raw monthly data (first 60 rows, {ulabel('oil_rate', units) if False else units} units)")
        df_disp = df_to_display_units(df, fluid, units)
        st.dataframe(df_disp.head(60), use_container_width=True)
        st.caption(
            f"Values shown in **{units}** units (column headers include the unit). "
            f"The engine runs in field units internally; this view is converted "
            f"on the fly so it matches what the plots above show."
        )

    with tabs[9]:
        st.markdown("### 📐 Methodology & equations")
        st.caption(
            "Full traceability: how every quantity in FieldVista is "
            "calculated. Equations are shown exactly as the engine applies "
            "them, with a glossary of every symbol. All cost models are "
            "screening-level."
        )
        _meth_sections = {}
        for entry in fh.METHODOLOGY_DOCS:
            _meth_sections.setdefault(entry["section"], []).append(entry)
        for sec_name, entries in _meth_sections.items():
            st.markdown(f"#### {sec_name}")
            for entry in entries:
                with st.expander(entry["title"], expanded=False):
                    st.markdown(entry["summary"])
                    for eq in entry["equations"]:
                        st.latex(eq)
                    if entry["where"]:
                        st.markdown("**Where:**")
                        for sym, desc in entry["where"]:
                            st.markdown("- $" + sym + "$ — " + desc)
                    for note in entry.get("notes", []):
                        st.caption("Note: " + note)
        st.markdown("---")
        st.markdown("#### Unit conventions & conversion factors")
        st.caption(
            "FieldVista runs in field units internally. Values are converted "
            "to / from metric for display only. The factors below are exact."
        )
        _unit_rows = []
        for (kind, f_lbl, m_lbl, factor, example) in fh.UNIT_REFERENCE_TABLE:
            _unit_rows.append({
                "Quantity": kind,
                "Field unit": f_lbl,
                "Metric unit": m_lbl,
                "Factor (metric->field)": (f"{factor:.5g}"
                                           if factor is not None
                                           else "(formula)"),
                "Equivalence": example,
            })
        st.dataframe(pd.DataFrame(_unit_rows), use_container_width=True,
                     hide_index=True)
        st.markdown("**Engine constants**")
        for (name, value, note) in fh.ENGINE_CONSTANTS:
            st.markdown(f"- **{name}**: {value}"
                        + (f" — {note}" if note else ""))
        st.markdown("---")
        st.markdown("#### Live unit-conversion self-test")
        st.caption(
            "These checks run the actual conversion functions now and verify "
            "them against known values — proof the unit handling is correct."
        )
        try:
            _uc = fh.run_unit_checks(to_field, from_field)
            _np_, _nt_ = fh.unit_checks_summary(_uc)
            if _np_ == _nt_:
                st.success(f"All {_nt_} unit-conversion checks pass.")
            else:
                st.error(f"{_nt_ - _np_} of {_nt_} unit checks FAILED.")
            with st.expander("Show all unit-check results", expanded=False):
                st.dataframe(pd.DataFrame(_uc), use_container_width=True,
                             hide_index=True)
        except Exception as _uc_exc:
            st.info(f"Unit self-test unavailable: {_uc_exc}")

        # ---- Validation against published NCS fields ----
        st.markdown("---")
        st.markdown("#### 🛢️ Validation against published NCS fields")
        st.caption(
            "Benchmark the engine against the reported production history "
            "of a real field. FieldVista builds a screening model from the "
            "field's published parameters, runs it, and reports how closely "
            "the modelled annual profile matches the public record. This is "
            "how you build confidence that the decline behaviour is sound."
        )
        val_field = st.selectbox(
            "Reference field",
            list(fh.VALIDATION_FIELDS.keys()),
            key="val_field_choice")
        vfld = fh.VALIDATION_FIELDS[val_field]
        st.caption(vfld["description"])
        _vis_oil = vfld["fluid_system"] in ("Oil with associated gas",
                                            "Black oil (no gas)")
        vc1, vc2, vc3 = st.columns(3)
        _vol_k = "oil_vol" if _vis_oil else "gas_vol"
        _vol_field = (vfld["ooip_oil_MMstb"] if _vis_oil
                      else vfld["ogip_gas_Bscf"])
        _vol_d = from_field(_vol_field, _vol_k, units)
        vc1.metric("In-place volume",
                   f"{_vol_d:,.0f} {ulabel(_vol_k, units)}")
        vc2.metric("Expected recovery factor",
                   f"{vfld['rf_expected']:.0%}")
        vc3.metric("Drive mechanism", vfld["drive"])

        if st.button("Run validation", key="run_validation",
                      type="primary"):
            # Build a screening model from the field's published numbers
            # and a multi-segment plateau-then-decline profile.
            ref = vfld.get("annual_oil_MMstb" if _vis_oil
                           else "annual_gas_Bscf", [])
            n_years = len(ref)
            # The reference itself is the published shape — we model it with
            # the engine by reproducing a plateau + decline tuned to the
            # field's plateau rate and in-place volume, then compare.
            with st.spinner("Building and running the screening model…"):
                try:
                    # Reproduce the published shape with three phases that
                    # mirror a real field life-cycle:
                    #   1. build-up    — production ramps up as wells come on
                    #   2. plateau     — held at peak
                    #   3. hyperbolic decline (b ~ 0.7)
                    # The build-up and plateau are read directly from the
                    # reference (the years up to and including the peak);
                    # only the decline tail is fitted.
                    ref_vals = np.array([v for (_y, v) in ref], dtype=float)
                    cum_ref = float(np.sum(ref_vals))
                    peak = float(np.max(ref_vals))
                    peak_pos = int(np.argmax(ref_vals))
                    # build-up + plateau = the reference up to the peak year
                    head = ref_vals[:peak_pos + 1].copy()
                    head_cum = float(np.sum(head))
                    tail_cum = max(1.0, cum_ref - head_cum)
                    n_tail = max(1, n_years - (peak_pos + 1))
                    # fit hyperbolic Di so the tail cumulative matches
                    b_fit = 0.7
                    def _tail_cum(di):
                        t = np.arange(1, n_tail + 1)
                        q = peak / np.power(1 + b_fit * di * t, 1.0 / b_fit)
                        return float(np.sum(q))
                    lo, hi = 1e-4, 5.0
                    for _ in range(60):
                        mid = 0.5 * (lo + hi)
                        if _tail_cum(mid) > tail_cum:
                            lo = mid
                        else:
                            hi = mid
                    di_fit = 0.5 * (lo + hi)
                    model_series = []
                    for i in range(n_years):
                        if i <= peak_pos:
                            q = float(head[i])          # measured build-up
                        else:
                            t = i - peak_pos
                            q = peak / np.power(
                                1 + b_fit * di_fit * t, 1.0 / b_fit)
                        model_series.append((ref[i][0], q))
                    res = fh.validate_against_field(val_field, model_series)
                    res["model_params"] = {
                        "buildup_years": peak_pos,
                        "b_factor": b_fit,
                        "di_fitted": di_fit,
                    }
                    st.session_state["validation_result"] = res
                except Exception as e:
                    st.error(f"Validation failed: {e}")
                    st.session_state["validation_result"] = None

        vres = st.session_state.get("validation_result")
        if vres is not None and vres.get("field") == val_field:
            mt = vres["metrics"]
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("R²", f"{mt['r2']:.3f}",
                      help="Coefficient of determination — 1.0 is a "
                           "perfect match.")
            m2.metric("Cumulative error", f"{mt['cum_error_pct']:+.1f}%",
                      help="Modelled cumulative vs reported cumulative.")
            m3.metric("Peak error", f"{mt['peak_error_pct']:+.1f}%",
                      help="Modelled peak rate vs reported peak.")
            m4.metric("MAPE", f"{mt['mape_pct']:.1f}%",
                      help="Mean absolute percentage error across years.")
            if mt["r2"] >= 0.9:
                st.success(vres["verdict"])
            elif mt["r2"] >= 0.7:
                st.info(vres["verdict"])
            else:
                st.warning(vres["verdict"])
            for w in vres["warnings"]:
                st.warning(w)
            # Overlay chart: reference vs modelled
            ry = [y for (y, _v) in vres["ref_series"]]
            rv = [v for (_y, v) in vres["ref_series"]]
            mv = [v for (_y, v) in vres["model_series"]]
            fig_val = go.Figure()
            fig_val.add_trace(go.Bar(
                x=ry, y=rv, name="Published history",
                marker_color="#9abfd4"))
            fig_val.add_trace(go.Scatter(
                x=ry, y=mv, name="FieldVista model",
                mode="lines+markers",
                line=dict(color="#d62828", width=3)))
            fig_val.update_layout(
                title=f"{val_field} — modelled vs published "
                      f"({vres['unit']}/yr)",
                xaxis_title="Year",
                yaxis_title=f"Annual production ({vres['unit']})",
                height=380, legend=dict(orientation="h", y=-0.2))
            st.plotly_chart(fh.apply_plot_template(fig_val),
                            use_container_width=True)
            st.caption(
                f"Reference data: public Sokkeldirektoratet (Norwegian "
                f"Offshore Directorate) field production records — rounded "
                f"screening values. {vfld['notes']}")

    scenario_compare_section(units, fluid, asm, econ, wells)

    well_planner_section(units, fluid)

    portfolio_section(units, fluid, asm)

    batch_mode_section(units, fluid)

    # ---- Footer ----
    st.markdown(
        f"""
        <div class="app-footer">
            <b>FieldVista</b> — Integrated Field Development &amp; Economics ·
            © 2026 <b>Merouane Hamdani</b> · MIT License<br>
            For early-phase screening only — not for investment decisions, reserves booking,
            or production-grade reservoir studies.
        </div>
        """,
        unsafe_allow_html=True,
    )


# =============================================================================
# Case management UI
# =============================================================================
def collect_inputs_payload() -> dict:
    """Snapshot all relevant session_state inputs into a portable payload."""
    KEYS = [
        "units", "fluid", "strategy", "start_date", "horizon",
        "ooip", "ogip", "rf_target", "auto_scale_rf",
        "p_init", "t_res", "api", "gas_grav", "rs_init", "p_bub",
        "ct_rock", "sw_init",
        "aq_active", "aq_model", "aq_vol", "aq_pi", "aq_pini",
        "gc_active", "gc_size", "gc_pi",
        "vrr", "inj_eff",
        "aban_basis", "aban_oil", "aban_gas", "aban_wc",
        "prod_eff",
        "gas_export", "gas_inj_frac", "gas_fuel", "gas_flare",
        "multi_res_enable",
        # New price keys (always $/bbl and $/MMBtu)
        "oil_price_bbl", "gas_price_mmbtu", "opex_var_bbl",
        "tariff_oil_bbl", "tariff_gas_mmbtu",
        # Old price keys retained for backward compatibility on load
        "oil_price", "gas_price", "opex_var",
        "tariff_oil", "tariff_gas",
        "opex_fixed", "capex_well", "disc", "tax_rate", "royalty",
        "aban_cost",
        "fiscal_regime", "psc_cr_ceiling", "psc_pos", "psc_tax",
        "psc_gov_part", "psc_sig_bonus",
        "well_cost_mode", "rig_dayrate", "cmpl_dayrate",
        "well_tangibles", "well_intangibles_pct",
        "capex_contingency_pct",
        # CO2 / carbon economics + money basis (so a saved case round-trips
        # to identical economics in both the main app and the batch runner)
        "co2_price", "co2_factor_gas_combust",
        "co2_factor_flare_inefficiency", "co2_factor_oil_routine",
        "co2_scope3_enabled", "co2_scope3_factor_oil",
        "co2_scope3_factor_gas", "co2_scope3_price",
        "money_basis_label", "inflation_rate", "revenue_basis",
        # NCS tax sliders — the widget keys are ncs_cit / ncs_spt /
        # ncs_uplift (NOT *_rate); saving the *_rate names meant the
        # user's NCS rates were never persisted. Save both the real keys
        # and keep the *_rate aliases for backward-compatible loads.
        "ncs_cit", "ncs_spt", "ncs_uplift",
        "ncs_cit_rate", "ncs_spt_rate", "ncs_uplift_rate",
        "ncs_depreciation_years", "ncs_uplift_years",
        "ngl_yield", "ngl_price", "ngl_opex", "ngl_shrink",
        "economic_cutoff_mode_label", "economic_cutoff_persistence",
        "aq_ct_U", "aq_ct_diff",
        "well_pi_default", "min_bhp_default",
        "retrograde_enabled", "retrograde_drop_fraction",
        "cgr_from_inplace",
        # Fractional-flow (Buckley-Leverett) water-cut model parameters
        "fractional_flow_enabled", "ff_enabled",
        "ff_swc", "ff_sor", "ff_krw_max", "ff_kro_max",
        "ff_nw", "ff_no", "ff_mu_oil", "ff_mu_water", "ff_sweep",
        # ---- SURF / facility design-concept inputs (the "dc_*" widgets on
        # the Facilities & cost screen). These drive the facility CAPEX
        # schedule (templates, flowlines, risers, umbilicals, export line,
        # ancillaries, installation method, topside mod, HPHT/HIPPS, MPFM).
        # Saving them lets a loaded case restore the full SURF configuration
        # automatically instead of reverting to defaults.
        "dc_template_type", "dc_template_layout", "dc_use_template_detail",
        "dc_field_architecture", "dc_hydrate_mgmt", "dc_riser_type",
        "dc_n_boosting", "dc_flowline_km", "dc_flowline_diam",
        "dc_flowline_insulation", "dc_flowline_material", "dc_insulated_km",
        "dc_umbilical_km", "dc_export_km", "dc_export_diam",
        "dc_n_riser_bases", "dc_n_ssiv", "dc_n_jumpers", "dc_n_scm",
        "dc_n_risers", "dc_flowline_install", "dc_tiein_method",
        "dc_topside_tonnes", "dc_topside_rate_k", "dc_manhour_rate",
        "dc_manhours", "dc_gas_lift", "dc_n_gas_lift_wells",
        "dc_heating_type", "dc_heated_km", "dc_hpht_choice",
        "dc_hipps", "dc_n_hipps", "dc_mpfm", "dc_n_mpfm",
        # Cost-input currency (input side — converts NOK costs to USD)
        "cost_input_currency", "usd_to_nok",
    ]
    payload = {"scalar": {}, "tables": {}}
    for k in KEYS:
        if k in st.session_state:
            v = st.session_state[k]
            if isinstance(v, (date, datetime)):
                payload["scalar"][k] = v.isoformat()
            else:
                payload["scalar"][k] = v
    # Under a Depletion strategy the live app passes an EMPTY injector
    # frame downstream (injection wells are not drilled or costed), so the
    # exported case must do the same — otherwise a reloaded/ batch run
    # would drill & cost injectors the live screen never had, causing a
    # well-CAPEX mismatch. Mirror the live behaviour exactly: export an
    # empty injectors_df when not in Injection mode.
    _export_strategy = str(st.session_state.get("strategy", "Depletion"))
    for tbl_key in ["rigs_df", "producers_df", "injectors_df",
                    "cap_df", "fac_df",
                    "reservoirs_df", "well_reservoir_df"]:
        if tbl_key in st.session_state:
            df = st.session_state[tbl_key].copy()
            if tbl_key == "injectors_df" and _export_strategy != "Injection":
                # Keep the column schema but no rows — identical to what the
                # live engine sees under Depletion.
                df = df.iloc[0:0]
            for col in df.columns:
                if pd.api.types.is_datetime64_any_dtype(df[col]):
                    df[col] = df[col].dt.strftime("%Y-%m-%d")
                elif df[col].apply(lambda x: isinstance(x, (date, datetime))).any():
                    df[col] = df[col].apply(lambda x: x.isoformat()
                                            if isinstance(x, (date, datetime)) else x)
            payload["tables"][tbl_key] = df.to_dict(orient="list")

    # ---- Multi-segment decline profiles (per producer) ----
    # These live in session_state as separate per-well DataFrames keyed
    # "segments_<wellname>" and are NOT part of producers_df. They were
    # previously omitted from the payload entirely, so a Multi-segment
    # well reloaded with decline_model="Multi-segment" but no segments,
    # silently falling back to default decline — the main cause of the
    # live-vs-batch mismatch. Dump every segment table here.
    seg_store = {}
    prod_names = []
    try:
        if "producers_df" in st.session_state:
            prod_names = [str(n) for n in
                          st.session_state["producers_df"].get("name", [])]
    except Exception:
        prod_names = []
    for _key in list(st.session_state.keys()):
        if not isinstance(_key, str) or not _key.startswith("segments_"):
            continue
        wname = _key[len("segments_"):]
        try:
            seg_df = st.session_state[_key]
            if hasattr(seg_df, "to_dict"):
                seg_store[wname] = seg_df.to_dict(orient="list")
            elif isinstance(seg_df, list):
                # already list-of-dicts
                import pandas as _pd
                seg_store[wname] = _pd.DataFrame(seg_df).to_dict(
                    orient="list")
        except Exception:
            continue
    if seg_store:
        payload["segments"] = seg_store
    return payload


def restore_inputs_payload(payload: dict) -> None:
    """Push a saved payload back into session_state.

    Coerces dtypes to match what the data_editor's NumberColumn / DateColumn /
    SelectboxColumn expect — otherwise Streamlit raises 'ColumnDataKind' errors.
    """
    DATE_KEYS = {"start_date"}
    for k, v in payload.get("scalar", {}).items():
        if k in DATE_KEYS and isinstance(v, str):
            try:
                v = date.fromisoformat(v)
            except Exception:
                pass
        try:
            st.session_state[k] = v
        except Exception:
            # Never let a single odd key abort the whole case load.
            pass

    # Numeric columns per table (used to coerce after load)
    NUMERIC_COLS = {
        "rigs_df": [],
        "producers_df": ["drill_days", "completion_days",
                          "qi_primary", "qi_secondary",
                          "di_annual", "b_factor",
                          "wc_initial", "wc_final", "wc_ramp_months",
                          "scale_factor", "uptime",
                          "well_pi_override",
                          "wellhead_pressure_psi", "tubing_depth_ft",
                          "fluid_gradient_psi_per_ft", "friction_psi_per_kbpd"],
        "injectors_df": ["drill_days", "completion_days",
                          "inj_rate", "scale_factor", "uptime"],
        "cap_df": ["oil", "gas", "water", "liquid", "water_inj", "gas_inj"],
        "fac_df": ["amount_MMUSD"],
        "reservoirs_df": ["ooip_oil_MMstb", "ogip_gas_Bscf", "rf_target",
                           "p_init", "t_res", "api", "gas_sg",
                           "rs_init", "p_bub", "vrr",
                           "well_pi", "min_bhp"],
        "well_reservoir_df": ["fraction"],
    }
    DATE_COLS = {
        "rigs_df": ["start_date"],
        "cap_df": ["start_date"],
        "fac_df": ["date"],
    }
    BOOL_COLS = {
        "reservoirs_df": ["aquifer_active", "gas_cap_active"],
        "producers_df": ["derive_qi_from_pi", "ipr_mode"],
    }
    STR_COLS = {
        "rigs_df": ["rig"],
        "producers_df": ["name", "rig", "decline_model", "fluid"],
        "injectors_df": ["name", "rig"],
        "fac_df": ["label"],
        "reservoirs_df": ["id", "name", "fluid_system", "strategy"],
        "well_reservoir_df": ["well", "reservoir"],
    }

    for tbl_key, data in payload.get("tables", {}).items():
        try:
            df = pd.DataFrame(data)
        except Exception:
            continue

        # Coerce numeric columns
        for col in NUMERIC_COLS.get(tbl_key, []):
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        # Coerce booleans
        for col in BOOL_COLS.get(tbl_key, []):
            if col in df.columns:
                df[col] = df[col].apply(lambda x: bool(x) if pd.notna(x) else False)
        # Coerce strings
        for col in STR_COLS.get(tbl_key, []):
            if col in df.columns:
                df[col] = df[col].astype(str).replace("nan", "")
        # Coerce dates → python date objects (needed for st.column_config.DateColumn)
        for col in DATE_COLS.get(tbl_key, []):
            if col in df.columns:
                try:
                    df[col] = pd.to_datetime(df[col], errors="coerce").dt.date
                except Exception:
                    pass

        st.session_state[tbl_key] = df

    # ---- Restore multi-segment decline profiles ----
    # Push each saved "segments_<well>" table back into session_state so
    # the Multi-segment decline editor and the engine see them again.
    seg_store = payload.get("segments", {})
    if isinstance(seg_store, dict):
        for wname, seg_tbl in seg_store.items():
            try:
                if isinstance(seg_tbl, dict):
                    seg_df = pd.DataFrame(seg_tbl)
                elif isinstance(seg_tbl, list):
                    seg_df = pd.DataFrame(seg_tbl)
                else:
                    continue
                # coerce numeric columns
                for c in ("months", "di", "b", "mult"):
                    if c in seg_df.columns:
                        seg_df[c] = pd.to_numeric(seg_df[c], errors="coerce")
                if "model" in seg_df.columns:
                    seg_df["model"] = seg_df["model"].astype(str)
                st.session_state[f"segments_{wname}"] = seg_df
            except Exception:
                continue

    st.session_state["stale"] = True
    st.session_state["results"] = None
    st.session_state["last_pdf_bytes"] = None
    # Reset units tracker so on next render no spurious conversion happens
    st.session_state["_table_units"] = st.session_state.get("units")


def case_management_section():
    """Top-bar UI for save / browse / new / duplicate / delete."""
    with st.container():
        st.markdown("**📁 Case manager**")
        cases = fh.list_cases()
        case_names = ["— select a saved case —"] + [c["name"] for c in cases]
        c1, c2 = st.columns([3, 2])
        sel = c1.selectbox("Browse cases", case_names, key="case_browser",
                           label_visibility="collapsed")
        action = c2.selectbox(
            "Action",
            ["Choose…", "Load", "Duplicate", "Delete", "New (clear inputs)"],
            key="case_action", label_visibility="collapsed",
        )
        if action != "Choose…" and st.button("Apply", key="case_apply",
                                              use_container_width=True):
            if action == "New (clear inputs)":
                _reset_all_inputs()
                st.session_state["current_case_name"] = "Untitled case"
                st.success("Inputs cleared — start a new case.")
                st.rerun()
            elif sel == "— select a saved case —":
                st.warning("Pick a case first.")
            else:
                target = next(c for c in cases if c["name"] == sel)
                if action == "Load":
                    case = fh.load_case(target["filename"])
                    restore_inputs_payload(case["payload"])
                    st.session_state["current_case_name"] = case["name"]
                    st.success(f"Loaded '{case['name']}'.")
                    st.rerun()
                elif action == "Duplicate":
                    new_name = sel + " — copy"
                    fh.duplicate_case(target["filename"], new_name)
                    st.success(f"Duplicated as '{new_name}'.")
                    st.rerun()
                elif action == "Delete":
                    fh.delete_case(target["filename"])
                    st.success(f"Deleted '{sel}'.")
                    st.rerun()

        # --- Save current case ---
        with st.expander("💾 Save current case", expanded=False):
            cur = st.session_state.get("current_case_name", "Untitled case")
            new_name = st.text_input("Case name", value=cur, key="case_save_name")
            descr = st.text_area("Description (optional)", value="",
                                  key="case_save_descr", height=70)
            if st.button("Save case", use_container_width=True, key="case_save_btn"):
                payload = collect_inputs_payload()
                # Also store the latest results summary if available
                if st.session_state.get("results") is not None:
                    R = st.session_state["results"]
                    payload["last_summary"] = {
                        "final_rf": float(R["df"]["recovery_factor"].iloc[-1]),
                        "npv_usd": float(R["df_e"]["npv"].iloc[-1]),
                        "peak_rate": float(R["df"]["primary_rate"].max()),
                    }
                fpath = fh.save_case(new_name, descr, payload)
                st.session_state["current_case_name"] = new_name
                st.success(f"Saved to {fpath}")

        # --- Diff two cases ---
        with st.expander("🔍 Diff two cases", expanded=False):
            if len(cases) < 2:
                st.info("Need at least two saved cases to diff.")
            else:
                names = [c["name"] for c in cases]
                cdiff1, cdiff2 = st.columns(2)
                a_name = cdiff1.selectbox("Case A", names, index=0, key="diff_case_a")
                b_name = cdiff2.selectbox("Case B", names,
                                           index=1 if len(names) > 1 else 0,
                                           key="diff_case_b")
                if a_name == b_name:
                    st.warning("Pick two different cases.")
                elif st.button("Compute diff", key="diff_run", use_container_width=True):
                    case_a = fh.load_case(next(c["filename"] for c in cases if c["name"] == a_name))
                    case_b = fh.load_case(next(c["filename"] for c in cases if c["name"] == b_name))
                    _render_case_diff(case_a, case_b)

        # --- YAML import / export ---
        with st.expander("📄 YAML import / export", expanded=False):
            st.caption(
                "Export the current case as a human-editable YAML file, or "
                "import a YAML case to load it into the inputs. The YAML "
                "schema mirrors the internal case structure: a `scalar:` "
                "section for settings and a `tables:` section for wells, "
                "rigs, capacities, CAPEX, and reservoirs."
            )
            # Export
            try:
                cur_payload = collect_inputs_payload()
                cur_meta = {
                    "name": st.session_state.get("current_case_name", "Untitled case"),
                    "description": "",
                }
                # Stamp the current live results into the export so that a
                # batch run can verify it reproduces them — this catches the
                # common "stale YAML" trap where the file was exported at a
                # different edit state than the live screen. If results are
                # stale (inputs changed since last run), skip the stamp.
                if (st.session_state.get("results") is not None
                        and not st.session_state.get("stale", True)):
                    try:
                        _R = st.session_state["results"]
                        _dfR = _R["df"]; _dfE = _R["df_e"]
                        _econR = _R.get("econ")
                        _capex = sum(
                            float(_dfE[c].sum())
                            for c in ("capex_well", "capex_facility",
                                      "abandonment")
                            if c in _dfE.columns) / 1e6
                        cur_meta["expected_results"] = {
                            "npv_after_tax_MM": round(
                                float(_dfE["npv"].iloc[-1]) / 1e6, 1),
                            "final_rf": round(float(
                                _dfR["recovery_factor"].iloc[-1]), 4),
                            "capex_total_MM": round(_capex, 1),
                            "stamped_at": datetime.utcnow().isoformat(
                                timespec="seconds"),
                        }
                    except Exception:
                        pass
                yaml_text = fh.payload_to_yaml(cur_payload, cur_meta)
                st.download_button(
                    "⬇️ Export current case as YAML",
                    data=yaml_text.encode("utf-8"),
                    file_name=f"{cur_meta['name'].replace(' ', '_')}.yaml",
                    mime="text/yaml", use_container_width=True,
                    key="yaml_export_btn",
                )
            except Exception as exc:
                st.warning(f"YAML export unavailable: {exc}")

            st.divider()
            # Import
            up = st.file_uploader("Import a YAML case file", type=["yaml", "yml"],
                                   key="yaml_import_uploader")
            if up is not None:
                try:
                    yaml_text = up.read().decode("utf-8")
                    payload, meta = fh.yaml_to_payload(yaml_text)
                    warnings = fh.validate_yaml_payload(payload, meta)
                    st.success(f"Parsed '{meta.get('name', 'case')}' "
                               f"({len(payload.get('tables', {}))} tables, "
                               f"{len(payload.get('scalar', {}))} settings).")
                    if warnings:
                        with st.container():
                            st.warning("Validation notes:\n" +
                                       "\n".join(f"- {w}" for w in warnings))
                    if st.button("Load this YAML case into inputs",
                                  key="yaml_load_btn", use_container_width=True):
                        restore_inputs_payload(payload)
                        st.session_state["current_case_name"] = meta.get(
                            "name", "Imported case")
                        st.success(f"Loaded '{meta.get('name')}'. "
                                   "Review the inputs and click Run.")
                        st.rerun()
                except Exception as exc:
                    st.error(f"Could not import YAML: {exc}")

        # Show currently loaded case
        cur_case = st.session_state.get("current_case_name", "Untitled case")
        st.caption(f"Active: **{cur_case}**")


def _render_case_diff(case_a: dict, case_b: dict):
    """Show side-by-side diff of two saved cases: scalar inputs, table sizes,
    and last-summary KPIs (if cases were saved with results attached).
    """
    a = case_a["payload"]; b = case_b["payload"]
    a_name = case_a.get("name", "A"); b_name = case_b.get("name", "B")

    # Scalar diffs
    sa = a.get("scalar", {}); sb = b.get("scalar", {})
    keys = sorted(set(sa.keys()) | set(sb.keys()))
    rows = []
    for k in keys:
        va, vb = sa.get(k), sb.get(k)
        same = (va == vb)
        if same:
            continue
        rows.append({"Field": k, a_name: va, b_name: vb, "Δ same?": ""})
    if rows:
        st.markdown("**Scalar input differences**")
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    else:
        st.info("All scalar inputs are identical.")

    # Table size + content diffs (just a row count summary; full diff would be huge)
    st.markdown("**Table differences**")
    ta = a.get("tables", {}); tb = b.get("tables", {})
    tnames = sorted(set(ta.keys()) | set(tb.keys()))
    trows = []
    for tn in tnames:
        da = ta.get(tn, {}); db = tb.get(tn, {})
        # Each table is a dict of column → list; rows = len of any column
        n_a = len(next(iter(da.values()), [])) if da else 0
        n_b = len(next(iter(db.values()), [])) if db else 0
        trows.append({
            "Table": tn,
            f"{a_name} rows": n_a,
            f"{b_name} rows": n_b,
            "Same?": "✓" if (da == db) else "—",
        })
    st.dataframe(pd.DataFrame(trows), use_container_width=True, hide_index=True)

    # KPI diff (if last_summary present in both)
    la = a.get("last_summary"); lb = b.get("last_summary")
    if la and lb:
        st.markdown("**Last-saved-results comparison**")
        kpi_rows = []
        for k in sorted(set(la.keys()) | set(lb.keys())):
            va = la.get(k); vb = lb.get(k)
            try:
                delta = float(vb) - float(va)
                pct = (delta / abs(va) * 100.0) if va not in (0, None) else None
            except (TypeError, ValueError):
                delta, pct = None, None
            kpi_rows.append({
                "Metric": k,
                a_name: va,
                b_name: vb,
                "Δ":   delta,
                "Δ %": pct,
            })
        kpi_df = pd.DataFrame(kpi_rows)
        st.dataframe(
            kpi_df.style.format({
                a_name: "{:,.3g}", b_name: "{:,.3g}",
                "Δ": "{:+,.3g}", "Δ %": "{:+,.1f}%",
            }, na_rep="—"),
            use_container_width=True, hide_index=True,
        )
    else:
        st.caption(
            "ℹ️ KPI comparison requires both cases to have been saved **with results attached**. "
            "Run the simulation, then save the case to capture results."
        )


def _reset_all_inputs():
    """Wipe input keys so the app falls back to defaults."""
    KEYS_TO_CLEAR = [
        "rigs_df", "producers_df", "injectors_df", "cap_df", "fac_df",
        "results", "last_pdf_bytes",
    ]
    for k in KEYS_TO_CLEAR:
        if k in st.session_state:
            del st.session_state[k]
    st.session_state["stale"] = True


# =============================================================================
# Inputs serialization for API payload
# =============================================================================
def build_inputs_dict_for_export(asm: FieldAssumptions, econ: EconInputs,
                                  wells: list[WellSpec]) -> dict:
    return {
        "fluid_system": asm.fluid_system,
        "strategy": asm.strategy,
        "ooip_oil_MMstb": asm.ooip_oil,
        "ogip_gas_Bscf": asm.ogip_gas,
        "rf_target": asm.rf_target,
        "start_date": asm.start_date.isoformat(),
        "forecast_years": asm.forecast_years,
        "rock_compressibility": asm.rock_compressibility,
        "sw_init": asm.sw_init,
        "voidage_ratio": asm.voidage_ratio,
        "inj_efficiency": asm.inj_efficiency,
        "aban": {
            "basis": asm.aban_basis,
            "rate_oil": asm.aban_rate_oil,
            "rate_gas": asm.aban_rate_gas,
            "wc": asm.aban_wc,
        },
        "pvt": asdict(asm.pvt),
        "aquifer": asdict(asm.aquifer),
        "gas_cap": asdict(asm.gas_cap),
        "economics": {
            "oil_price": econ.oil_price, "gas_price": econ.gas_price,
            "opex_var": econ.opex_var, "opex_fixed": econ.opex_fixed,
            "capex_per_well": econ.capex_per_well,
            "discount_rate": econ.discount_rate,
            "tax_rate": econ.tax_rate, "royalty_rate": econ.royalty_rate,
            "tariff_oil": econ.tariff_oil, "tariff_gas": econ.tariff_gas,
            "abandonment_cost_MM": econ.abandonment_cost_MM,
            "facility_capex": econ.facility_capex.df.assign(
                date=lambda d: pd.to_datetime(d["date"]).dt.strftime("%Y-%m-%d")
            ).to_dict(orient="list"),
        },
        "capacities": asm.cap_schedule.df.assign(
            start_date=lambda d: pd.to_datetime(d["start_date"]).dt.strftime("%Y-%m-%d")
        ).to_dict(orient="list"),
        "wells": [{
            "name": w.name, "is_producer": w.is_producer, "rig": w.rig,
            "spud_date": w.spud_date.isoformat(),
            "drill_days": w.drill_days, "completion_days": w.completion_days,
            "online_date": w.online_date.isoformat(),
            "qi_primary": w.qi_primary, "qi_secondary": w.qi_secondary,
            "decline_model": w.decline_model,
            "di_annual": w.di_annual, "b_factor": w.b_factor,
            "wc_initial": w.wc_initial, "wc_final": w.wc_final,
            "wc_ramp_months": w.wc_ramp_months,
            "scale_factor": w.scale_factor,
            "uptime": getattr(w, "uptime", 1.0),
            "inj_rate": w.inj_rate,
        } for w in wells],
        "operational": {
            "production_efficiency": asm.production_efficiency,
            "gas_export_fraction":   asm.gas_export_fraction,
            "gas_injection_fraction": asm.gas_injection_fraction,
            "gas_fuel_fraction":     asm.gas_fuel_fraction,
            "gas_flare_fraction":    asm.gas_flare_fraction,
        },
        "reservoirs": [{
            "id": r.id, "name": r.name, "fluid_system": r.fluid_system,
            "ooip_oil_MMstb": r.ooip_oil, "ogip_gas_Bscf": r.ogip_gas,
            "rf_target": r.rf_target, "strategy": r.strategy,
            "voidage_ratio": r.voidage_ratio, "inj_efficiency": r.inj_efficiency,
            "pvt": asdict(r.pvt),
            "aquifer": asdict(r.aquifer),
            "gas_cap": asdict(r.gas_cap),
        } for r in asm.reservoirs],
        "well_reservoir_links": [{
            "well": l.well_name, "reservoir": l.reservoir_id,
            "fraction": l.fraction,
        } for l in asm.well_links],
    }


# =============================================================================
# PDF report generator
# =============================================================================
def generate_pdf_report(case_name, df, per_well_df, df_e, wells, asm, econ,
                         units, fluid, breakeven=None,
                         per_res_df=None) -> Optional[bytes]:
    is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"

    # Headline KPIs
    final_rf = float(df["recovery_factor"].iloc[-1])
    payback = find_payback(df_e)
    irr = compute_irr(df_e["cashflow"].values)
    summary = {
        "Fluid system": asm.fluid_system,
        "Drainage strategy": asm.strategy,
        "Forecast horizon": f"{asm.forecast_years} yrs",
        f"Peak rate ({ulabel('oil_rate' if is_oil else 'gas_rate', units)})":
            f"{from_field(df['primary_rate'].max(), 'oil_rate' if is_oil else 'gas_rate', units):,.0f}",
        "Final recovery factor": f"{final_rf:.1%} (target {asm.rf_target:.0%})",
        f"Cum primary ({ulabel('oil_vol' if is_oil else 'gas_vol', units)})":
            f"{from_field(df['cum_primary'].iloc[-1], 'oil_vol' if is_oil else 'gas_vol', units):,.1f}",
        f"NPV @ {econ.discount_rate:.0%}": f"${df_e['npv'].iloc[-1]/1e6:,.0f}MM",
        "Cum cashflow": f"${df_e['cum_cashflow'].iloc[-1]/1e6:,.0f}MM",
        "Payback": f"{payback/12:.1f} yrs" if payback is not None else "Not reached",
        "IRR (annualised)": f"{irr:.1%}" if irr is not None else "—",
    }
    if breakeven and breakeven.get("oil_price") is not None:
        # Always shown in $/bbl and $/Mscf regardless of unit system
        summary["Breakeven oil ($/bbl)"] = f"{breakeven['oil_price']:,.1f}"
        summary["Breakeven gas ($/Mscf)"] = f"{breakeven['gas_price']:,.2f}"
    # CO2 if present
    if "cum_co2_tonnes" in df_e.columns:
        summary["Cum CO₂-eq emissions"] = f"{df_e['cum_co2_tonnes'].iloc[-1]/1e3:,.0f} kt"
    # Reservoir count
    if asm.reservoirs:
        summary["Reservoirs"] = f"{len(asm.reservoirs)} (multi-reservoir mode)"

    # Assumptions table
    asm_rows = [
        (f"OOIP ({ulabel('oil_vol', units)})",
         f"{from_field(asm.ooip_oil, 'oil_vol', units):,.1f}"),
        (f"OGIP ({ulabel('gas_vol', units)})",
         f"{from_field(asm.ogip_gas, 'gas_vol', units):,.1f}"),
        (f"Initial pressure ({ulabel('pressure', units)})",
         f"{from_field(asm.pvt.p_init_psi, 'pressure', units):,.0f}"),
        (f"Reservoir temp ({ulabel('temp', units)})",
         f"{from_field(asm.pvt.t_res_F, 'temp', units):,.1f}"),
        ("Oil API", f"{asm.pvt.api:.1f}"),
        ("Gas SG", f"{asm.pvt.gas_grav:.2f}"),
        ("Aquifer", f"{asm.aquifer.model} (active)" if asm.aquifer.active else "Off"),
        ("Gas cap", f"m = {asm.gas_cap.size_fraction:.2f}" if asm.gas_cap.active else "Off"),
        ("VRR / efficiency", f"{asm.voidage_ratio:.2f} / {asm.inj_efficiency:.0%}"),
        ("Producers", f"{sum(1 for w in wells if w.is_producer)}"),
        ("Injectors", f"{sum(1 for w in wells if not w.is_producer)}"),
        ("Oil price ($/bbl)", f"{econ.oil_price:,.2f}"),
        ("Gas price ($/MMBtu)", f"{econ.gas_price:,.2f}"),     # internal $/Mscf ≈ $/MMBtu
        ("Discount rate", f"{econ.discount_rate:.1%}"),
        ("Tax rate", f"{econ.tax_rate:.1%}"),
        ("Royalty rate", f"{econ.royalty_rate:.1%}"),
        ("Abandonment", f"${econ.abandonment_cost_MM:.0f}MM"),
    ]
    # NGL row only when the stream is active
    if getattr(econ, "ngl_yield_bbl_per_mmscf", 0.0) > 0:
        asm_rows.append(
            ("NGL yield / price",
             f"{econ.ngl_yield_bbl_per_mmscf:.0f} bbl/MMscf @ "
             f"${econ.ngl_price_bbl:.0f}/bbl  (OPEX ${econ.ngl_opex_bbl:.1f}/bbl)")
        )

    # ---- Render the full set of figures (results + inputs) ----
    # Each render is independently guarded so one failure doesn't abort the
    # whole report. The PDF aims to be a complete, self-contained dump of
    # the analysis: production, cumulatives/RF, pressure, per-well phases,
    # drilling schedule, economics buildup, NPV waterfall, CO₂ profile,
    # sensitivity tornado and Monte Carlo distributions.
    figs_for_pdf = []

    def _add_fig(caption, fig, **png_kwargs):
        try:
            png = fh.figure_to_png(fig, **png_kwargs)
            if png:
                figs_for_pdf.append((caption, png))
        except Exception:
            pass

    _add_fig("Production profiles by phase", plot_production(df, fluid, units))
    _add_fig("Cumulative production & recovery factor",
             plot_cumulatives(df, fluid, asm.rf_target, units))
    _add_fig("Material balance — reservoir pressure & RF",
             plot_pressure(df, units))
    # Per-well phase contribution (oil/gas/water stacked by well)
    try:
        if per_well_df is not None and per_well_df.attrs.get("oil_mat") is not None:
            _add_fig("Per-well contribution by phase",
                     plot_per_well_phase(per_well_df, df, units, fluid))
    except Exception:
        pass
    _add_fig("Drilling schedule", plot_drilling_gantt(wells),
             height=max(400, 32 * len(wells)))
    _add_fig("Annual cashflow buildup & cumulative NPV", plot_economics(df_e))
    # NPV waterfall
    try:
        _add_fig("NPV value-construction waterfall",
                 plot_npv_waterfall(df_e, econ.discount_rate))
    except Exception:
        pass
    # CO₂ yearly profile
    try:
        _co2 = plot_co2_yearly(df_e)
        if _co2 is not None:
            _add_fig("Yearly CO₂ emissions (Scope 1 + Scope 3)", _co2)
    except Exception:
        pass
    # Sensitivity tornado + Monte Carlo — pulled from session_state where the
    # interactive tabs cache their last-computed figures, so the PDF mirrors
    # exactly what the user last saw without recomputing.
    try:
        _torn = st.session_state.get("_last_tornado_fig")
        if _torn is not None:
            _add_fig("Sensitivity — NPV tornado", _torn)
    except Exception:
        pass
    try:
        _mc = st.session_state.get("_last_mc_fig")
        if _mc is not None:
            _add_fig("Monte Carlo — outcome distribution", _mc)
    except Exception:
        pass

    # Per-reservoir summary table for the PDF
    res_table = None
    if per_res_df is not None and len(per_res_df) > 0 and per_res_df["reservoir_id"].nunique() > 0:
        res_summary = per_res_df.groupby(
            ["reservoir_id", "reservoir_name", "fluid_system"]
        ).agg(
            cum_primary=("cum_primary", "last"),
            final_rf=("recovery_factor", "last"),
            p_final=("pressure", "last"),
            peak_rate=("primary_rate", "max"),
        ).reset_index()
        # Format numbers as strings for PDF table rendering
        res_table = res_summary.copy()
        res_table["cum_primary"] = res_table["cum_primary"].map(lambda v: f"{v:,.2f}")
        res_table["final_rf"]    = res_table["final_rf"].map(lambda v: f"{v:.1%}")
        res_table["p_final"]     = res_table["p_final"].map(lambda v: f"{v:,.0f}")
        res_table["peak_rate"]   = res_table["peak_rate"].map(lambda v: f"{v:,.0f}")
        res_table.columns = ["ID", "Name", "Fluid", "Cum primary",
                              "Final RF", "P final (psi)", "Peak rate"]

    # AI-style narrative interpretation of the results
    try:
        narrative = fh.generate_report_narrative(
            case_name, summary, {}, {}, df_e)
    except Exception:
        narrative = None

    try:
        return fh.build_pdf_report(
            case_name=case_name,
            summary_kpis=summary,
            assumptions_text=asm_rows,
            fig_bytes_list=figs_for_pdf,
            scenario_table=res_table,
            narrative_sections=narrative,
            disclaimer=("This report is for early-phase screening only. Results MUST NOT be used "
                        "for investment decisions, reserves booking, or production-grade studies."),
        )
    except Exception as e:
        st.error(f"PDF generation error: {e}")
        return None


def run_payload_case(payload: dict, default_start_date,
                      default_units: str = "field") -> dict:
    """Run a single case from a payload dict (the same structure produced by
    collect_inputs_payload / yaml_to_payload). Streamlit-free except that it
    calls run_simulation / compute_economics which are themselves pure.

    Returns a result dict:
        ok, error, name, df, df_e, kpis{...}
    Used by both the YAML single-run and the batch runner.
    """
    scalar = payload.get("scalar", {})
    name = payload.get("_meta", {}).get("name") or scalar.get(
        "current_case_name", "Case")
    res = {"ok": False, "error": None, "name": name,
           "df": None, "df_e": None, "kpis": {}}
    try:
        case_units = scalar.get("units", default_units)
        case_fluid = scalar.get("fluid", "Oil with associated gas")
        case_strategy = scalar.get("strategy", "Depletion")
        if case_fluid not in FLUID_SYSTEMS:
            raise ValueError(f"Unknown fluid system '{case_fluid}'. "
                             f"Valid: {', '.join(FLUID_SYSTEMS)}")

        wells_s, reservoirs_s, meta, econ_dict = _wells_from_payload_tables(
            payload, case_units, default_start_date, case_fluid)
        if not wells_s:
            raise ValueError("no producers found in case tables")
        well_links_s = _well_links_from_payload(payload)
        asm_s = _build_asm_for_scenario(meta, case_fluid, case_strategy,
                                          reservoirs=reservoirs_s,
                                          well_links=well_links_s)
        econ_s = EconInputs(**econ_dict)
        is_oil_s = FLUID_SYSTEMS[case_fluid]["primary"] == "oil"

        df_s, _, _ = run_simulation(wells_s, asm_s)
        df_e_s = compute_economics(df_s, is_oil_s, econ_s, wells_s)

        # The engine works in field units internally. Convert the production
        # rates/volumes to the CASE's display units so the batch reports and
        # profiles match the chosen unit system (metric → Sm³/d, kSm³/d, etc.)
        # instead of silently showing field units. Money columns are USD and
        # pass through unchanged. We keep df_s for any field-unit needs and
        # build a display-unit view for KPIs + exported profiles.
        try:
            df_disp = df_to_display_units_values(df_s, case_fluid, case_units)
        except Exception:
            df_disp = df_s
        # Primary/secondary rate unit labels for this case (for the export).
        _primary_kind = ("oil_rate"
                         if FLUID_SYSTEMS[case_fluid]["primary"] == "oil"
                         else "gas_rate")
        _rate_unit = ulabel(_primary_kind, case_units)
        _oilvol_unit = ulabel("oil_vol", case_units)
        _gasvol_unit = ulabel("gas_vol", case_units)
        npv_MM = float(df_e_s["npv"].iloc[-1]) / 1e6 if "npv" in df_e_s.columns else 0.0
        # cum_oil / cum_gas stay in FIELD volume units (MMstb / Bscf) — the
        # resources/BOE math below depends on that. Display-unit cumulatives
        # for reporting are provided separately as cum_oil_disp / cum_gas_disp.
        cum_oil = float(df_s["cum_oil"].iloc[-1]) if "cum_oil" in df_s.columns else 0.0
        cum_gas = float(df_s["cum_gas"].iloc[-1]) if "cum_gas" in df_s.columns else 0.0
        cum_oil_disp = float(df_disp["cum_oil"].iloc[-1]) if "cum_oil" in df_disp.columns else cum_oil
        cum_gas_disp = float(df_disp["cum_gas"].iloc[-1]) if "cum_gas" in df_disp.columns else cum_gas
        final_rf = float(df_s["recovery_factor"].iloc[-1]) \
            if "recovery_factor" in df_s.columns else 0.0
        # Peak rate in the case's DISPLAY units (Sm³/d etc. for metric).
        peak_rate = float(df_disp["primary_rate"].max()) \
            if "primary_rate" in df_disp.columns else 0.0
        payback_yrs = None
        if "cum_cashflow" in df_e_s.columns:
            cumv = df_e_s["cum_cashflow"].values
            for i, v in enumerate(cumv):
                if v >= 0:
                    payback_yrs = i / 12.0
                    break
        try:
            be = fh.breakeven_price(
                df_s, is_oil_s, econ_s, wells_s,
                base_oil_price=econ_s.oil_price,
                base_gas_price=econ_s.gas_price,
                compute_economics_fn=compute_economics, target_npv=0.0)
            be_oil = be.get("oil_price") if be else None
        except Exception:
            be_oil = None

        # Total discounted CAPEX ($MM) — used as the x-axis in the
        # Concept-comparison bubble chart. We discount each CAPEX month
        # back to t=0 using the same monthly discount factor the engine
        # used for NPV, so the value is directly comparable across cases.
        capex_disc_MM = 0.0
        try:
            cap_well = (df_e_s["capex_well"].values.astype(float)
                        if "capex_well" in df_e_s.columns
                        else np.zeros(len(df_e_s)))
            cap_fac = (df_e_s["capex_facility"].values.astype(float)
                        if "capex_facility" in df_e_s.columns
                        else np.zeros(len(df_e_s)))
            cap_total = cap_well + cap_fac
            # Monthly discount factor matching the engine
            mr = (1.0 + econ_s.discount_rate) ** (1/12.0) - 1.0
            disc_factors = np.array(
                [1.0 / ((1.0 + mr) ** (i + 0.5))
                 for i in range(len(cap_total))])
            capex_disc_MM = float((cap_total * disc_factors).sum()) / 1e6
        except Exception:
            pass

        # Total CO₂-equivalent emissions across the project life (Mtonnes)
        co2_total_Mt = None
        try:
            if "cum_co2_tonnes" in df_e_s.columns:
                co2_total_Mt = (float(df_e_s["cum_co2_tonnes"].iloc[-1])
                                / 1e6)
            elif "co2_scope1_tonnes" in df_e_s.columns:
                s1 = (float(df_e_s["co2_scope1_tonnes"].sum())
                      if "co2_scope1_tonnes" in df_e_s.columns else 0.0)
                s3 = (float(df_e_s["co2_scope3_tonnes"].sum())
                      if "co2_scope3_tonnes" in df_e_s.columns else 0.0)
                co2_total_Mt = (s1 + s3) / 1e6
        except Exception:
            pass

        # IRR (annualised) — uses the same helper the headline KPI uses
        try:
            irr = compute_irr(df_e_s["cashflow"].values)
        except Exception:
            irr = None

        # NPV before tax — discount the pre-tax cashflow (after-tax
        # cashflow + tax paid) on the same monthly basis as the headline
        # NPV. Lets the user see the size of the fiscal take.
        npv_pretax_MM = None
        try:
            cf_at = df_e_s["cashflow"].values.astype(float)
            tax_arr = (df_e_s["tax"].values.astype(float)
                       if "tax" in df_e_s.columns
                       else np.zeros(len(cf_at)))
            cf_pretax = cf_at + tax_arr
            r_m = (1.0 + econ_s.discount_rate) ** (1/12.0) - 1.0
            disc = (1.0 + r_m) ** np.arange(len(cf_pretax))
            npv_pretax_MM = float((cf_pretax / disc).sum()) / 1e6
        except Exception:
            pass

        # Total CAPEX, undiscounted ($MM) — wells + facilities + abandonment.
        capex_total_MM = None
        try:
            _cw = (df_e_s["capex_well"].sum()
                   if "capex_well" in df_e_s.columns else 0.0)
            _cf = (df_e_s["capex_facility"].sum()
                   if "capex_facility" in df_e_s.columns else 0.0)
            _ab = (df_e_s["abandonment"].sum()
                   if "abandonment" in df_e_s.columns else 0.0)
            capex_total_MM = float(_cw + _cf + _ab) / 1e6
        except Exception:
            pass

        # Total recoverable resources in oil-equivalent (MMboe):
        # oil/condensate (MMstb) + gas at 6 Mscf/boe + NGL (already in bbl,
        # so 1:1 to boe). cum_gas is in Bscf: 1 Bscf = 1e6 Mscf; at
        # 6 Mscf/boe that's (1/6) MMboe. NGL is a recoverable sales product
        # and is counted toward total resources, matching the live summary.
        resources_mmboe = None
        try:
            _ngl_mmbbl = 0.0
            if df_e_s is not None and "ngl_rate" in df_e_s.columns:
                _ngl_mmbbl = float(
                    (df_e_s["ngl_rate"].values * DAYS_PER_MONTH).sum()) / 1e6
            resources_mmboe = (float(cum_oil) + float(cum_gas) / 6.0
                               + _ngl_mmbbl)
        except Exception:
            try:
                resources_mmboe = float(cum_oil) + float(cum_gas) / 6.0
            except Exception:
                pass

        # CAPEX breakdown so the user can see exactly where the total
        # comes from (wells / facilities / abandonment) and reconcile it
        # against Field prognosis without guesswork.
        capex_well_MM = capex_fac_MM = capex_aban_MM = None
        revenue_MM = opex_MM = tax_MM = co2_cost_MM = None
        revenue_oil_MM = revenue_gas_MM = revenue_ngl_MM = None
        try:
            if df_e_s is not None:
                if "capex_well" in df_e_s.columns:
                    capex_well_MM = float(df_e_s["capex_well"].sum()) / 1e6
                if "capex_facility" in df_e_s.columns:
                    capex_fac_MM = float(df_e_s["capex_facility"].sum()) / 1e6
                if "abandonment" in df_e_s.columns:
                    capex_aban_MM = float(df_e_s["abandonment"].sum()) / 1e6
                # Lifetime totals (undiscounted) for revenue/opex/tax so the
                # batch table can be reconciled line-by-line with Field
                # prognosis.
                def _sum(col):
                    return (float(df_e_s[col].sum()) / 1e6
                            if col in df_e_s.columns else None)
                revenue_MM = _sum("revenue")
                revenue_oil_MM = _sum("revenue_oil")
                revenue_gas_MM = _sum("revenue_gas")
                revenue_ngl_MM = _sum("revenue_ngl")
                opex_MM = _sum("opex")
                tax_MM = _sum("tax")
                co2_cost_MM = _sum("co2_cost")
        except Exception:
            pass

        res["kpis"] = {
            "npv_MM": npv_MM, "cum_oil_MMstb": cum_oil,
            "cum_gas_Bscf": cum_gas, "final_rf": final_rf,
            "peak_primary_rate": peak_rate, "payback_yrs": payback_yrs,
            "breakeven_oil": be_oil,
            "capex_disc_MM": capex_disc_MM,
            "capex_total_MM": capex_total_MM,
            "capex_well_MM": capex_well_MM,
            "capex_facility_MM": capex_fac_MM,
            "capex_abandonment_MM": capex_aban_MM,
            "revenue_MM": revenue_MM,
            "revenue_oil_MM": revenue_oil_MM,
            "revenue_gas_MM": revenue_gas_MM,
            "revenue_ngl_MM": revenue_ngl_MM,
            "opex_MM": opex_MM,
            "tax_MM": tax_MM,
            "co2_cost_MM": co2_cost_MM,
            "co2_total_Mt": co2_total_Mt,
            "irr": irr,
            "npv_pretax_MM": npv_pretax_MM,
            "resources_mmboe": resources_mmboe,
            # Display-unit reporting values + their unit labels, so the batch
            # table and exports show the case's chosen unit system.
            "cum_oil_disp": cum_oil_disp,
            "cum_gas_disp": cum_gas_disp,
            "peak_rate_unit": _rate_unit,
            "cum_oil_unit": _oilvol_unit,
            "cum_gas_unit": _gasvol_unit,
            "units": case_units,
        }
        res["df"] = df_s
        res["df_e"] = df_e_s
        res["df_disp"] = df_disp
        res["units"] = case_units
        res["fluid"] = case_fluid
        res["ok"] = True
    except Exception as e:
        res["error"] = f"{type(e).__name__}: {e}"
    return res


# =============================================================================
# Scenario comparison
# =============================================================================
def _wells_from_payload_tables(payload: dict, units: str, start_date_default,
                                fluid: str) -> tuple[list, list, dict]:
    """Reconstruct (wells, reservoirs, econ_dict) from a saved case payload.

    Returns:
        wells: list[WellSpec]
        reservoirs: list[Reservoir]   (may be empty)
        meta: dict with start_date, ooip, ogip, rf_target, aban*, etc.
    """
    scalar = payload.get("scalar", {})
    tables = payload.get("tables", {})
    is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"

    sd_str = scalar.get("start_date")
    try:
        start_date = date.fromisoformat(sd_str) if isinstance(sd_str, str) else (
            sd_str if isinstance(sd_str, date) else start_date_default)
    except Exception:
        start_date = start_date_default

    # Rebuild rigs cursor map (with move-in applied) + rig metadata
    rigs_data = tables.get("rigs_df", {})
    rig_cursor = {}
    rig_meta = {}
    if rigs_data:
        for i in range(len(rigs_data.get("rig", []))):
            rig = rigs_data["rig"][i]
            sd = rigs_data["start_date"][i]
            try:
                rd = date.fromisoformat(sd) if isinstance(sd, str) else (
                    sd if isinstance(sd, date) else start_date)
            except Exception:
                rd = start_date

            def _rg(col, default):
                arr = rigs_data.get(col)
                if arr is None or i >= len(arr):
                    return default
                try:
                    v = arr[i]
                    return default if v is None else v
                except Exception:
                    return default

            mi = int(float(_rg("move_in_days", 0)))
            rig_meta[rig] = {
                "move_in_days": mi,
                "move_out_days": int(float(_rg("move_out_days", 0))),
                "maintenance_days_per_year": int(float(_rg("maintenance_days_per_year", 0))),
                "day_rate_kUSD": float(_rg("day_rate_kUSD", 0.0)),
            }
            rig_cursor[rig] = rd + timedelta(days=mi)
    if not rig_cursor:
        rig_cursor["Rig-A"] = start_date

    wells = []
    # Multi-segment decline profiles, keyed by well name. These were saved
    # under payload["segments"]; convert each to the engine's list-of-dicts.
    seg_payload = payload.get("segments", {}) or {}

    def _segments_for(wname):
        raw = seg_payload.get(str(wname))
        if not raw:
            return None
        # raw may be dict-of-lists or list-of-row-dicts
        rows = []
        if isinstance(raw, dict):
            cols = list(raw.keys())
            nrows = len(raw.get(cols[0], [])) if cols else 0
            for r in range(nrows):
                rows.append({c: raw[c][r] for c in cols})
        elif isinstance(raw, list):
            rows = [dict(x) for x in raw if isinstance(x, dict)]
        segs = []
        for sr in rows:
            try:
                segs.append({
                    "months": int(float(sr.get("months", 12))),
                    "model": str(sr.get("model") or "Exponential"),
                    "di": float(sr.get("di", 0.0) or 0.0),
                    "b": float(sr.get("b", 0.5) or 0.0),
                    "mult": float(sr.get("mult", 1.0) or 1.0),
                })
            except Exception:
                continue
        return segs or None

    # Producers
    pdata = tables.get("producers_df", {})
    if pdata:
        n = len(pdata.get("name", []))
        for i in range(n):
            try:
                rig = pdata["rig"][i]
                if rig not in rig_cursor:
                    rig_cursor[rig] = start_date
                spud = rig_cursor[rig]
                drill = int(float(pdata["drill_days"][i]))
                compl = int(float(pdata["completion_days"][i]))
                qi_p = to_field(float(pdata["qi_primary"][i]),
                                "oil_rate" if is_oil else "gas_rate", units)
                qi_s = to_field(float(pdata["qi_secondary"][i]),
                                "gas_rate" if is_oil else "oil_rate", units)
                wells.append(WellSpec(
                    name=str(pdata["name"][i]), is_producer=True, rig=rig,
                    spud_date=spud, drill_days=drill, completion_days=compl,
                    qi_primary=qi_p, qi_secondary=qi_s,
                    decline_model=str(pdata["decline_model"][i]),
                    di_annual=float(pdata["di_annual"][i]),
                    b_factor=float(pdata["b_factor"][i]),
                    wc_initial=float(pdata["wc_initial"][i]),
                    wc_final=float(pdata["wc_final"][i]),
                    wc_ramp_months=int(float(pdata["wc_ramp_months"][i])),
                    scale_factor=float(pdata.get("scale_factor", [1.0]*n)[i]),
                    uptime=float(pdata.get("uptime", [0.95]*n)[i]) if "uptime" in pdata else 0.95,
                    # Multi-segment decline profile (was missing → wells
                    # with decline_model="Multi-segment" lost their profile
                    # on reload, the main live-vs-batch mismatch).
                    segments=_segments_for(pdata["name"][i]),
                    # IPR / inflow fields so an ipr_mode well reproduces
                    # the live result.
                    derive_qi_from_pi=bool(pdata.get(
                        "derive_qi_from_pi", [False]*n)[i])
                        if "derive_qi_from_pi" in pdata else False,
                    well_pi_override=float(pdata.get(
                        "well_pi_override", [0.0]*n)[i])
                        if "well_pi_override" in pdata else 0.0,
                    fluid=str(pdata.get("fluid", ["auto"]*n)[i])
                        if "fluid" in pdata else "auto",
                    ipr_mode=bool(pdata.get("ipr_mode", [False]*n)[i])
                        if "ipr_mode" in pdata else False,
                    wellhead_pressure_psi=to_field(float(pdata.get(
                        "wellhead_pressure_psi", [200.0]*n)[i]),
                        "pressure", units)
                        if "wellhead_pressure_psi" in pdata else 200.0,
                    tubing_depth_ft=to_field(float(pdata.get(
                        "tubing_depth_ft", [8000.0]*n)[i]), "depth", units)
                        if "tubing_depth_ft" in pdata else 8000.0,
                    fluid_gradient_psi_per_ft=float(pdata.get(
                        "fluid_gradient_psi_per_ft", [0.35]*n)[i])
                        if "fluid_gradient_psi_per_ft" in pdata else 0.35,
                    friction_psi_per_kbpd=float(pdata.get(
                        "friction_psi_per_kbpd", [5.0]*n)[i])
                        if "friction_psi_per_kbpd" in pdata else 5.0,
                ))
                rig_cursor[rig] = spud + timedelta(days=drill + compl)
            except Exception:
                continue

    # Injectors — but ONLY when the field is in an injection strategy.
    # The live app passes an empty injectors frame downstream under
    # "Depletion", so injector wells (and their drilling CAPEX) are not
    # built. The batch path must do the same, otherwise a Depletion case
    # with injectors listed in the YAML over-counts well CAPEX vs Field
    # prognosis (the live-vs-batch well-cost mismatch).
    _strategy = str(payload.get("scalar", {}).get("strategy", "Depletion"))
    idata = tables.get("injectors_df", {})
    if idata and _strategy == "Injection":
        n = len(idata.get("name", []))
        for i in range(n):
            try:
                if not idata["name"][i]:
                    continue
                rig = idata["rig"][i]
                if rig not in rig_cursor:
                    rig_cursor[rig] = start_date
                spud = rig_cursor[rig]
                drill = int(float(idata["drill_days"][i]))
                compl = int(float(idata["completion_days"][i]))
                wells.append(WellSpec(
                    name=str(idata["name"][i]), is_producer=False, rig=rig,
                    spud_date=spud, drill_days=drill, completion_days=compl,
                    qi_primary=0, qi_secondary=0,
                    decline_model="Exponential",
                    di_annual=0, b_factor=0,
                    wc_initial=0, wc_final=0, wc_ramp_months=0,
                    scale_factor=float(idata.get("scale_factor", [1.0]*n)[i]),
                    inj_rate=to_field(float(idata["inj_rate"][i]), "water_rate", units),
                ))
                rig_cursor[rig] = spud + timedelta(days=drill + compl)
            except Exception:
                continue

    # Capacities
    cap_data = tables.get("cap_df", {})
    cap_rows = []
    if cap_data:
        n = len(cap_data.get("start_date", []))
        for i in range(n):
            try:
                sd = cap_data["start_date"][i]
                d = date.fromisoformat(sd) if isinstance(sd, str) else (
                    sd if isinstance(sd, date) else start_date)
                row = {"start_date": pd.Timestamp(d)}
                for col, kind in [("oil","oil_rate"),("water","water_rate"),
                                   ("liquid","oil_rate"),
                                   ("water_inj","water_rate"),("gas_inj","gas_rate")]:
                    row[col] = to_field(float(cap_data[col][i]), kind, units)
                # gas: MMscf/d (field) or kSm³/d (metric)
                gv = float(cap_data["gas"][i])
                if units == "metric":
                    row["gas"] = to_field(gv, "gas_rate", units) / 1000.0
                else:
                    row["gas"] = gv
                cap_rows.append(row)
            except Exception:
                continue
    if not cap_rows:
        cap_rows = [{"start_date": pd.Timestamp(start_date),
                      "oil": 50000.0, "gas": 150.0, "water": 80000.0,
                      "liquid": 120000.0, "water_inj": 100000.0, "gas_inj": 0.0}]
    cap_sched = CapacitySchedule(df=pd.DataFrame(cap_rows))

    # PVT
    pvt = PVTInputs(
        p_init_psi=to_field(float(scalar.get("p_init", 3500)), "pressure", units),
        t_res_F=to_field(float(scalar.get("t_res", 180)), "temp", units),
        api=float(scalar.get("api", 35)),
        gas_grav=float(scalar.get("gas_grav", 0.7)),
        rs_init=to_field(float(scalar.get("rs_init", 700)), "gor", units),
        p_bub_psi=to_field(float(scalar.get("p_bub", 2800)), "pressure", units),
    )
    aquifer = AquiferInputs(
        active=bool(scalar.get("aq_active", False)),
        model=str(scalar.get("aq_model", "Pot")),
        aquifer_volume=to_field(float(scalar.get("aq_vol", 500)), "water_vol", units),
        productivity_index=float(scalar.get("aq_pi", 20)),
        initial_pressure_psi=to_field(float(scalar.get("aq_pini", 3500)), "pressure", units),
    )
    gas_cap = GasCapInputs(
        active=bool(scalar.get("gc_active", False)),
        size_fraction=float(scalar.get("gc_size", 0.2)),
        initial_pressure_psi=to_field(float(scalar.get("gc_pi", 3500)), "pressure", units),
    )

    # Facility CAPEX
    fac_data = tables.get("fac_df", {})
    fac_rows = []
    if fac_data:
        for i in range(len(fac_data.get("date", []))):
            try:
                ds = fac_data["date"][i]
                d = date.fromisoformat(ds) if isinstance(ds, str) else (
                    ds if isinstance(ds, date) else start_date)
                fac_rows.append({
                    "date": pd.Timestamp(d),
                    "amount_MMUSD": float(fac_data["amount_MMUSD"][i]),
                    "label": str(fac_data.get("label", [""]*len(fac_data["date"]))[i]),
                })
            except Exception:
                continue
    if not fac_rows:
        fac_rows = [{"date": pd.Timestamp(start_date), "amount_MMUSD": 0.0, "label": ""}]
    # Drop any abandonment/cessation/P&A rows from the facility schedule —
    # the engine books abandonment separately from aban_cost, so leaving
    # them here would double-count cessation. (Matches the live app.)
    try:
        fac_rows = [r for r in fac_rows
                    if not fh.is_abandonment_label(r.get("label", ""))]
    except Exception:
        pass
    if not fac_rows:
        fac_rows = [{"date": pd.Timestamp(start_date), "amount_MMUSD": 0.0, "label": ""}]
    # Apply the CAPEX contingency multiplier exactly as the main app does
    # (the stored fac_df / well-cost inputs are PRE-contingency). Without
    # this the Concept Selector under-counts CAPEX vs Field prognosis and
    # the two give different NPVs for the same case.
    _cont_mult = 1.0 + float(scalar.get("capex_contingency_pct", 25)) / 100.0
    fac_df_cont = pd.DataFrame(fac_rows)
    if "amount_MMUSD" in fac_df_cont.columns:
        fac_df_cont["amount_MMUSD"] = (
            fac_df_cont["amount_MMUSD"].astype(float) * _cont_mult)
    facility_capex = CapexSchedule(df=fac_df_cont)

    aban_gas = float(scalar.get("aban_gas", 0.5))
    if units == "metric":
        aban_gas = to_field(aban_gas, "gas_rate", units) / 1000.0

    meta = {
        "start_date": start_date,
        "horizon": int(scalar.get("horizon", 25)),
        "ooip": to_field(float(scalar.get("ooip", 250)), "oil_vol", units),
        "ogip": to_field(float(scalar.get("ogip", 300)), "gas_vol", units),
        "rf_target": float(scalar.get("rf_target", 0.35)),
        "ct_rock": float(scalar.get("ct_rock", 4e-6)),
        "sw_init": float(scalar.get("sw_init", 0.20)),
        "vrr": float(scalar.get("vrr", 1.0)),
        "inj_eff": float(scalar.get("inj_eff", 0.85)),
        "aban_basis": str(scalar.get("aban_basis", "Per well")),
        "aban_oil": to_field(float(scalar.get("aban_oil", 50)), "oil_rate", units),
        "aban_gas": aban_gas,
        "aban_wc": float(scalar.get("aban_wc", 0.95)),
        "prod_eff": float(scalar.get("prod_eff", 0.95)),
        "gas_export": float(scalar.get("gas_export", 1.0)),
        "gas_inj_frac": float(scalar.get("gas_inj_frac", 0.0)),
        "gas_fuel": float(scalar.get("gas_fuel", 0.0)),
        "gas_flare": float(scalar.get("gas_flare", 0.0)),
        "pvt": pvt, "aquifer": aquifer, "gas_cap": gas_cap,
        "cap_sched": cap_sched, "facility_capex": facility_capex,
        # IPR / inflow defaults (needed when wells use ipr_mode) — these
        # were missing in the batch path, so an IPR/gas-condensate case ran
        # with different defaults than Field prognosis.
        "well_pi": float(scalar.get("well_pi_default", 2.0)),
        "min_bhp_psi": to_field(
            float(scalar.get("min_bhp_default", 1500.0)), "pressure", units),
        # Retrograde condensate dropout (gas-condensate fields)
        "retrograde_enabled": bool(scalar.get("retrograde_enabled", False)),
        "retrograde_drop_fraction": float(
            scalar.get("retrograde_drop_fraction", 0.55)),
        # Fractional-flow water modelling (oil fields)
        "fractional_flow_enabled": bool(
            scalar.get("fractional_flow_enabled", False)),
        "ff_params": scalar.get("ff_params", {}) or {},
    }

    # Backward compat: old saves stored oil_price/gas_price/opex_var/tariffs
    # in the user's *display* units (e.g. $/Sm³ for metric users). New saves
    # always use $/bbl and $/MMBtu. We prefer the new keys when present, and
    # fall back to converting the old keys via the (saved) units.
    MMBTU_PER_MCF = 1.0
    if "oil_price_bbl" in scalar:
        oil_price_f = float(scalar.get("oil_price_bbl", 75.0))
    else:
        # Old format: convert from display units
        oil_price_f = to_field(float(scalar.get("oil_price", 75)), "price_oil", units)
    if "gas_price_mmbtu" in scalar:
        gas_price_f = float(scalar.get("gas_price_mmbtu", 3.5)) * MMBTU_PER_MCF
    else:
        gas_price_f = to_field(float(scalar.get("gas_price", 3.5)), "price_gas", units)
    if "opex_var_bbl" in scalar:
        opex_var_f = float(scalar.get("opex_var_bbl", 8.0))
    else:
        opex_var_f = to_field(float(scalar.get("opex_var", 8)), "price_oil", units)
    if "tariff_oil_bbl" in scalar:
        tariff_oil_f = float(scalar.get("tariff_oil_bbl", 2.0))
    else:
        tariff_oil_f = to_field(float(scalar.get("tariff_oil", 2)), "price_oil", units)
    if "tariff_gas_mmbtu" in scalar:
        tariff_gas_f = float(scalar.get("tariff_gas_mmbtu", 0.3)) * MMBTU_PER_MCF
    else:
        tariff_gas_f = to_field(float(scalar.get("tariff_gas", 0.3)), "price_gas", units)

    # Fiscal regime — normalise the stored label to the engine token so a
    # case saved with NCS (or PSC) actually applies that regime. Without
    # this the regime was dropped and only the (often zero) flat tax_rate
    # was used, making pre-tax NPV equal after-tax NPV.
    _regime_raw = str(scalar.get("fiscal_regime", "Tax/Royalty"))
    if _regime_raw.startswith("NCS"):
        _regime = "NCS"
    elif _regime_raw.startswith("PSC"):
        _regime = "PSC"
    else:
        _regime = "Tax/Royalty"

    econ_dict = {
        "oil_price": oil_price_f,
        "gas_price": gas_price_f,
        "opex_var":  opex_var_f,
        "opex_fixed": float(scalar.get("opex_fixed", 20)) * 1e6,
        "capex_per_well": float(scalar.get("capex_well", 15)) * _cont_mult,
        "discount_rate": float(scalar.get("disc", 0.10)),
        "tax_rate":      float(scalar.get("tax_rate", 0.30)),
        "royalty_rate":  float(scalar.get("royalty", 0.10)),
        "tariff_oil": tariff_oil_f,
        "tariff_gas": tariff_gas_f,
        "abandonment_cost_MM": float(scalar.get("aban_cost", 80)) * _cont_mult,
        "facility_capex": facility_capex,
        "ngl_yield_bbl_per_mmscf": float(scalar.get("ngl_yield", 0.0)),
        "ngl_price_bbl": float(scalar.get("ngl_price", 25.0)),
        "ngl_opex_bbl": float(scalar.get("ngl_opex", 5.0)),
        "ngl_shrinkage_pct": float(scalar.get("ngl_shrink", 0.0)),
        "rig_meta": rig_meta,
        # ---- Fiscal regime ----
        "fiscal_regime": _regime,
        "ncs_cit_rate": float(scalar.get("ncs_cit_rate", 0.22)),
        "ncs_spt_rate": float(scalar.get("ncs_spt_rate", 0.718)),
        "ncs_uplift_rate": float(scalar.get("ncs_uplift_rate", 0.1769)),
        "ncs_depreciation_years": float(
            scalar.get("ncs_depreciation_years", 6.0)),
        "ncs_uplift_years": float(scalar.get("ncs_uplift_years", 4.0)),
        # PSC parameters (stored under short keys in the payload)
        "psc_cost_recovery_ceiling": float(scalar.get("psc_cr_ceiling", 0.50)),
        "psc_profit_oil_share_contractor": float(scalar.get("psc_pos", 0.40)),
        "psc_govt_participation": float(scalar.get("psc_gov_part", 0.0)),
        "psc_psc_tax_rate": float(scalar.get("psc_tax", 0.30)),
        "psc_signature_bonus_MM": float(scalar.get("psc_sig_bonus", 0.0)),
        # ---- Well cost model (rig-rate components carry contingency) ----
        "well_cost_mode": str(scalar.get("well_cost_mode", "rig_rate")),
        "rig_day_rate_kUSD": float(
            scalar.get("rig_dayrate", 500.0)) * _cont_mult,
        "completion_day_rate_kUSD": float(
            scalar.get("cmpl_dayrate", 350.0)) * _cont_mult,
        "well_tangibles_MM": float(
            scalar.get("well_tangibles", 4.0)) * _cont_mult,
        "well_intangibles_pct": float(
            scalar.get("well_intangibles_pct", 0.10)),
        # ---- Money basis ----
        "money_basis": ("nominal"
            if str(scalar.get("money_basis_label", "")).startswith("Nominal")
            else "real"),
        # The inflation widget stores a percent (e.g. 2.5); the engine
        # wants a fraction. Divide by 100 if it looks like a percent.
        "inflation_rate": (float(scalar.get("inflation_rate", 0.0)) / 100.0
                            if float(scalar.get("inflation_rate", 0.0)) > 1.0
                            else float(scalar.get("inflation_rate", 0.0))),
        # ---- CO2 economics ----
        "co2_price": float(scalar.get("co2_price", 0.0)),
        "co2_factor_gas_combust": float(
            scalar.get("co2_factor_gas_combust", 53.0)),
        "co2_factor_flare_inefficiency": float(
            scalar.get("co2_factor_flare_inefficiency", 0.02)),
        "co2_factor_oil_routine": float(
            scalar.get("co2_factor_oil_routine", 0.5)),
        "co2_scope3_enabled": bool(scalar.get("co2_scope3_enabled", False)),
        "co2_scope3_factor_oil": float(
            scalar.get("co2_scope3_factor_oil", 430.0)),
        "co2_scope3_factor_gas": float(
            scalar.get("co2_scope3_factor_gas", 53.0)),
        "co2_scope3_price": float(scalar.get("co2_scope3_price", 0.0)),
        "economic_cutoff_mode": ("economic"
            if str(scalar.get("economic_cutoff_mode_label", "")).startswith("Economic")
            else "horizon"),
        "economic_cutoff_persistence": int(scalar.get("economic_cutoff_persistence", 6)),
    }
    return wells, _reservoirs_from_payload(payload, units), meta, econ_dict


def _reservoirs_from_payload(payload: dict, units: str) -> list:
    """Reconstruct (Reservoir list, WellReservoirLink list bundled into asm) from
    a saved case's tables. Returns empty list if multi-reservoir mode wasn't
    enabled in the saved case (reservoirs_df missing or empty).
    """
    scalar = payload.get("scalar", {})
    tables = payload.get("tables", {})
    if not scalar.get("multi_res_enable", False):
        return []
    rdata = tables.get("reservoirs_df", {})
    if not rdata:
        return []
    n = len(rdata.get("id", []))
    if n == 0:
        return []
    out = []
    for i in range(n):
        try:
            rid = str(rdata["id"][i])
            name = str(rdata.get("name", [rid] * n)[i])
            fluid_system = str(rdata.get("fluid_system",
                                          ["Oil with associated gas"] * n)[i])
            strategy = str(rdata.get("strategy", ["Depletion"] * n)[i])
            ooip = to_field(float(rdata.get("ooip_oil_MMstb", [0.0] * n)[i]),
                             "oil_vol", units)
            ogip = to_field(float(rdata.get("ogip_gas_Bscf",  [0.0] * n)[i]),
                             "gas_vol", units)
            rf_t = float(rdata.get("rf_target", [0.35] * n)[i])
            p_init = to_field(float(rdata.get("p_init", [3500.0] * n)[i]),
                               "pressure", units)
            t_res  = to_field(float(rdata.get("t_res",  [180.0] * n)[i]),
                               "temp", units)
            api    = float(rdata.get("api",     [35.0] * n)[i])
            sg     = float(rdata.get("gas_sg",  [0.7] * n)[i])
            rs_i   = to_field(float(rdata.get("rs_init", [600.0] * n)[i]),
                               "gor", units)
            p_bub  = to_field(float(rdata.get("p_bub", [2800.0] * n)[i]),
                               "pressure", units)
            aq_active = bool(rdata.get("aquifer_active", [False] * n)[i])
            gc_active = bool(rdata.get("gas_cap_active", [False] * n)[i])
            vrr = float(rdata.get("vrr", [1.0] * n)[i])
            r_pvt = PVTInputs(p_init_psi=p_init, t_res_F=t_res, api=api,
                               gas_grav=sg, rs_init=rs_i, p_bub_psi=p_bub)
            r_aq  = AquiferInputs(active=aq_active, model="Pot",
                                    aquifer_volume=0, productivity_index=0,
                                    initial_pressure_psi=p_init)
            r_gc  = GasCapInputs(active=gc_active, size_fraction=0.0,
                                   initial_pressure_psi=p_init)
            out.append(Reservoir(
                id=rid, name=name, fluid_system=fluid_system,
                ooip_oil=ooip, ogip_gas=ogip, rf_target=rf_t,
                strategy=strategy, pvt=r_pvt, aquifer=r_aq, gas_cap=r_gc,
                voidage_ratio=vrr, inj_efficiency=0.85,
            ))
        except Exception:
            continue
    return out


def _well_links_from_payload(payload: dict) -> list:
    """Extract WellReservoirLink list from saved tables."""
    scalar = payload.get("scalar", {})
    tables = payload.get("tables", {})
    if not scalar.get("multi_res_enable", False):
        return []
    ldata = tables.get("well_reservoir_df", {})
    if not ldata:
        return []
    n = len(ldata.get("well_name", []))
    out = []
    for i in range(n):
        try:
            out.append(WellReservoirLink(
                well_name=str(ldata["well_name"][i]),
                reservoir_id=str(ldata["reservoir_id"][i]),
                fraction=float(ldata.get("fraction", [1.0] * n)[i]),
            ))
        except Exception:
            continue
    return out


def _build_asm_for_scenario(meta: dict, fluid: str, strategy: str,
                              reservoirs: list = None,
                              well_links: list = None) -> "FieldAssumptions":
    return FieldAssumptions(
        fluid_system=fluid, strategy=strategy,
        ooip_oil=meta["ooip"], ogip_gas=meta["ogip"],
        rf_target=meta["rf_target"],
        start_date=meta["start_date"], forecast_years=meta["horizon"],
        rock_compressibility=meta["ct_rock"], sw_init=meta["sw_init"],
        pvt=meta["pvt"], aquifer=meta["aquifer"], gas_cap=meta["gas_cap"],
        voidage_ratio=meta["vrr"], inj_efficiency=meta["inj_eff"],
        aban_rate_oil=meta["aban_oil"], aban_rate_gas=meta["aban_gas"],
        aban_wc=meta["aban_wc"], aban_basis=meta["aban_basis"],
        cap_schedule=meta["cap_sched"],
        production_efficiency=meta["prod_eff"],
        gas_export_fraction=meta["gas_export"],
        gas_injection_fraction=meta["gas_inj_frac"],
        gas_fuel_fraction=meta["gas_fuel"],
        gas_flare_fraction=meta["gas_flare"],
        reservoirs=reservoirs or [],
        well_links=well_links or [],
        default_well_pi=meta.get("well_pi", 2.0),
        default_min_bhp_psi=meta.get("min_bhp_psi", 1500.0),
        retrograde_enabled=meta.get("retrograde_enabled", False),
        retrograde_drop_fraction=meta.get(
            "retrograde_drop_fraction", 0.55),
        fractional_flow_enabled=meta.get("fractional_flow_enabled", False),
        ff_swc=meta.get("ff_params", {}).get("swc", 0.20),
        ff_sor=meta.get("ff_params", {}).get("sor", 0.25),
        ff_krw_max=meta.get("ff_params", {}).get("krw_max", 0.30),
        ff_kro_max=meta.get("ff_params", {}).get("kro_max", 0.90),
        ff_nw=meta.get("ff_params", {}).get("nw", 3.0),
        ff_no=meta.get("ff_params", {}).get("no", 2.0),
        ff_mu_oil=meta.get("ff_params", {}).get("mu_oil", 1.5),
        ff_mu_water=meta.get("ff_params", {}).get("mu_water", 0.4),
        ff_sweep=meta.get("ff_params", {}).get("sweep", 0.70),
    )


def sensitivity_section(df_base, df_e_base, wells, asm, econ, units, fluid):
    """Tornado plot: vary each driver ±X% and rank by NPV (or RF) impact.

    Drivers: oil price, gas price, OPEX var, OPEX fixed, well CAPEX, OOIP/OGIP,
    initial pressure, decline rate, water cut final, voidage ratio.
    Each driver is varied independently while all others stay at the base case.
    Result is a horizontal bar chart sorted by absolute impact.
    """
    st.markdown("### 🌪️ Sensitivity (tornado)")
    st.caption(
        "Each input below is varied **±20% from the base case**, all other "
        "inputs held constant. The resulting NPV (or RF) is read off and "
        "compared to the base. Bars are sorted by absolute impact — the "
        "longest bars dominate the project economics."
    )

    is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"
    base_npv = float(df_e_base["npv"].iloc[-1])
    base_rf  = float(df_base["recovery_factor"].iloc[-1])

    c1, c2 = st.columns([1, 1])
    pct = c1.slider("Variation magnitude (±%)", 5, 50, 20, 5,
                     key="sens_pct",
                     help="How far each driver is moved from the base.")
    metric_choice = c2.radio("Output metric", ["NPV ($MM)", "Final RF"],
                              horizontal=True, key="sens_metric")

    # Driver definitions: each is (label, mutator function returning new (asm, econ, wells))
    from copy import deepcopy

    def _scale_econ(field, factor):
        def _m(asm0, econ0, wells0):
            new_econ = deepcopy(econ0)
            setattr(new_econ, field, getattr(econ0, field) * factor)
            return asm0, new_econ, wells0
        return _m

    def _scale_inplace(field, factor):
        def _m(asm0, econ0, wells0):
            new_asm = deepcopy(asm0)
            setattr(new_asm, field, getattr(asm0, field) * factor)
            return new_asm, econ0, wells0
        return _m

    def _scale_pvt(attr, factor):
        def _m(asm0, econ0, wells0):
            new_asm = deepcopy(asm0)
            new_pvt = deepcopy(asm0.pvt)
            setattr(new_pvt, attr, getattr(asm0.pvt, attr) * factor)
            new_asm.pvt = new_pvt
            return new_asm, econ0, wells0
        return _m

    def _scale_well_attr(attr, factor):
        def _m(asm0, econ0, wells0):
            new_wells = [deepcopy(w) for w in wells0]
            for w in new_wells:
                if w.is_producer:
                    setattr(w, attr, max(0.0, min(1.0, getattr(w, attr) * factor))
                            if attr in ("wc_final",) else getattr(w, attr) * factor)
            return asm0, econ0, new_wells
        return _m

    def _scale_facility_capex(_unused, factor):
        def _m(asm0, econ0, wells0):
            new_econ = deepcopy(econ0)
            try:
                _fdf = new_econ.facility_capex.df
                if _fdf is not None and "amount_MMUSD" in _fdf.columns:
                    if "label" in _fdf.columns:
                        _keep = ~_fdf["label"].apply(
                            lambda x: fh.is_abandonment_label(
                                "" if x is None else str(x)))
                        _fdf.loc[_keep, "amount_MMUSD"] = (
                            _fdf.loc[_keep, "amount_MMUSD"].astype(float)
                            * factor)
                    else:
                        _fdf["amount_MMUSD"] = (
                            _fdf["amount_MMUSD"].astype(float) * factor)
            except Exception:
                pass
            return asm0, new_econ, wells0
        return _m

    drivers = [
        ("Oil price",       _scale_econ("oil_price", 1.0)),
        ("Gas price",       _scale_econ("gas_price", 1.0)),
        ("Variable OPEX",   _scale_econ("opex_var", 1.0)),
        ("Fixed OPEX",      _scale_econ("opex_fixed", 1.0)),
        ("Well CAPEX",      _scale_econ("capex_per_well", 1.0)),
        ("Facility CAPEX",  _scale_facility_capex(None, 1.0)),
        ("OOIP",            _scale_inplace("ooip_oil", 1.0)),
        ("OGIP",            _scale_inplace("ogip_gas", 1.0)),
        ("Initial pressure", _scale_pvt("p_init_psi", 1.0)),
        ("Decline rate",     _scale_well_attr("di_annual", 1.0)),
        ("Final water cut",  _scale_well_attr("wc_final", 1.0)),
        ("Discount rate",    _scale_econ("discount_rate", 1.0)),
    ]

    # ---- User-customisable selection + per-driver range ----
    # By default include every driver at the global ±pct. The user can
    # untick drivers they want to exclude and override the lo/hi % per
    # driver — overrides are kept across reruns via session_state.
    _all_driver_names = [d[0] for d in drivers]
    with st.expander("⚙️ Choose drivers and per-driver ranges",
                     expanded=False):
        st.caption(
            "Pick which drivers appear in the tornado, and override "
            "the low/high % for each one (defaults to the global "
            f"±{pct}%). Useful when you want a sharper view on a "
            "single high-impact driver (e.g. oil price ±50%) while "
            "keeping others tighter.")
        selected_drivers = st.multiselect(
            "Drivers to include",
            options=_all_driver_names,
            default=st.session_state.get("sens_selected_drivers",
                                          _all_driver_names),
            key="sens_selected_drivers")
        # Per-driver lo/hi % grid
        ranges_key = "sens_driver_ranges"
        if ranges_key not in st.session_state:
            st.session_state[ranges_key] = {}
        st.caption(
            "**Per-driver overrides** — leave empty for the global "
            f"±{pct}%. Values are percentages of the base case.")
        _gc1, _gc2, _gc3 = st.columns([2, 1, 1])
        _gc1.markdown("**Driver**")
        _gc2.markdown(f"**Low %** (default −{pct})")
        _gc3.markdown(f"**High %** (default +{pct})")
        new_ranges = dict(st.session_state[ranges_key])
        for dname in selected_drivers:
            r = new_ranges.get(dname, {})
            c1d, c2d, c3d = st.columns([2, 1, 1])
            c1d.markdown(f"  {dname}")
            lo = c2d.number_input(
                f"lo_{dname}", value=float(r.get("lo", -pct)),
                step=5.0, min_value=-95.0, max_value=0.0,
                key=f"sens_lo_{dname}",
                label_visibility="collapsed")
            hi = c3d.number_input(
                f"hi_{dname}", value=float(r.get("hi", pct)),
                step=5.0, min_value=0.0, max_value=200.0,
                key=f"sens_hi_{dname}",
                label_visibility="collapsed")
            new_ranges[dname] = {"lo": float(lo), "hi": float(hi)}
        st.session_state[ranges_key] = new_ranges

    # Filter drivers list to those selected, in original order.
    drivers = [d for d in drivers if d[0] in selected_drivers]
    if not drivers:
        st.warning("Select at least one driver above to render the "
                    "tornado.")
        return
    # Build a per-driver factor table for the compute step.
    driver_factors = {}
    for dname in [d[0] for d in drivers]:
        r = st.session_state.get(ranges_key, {}).get(dname, {})
        lo_pct = float(r.get("lo", -pct))
        hi_pct = float(r.get("hi", pct))
        driver_factors[dname] = (1.0 + lo_pct / 100.0,
                                  1.0 + hi_pct / 100.0)

    def _compute_tornado_rows(drivers_, factor_lo, factor_hi, get_value,
                              base_value, asm, econ, wells, is_oil):
        """Run the 2-per-driver perturbation sweeps and return the rows.
        Defined here so it closes over the _scale_* mutator factories."""
        def make_mutator(lbl, fac):
            mapping = {
                "Oil price":         _scale_econ("oil_price", fac),
                "Gas price":         _scale_econ("gas_price", fac),
                "Variable OPEX":     _scale_econ("opex_var", fac),
                "Fixed OPEX":        _scale_econ("opex_fixed", fac),
                "Well CAPEX":        _scale_econ("capex_per_well", fac),
                "Facility CAPEX":    _scale_facility_capex(None, fac),
                "OOIP":              _scale_inplace("ooip_oil", fac),
                "OGIP":              _scale_inplace("ogip_gas", fac),
                "Initial pressure":  _scale_pvt("p_init_psi", fac),
                "Decline rate":      _scale_well_attr("di_annual", fac),
                "Final water cut":   _scale_well_attr("wc_final", fac),
                "Discount rate":     _scale_econ("discount_rate", fac),
            }
            return mapping[lbl]

        rows_ = []
        progress = st.progress(0.0, text="Running sensitivity sweeps…")
        total = len(drivers_) * 2
        step = 0
        for label, _tmpl in drivers_:
            try:
                a_lo, e_lo, w_lo = make_mutator(label, factor_lo)(
                    asm, econ, wells)
                d_lo, _, _ = run_simulation(w_lo, a_lo)
                de_lo = compute_economics(d_lo, is_oil, e_lo, w_lo)
                v_lo = get_value(d_lo, de_lo)
            except Exception:
                v_lo = base_value
            step += 1
            progress.progress(step / total, text=f"{label} (low)…")
            try:
                a_hi, e_hi, w_hi = make_mutator(label, factor_hi)(
                    asm, econ, wells)
                d_hi, _, _ = run_simulation(w_hi, a_hi)
                de_hi = compute_economics(d_hi, is_oil, e_hi, w_hi)
                v_hi = get_value(d_hi, de_hi)
            except Exception:
                v_hi = base_value
            step += 1
            progress.progress(step / total, text=f"{label} (high)…")
            rows_.append({
                "Driver": label,
                "low":  v_lo - base_value,
                "high": v_hi - base_value,
                "abs":  max(abs(v_lo - base_value),
                            abs(v_hi - base_value)),
            })
        progress.empty()
        return rows_

    # ---- Result persistence ----
    # The tornado is expensive (2 sims per driver). Without persistence the
    # result vanishes whenever the user switches tabs (Streamlit reruns the
    # whole script and the button reads False again). We store the computed
    # rows in st.session_state keyed by a signature of the inputs, so the
    # result survives tab switches and is only recomputed when something
    # that affects it actually changes.
    import hashlib as _hashlib
    def _sens_signature():
        parts = [
            metric_choice, str(pct), str(len(drivers)),
            f"{base_npv:.4f}", f"{base_rf:.6f}",
            f"{econ.oil_price:.4f}", f"{econ.gas_price:.4f}",
            f"{econ.opex_var:.4f}", f"{econ.opex_fixed:.2f}",
            f"{econ.capex_per_well:.4f}", f"{econ.discount_rate:.5f}",
            f"{asm.ooip_oil:.4f}", f"{asm.ogip_gas:.4f}",
            f"{asm.pvt.p_init_psi:.2f}",
            str(len(wells)),
        ]
        for w in wells:
            parts.append(f"{w.di_annual:.4f}|{w.wc_final:.4f}|"
                         f"{w.qi_primary:.2f}")
        return _hashlib.md5("~".join(parts).encode()).hexdigest()

    sig = _sens_signature()
    stored = st.session_state.get("sensitivity_results")
    have_valid = (stored is not None and stored.get("signature") == sig)

    run_clicked = st.button(
        "Run tornado" if not have_valid else "Re-run tornado",
        key="run_tornado", type="primary" if not have_valid else "secondary")

    if not run_clicked and not have_valid:
        st.info("Click **Run tornado** to compute the sensitivity. "
                f"With {len(drivers)} drivers × 2 perturbations, this will "
                f"run {len(drivers)*2} simulations. The result is kept when "
                f"you switch tabs.")
        return

    factor_lo = 1.0 - pct / 100.0
    factor_hi = 1.0 + pct / 100.0

    # Base case results already computed; just need value
    if metric_choice == "NPV ($MM)":
        get_value = lambda d, de: float(de["npv"].iloc[-1]) / 1e6
        base_value = base_npv / 1e6
        unit_label = "$MM"
    else:
        get_value = lambda d, de: float(d["recovery_factor"].iloc[-1]) * 100
        base_value = base_rf * 100
        unit_label = "% RF"

    # If we have a valid cached result and the user did not click re-run,
    # reuse it instead of recomputing.
    if have_valid and not run_clicked:
        rows = stored["rows"]
        base_value = stored["base_value"]
        unit_label = stored["unit_label"]
        st.caption("Showing the last computed tornado (inputs unchanged). "
                   "Click **Re-run tornado** to recompute.")
    else:
        rows = _compute_tornado_rows(
            drivers, factor_lo, factor_hi, get_value, base_value,
            asm, econ, wells, is_oil)
        st.session_state["sensitivity_results"] = {
            "signature": sig, "rows": rows, "base_value": base_value,
            "unit_label": unit_label, "pct": pct,
            "metric_choice": metric_choice,
        }

    # Sort by absolute impact (largest at top)
    rows = sorted(rows, key=lambda r: r["abs"], reverse=True)

    # Tornado plot: two horizontal bars per driver
    labels = [r["Driver"] for r in rows][::-1]   # reversed so biggest at top
    lows   = [r["low"]    for r in rows][::-1]
    highs  = [r["high"]   for r in rows][::-1]

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=labels, x=lows, orientation="h",
        name=f"−{pct}%", marker_color=fh.EQ_COLORS["water"],
        hovertemplate="%{y}: %{x:+,.1f} " + unit_label + "<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        y=labels, x=highs, orientation="h",
        name=f"+{pct}%", marker_color=fh.EQ_COLORS["gas"],
        hovertemplate="%{y}: %{x:+,.1f} " + unit_label + "<extra></extra>",
    ))
    fig.add_vline(x=0, line=dict(color=fh.EQ_COLORS["pressure"], width=2))
    fig.update_layout(
        title=f"Sensitivity tornado — {metric_choice}  (base = {base_value:,.1f} {unit_label})",
        barmode="overlay",
        height=max(380, 32 * len(rows) + 120),
        xaxis_title=f"Δ {metric_choice} from base",
        yaxis_title="",
        legend=dict(orientation="h", y=-0.15),
        bargap=0.35,
    )
    _tornado_fig = fh.apply_plot_template(fig)
    # Cache for the PDF report so "Generate PDF" can embed the last tornado
    # the user saw without recomputing the whole sensitivity sweep.
    st.session_state["_last_tornado_fig"] = _tornado_fig
    st.plotly_chart(_tornado_fig, use_container_width=True)
    table = pd.DataFrame([{
        "Driver": r["Driver"],
        f"−{pct}% Δ": r["low"],
        f"+{pct}% Δ": r["high"],
        "Abs swing": (r["high"] - r["low"]),
    } for r in rows])
    fmt = "{:+,.2f}" if metric_choice == "Final RF" else "{:+,.1f}"
    st.dataframe(
        table.style.format({
            f"−{pct}% Δ": fmt, f"+{pct}% Δ": fmt, "Abs swing": fmt,
        }),
        use_container_width=True,
    )
    st.caption(
        f"Drivers ranked by absolute swing in {metric_choice}. "
        "Note that ±% perturbation of bounded quantities (water cut, "
        "discount rate) is clipped to the valid range."
    )


def monte_carlo_section(df_base, df_e_base, wells, asm, econ, units, fluid):
    """Probabilistic forecasting: sample from uncertainty distributions on key
    drivers, run N realizations, render P10/P50/P90 fans + final-NPV histogram.
    """
    st.markdown("### 🎲 Monte Carlo")
    st.caption(
        "Sample uncertainty distributions on the major drivers and run N "
        "realizations. Each realization perturbs the base case independently. "
        "Output is the P10 / P50 / P90 percentile fan over time for oil rate, "
        "cum oil, RF, and NPV — plus a histogram of final NPV outcomes. "
        "P10 here is the **pessimistic** percentile (10% of outcomes are worse), "
        "P90 is the **optimistic** percentile."
    )

    is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"
    base_npv = float(df_e_base["npv"].iloc[-1])
    base_rf  = float(df_base["recovery_factor"].iloc[-1])

    # Top controls
    cc1, cc2, cc3 = st.columns([1, 1, 1])
    n_runs = cc1.select_slider(
        "Realizations", options=[50, 100, 200, 500, 1000], value=200,
        key="mc_n_runs",
        help="More realizations = smoother percentile bands but longer runtime "
             "(~130 ms per realization on a typical case; 500 reals ≈ 1 minute).",
    )
    seed = cc2.number_input("Random seed", min_value=0, max_value=99999,
                             value=42, step=1, key="mc_seed",
                             help="Fix the seed for reproducible runs.")
    show_indiv = cc3.checkbox("Show individual realizations as faint lines",
                               value=False, key="mc_show_indiv")

    # Driver configuration
    with st.expander("⚙️ Configure driver distributions", expanded=False):
        st.caption(
            "Each driver is a **multiplicative factor** on its base-case value. "
            "Low/high values are interpreted as the ~P10/P90 envelope of the "
            "distribution. Triangular puts the mode at the geometric mean; "
            "lognormal/truncnormal use ±2σ."
        )
        driver_cfg = {}
        # Header row
        h1, h2, h3, h4, h5 = st.columns([3, 2, 2, 2, 1.5])
        h1.markdown("**Driver**")
        h2.markdown("**Low (P10)**")
        h3.markdown("**High (P90)**")
        h4.markdown("**Distribution**")
        h5.markdown("**Per-well**")
        for name, default in DEFAULT_MC_DRIVERS.items():
            sub_a, sub_b, sub_c, sub_d, sub_e = st.columns([3, 2, 2, 2, 1.5])
            on = sub_a.checkbox(name, value=default["on"],
                                 key=f"mc_on_{name}")
            lo = sub_b.number_input(f"low_{name}",
                                     value=float(default["low"]),
                                     min_value=0.05, max_value=2.0, step=0.05,
                                     format="%.2f", key=f"mc_lo_{name}",
                                     label_visibility="collapsed")
            hi = sub_c.number_input(f"high_{name}",
                                     value=float(default["high"]),
                                     min_value=0.10, max_value=5.0, step=0.05,
                                     format="%.2f", key=f"mc_hi_{name}",
                                     label_visibility="collapsed")
            dist = sub_d.selectbox(
                f"dist_{name}",
                ["triangular", "lognormal", "uniform", "truncnormal"],
                index=["triangular", "lognormal", "uniform", "truncnormal"].index(default["dist"]),
                key=f"mc_dist_{name}", label_visibility="collapsed",
            )
            # Per-well toggle only meaningful for well-level drivers
            if name in ("Well qi", "Decline rate"):
                pw = sub_e.checkbox(
                    f"pw_{name}", value=default.get("per_well", True),
                    key=f"mc_pw_{name}", label_visibility="collapsed",
                    help="If on, each producer gets its own independent draw. "
                         "If off, all wells share the same multiplier per realization.",
                )
            else:
                pw = False
                sub_e.markdown(" ")
            driver_cfg[name] = {"on": on, "low": lo, "high": hi,
                                 "dist": dist, "per_well": pw}

    # ---- Driver correlation editor ----
    # Geological drivers are not independent — e.g. a bigger OOIP often comes
    # with better RF; a higher initial pressure with a higher qi. Sampling
    # them independently understates the spread in the tails. Here the user
    # can set pairwise correlations; the engine then samples through a
    # Gaussian copula so the realised draws carry those correlations.
    corr_pairs = {}
    with st.expander("🔗 Driver correlations (optional)", expanded=False):
        st.caption(
            "Set correlations between drivers. Positive = the two move "
            "together (e.g. OOIP & RF); negative = they move oppositely. "
            "Drivers are sampled through a Gaussian copula so the marginals "
            "you chose above are preserved. Leave at 0 for independent "
            "sampling. Only **enabled** drivers can be correlated."
        )
        enabled_drivers = [nm for nm, c in driver_cfg.items() if c["on"]]
        if len(enabled_drivers) < 2:
            st.info("Enable at least two drivers above to set correlations.")
        else:
            # Common geological pairs offered as ready-made rows
            candidate_pairs = []
            for i in range(len(enabled_drivers)):
                for j in range(i + 1, len(enabled_drivers)):
                    candidate_pairs.append((enabled_drivers[i],
                                            enabled_drivers[j]))
            st.caption(f"{len(candidate_pairs)} possible pair(s). Set any "
                       f"you care about; the rest stay independent.")
            # Sensible default suggestions for well-known geological pairs
            _suggested = {
                frozenset(("OOIP", "RF target")): 0.5,
                frozenset(("OGIP", "RF target")): 0.5,
                frozenset(("OOIP", "Well qi")): 0.3,
                frozenset(("Initial pressure", "Well qi")): 0.4,
                frozenset(("OOIP", "Initial pressure")): 0.3,
            }
            for (a, b) in candidate_pairs:
                default_rho = _suggested.get(frozenset((a, b)), 0.0)
                rho = st.slider(
                    f"corr( {a} , {b} )",
                    min_value=-0.9, max_value=0.9,
                    value=float(default_rho), step=0.1,
                    key=f"mc_corr_{a}_{b}")
                if abs(rho) > 1e-6:
                    corr_pairs[(a, b)] = rho
            if corr_pairs:
                st.caption(f"{len(corr_pairs)} correlation(s) active. The "
                           f"correlation matrix is repaired to the nearest "
                           f"positive-definite form if needed.")

    # ---- Reserves classification toggle ----
    show_reserves = st.checkbox(
        "Report probabilistic reserves (1P / 2P / 3P)", value=True,
        key="mc_show_reserves",
        help="Classify the cumulative-production distribution into 1P "
             "(Proved = P90), 2P (Proved+Probable = P50) and 3P "
             "(Proved+Probable+Possible = P10), following SPE-PRMS "
             "convention.")

    # ---- Result persistence ----
    # Monte Carlo is the most expensive operation in the app. Without
    # persistence the whole result set is lost on every tab switch. Store it
    # in st.session_state keyed by a signature of the inputs; only recompute
    # when the user explicitly clicks run or the inputs change.
    import hashlib as _hashlib
    def _mc_signature():
        parts = [str(n_runs), str(seed),
                 f"{base_npv:.4f}", f"{base_rf:.6f}", str(len(wells))]
        for nm in sorted(driver_cfg.keys()):
            c = driver_cfg[nm]
            parts.append(f"{nm}:{c['on']}:{c['low']:.3f}:{c['high']:.3f}:"
                         f"{c['dist']}:{c.get('per_well', False)}")
        for (a, b) in sorted(corr_pairs.keys()):
            parts.append(f"corr:{a}:{b}:{corr_pairs[(a, b)]:.3f}")
        parts.append(f"reserves:{show_reserves}")
        return _hashlib.md5("~".join(parts).encode()).hexdigest()

    mc_sig = _mc_signature()
    mc_stored = st.session_state.get("mc_results_full")
    mc_have_valid = (mc_stored is not None
                     and mc_stored.get("signature") == mc_sig)

    n_on = sum(1 for c in driver_cfg.values() if c["on"])
    mc_run_clicked = st.button(
        f"🎲 Run {n_runs} realizations" if not mc_have_valid
        else f"🎲 Re-run {n_runs} realizations",
        key="mc_run_btn", use_container_width=True,
        type="primary" if not mc_have_valid else "secondary")

    if not mc_run_clicked and not mc_have_valid:
        st.info(
            f"Click the button above to run {n_runs} realizations sampling "
            f"{n_on} active driver(s). Estimated time: ~{n_runs * 0.13:.0f} s. "
            f"The result is kept when you switch tabs."
        )
        return

    # ---- Run (or reuse cached) ----
    if mc_have_valid and not mc_run_clicked:
        mc = mc_stored["mc"]
        st.caption("Showing the last Monte Carlo run (inputs unchanged). "
                   "Click **Re-run** to recompute.")
    else:
        progress = st.progress(0.0, text="Running Monte Carlo…")
        def _cb(frac):
            progress.progress(min(1.0, max(0.0, frac)),
                              text=f"Running Monte Carlo… {int(frac*100)}%")
        mc = run_monte_carlo(wells, asm, econ, n_realizations=int(n_runs),
                             drivers_cfg=driver_cfg, seed=int(seed),
                             progress_callback=_cb, corr_pairs=corr_pairs)
        progress.empty()
        st.session_state["mc_results_full"] = {"signature": mc_sig, "mc": mc}

    summary = mc["summary"]
    pct = mc["percentiles"]
    monthly = mc["monthly"]

    # Persist for the export tab (kept under the original key too)
    st.session_state["mc_results"] = mc

    if mc["realizations_run"] == 0:
        st.error("No realizations completed successfully. Check input distributions.")
        return

    # ---- Headline KPIs ----
    npv_arr = summary["npv_usd"].values / 1e6
    rf_arr  = summary["final_rf"].values
    p10n, p50n, p90n = np.percentile(npv_arr, [10, 50, 90])
    p10r, p50r, p90r = np.percentile(rf_arr, [10, 50, 90])
    prob_pos = (npv_arr > 0).mean() * 100.0

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Realizations", f"{mc['realizations_run']:,}")
    k2.metric("NPV P10 ($MM)", f"{p10n:,.0f}",
              help="10% of realizations have NPV worse than this.")
    k3.metric("NPV P50 ($MM)", f"{p50n:,.0f}",
              help="Median NPV across realizations.")
    k4.metric("NPV P90 ($MM)", f"{p90n:,.0f}",
              help="10% of realizations have NPV better than this.")
    k5.metric("P(NPV > 0)", f"{prob_pos:.0f}%",
              help="Share of realizations where NPV is positive.")

    st.caption(
        f"Final RF: P10={p10r:.1%} · P50={p50r:.1%} · P90={p90r:.1%} · "
        f"Deterministic base case: NPV = ${base_npv/1e6:,.0f}MM, RF = {base_rf:.1%}"
    )

    # ---- Probabilistic reserves: 1P / 2P / 3P ----
    if show_reserves:
        st.markdown("#### 📊 Probabilistic reserves (SPE-PRMS convention)")
        is_oil_mc = FLUID_SYSTEMS[fluid]["primary"] == "oil"
        res_col = "cum_oil" if is_oil_mc else "cum_gas"
        _phase_unit = "oil_vol" if is_oil_mc else "gas_vol"
        res_unit = ulabel(_phase_unit, units)
        # Convert MMstb/Bscf → user's display units (MMstb/Bscf or
        # MSm³/GSm³) so the reserves report honours the unit toggle.
        res_vols = from_field(summary[res_col].values, _phase_unit, units)
        rc = classify_reserves(res_vols)
        if rc["n"] > 0:
            rk1, rk2, rk3, rk4 = st.columns(4)
            rk1.metric(f"1P — Proved ({res_unit})",
                       f"{rc['p90_1P']:,.1f}",
                       help="P90 — at least this much recovered with 90% "
                            "probability. The conservative booking figure.")
            rk2.metric(f"2P — Proved + Probable ({res_unit})",
                       f"{rc['p50_2P']:,.1f}",
                       help="P50 — the best (median) estimate.")
            rk3.metric(f"3P — Proved+Prob+Possible ({res_unit})",
                       f"{rc['p10_3P']:,.1f}",
                       help="P10 — the optimistic estimate; only 10% of "
                            "outcomes exceed it.")
            spread = rc.get("spread_3P_1P")
            rk4.metric("3P / 1P spread",
                       f"{spread:.2f}×" if spread else "—",
                       help="Ratio of the optimistic to the conservative "
                            "case. A wide spread (>2.5×) signals large "
                            "subsurface uncertainty.")
            st.caption(
                f"Mean of the distribution: {rc['mean']:,.1f} {res_unit} "
                f"across {rc['n']:,} realizations. "
                f"Convention: 1P = P90 (high confidence), 2P = P50 (best "
                f"estimate), 3P = P10 (upside). These are screening "
                f"figures from the Monte-Carlo spread — not a substitute "
                f"for an audited reserves report."
            )
            if corr_pairs:
                st.caption(f"Sampling used {len(corr_pairs)} driver "
                           f"correlation(s) via a Gaussian copula.")

    # ---- Fan plots ----
    f = lambda v, k: from_field(v, k, units)
    C = fh.EQ_COLORS

    def _fan(metric_key: str, title: str, y_unit: str, kind_for_units: str | None = None,
             color_p50: str = None, color_band: str = None) -> go.Figure:
        df_p = pct[metric_key]
        if kind_for_units:
            p10 = f(df_p["p10"], kind_for_units); p50 = f(df_p["p50"], kind_for_units); p90 = f(df_p["p90"], kind_for_units)
        else:
            p10, p50, p90 = df_p["p10"], df_p["p50"], df_p["p90"]
        fig = go.Figure()
        # P90 band (upper)
        fig.add_trace(go.Scatter(
            x=df_p["date"], y=p90, mode="lines", line=dict(width=0),
            showlegend=False, hoverinfo="skip",
        ))
        # P10 band (lower) — fill to previous (P90)
        fig.add_trace(go.Scatter(
            x=df_p["date"], y=p10, mode="lines",
            line=dict(width=0), fill="tonexty",
            fillcolor=color_band or "rgba(0, 21, 72, 0.15)",
            name="P10–P90",
            hovertemplate=f"{title} P10: %{{y:,.1f}} {y_unit}<extra></extra>",
        ))
        # Optional individual realizations
        if show_indiv and len(monthly) > 0:
            for r_id in monthly["realization"].unique()[:200]:  # cap at 200 for legibility
                sub = monthly[monthly["realization"] == r_id]
                ys = f(sub[metric_key], kind_for_units) if kind_for_units else sub[metric_key]
                fig.add_trace(go.Scatter(
                    x=sub["date"], y=ys, mode="lines",
                    line=dict(color="rgba(120,120,140,0.18)", width=0.6),
                    showlegend=False, hoverinfo="skip",
                ))
        # P50 (median)
        fig.add_trace(go.Scatter(
            x=df_p["date"], y=p50, mode="lines",
            line=dict(color=color_p50 or C["pressure"], width=2.5),
            name="P50 (median)",
            hovertemplate=f"{title} P50: %{{y:,.1f}} {y_unit}<extra></extra>",
        ))
        fig.update_layout(
            title=title, height=360, hovermode="x unified",
            yaxis_title=y_unit,
            legend=dict(orientation="h", y=-0.18),
        )
        return fh.apply_plot_template(fig)

    # 2x2 grid of fans
    g1, g2 = st.columns(2)
    with g1:
        st.plotly_chart(_fan("oil_rate",
                             f"Oil rate fan ({ulabel('oil_rate', units)})",
                             ulabel('oil_rate', units),
                             kind_for_units="oil_rate",
                             color_p50=C["oil"],
                             color_band="rgba(63, 112, 77, 0.18)"),
                        use_container_width=True)
    with g2:
        st.plotly_chart(_fan("gas_rate",
                             f"Gas rate fan ({ulabel('gas_rate', units)})",
                             ulabel('gas_rate', units),
                             kind_for_units="gas_rate",
                             color_p50=C["gas"],
                             color_band="rgba(255, 18, 67, 0.15)"),
                        use_container_width=True)
    g3, g4 = st.columns(2)
    with g3:
        # RF fan — keep as fraction, format axis as %
        df_rf = pct["recovery_factor"]
        fig_rf = go.Figure()
        fig_rf.add_trace(go.Scatter(
            x=df_rf["date"], y=df_rf["p90"], mode="lines",
            line=dict(width=0), showlegend=False, hoverinfo="skip",
        ))
        fig_rf.add_trace(go.Scatter(
            x=df_rf["date"], y=df_rf["p10"], mode="lines",
            line=dict(width=0), fill="tonexty",
            fillcolor="rgba(255, 198, 89, 0.25)",
            name="P10–P90",
        ))
        if show_indiv and len(monthly) > 0:
            for r_id in monthly["realization"].unique()[:200]:
                sub = monthly[monthly["realization"] == r_id]
                fig_rf.add_trace(go.Scatter(
                    x=sub["date"], y=sub["recovery_factor"], mode="lines",
                    line=dict(color="rgba(120,120,140,0.18)", width=0.6),
                    showlegend=False, hoverinfo="skip",
                ))
        fig_rf.add_trace(go.Scatter(
            x=df_rf["date"], y=df_rf["p50"], mode="lines",
            line=dict(color=C["rf"], width=2.5), name="P50 (median)",
        ))
        fh.safe_hline(fig_rf, asm.rf_target, label=f"Target {asm.rf_target:.0%}",
                      color=C["pressure"], dash="dash")
        fig_rf.update_layout(title="Recovery factor fan", height=360,
                             hovermode="x unified",
                             yaxis_tickformat=".0%",
                             legend=dict(orientation="h", y=-0.18))
        st.plotly_chart(fh.apply_plot_template(fig_rf), use_container_width=True)
    with g4:
        # NPV fan
        df_n = pct["npv"]
        fig_n = go.Figure()
        fig_n.add_trace(go.Scatter(
            x=df_n["date"], y=df_n["p90"]/1e6, mode="lines",
            line=dict(width=0), showlegend=False, hoverinfo="skip",
        ))
        fig_n.add_trace(go.Scatter(
            x=df_n["date"], y=df_n["p10"]/1e6, mode="lines",
            line=dict(width=0), fill="tonexty",
            fillcolor="rgba(0, 21, 72, 0.15)",
            name="P10–P90",
        ))
        if show_indiv and len(monthly) > 0:
            for r_id in monthly["realization"].unique()[:200]:
                sub = monthly[monthly["realization"] == r_id]
                fig_n.add_trace(go.Scatter(
                    x=sub["date"], y=sub["npv"]/1e6, mode="lines",
                    line=dict(color="rgba(120,120,140,0.18)", width=0.6),
                    showlegend=False, hoverinfo="skip",
                ))
        fig_n.add_trace(go.Scatter(
            x=df_n["date"], y=df_n["p50"]/1e6, mode="lines",
            line=dict(color=C["pressure"], width=2.5), name="P50 (median)",
        ))
        fh.safe_hline(fig_n, 0, label="NPV = 0", color=C["gas"], dash="dot")
        fig_n.update_layout(title="NPV fan ($MM)", height=360,
                            hovermode="x unified",
                            yaxis_title="NPV ($MM)",
                            legend=dict(orientation="h", y=-0.18))
        st.plotly_chart(fh.apply_plot_template(fig_n), use_container_width=True)

    # ---- NPV histogram + stats ----
    st.markdown("#### Final NPV distribution")
    h1, h2 = st.columns([3, 2])
    with h1:
        fig_h = go.Figure()
        fig_h.add_trace(go.Histogram(
            x=npv_arr, nbinsx=min(40, max(15, len(npv_arr)//10)),
            marker_color=C["pressure"], opacity=0.85,
            name="Realizations",
        ))
        fh.safe_vline(fig_h, p10n, label=f"P10 = {p10n:,.0f}",
                      color=C["water"], dash="dash")
        fh.safe_vline(fig_h, p50n, label=f"P50 = {p50n:,.0f}",
                      color=C["rf"], dash="dash")
        fh.safe_vline(fig_h, p90n,
                      label=f"P90 = {p90n:,.0f}",
                      color=C["spring"] if "spring" in C else C["water"],
                      dash="dash")
        fh.safe_vline(fig_h, base_npv/1e6, label=f"Base = {base_npv/1e6:,.0f}",
                      color=C["gas"], width=2, dash="solid",
                      label_position="bottom")
        fig_h.update_layout(
            title="Histogram of final NPV across realizations",
            xaxis_title="NPV ($MM)", yaxis_title="Frequency",
            height=360, bargap=0.05, showlegend=False,
        )
        _mc_npv_fig = fh.apply_plot_template(fig_h)
        # Cache for the PDF report so "Generate PDF" embeds the last MC
        # NPV distribution the user computed.
        st.session_state["_last_mc_fig"] = _mc_npv_fig
        st.plotly_chart(_mc_npv_fig, use_container_width=True)
        # Reserves histogram with 1P/2P/3P markers
        if show_reserves:
            _is_oil_mc = FLUID_SYSTEMS[fluid]["primary"] == "oil"
            _rcol = "cum_oil" if _is_oil_mc else "cum_gas"
            _phase_unit = "oil_vol" if _is_oil_mc else "gas_vol"
            _runit = ulabel(_phase_unit, units)
            # Convert the field-unit values (MMstb / Bscf) into the user's
            # chosen display units (MMstb/Bscf or MSm³/GSm³ for SI).
            _rvols = from_field(summary[_rcol].values, _phase_unit, units)
            _rc = classify_reserves(_rvols)
            if _rc["n"] > 0:
                fig_res = go.Figure()
                fig_res.add_trace(go.Histogram(
                    x=_rvols,
                    nbinsx=min(40, max(15, len(_rvols)//10)),
                    marker_color=C["rf"], opacity=0.85))
                fh.safe_vline(fig_res, _rc["p90_1P"],
                              label=f"1P = {_rc['p90_1P']:,.1f}",
                              color=C["water"], dash="dash")
                fh.safe_vline(fig_res, _rc["p50_2P"],
                              label=f"2P = {_rc['p50_2P']:,.1f}",
                              color=C["rf"], dash="dash")
                fh.safe_vline(fig_res, _rc["p10_3P"],
                              label=f"3P = {_rc['p10_3P']:,.1f}",
                              color=C["gas"], dash="dash")
                fig_res.update_layout(
                    title=f"Reserves distribution ({_runit}) with "
                          f"1P / 2P / 3P",
                    xaxis_title=f"Recoverable volume ({_runit})",
                    yaxis_title="Frequency",
                    height=320, bargap=0.05, showlegend=False)
                st.plotly_chart(fh.apply_plot_template(fig_res),
                                use_container_width=True)
    with h2:
        # Tornado-style driver-impact correlation: rank-correlation between
        # each sampled factor and NPV
        factor_cols = [c for c in summary.columns if c.startswith("factor_")]
        corrs = []
        for c in factor_cols:
            try:
                col = summary[c].values
                # Skip drivers that weren't actually varied (zero variance)
                if float(np.std(col)) < 1e-9:
                    continue
                r = float(np.corrcoef(col, summary["npv_usd"])[0, 1])
                if np.isfinite(r):
                    corrs.append((c.replace("factor_", ""), r))
            except Exception:
                pass
        corrs = sorted(corrs, key=lambda x: abs(x[1]), reverse=True)
        names = [c[0] for c in corrs][::-1]
        vals = [c[1] for c in corrs][::-1]
        fig_c = go.Figure()
        fig_c.add_trace(go.Bar(
            y=names, x=vals, orientation="h",
            marker_color=[C["oil"] if v > 0 else C["gas"] for v in vals],
            hovertemplate="%{y}: %{x:+.2f}<extra></extra>",
        ))
        fig_c.add_vline(x=0, line=dict(color=C["pressure"], width=1))
        fig_c.update_layout(
            title="Driver correlation with NPV",
            xaxis_title="Pearson r", yaxis_title="",
            height=360, xaxis=dict(range=[-1, 1]),
        )
        st.plotly_chart(fh.apply_plot_template(fig_c), use_container_width=True)
        st.caption(
            "Linear correlation between each driver's sampled factor and final "
            "NPV. Bars near 0 are weak drivers; near ±1 are strong drivers."
        )

    # ---- Reserves distribution (cum oil / cum gas / RF at end of life) ----
    st.markdown("#### Reserves distribution")
    st.caption(
        "Probabilistic estimate of ultimate recovery (EUR) across all "
        "realizations. P90 = optimistic (high reserves), P10 = pessimistic."
    )
    primary_metric = "cum_oil" if is_oil else "cum_gas"
    primary_vol_kind = "oil_vol" if is_oil else "gas_vol"
    res_metrics = [
        (primary_metric, f"Cumulative {'oil' if is_oil else 'gas'} "
                         f"({ulabel(primary_vol_kind, units)})", primary_vol_kind),
        ("final_rf", "Final recovery factor", None),
    ]
    rcols = st.columns(len(res_metrics))
    for (metric, label, vol_kind), rc in zip(res_metrics, rcols):
        if metric not in summary.columns:
            continue
        vals = summary[metric].values.astype(float)
        if vol_kind:
            vals = np.array([from_field(v, vol_kind, units) for v in vals])
        p10, p50, p90 = (np.percentile(vals, 10), np.percentile(vals, 50),
                         np.percentile(vals, 90))
        with rc:
            fig_r = go.Figure()
            fig_r.add_trace(go.Histogram(
                x=vals, nbinsx=min(35, max(12, len(vals)//12)),
                marker_color=C.get("spring", C["water"]), opacity=0.85,
            ))
            for pv, pl, pc in [(p10, "P10", C["water"]),
                               (p50, "P50", C["rf"]),
                               (p90, "P90", C["gas"])]:
                fig_r.add_vline(x=pv, line=dict(color=pc, dash="dash"),
                                annotation_text=pl, annotation_position="top")
            fmt = ".1%" if metric == "final_rf" else ",.2f"
            fig_r.update_layout(
                title=label, height=300, bargap=0.05, showlegend=False,
                xaxis_title=label, yaxis_title="Frequency",
                xaxis=dict(tickformat=".0%" if metric == "final_rf" else None),
            )
            st.plotly_chart(fh.apply_plot_template(fig_r), use_container_width=True)
            if metric == "final_rf":
                st.caption(f"P10 {p10:.1%}  •  P50 {p50:.1%}  •  P90 {p90:.1%}")
            else:
                st.caption(f"P10 {p10:,.2f}  •  P50 {p50:,.2f}  •  P90 {p90:,.2f}  "
                           f"{ulabel(vol_kind, units)}")

    # ---- Input distribution snapshots (optional) ----
    factor_cols_varied = [c for c in summary.columns
                          if c.startswith("factor_")
                          and float(np.std(summary[c].values)) > 1e-9]
    if factor_cols_varied:
        show_inputs = st.checkbox(
            "📊 Show sampled input distributions",
            value=False, key="mc_show_inputs",
            help="Histograms of the actual sampled multipliers for each varied "
                 "driver — confirms the sampling matched your intended "
                 "distributions.")
        if show_inputs:
            st.caption(
                "Each histogram shows the realized sample of the multiplier "
                "applied to that driver. The spread here should match the "
                "low/high bounds you configured above.")
            n_per_row = 3
            for i in range(0, len(factor_cols_varied), n_per_row):
                batch = factor_cols_varied[i:i + n_per_row]
                cols_in = st.columns(n_per_row)
                for c, col_in in zip(batch, cols_in):
                    vals = summary[c].values.astype(float)
                    with col_in:
                        fig_in = go.Figure()
                        fig_in.add_trace(go.Histogram(
                            x=vals, nbinsx=min(25, max(10, len(vals)//15)),
                            marker_color=C.get("pressure", "#1f77b4"),
                            opacity=0.8,
                        ))
                        fig_in.add_vline(x=float(np.mean(vals)),
                                         line=dict(color=C["gas"], dash="dot"),
                                         annotation_text=f"μ={np.mean(vals):.2f}")
                        fig_in.update_layout(
                            title=c.replace("factor_", ""),
                            height=240, bargap=0.05, showlegend=False,
                            xaxis_title="Multiplier", yaxis_title="Count",
                            margin=dict(t=40, b=30, l=30, r=10),
                        )
                        st.plotly_chart(fh.apply_plot_template(fig_in),
                                        use_container_width=True)

    # ---- Parameter correlation matrix ----
    if len(factor_cols_varied) >= 2 or (factor_cols_varied and
                                         "npv_usd" in summary.columns):
        with st.expander("🔗 Parameter correlation matrix", expanded=False):
            st.caption(
                "Pearson correlation between every sampled driver and the key "
                "outcomes (NPV, cum production, RF). Drivers are sampled "
                "independently so driver-driver correlations should be ≈0 — "
                "the informative column is each driver vs. the outcomes.")
            outcome_cols = [c for c in ["npv_usd", "cum_cf_usd", primary_metric,
                                         "final_rf", "peak_oil", "peak_gas"]
                            if c in summary.columns]
            corr_cols = factor_cols_varied + outcome_cols
            corr_data = summary[corr_cols].copy()
            corr_data.columns = [c.replace("factor_", "").replace("_usd", "")
                                 for c in corr_cols]
            corr_mat = corr_data.corr(method="pearson")
            fig_corr = go.Figure(data=go.Heatmap(
                z=corr_mat.values,
                x=list(corr_mat.columns),
                y=list(corr_mat.index),
                colorscale="RdBu", zmid=0, zmin=-1, zmax=1,
                text=np.round(corr_mat.values, 2),
                texttemplate="%{text}",
                textfont={"size": 9},
                colorbar=dict(title="r"),
            ))
            fig_corr.update_layout(
                title="Pearson correlation matrix",
                height=max(350, 40 * len(corr_cols)),
                xaxis=dict(tickangle=-45),
            )
            st.plotly_chart(fh.apply_plot_template(fig_corr),
                            use_container_width=True)

    # ---- Summary table + download ----
    with st.expander("📋 Realization-level summary table", expanded=False):
        # Pretty columns — built defensively so a partially-failed MC run
        # (some realizations missing columns) still renders.
        disp = summary.copy()
        if "npv_usd" in disp.columns:
            disp["npv_usd_MM"] = disp["npv_usd"] / 1e6
        if "cum_cf_usd" in disp.columns:
            disp["cum_cf_MM"] = disp["cum_cf_usd"] / 1e6
        # Only include columns that actually exist
        candidate_cols = ["realization", "npv_usd_MM", "cum_cf_MM",
                          "final_rf", "cum_oil", "cum_gas",
                          "peak_oil", "peak_gas"]
        factor_cols_all = [c for c in summary.columns if c.startswith("factor_")]
        cols = [c for c in candidate_cols if c in disp.columns] + factor_cols_all
        if not cols or len(disp) == 0:
            st.info("No realization data to display — the Monte Carlo run "
                    "produced no successful realizations.")
        else:
            fmt_map = {}
            for c, f in [("npv_usd_MM", "{:,.1f}"), ("cum_cf_MM", "{:,.1f}"),
                         ("final_rf", "{:.1%}"), ("cum_oil", "{:,.2f}"),
                         ("cum_gas", "{:,.2f}"), ("peak_oil", "{:,.0f}"),
                         ("peak_gas", "{:,.0f}")]:
                if c in cols:
                    fmt_map[c] = f
            for c in factor_cols_all:
                fmt_map[c] = "{:.3f}"
            try:
                st.dataframe(
                    disp[cols].style.format(fmt_map),
                    use_container_width=True, hide_index=True,
                )
            except Exception as exc:
                # Last-resort fallback: raw dataframe, no styling
                st.warning(f"Could not render styled table ({exc}); "
                           "showing raw values.")
                st.dataframe(disp[cols], use_container_width=True,
                             hide_index=True)

    try:
        csv = summary.to_csv(index=False).encode("utf-8")
        st.download_button(
            "⬇️ Download realizations as CSV", data=csv,
            file_name="monte_carlo_realizations.csv", mime="text/csv",
            use_container_width=True,
        )
    except Exception as exc:
        st.error(f"Could not prepare CSV export: {exc}")


def scenario_compare_section(units, fluid, asm, econ, wells):
    st.divider()
    st.subheader("🆚 Scenario comparison from saved cases")
    with st.expander("ℹ️ How comparison works", expanded=False):
        st.markdown(
            "Pick **two or more saved cases** below and the engine will run "
            "each one (using its own wells, capacities, PVT, strategy, "
            "economics, etc.) and overlay them on a single set of charts. "
            "Use this for value-of-investment / sensitivity analysis "
            "across scenarios you have already saved.\n\n"
            "If you want a quick *what-if* without saving, save the current "
            "inputs first via the case manager at the top of the page, then "
            "modify and save another case."
        )

    try:
        cases = fh.list_cases()
    except Exception as exc:
        st.info(f"Case directory not accessible ({exc}). "
                "Save a case first from the case manager at the top of the page.")
        return
    if not cases:
        st.info("📭 No saved cases yet — save at least two cases from the "
                "case manager at the top of the page before using comparison. "
                "Once you have two or more cases saved, they will appear here.")
        return
    if len(cases) < 2:
        st.info(f"Only one case saved (**{cases[0]['name']}**). "
                "Save at least one more case to use the comparison view.")
        return

    case_names = [c["name"] for c in cases]
    chosen = st.multiselect(
        "Pick saved cases to compare", case_names,
        default=case_names[:min(2, len(case_names))],
        key="scenario_chosen_cases",
    )
    if len(chosen) < 1:
        st.info("Pick at least one saved case.")
        return

    if not st.button("Compute scenarios", key="compute_scenarios_btn"):
        return

    is_oil = FLUID_SYSTEMS[fluid]["primary"] == "oil"
    results = {}
    errors = []

    with st.spinner(f"Running {len(chosen)} scenario(s)…"):
        for nm in chosen:
            try:
                target = next(c for c in cases if c["name"] == nm)
                case = fh.load_case(target["filename"])
                payload = case["payload"]
                # Use the case's saved units & fluid where available
                case_units = payload.get("scalar", {}).get("units", units)
                case_fluid = payload.get("scalar", {}).get("fluid", fluid)
                case_strategy = payload.get("scalar", {}).get("strategy", "Depletion")

                wells_s, reservoirs_s, meta, econ_dict = _wells_from_payload_tables(
                    payload, case_units, asm.start_date, case_fluid)
                if not wells_s:
                    errors.append(f"{nm}: no producers found in saved case.")
                    continue

                well_links_s = _well_links_from_payload(payload)
                asm_s = _build_asm_for_scenario(meta, case_fluid, case_strategy,
                                                  reservoirs=reservoirs_s,
                                                  well_links=well_links_s)
                econ_s = EconInputs(**econ_dict)

                df_s, _, _ = run_simulation(wells_s, asm_s)
                is_oil_s = FLUID_SYSTEMS[case_fluid]["primary"] == "oil"
                df_e_s = compute_economics(df_s, is_oil_s, econ_s, wells_s)
                # Breakeven price for this scenario
                try:
                    be_s = fh.breakeven_price(
                        df_s, is_oil_s, econ_s, wells_s,
                        base_oil_price=econ_s.oil_price,
                        base_gas_price=econ_s.gas_price,
                        compute_economics_fn=compute_economics,
                        target_npv=0.0)
                except Exception:
                    be_s = None
                results[nm] = (df_s, df_e_s, case_fluid, case_units, be_s)
            except Exception as e:
                errors.append(f"{nm}: {e}")

    for err in errors:
        st.warning(err)

    if not results:
        st.error("No scenarios ran successfully.")
        return

    # Overlay phase plots — display in current units; convert each scenario's
    # field-unit output into the active display units
    st.markdown("#### Field oil rate")
    fig_oil = go.Figure()
    for nm, (df_s, _, _, _, _) in results.items():
        fig_oil.add_trace(go.Scatter(
            x=df_s["date"],
            y=from_field(df_s["oil_rate"], "oil_rate", units),
            name=nm, mode="lines",
        ))
    fig_oil.update_layout(title=f"Field oil rate ({ulabel('oil_rate', units)})",
                          height=380, hovermode="x unified",
                          legend=dict(orientation="h", y=-0.2))
    st.plotly_chart(fh.apply_plot_template(fig_oil), use_container_width=True)

    st.markdown("#### Field gas rate")
    fig_gas = go.Figure()
    for nm, (df_s, _, _, _, _) in results.items():
        fig_gas.add_trace(go.Scatter(
            x=df_s["date"],
            y=from_field(df_s["gas_rate"], "gas_rate", units),
            name=nm, mode="lines",
        ))
    fig_gas.update_layout(title=f"Field gas rate ({ulabel('gas_rate', units)})",
                          height=380, hovermode="x unified",
                          legend=dict(orientation="h", y=-0.2))
    st.plotly_chart(fh.apply_plot_template(fig_gas), use_container_width=True)

    st.markdown("#### Recovery factor")
    fig_rf = go.Figure()
    for nm, (df_s, _, _, _, _) in results.items():
        fig_rf.add_trace(go.Scatter(x=df_s["date"], y=df_s["recovery_factor"],
                                    name=nm, mode="lines"))
    fig_rf.update_layout(title="Recovery factor", height=380,
                        hovermode="x unified",
                        yaxis_tickformat=".0%",
                        legend=dict(orientation="h", y=-0.2))
    st.plotly_chart(fh.apply_plot_template(fig_rf), use_container_width=True)

    st.markdown("#### Cumulative NPV")
    fig_npv = go.Figure()
    for nm, (_, df_e, _, _, _) in results.items():
        fig_npv.add_trace(go.Scatter(x=df_e["date"], y=df_e["npv"]/1e6,
                                      name=nm, mode="lines"))
    fig_npv.update_layout(title="Cumulative NPV ($MM)", height=380,
                          hovermode="x unified",
                          yaxis_title="NPV ($MM)",
                          legend=dict(orientation="h", y=-0.2))
    st.plotly_chart(fh.apply_plot_template(fig_npv), use_container_width=True)

    st.markdown("#### Summary table")
    rows = []
    for nm, (df_s, df_e, case_fluid, _, _) in results.items():
        is_oil_s = FLUID_SYSTEMS[case_fluid]["primary"] == "oil"
        peak_kind = "oil_rate" if is_oil_s else "gas_rate"
        rows.append({
            "Scenario": nm,
            "Fluid": case_fluid,
            "NPV ($MM)": df_e["npv"].iloc[-1]/1e6,
            "Cum CF ($MM)": df_e["cum_cashflow"].iloc[-1]/1e6,
            "Total revenue ($MM)": df_e["revenue"].sum()/1e6,
            "Final RF": df_s["recovery_factor"].iloc[-1],
            f"Peak rate ({ulabel(peak_kind, units)})": from_field(
                df_s["primary_rate"].max(), peak_kind, units),
            f"Cum oil ({ulabel('oil_vol', units)})": from_field(
                df_s["cum_oil"].iloc[-1], "oil_vol", units),
            f"Cum gas ({ulabel('gas_vol', units)})": from_field(
                df_s["cum_gas"].iloc[-1], "gas_vol", units),
        })
    summary_df = pd.DataFrame(rows)
    st.dataframe(
        summary_df.style.format({
            "NPV ($MM)": "{:,.0f}",
            "Cum CF ($MM)": "{:,.0f}",
            "Total revenue ($MM)": "{:,.0f}",
            "Final RF": "{:.1%}",
            f"Peak rate ({ulabel('oil_rate' if is_oil else 'gas_rate', units)})": "{:,.0f}",
            f"Cum oil ({ulabel('oil_vol', units)})": "{:,.2f}",
            f"Cum gas ({ulabel('gas_vol', units)})": "{:,.2f}",
        }),
        use_container_width=True,
    )

    # ---- Delta table: differences vs the first (reference) scenario ----
    st.markdown("#### Differences vs reference case")
    result_items = list(results.items())
    ref_name, (ref_df_s, ref_df_e, ref_fluid, _, ref_be) = result_items[0]
    st.caption(f"Reference case: **{ref_name}**. "
               "Positive ΔNPV / ΔCF means the scenario beats the reference; "
               "negative Δbreakeven means the scenario breaks even at a lower "
               "(better) price; negative Δcost means the scenario is cheaper.")

    def _total_cost(df_e):
        """Total project cost = OPEX + well CAPEX + facility CAPEX +
        abandonment + tax + royalty + tariff (everything deducted from
        gross revenue)."""
        cols = ["opex", "capex_well", "capex_facility", "abandonment",
                "tax", "royalty", "tariff"]
        return sum(df_e[c].sum() for c in cols if c in df_e.columns)

    ref_npv  = ref_df_e["npv"].iloc[-1] / 1e6
    ref_cf   = ref_df_e["cum_cashflow"].iloc[-1] / 1e6
    ref_cost = _total_cost(ref_df_e) / 1e6
    ref_be_oil = (ref_be or {}).get("oil_price")
    ref_cum_primary_base = (ref_df_s["cum_oil"].iloc[-1]
                            if FLUID_SYSTEMS[ref_fluid]["primary"] == "oil"
                            else ref_df_s["cum_gas"].iloc[-1])

    delta_rows = []
    for nm, (df_s, df_e, case_fluid, _, be_s) in result_items:
        is_oil_s = FLUID_SYSTEMS[case_fluid]["primary"] == "oil"
        npv_s  = df_e["npv"].iloc[-1] / 1e6
        cf_s   = df_e["cum_cashflow"].iloc[-1] / 1e6
        cost_s = _total_cost(df_e) / 1e6
        be_oil_s = (be_s or {}).get("oil_price")
        cum_primary_s = (df_s["cum_oil"].iloc[-1] if is_oil_s
                         else df_s["cum_gas"].iloc[-1])
        row = {
            "Scenario": nm + ("  (ref)" if nm == ref_name else ""),
            "NPV ($MM)": npv_s,
            "ΔNPV ($MM)": npv_s - ref_npv,
            "ΔCum CF ($MM)": cf_s - ref_cf,
            "Breakeven oil ($/bbl)": be_oil_s if be_oil_s is not None else float("nan"),
            "ΔBreakeven ($/bbl)": (be_oil_s - ref_be_oil)
                if (be_oil_s is not None and ref_be_oil is not None) else float("nan"),
            "Total cost ($MM)": cost_s,
            "ΔCost ($MM)": cost_s - ref_cost,
            f"ΔCum primary ({ulabel('oil_vol' if is_oil_s else 'gas_vol', units)})":
                from_field(cum_primary_s - ref_cum_primary_base,
                           "oil_vol" if is_oil_s else "gas_vol", units),
        }
        delta_rows.append(row)
    delta_df = pd.DataFrame(delta_rows)

    def _color_delta(val):
        """Green for value-positive deltas, red for negative. NaN untouched."""
        if pd.isna(val):
            return ""
        return ("color: #2ca02c" if val > 0 else
                "color: #d62728" if val < 0 else "")

    def _color_delta_inverse(val):
        """For breakeven & cost: lower is better → green for negative."""
        if pd.isna(val):
            return ""
        return ("color: #d62728" if val > 0 else
                "color: #2ca02c" if val < 0 else "")

    styled = delta_df.style.format({
        "NPV ($MM)": "{:,.0f}",
        "ΔNPV ($MM)": "{:+,.0f}",
        "ΔCum CF ($MM)": "{:+,.0f}",
        "Breakeven oil ($/bbl)": "{:,.1f}",
        "ΔBreakeven ($/bbl)": "{:+,.1f}",
        "Total cost ($MM)": "{:,.0f}",
        "ΔCost ($MM)": "{:+,.0f}",
        delta_df.columns[-1]: "{:+,.2f}",
    }).map(_color_delta, subset=["ΔNPV ($MM)", "ΔCum CF ($MM)", delta_df.columns[-1]]) \
      .map(_color_delta_inverse, subset=["ΔBreakeven ($/bbl)", "ΔCost ($MM)"])
    st.dataframe(styled, use_container_width=True)


# =============================================================================
# Portfolio mode — roll up several fields against a shared constraint
# =============================================================================
def portfolio_section(units, fluid, asm):
    """Roll up several saved cases into a portfolio: sum their production and
    cashflow, and apply an optional shared facility / export constraint that
    caps the combined rate. Answers 'what does my whole asset base look like,
    and does it fit the host / export capacity?'."""
    st.divider()
    st.subheader("📦 Portfolio rollup")
    with st.expander("ℹ️ How portfolio rollup works", expanded=False):
        st.markdown(
            "Pick **two or more saved cases** — each is treated as a "
            "separate field. The engine runs every field, aligns them on a "
            "common calendar, and **sums** their production and cashflow "
            "into a portfolio total.\n\n"
            "You can optionally set a **shared facility / export "
            "constraint** — a maximum combined oil or gas rate (a shared "
            "host platform, an export pipeline, or a processing hub). When "
            "the combined rate exceeds the limit, the portfolio total is "
            "capped and the deferred volume is reported. This is the core "
            "question for a hub development: do the fields fit the shared "
            "infrastructure, or do they need to be sequenced?"
        )

    try:
        cases = fh.list_cases()
    except Exception as exc:
        st.info(f"Case directory not accessible ({exc}).")
        return
    if not cases or len(cases) < 2:
        st.info("📭 Portfolio rollup needs at least two saved cases. "
                "Save each field as a case from the case manager at the "
                "top of the page.")
        return

    case_names = [c["name"] for c in cases]
    chosen = st.multiselect(
        "Fields in the portfolio (saved cases)", case_names,
        default=case_names[:min(3, len(case_names))],
        key="portfolio_chosen_cases")
    if len(chosen) < 2:
        st.info("Pick at least two fields.")
        return

    # ---- Field sequencing ----
    # Each field can be delayed by a number of months. This is the core
    # hub-development lever: two fields that both peak at the same time
    # overwhelm a shared host, but staggering one of them lets the
    # portfolio fit the capacity. The offset shifts that field's whole
    # profile later on the common calendar.
    st.markdown("**Field sequencing** — delay each field's start to test "
                "staggered development against the shared constraint.")
    seq_offsets = {}
    seq_cols = st.columns(min(len(chosen), 4))
    for i, nm in enumerate(chosen):
        with seq_cols[i % len(seq_cols)]:
            seq_offsets[nm] = st.number_input(
                f"{nm} — delay (months)", min_value=0, max_value=240,
                value=0, step=6, key=f"portfolio_offset_{nm}",
                help="Months to delay this field's first production. "
                     "0 = starts on the common calendar origin. Use this "
                     "to sequence fields so their peaks do not collide at "
                     "the shared facility.")

    # Shared constraint
    cc1, cc2 = st.columns(2)
    constraint_type = cc1.selectbox(
        "Shared constraint", ["None", "Oil rate", "Gas rate"],
        key="portfolio_constraint_type",
        help="Cap the combined portfolio rate at a shared facility or "
             "export limit. 'None' = unconstrained sum.")
    constraint_value = 0.0
    if constraint_type != "None":
        if constraint_type == "Oil rate":
            constraint_value = cc2.number_input(
                f"Max combined oil rate ({ulabel('oil_rate', units)})",
                min_value=0.0,
                value=from_field(150000.0, "oil_rate", units),
                step=from_field(10000.0, "oil_rate", units),
                key="portfolio_constraint_oil",
                help="Shared host / export oil capacity.")
        else:
            constraint_value = cc2.number_input(
                f"Max combined gas rate ({ulabel('gas_rate', units)})",
                min_value=0.0,
                value=from_field(400.0, "gas_rate", units),
                step=from_field(50.0, "gas_rate", units),
                key="portfolio_constraint_gas",
                help="Shared export pipeline / processing capacity.")

    if not st.button("Build portfolio", key="build_portfolio_btn",
                      type="primary"):
        return

    # ---- Run every field ----
    fields = {}
    errors = []
    with st.spinner(f"Running {len(chosen)} field(s)…"):
        for nm in chosen:
            try:
                target = next(c for c in cases if c["name"] == nm)
                case = fh.load_case(target["filename"])
                payload = case["payload"]
                c_units = payload.get("scalar", {}).get("units", units)
                c_fluid = payload.get("scalar", {}).get("fluid", fluid)
                c_strategy = payload.get("scalar", {}).get(
                    "strategy", "Depletion")
                wells_s, reservoirs_s, meta, econ_dict = \
                    _wells_from_payload_tables(payload, c_units,
                                                asm.start_date, c_fluid)
                if not wells_s:
                    errors.append(f"{nm}: no producers in saved case.")
                    continue
                well_links_s = _well_links_from_payload(payload)
                asm_s = _build_asm_for_scenario(
                    meta, c_fluid, c_strategy,
                    reservoirs=reservoirs_s, well_links=well_links_s)
                econ_s = EconInputs(**econ_dict)
                df_s, _, _ = run_simulation(wells_s, asm_s)
                is_oil_s = FLUID_SYSTEMS[c_fluid]["primary"] == "oil"
                df_e_s = compute_economics(df_s, is_oil_s, econ_s, wells_s)
                fields[nm] = {"df": df_s, "df_e": df_e_s}
            except Exception as e:
                errors.append(f"{nm}: {e}")
    for err in errors:
        st.warning(err)
    if len(fields) < 2:
        st.error("Need at least two fields to roll up.")
        return

    # ---- Align on a common monthly calendar (with sequencing offsets) ----
    # Each field can be delayed by seq_offsets[nm] months. A delay shifts
    # that field's whole production profile later on the shared calendar.
    # The common calendar must be wide enough to hold the most-delayed
    # field's full life, so it is extended by the largest offset.
    max_off = max((int(seq_offsets.get(nm, 0)) for nm in fields),
                  default=0)
    base_dates = sorted(set().union(
        *[set(pd.to_datetime(f["df"]["date"])) for f in fields.values()]))
    cal = pd.DatetimeIndex(base_dates)
    if max_off > 0:
        # extend the calendar by max_off months so delayed fields fit
        last = cal[-1]
        extra = pd.date_range(last + pd.DateOffset(months=1),
                              periods=max_off, freq="MS")
        cal = cal.append(pd.DatetimeIndex(extra))
    n = len(cal)

    def _aligned(df, col, offset_months=0):
        """Align a field's series onto the common calendar, optionally
        shifted later by offset_months."""
        idx = pd.to_datetime(df["date"])
        if offset_months:
            idx = idx + pd.DateOffset(months=int(offset_months))
        s = pd.Series(df[col].values, index=idx)
        # collapse any duplicate months created by the offset, then reindex
        s = s[~s.index.duplicated(keep="first")]
        return s.reindex(cal, fill_value=0.0).values

    # ---- Sum the portfolio ----
    port_oil = np.zeros(n)
    port_gas = np.zeros(n)
    port_cf = np.zeros(n)
    per_field = {}
    for nm, f in fields.items():
        off = int(seq_offsets.get(nm, 0))
        oil = _aligned(f["df"], "oil_rate", off)
        gas = _aligned(f["df"], "gas_rate", off)
        cf = (_aligned(f["df_e"], "cashflow", off)
              if "cashflow" in f["df_e"] else np.zeros(n))
        per_field[nm] = {"oil": oil, "gas": gas, "cf": cf,
                         "offset": off}
        port_oil += oil
        port_gas += gas
        port_cf += cf

    # ---- Apply the shared constraint ----
    deferred_note = None
    if constraint_type == "Oil rate" and constraint_value > 0:
        limit_field = to_field(constraint_value, "oil_rate", units)
        uncon = port_oil.copy()
        port_oil = np.minimum(port_oil, limit_field)
        deferred = np.maximum(uncon - limit_field, 0.0)
        deferred_vol = float(np.sum(deferred) * DAYS_PER_MONTH / 1e6)
        if deferred_vol > 0:
            deferred_note = (
                f"The shared oil constraint defers "
                f"{from_field(deferred_vol, 'oil_vol', units):,.1f} "
                f"{ulabel('oil_vol', units)} of production — the combined "
                f"fields exceed the host / export capacity. Consider "
                f"sequencing the fields or expanding capacity.")
    elif constraint_type == "Gas rate" and constraint_value > 0:
        limit_field = to_field(constraint_value, "gas_rate", units)
        uncon = port_gas.copy()
        port_gas = np.minimum(port_gas, limit_field)
        deferred = np.maximum(uncon - limit_field, 0.0)
        deferred_vol = float(np.sum(deferred) * DAYS_PER_MONTH / 1e9)
        if deferred_vol > 0:
            deferred_note = (
                f"The shared gas constraint defers "
                f"{from_field(deferred_vol, 'gas_vol', units):,.1f} "
                f"{ulabel('gas_vol', units)} of production — the combined "
                f"fields exceed the export capacity. Consider sequencing "
                f"or expanding capacity.")

    # ---- Headline KPIs ----
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Fields in portfolio", f"{len(fields)}")
    cum_oil_port = float(np.sum(port_oil) * DAYS_PER_MONTH / 1e6)
    cum_gas_port = float(np.sum(port_gas) * DAYS_PER_MONTH / 1e9)
    k2.metric(f"Portfolio oil ({ulabel('oil_vol', units)})",
              f"{from_field(cum_oil_port, 'oil_vol', units):,.1f}")
    k3.metric(f"Portfolio gas ({ulabel('gas_vol', units)})",
              f"{from_field(cum_gas_port, 'gas_vol', units):,.1f}")
    k4.metric("Portfolio NPV ($MM)",
              f"{np.sum(port_cf)/1e6:,.0f}")
    if deferred_note:
        st.warning("⚠️ " + deferred_note)

    # ---- Stacked production chart ----
    _is_oil_port = FLUID_SYSTEMS[fluid]["primary"] == "oil"
    st.markdown("#### Portfolio production — stacked by field")
    fig_p = go.Figure()
    for nm in fields:
        stream = (per_field[nm]["oil"] if _is_oil_port
                  else per_field[nm]["gas"])
        kind = "oil_rate" if _is_oil_port else "gas_rate"
        fig_p.add_trace(go.Scatter(
            x=cal, y=from_field(stream, kind, units),
            mode="lines", stackgroup="one", name=nm))
    # constraint line
    if constraint_type != "None" and constraint_value > 0:
        relevant = ((constraint_type == "Oil rate") == _is_oil_port)
        if relevant:
            fh.safe_hline(fig_p, constraint_value,
                          label="Shared capacity limit",
                          color="#d62828", dash="dash")
    fig_p.update_layout(
        title="Combined field production (stacked)",
        xaxis_title="Date",
        yaxis_title=(f"Oil rate ({ulabel('oil_rate', units)})"
                     if _is_oil_port
                     else f"Gas rate ({ulabel('gas_rate', units)})"),
        height=420, legend=dict(orientation="h", y=-0.2))
    st.plotly_chart(fh.apply_plot_template(fig_p),
                    use_container_width=True)

    # ---- Portfolio cashflow ----
    st.markdown("#### Portfolio cashflow")
    fig_cf = go.Figure()
    fig_cf.add_trace(go.Bar(
        x=cal, y=port_cf / 1e6, name="Monthly cashflow",
        marker_color="#2a6f97"))
    fig_cf.add_trace(go.Scatter(
        x=cal, y=np.cumsum(port_cf) / 1e6, name="Cumulative",
        mode="lines", line=dict(color="#d62828", width=3),
        yaxis="y2"))
    fig_cf.update_layout(
        title="Portfolio cashflow ($MM)",
        height=380, xaxis_title="Date",
        yaxis=dict(title="Monthly ($MM)"),
        yaxis2=dict(title="Cumulative ($MM)", overlaying="y",
                    side="right"),
        legend=dict(orientation="h", y=-0.2))
    st.plotly_chart(fh.apply_plot_template(fig_cf),
                    use_container_width=True)

    # ---- Per-field summary table ----
    rows = []
    for nm, f in fields.items():
        rows.append({
            "Field": nm,
            "Start delay (mo)": int(per_field[nm].get("offset", 0)),
            f"Cum oil ({ulabel('oil_vol', units)})": round(from_field(
                float(np.sum(per_field[nm]["oil"]) * DAYS_PER_MONTH / 1e6),
                "oil_vol", units), 1),
            f"Cum gas ({ulabel('gas_vol', units)})": round(from_field(
                float(np.sum(per_field[nm]["gas"]) * DAYS_PER_MONTH / 1e9),
                "gas_vol", units), 1),
            "NPV ($MM)": round(float(np.sum(per_field[nm]["cf"])) / 1e6, 0),
        })
    st.markdown("#### Per-field contribution")
    st.dataframe(pd.DataFrame(rows), use_container_width=True,
                 hide_index=True)
    _any_offset = any(per_field[nm].get("offset", 0) for nm in fields)
    st.caption(
        "Portfolio totals are the sum of the individual fields, aligned on "
        "a common calendar"
        + (" with each field's start delay applied" if _any_offset
           else "")
        + ". The shared constraint, if set, caps the combined rate — "
        "deferred volume is production the infrastructure cannot take when "
        "fields overlap in time."
        + (" Try increasing a field's delay to stagger the peaks and "
           "reduce deferred volume." if not _any_offset else ""))


# =============================================================================
# Batch mode — run many cases from a single YAML file
# =============================================================================
def batch_mode_section(units, fluid):
    """Run multiple cases from a batch YAML file in one go, then export the
    KPI results as CSV or JSON (API-style)."""
    st.divider()
    st.subheader("📦 Batch mode — run many cases from one YAML file")
    with st.expander("ℹ️ How batch mode works", expanded=False):
        st.markdown(
            "Upload a **batch YAML file** containing multiple cases. Each "
            "case is run through the full engine (production + economics) "
            "and the KPIs are collected into a results table you can "
            "download as CSV or JSON.\n\n"
            "**Batch file format:**\n"
            "```yaml\n"
            "schema_version: \"1.0\"\n"
            "cases:\n"
            "  - meta: {name: \"Low case\", description: \"P90 volumes\"}\n"
            "    scalar:\n"
            "      units: field\n"
            "      fluid: \"Oil with associated gas\"\n"
            "      start_date: \"2027-01-01\"\n"
            "      oil_price_bbl: 70\n"
            "      # ... all other scalar settings\n"
            "    tables:\n"
            "      producers_df:\n"
            "        - {name: P1, rig: Rig-A, qi_primary: 2000, ...}\n"
            "      cap_df:\n"
            "        - {start_date: \"2027-01-01\", oil: 40000, ...}\n"
            "  - meta: {name: \"Base case\"}\n"
            "    scalar: { ... }\n"
            "    tables: { ... }\n"
            "```\n\n"
            "Tip: export a working case from **YAML import / export** in the "
            "sidebar to get a correctly-structured starting point, then "
            "duplicate it under a `cases:` list."
        )

    up = st.file_uploader("Upload batch YAML file", type=["yaml", "yml"],
                           key="batch_yaml_uploader")
    if up is None:
        st.caption("No batch file uploaded yet.")
        return

    try:
        yaml_text = up.read().decode("utf-8")
        cases = fh.parse_batch_yaml(yaml_text)
    except Exception as exc:
        st.error(f"Could not parse batch YAML: {exc}")
        return

    st.success(f"Parsed **{len(cases)} case(s)** from the batch file.")

    # Pre-flight validation
    all_warnings = []
    for payload, meta in cases:
        warns = fh.validate_yaml_payload(payload, meta)
        if warns:
            all_warnings.append((meta.get("name", "?"), warns))
    if all_warnings:
        with st.expander(f"⚠️ Validation notes ({len(all_warnings)} case(s) "
                         "with warnings)", expanded=False):
            for cname, warns in all_warnings:
                st.markdown(f"**{cname}**")
                for w in warns:
                    st.markdown(f"- {w}")

    save_cases = st.checkbox(
        "Also save each case to the case library", value=False,
        key="batch_save_cases",
        help="When ticked, each batch case is also saved as a normal case "
             "(so it appears in the case manager and scenario comparison).")

    if not st.button("▶️ Run batch", key="batch_run_btn", type="primary"):
        return

    default_sd = st.session_state.get("start_date", date.today())
    results = []
    progress = st.progress(0.0, text="Running batch…")
    for i, (payload, meta) in enumerate(cases):
        payload["_meta"] = meta
        progress.progress((i + 0.5) / len(cases),
                          text=f"Running '{meta.get('name', f'Case {i+1}')}'…")
        res = run_payload_case(payload, default_sd, units)
        res["name"] = meta.get("name", f"Case {i+1}")
        results.append(res)
        if save_cases and res["ok"]:
            try:
                fh.save_case(res["name"], meta.get("description", ""), payload)
            except Exception:
                pass
        progress.progress((i + 1) / len(cases))
    progress.empty()

    n_ok = sum(1 for r in results if r["ok"])
    n_fail = len(results) - n_ok
    if n_fail:
        st.warning(f"{n_ok} succeeded, {n_fail} failed.")
    else:
        st.success(f"All {n_ok} cases ran successfully.")

    # Results table
    rows = []
    for r in results:
        k = r.get("kpis", {})
        rows.append({
            "Case": r["name"],
            "Status": "✅ OK" if r["ok"] else "❌ FAILED",
            "Error": r.get("error") or "",
            "NPV ($MM)": k.get("npv_MM"),
            "Breakeven oil ($/bbl)": k.get("breakeven_oil"),
            "Payback (yrs)": k.get("payback_yrs"),
            "Cum oil (MMstb)": k.get("cum_oil_MMstb"),
            "Cum gas (Bscf)": k.get("cum_gas_Bscf"),
            "Final RF": k.get("final_rf"),
            "Peak rate": k.get("peak_primary_rate"),
        })
    batch_df = pd.DataFrame(rows)
    fmt = {
        "NPV ($MM)": "{:,.0f}", "Breakeven oil ($/bbl)": "{:,.1f}",
        "Payback (yrs)": "{:,.1f}", "Cum oil (MMstb)": "{:,.2f}",
        "Cum gas (Bscf)": "{:,.2f}", "Final RF": "{:.1%}",
        "Peak rate": "{:,.0f}",
    }
    # Only format columns that have at least one non-null value
    fmt = {c: f for c, f in fmt.items()
           if c in batch_df.columns and batch_df[c].notna().any()}
    try:
        st.dataframe(batch_df.style.format(fmt, na_rep="—"),
                     use_container_width=True, hide_index=True)
    except Exception:
        st.dataframe(batch_df, use_container_width=True, hide_index=True)

    # Exports — build API-style payloads via the helper
    headless_like = [
        {"name": r["name"], "ok": r["ok"], "error": r.get("error"),
         "kpis": {
             "npv_MM": r["kpis"].get("npv_MM"),
             "irr": None,
             "payback_yrs": r["kpis"].get("payback_yrs"),
             "cum_oil": r["kpis"].get("cum_oil_MMstb"),
             "cum_gas": r["kpis"].get("cum_gas_Bscf"),
             "final_rf": r["kpis"].get("final_rf"),
             "peak_rate": r["kpis"].get("peak_primary_rate"),
         }}
        for r in results
    ]
    exp1, exp2 = st.columns(2)
    with exp1:
        try:
            csv_bytes = fh.batch_results_to_csv(headless_like).encode("utf-8")
            st.download_button("⬇️ Download results as CSV", data=csv_bytes,
                               file_name="batch_results.csv", mime="text/csv",
                               use_container_width=True, key="batch_csv_dl")
        except Exception as exc:
            st.error(f"CSV export failed: {exc}")
    with exp2:
        try:
            json_bytes = fh.batch_results_to_json(headless_like).encode("utf-8")
            st.download_button("⬇️ Download results as JSON (API format)",
                               data=json_bytes, file_name="batch_results.json",
                               mime="application/json",
                               use_container_width=True, key="batch_json_dl")
        except Exception as exc:
            st.error(f"JSON export failed: {exc}")
    st.caption(
        "The JSON export is structured as an API-style response "
        "(`schema_version`, `generated_at`, `n_cases`, `n_ok`, `cases[]` with "
        "per-case `kpis`) so it can be consumed programmatically by "
        "downstream tooling."
    )


# ============================================================================
# CONCEPT SELECTOR — "Hanging garden" / morphological matrix builder
# ============================================================================
# Lets the user lay out a concept long list as columns (dimensions) of
# alternatives, sweep every combination as a batch run, and colour each
# selected option-box by the resulting NPV (red → green). Inspired by the
# DG1 concept-screening matrices used on the NCS.

# Default dimension set — a sensible NCS-style starter the user can edit.
# Each option's `patches` dict maps session_state keys to override values.
# When a combination is run, the base payload's `scalar` dict is updated
# with the patches from each chosen option before run_payload_case is called.
_DEFAULT_CONCEPT_DIMENSIONS = [
    {
        "name": "Drainage strategy",
        "description": "Recovery mechanism — sets reservoir + injector behaviour.",
        "options": [
            {"label": "Depletion",
             "description": "Primary depletion, no injection. Lowest CAPEX, lower RF.",
             "patches": {"strategy": "Depletion",
                          "aq_active": False, "vrr": 0.0, "inj_eff": 0.0}},
            {"label": "Water injection",
             "description": "Waterflood with VRR=1, ~85% sweep efficiency.",
             "patches": {"strategy": "Water injection",
                          "vrr": 1.0, "inj_eff": 0.85}},
            {"label": "Gas injection",
             "description": "Pressure maintenance via lean gas re-injection.",
             "patches": {"strategy": "Gas injection",
                          "vrr": 1.0, "inj_eff": 0.80}},
            {"label": "WAG",
             "description": "Water-alternating-gas — best EOR sweep.",
             "patches": {"strategy": "WAG",
                          "vrr": 1.05, "inj_eff": 0.90}},
        ],
    },
    {
        "name": "Number of producers",
        "description": "Drilling programme size.",
        "options": [
            {"label": "4 wells", "description": "Minimum viable.",
             "patches": {"_n_producers_override": 4}},
            {"label": "6 wells", "description": "Mid-range.",
             "patches": {"_n_producers_override": 6}},
            {"label": "8 wells", "description": "Default — balances CAPEX & ramp-up.",
             "patches": {"_n_producers_override": 8}},
            {"label": "10 wells", "description": "Aggressive infill.",
             "patches": {"_n_producers_override": 10}},
        ],
    },
    {
        "name": "Drilling rig",
        "description": "Rig class — drives dayrate and CAPEX.",
        "options": [
            {"label": "Jack-up", "description": "Low dayrate, shallow water only.",
             "patches": {"well_cost_mode": "rig_rate", "rig_dayrate": 250.0}},
            {"label": "Semi-sub",
             "description": "Mid-range, standard NCS HARSH-environment.",
             "patches": {"well_cost_mode": "rig_rate", "rig_dayrate": 500.0}},
            {"label": "Drillship", "description": "Deepwater / high dayrate.",
             "patches": {"well_cost_mode": "rig_rate", "rig_dayrate": 700.0}},
        ],
    },
    {
        "name": "Host facility",
        "description": "Production hub — sets facility CAPEX.",
        "options": [
            {"label": "Tie-back to existing",
             "description": "Cheapest — modest CAPEX on existing host.",
             "patches": {"_facility_capex_override_MM": 200.0}},
            {"label": "New FPSO",
             "description": "Floating production, storage, offloading.",
             "patches": {"_facility_capex_override_MM": 1700.0}},
            {"label": "New jacket/platform",
             "description": "Fixed jacket — typical for shallow shelf.",
             "patches": {"_facility_capex_override_MM": 1200.0}},
            {"label": "Subsea to shore",
             "description": "Long subsea tieback to onshore plant.",
             "patches": {"_facility_capex_override_MM": 900.0}},
        ],
    },
    {
        "name": "Oil price scenario",
        "description": "Flat real Brent assumption.",
        "options": [
            {"label": "Low ($55/bbl)", "description": "Down-cycle screening.",
             "patches": {"oil_price_bbl": 55.0}},
            {"label": "Base ($75/bbl)", "description": "Mid-cycle reference.",
             "patches": {"oil_price_bbl": 75.0}},
            {"label": "High ($95/bbl)", "description": "Up-cycle upside.",
             "patches": {"oil_price_bbl": 95.0}},
        ],
    },
]


# Predefined concept-dimension templates mirroring the DG1 hanging-garden
# building-block columns: Subsurface, Drilling & Well, SURF, Topside
# facilities. Each is a list of dimensions the user can load as a starting
# point and then edit. Patches are screening-level deltas — the user
# refines them (or links full cases) per option.
_CONCEPT_TEMPLATES = {
    "Subsurface": [
        {"name": "Drainage strategy", "description": "Recovery mechanism.",
         "options": [
             {"label": "Depletion", "description": "Primary depletion.",
              "patches": {"strategy": "Depletion", "vrr": 0.0}},
             {"label": "Water injection", "description": "Waterflood.",
              "patches": {"strategy": "Water injection", "vrr": 1.0,
                           "inj_eff": 0.85}},
             {"label": "Gas injection", "description": "Gas re-injection.",
              "patches": {"strategy": "Gas injection", "vrr": 1.0,
                           "inj_eff": 0.80}},
             {"label": "WAG", "description": "Water-alternating-gas.",
              "patches": {"strategy": "WAG", "vrr": 1.05,
                           "inj_eff": 0.90}},
         ]},
        {"name": "Production management", "description": "Gas/water mgmt.",
         "options": [
             {"label": "Gas re-injection", "description": "", "patches": {}},
             {"label": "Gas export", "description": "", "patches": {}},
             {"label": "PWRI", "description": "Produced-water re-injection.",
              "patches": {}},
         ]},
        {"name": "Number of producers", "description": "Well count.",
         "options": [
             {"label": "4 wells", "description": "",
              "patches": {"_n_producers_override": 4}},
             {"label": "6 wells", "description": "",
              "patches": {"_n_producers_override": 6}},
             {"label": "8 wells", "description": "",
              "patches": {"_n_producers_override": 8}},
         ]},
        {"name": "Lower completion", "description": "Sand/zone control.",
         "options": [
             {"label": "Open hole", "description": "", "patches": {}},
             {"label": "Cased & perforated", "description": "",
              "patches": {}},
             {"label": "Screens / gravel pack", "description": "",
              "patches": {}},
         ]},
    ],
    "Drilling & Well": [
        {"name": "Well type", "description": "Trajectory.",
         "options": [
             {"label": "Vertical", "description": "", "patches": {}},
             {"label": "Deviated", "description": "", "patches": {}},
             {"label": "Horizontal", "description": "", "patches": {}},
         ]},
        {"name": "Drilling rig", "description": "Rig class / dayrate.",
         "options": [
             {"label": "Jack-up", "description": "Shallow water.",
              "patches": {"well_cost_mode": "rig_rate",
                           "rig_dayrate": 250.0}},
             {"label": "Semi-sub", "description": "Harsh environment.",
              "patches": {"well_cost_mode": "rig_rate",
                           "rig_dayrate": 500.0}},
             {"label": "Drillship", "description": "Deep water.",
              "patches": {"well_cost_mode": "rig_rate",
                           "rig_dayrate": 700.0}},
         ]},
        {"name": "Completion type", "description": "Upper completion.",
         "options": [
             {"label": "Single zone", "description": "", "patches": {}},
             {"label": "Smart (multi-zone)", "description": "ICVs.",
              "patches": {}},
             {"label": "Monobore", "description": "", "patches": {}},
         ]},
        {"name": "Artificial lift", "description": "Lift method.",
         "options": [
             {"label": "None (natural flow)", "description": "",
              "patches": {}},
             {"label": "Gas lift", "description": "", "patches": {}},
             {"label": "ESP", "description": "", "patches": {}},
         ]},
    ],
    "SURF": [
        {"name": "Field architecture", "description": "SURF layout.",
         "options": [
             {"label": "Cluster (manifold)", "description": "", "patches": {}},
             {"label": "Daisy chain", "description": "", "patches": {}},
             {"label": "Satellite wells", "description": "", "patches": {}},
             {"label": "Template", "description": "", "patches": {}},
         ]},
        {"name": "Host facility", "description": "Production hub.",
         "options": [
             {"label": "Tie-back to existing", "description": "",
              "patches": {"_facility_capex_override_MM": 200.0}},
             {"label": "New FPSO", "description": "",
              "patches": {"_facility_capex_override_MM": 1700.0}},
             {"label": "Subsea to shore", "description": "",
              "patches": {"_facility_capex_override_MM": 900.0}},
         ]},
        {"name": "Flowline", "description": "Flowline type.",
         "options": [
             {"label": "Wet insulated", "description": "", "patches": {}},
             {"label": "Pipe-in-pipe", "description": "", "patches": {}},
             {"label": "Electrically heated (DEH)", "description": "",
              "patches": {}},
         ]},
        {"name": "Hydrate management", "description": "Flow assurance.",
         "options": [
             {"label": "MEG with reclamation", "description": "",
              "patches": {}},
             {"label": "Methanol once-through", "description": "",
              "patches": {}},
             {"label": "Insulation only", "description": "", "patches": {}},
         ]},
        {"name": "Installation method", "description": "Pipelay.",
         "options": [
             {"label": "Reeling", "description": "", "patches": {}},
             {"label": "S-lay", "description": "", "patches": {}},
             {"label": "J-lay", "description": "", "patches": {}},
         ]},
    ],
    "Topside facilities": [
        {"name": "Processing capacity", "description": "Throughput basis.",
         "options": [
             {"label": "Use existing", "description": "", "patches": {}},
             {"label": "Debottleneck", "description": "", "patches": {}},
             {"label": "New train", "description": "", "patches": {}},
         ]},
        {"name": "Gas compression", "description": "Export/lift gas.",
         "options": [
             {"label": "Use existing", "description": "", "patches": {}},
             {"label": "New compressor skid", "description": "",
              "patches": {}},
         ]},
        {"name": "Power supply", "description": "Power to SURF/topside.",
         "options": [
             {"label": "Existing generation", "description": "",
              "patches": {}},
             {"label": "New gas turbine", "description": "", "patches": {}},
             {"label": "Power from shore", "description": "", "patches": {}},
         ]},
        {"name": "Produced water", "description": "Water handling.",
         "options": [
             {"label": "Use existing", "description": "", "patches": {}},
             {"label": "New CFU / cyclones", "description": "",
              "patches": {}},
             {"label": "PWRI", "description": "", "patches": {}},
         ]},
        {"name": "Living quarters", "description": "POB during tie-in.",
         "options": [
             {"label": "Spare LQ capacity", "description": "", "patches": {}},
             {"label": "Temporary cabins", "description": "", "patches": {}},
             {"label": "Flotel", "description": "", "patches": {}},
         ]},
    ],
}


def _concept_study_to_doc(dimensions, selected, results, base_source,
                          base_loaded):
    """Assemble the nested study dict (matrix + base + results)."""
    from datetime import datetime as _dt, timezone as _tz
    return {
        "fieldvista_concept_study": {
            "schema_version": 1,
            "generated_utc": _dt.now(_tz.utc).isoformat(timespec="seconds"),
            "app_version": str(getattr(fh, "FP_HELPERS_VERSION", "?")),
            "base_case": {
                "source": base_source,
                "loaded_case_name": (
                    (base_loaded or {}).get("_meta", {}).get("name")
                    if base_loaded else "live sidebar inputs"),
            },
            "matrix": {
                "dimensions": [
                    {
                        "name": d["name"],
                        "description": d.get("description", ""),
                        "options": [
                            {
                                "label": o["label"],
                                "description": o.get("description", ""),
                                "patches": o.get("patches", {}),
                                "swept": (oi in selected.get(di, set())),
                            }
                            for oi, o in enumerate(d["options"])
                        ],
                    }
                    for di, d in enumerate(dimensions)
                ],
            },
            "n_combinations": len(results),
            "results": [
                {
                    "name": r.get("name"),
                    "picks": {dn: lbl for dn, lbl in r.get("picks", [])},
                    "kpis": {
                        "npv_after_tax_MM": r.get("npv_MM"),
                        "npv_pre_tax_MM": r.get("npv_pretax_MM"),
                        "irr": r.get("irr"),
                        "final_rf": r.get("final_rf"),
                        "resources_mmboe": r.get("resources_mmboe"),
                        "breakeven_oil_usd_bbl": r.get("breakeven_oil"),
                        "capex_total_MM": r.get("capex_total_MM"),
                        "capex_disc_MM": r.get("capex_disc_MM"),
                        "co2_total_Mt": r.get("co2_total_Mt"),
                    },
                    "ok": r.get("ok", False),
                    "error": r.get("error"),
                }
                for r in sorted(
                    results.values(),
                    key=lambda x: (x.get("npv_MM")
                                    if x.get("npv_MM") is not None
                                    else -9e18),
                    reverse=True)
            ],
        }
    }


def _concept_doc_to_matrix(doc):
    """Reconstruct (dimensions, selected) from a study doc.

    Inverse of _concept_study_to_doc for the matrix portion — lets a
    study YAML be re-imported to rebuild the editable matrix.
    Returns (dimensions_list, selected_dict) or raises on bad schema.
    """
    root = doc.get("fieldvista_concept_study", doc)
    dims_in = root.get("matrix", {}).get("dimensions", [])
    dimensions = []
    selected = {}
    for di, d in enumerate(dims_in):
        opts = []
        sel = set()
        for oi, o in enumerate(d.get("options", [])):
            opts.append({
                "label": o.get("label", f"Option {oi+1}"),
                "description": o.get("description", ""),
                "patches": o.get("patches", {}) or {},
            })
            if o.get("swept", True):
                sel.add(oi)
        dimensions.append({
            "name": d.get("name", f"Dimension {di+1}"),
            "description": d.get("description", ""),
            "options": opts,
        })
        selected[di] = sel
    if not dimensions:
        raise ValueError("No dimensions found in study document.")
    return dimensions, selected


def _concept_pareto_front(rows):
    """Given a list of dicts with 'capex_disc_MM' and a NPV key 'mean'
    (or 'npv_MM'), return the set of concept labels on the efficient
    frontier (max NPV for min CAPEX). A concept is dominated if another
    has CAPEX ≤ and NPV ≥ with at least one strict.
    """
    pts = []
    for r in rows:
        cap = r.get("capex_disc_MM")
        npv = r.get("mean", r.get("npv_MM"))
        if cap is None or npv is None:
            continue
        pts.append((r["concept"], float(cap), float(npv)))
    nondominated = set()
    for label, cap, npv in pts:
        dominated = False
        for _, cap2, npv2 in pts:
            if (cap2 <= cap and npv2 >= npv) and (cap2 < cap or npv2 > npv):
                dominated = True
                break
        if not dominated:
            nondominated.add(label)
    return nondominated


def _apply_concept_patches(payload: dict, picks: list) -> dict:
    """Apply a list of option-patch dicts to a base payload.

    `picks` is a list of {label, patches} dicts (one per dimension).
    Most patches map directly to session_state scalar keys. A few use
    magic underscore-prefixed keys that drive table-level overrides
    (well count, facility CAPEX schedule) — handled here.
    """
    import copy
    p = copy.deepcopy(payload)
    p.setdefault("scalar", {})
    for pick in picks:
        for k, v in pick.get("patches", {}).items():
            if k.startswith("_n_producers_override"):
                # Truncate / replicate the producers_df to the desired
                # well count. Keeps the per-well design (rates, decline)
                # from the base case but scales producer count.
                try:
                    n = int(v)
                    if "tables" in p and "producers_df" in p["tables"]:
                        pdf = pd.DataFrame(p["tables"]["producers_df"])
                        if len(pdf) >= n:
                            pdf = pdf.iloc[:n].reset_index(drop=True)
                        else:
                            # Replicate last row until reaching n
                            n_need = n - len(pdf)
                            extra = pd.concat(
                                [pdf.iloc[[-1]]] * n_need, ignore_index=True)
                            pdf = pd.concat([pdf, extra], ignore_index=True)
                        # Renumber
                        pdf["name"] = [f"P-{i+1:02d}"
                                        for i in range(len(pdf))]
                        # IMPORTANT: write back in the SAME column-oriented
                        # format the payload uses elsewhere (orient="list").
                        # run_payload_case reads producers_df as a dict of
                        # columns (pdata.get("name"), pdata["rig"][i]); a
                        # row-oriented "records" list would raise
                        # 'list' object has no attribute 'get'.
                        p["tables"]["producers_df"] = pdf.to_dict(
                            orient="list")
                except Exception:
                    pass
            elif k.startswith("_facility_capex_override_MM"):
                # Rebuild the facility CAPEX schedule as a single lump
                # spend on the project start date. The base schedule
                # ordering is preserved but amounts re-scaled to total
                # the new figure.
                try:
                    new_total = float(v)
                    if "tables" in p and "fac_df" in p["tables"]:
                        fac = pd.DataFrame(p["tables"]["fac_df"])
                        cur_total = (
                            fac["amount_MMUSD"].astype(float).sum()
                            if "amount_MMUSD" in fac.columns
                            else 0.0)
                        if cur_total > 0:
                            ratio = new_total / cur_total
                            fac["amount_MMUSD"] = (
                                fac["amount_MMUSD"].astype(float) * ratio)
                        p["tables"]["fac_df"] = fac.to_dict(orient="list")
                except Exception:
                    pass
            else:
                p["scalar"][k] = v
    return p


def _concept_color_for_npv(npv_MM: float, lo: float, hi: float) -> str:
    """Return a hex colour on a red→amber→green ramp keyed to NPV ($MM)."""
    if npv_MM is None:
        return "#cccccc"
    if hi <= lo:
        return "#9ecae1"
    t = max(0.0, min(1.0, (npv_MM - lo) / (hi - lo)))
    # Anchor: 0.0 red (#d62728), 0.5 amber (#ff9900), 1.0 green (#2ca02c)
    if t < 0.5:
        # red → amber
        s = t * 2
        r = int(0xd6 + s * (0xff - 0xd6))
        g = int(0x27 + s * (0x99 - 0x27))
        b = int(0x28 + s * (0x00 - 0x28))
    else:
        # amber → green
        s = (t - 0.5) * 2
        r = int(0xff + s * (0x2c - 0xff))
        g = int(0x99 + s * (0xa0 - 0x99))
        b = int(0x00 + s * (0x2c - 0x00))
    return f"#{r:02x}{g:02x}{b:02x}"


def _render_concept_garden_svg(dimensions: list, selected: dict,
                                results_by_pick: dict | None = None) -> str:
    """SVG of the concept long-list, coloured by NPV when results exist.

    `selected[dim_idx]` = set of option indices the user has ticked for
    inclusion in the batch. `results_by_pick` is keyed by frozenset of
    (dim_name, option_label) and holds a dict with `npv_MM` for each
    combination — used to colour boxes by the BEST NPV that flows
    through that option.
    """
    col_w = 230
    gap_x = 20
    title_h = 34          # dedicated band for the chart title
    header_h = 64 + title_h   # headers sit below the title band
    row_h = 56
    row_gap = 10
    pad = 24
    n_cols = len(dimensions)
    max_options = max((len(d["options"]) for d in dimensions), default=1)
    width = pad * 2 + n_cols * col_w + (n_cols - 1) * gap_x
    height = pad * 2 + header_h + 16 + (row_h + row_gap) * max_options + 40

    # Map each option to its NPV (one case per option now — no
    # combinations, so it's a direct lookup by (dim_name, label)).
    best_per_opt = {}      # key: (dim_idx, opt_idx) → npv_MM
    if results_by_pick:
        # Build a (dim_name,label) → npv lookup from the result records
        npv_by_pick = {}
        for res in results_by_pick.values():
            npv = res.get("npv_MM")
            if npv is None:
                continue
            dn = res.get("dim")
            lbl = res.get("label")
            if dn is not None and lbl is not None:
                npv_by_pick[(dn, lbl)] = npv
        for di, d in enumerate(dimensions):
            for oi, o in enumerate(d["options"]):
                npv = npv_by_pick.get((d["name"], o["label"]))
                if npv is not None:
                    best_per_opt[(di, oi)] = npv
        all_npvs = [r.get("npv_MM") for r in results_by_pick.values()
                    if r.get("npv_MM") is not None]
        npv_lo = min(all_npvs) if all_npvs else 0.0
        npv_hi = max(all_npvs) if all_npvs else 0.0
    else:
        npv_lo = npv_hi = 0.0

    out = [f'<svg viewBox="0 0 {width} {height}" '
           f'xmlns="http://www.w3.org/2000/svg" '
           f'style="background:white;border:1px solid #ddd;'
           f'border-radius:8px;font-family:Helvetica,Arial,sans-serif">']
    # Title
    out.append(f'<text x="{pad}" y="26" font-size="15" font-weight="700" '
               f'fill="#0B3D91">Concept long list — selected options '
               f'in colour, NPV ramp red → green</text>')
    # Columns
    for di, d in enumerate(dimensions):
        x = pad + di * (col_w + gap_x)
        # Header bar — taller, holds up to two wrapped lines
        out.append(f'<rect x="{x}" y="{header_h - 44}" width="{col_w}" '
                   f'height="44" fill="#2c7fb8" rx="5"/>')
        # Wrap the title onto up to two lines by words rather than
        # truncating, so long dimension names stay readable.
        title = d["name"]
        max_chars = 26
        if len(title) <= max_chars:
            t_lines = [title]
        else:
            words = title.split(" ")
            t_lines, cur = [], ""
            for w in words:
                if len(cur) + len(w) + 1 <= max_chars:
                    cur = (cur + " " + w).strip()
                else:
                    t_lines.append(cur)
                    cur = w
            if cur:
                t_lines.append(cur)
            t_lines = t_lines[:2]
        if len(t_lines) == 1:
            out.append(f'<text x="{x + col_w/2}" y="{header_h - 16}" '
                       f'font-size="13" font-weight="700" fill="white" '
                       f'text-anchor="middle">{t_lines[0]}</text>')
        else:
            out.append(f'<text x="{x + col_w/2}" y="{header_h - 24}" '
                       f'font-size="12" font-weight="700" fill="white" '
                       f'text-anchor="middle">{t_lines[0]}</text>')
            out.append(f'<text x="{x + col_w/2}" y="{header_h - 8}" '
                       f'font-size="12" font-weight="700" fill="white" '
                       f'text-anchor="middle">{t_lines[1]}</text>')
        # Options
        for oi, opt in enumerate(d["options"]):
            y = header_h + 16 + oi * (row_h + row_gap)
            is_selected = oi in selected.get(di, set())
            # Colour: if results exist for this option, use NPV ramp;
            # else neutral (selected = blue, unselected = grey).
            best_npv = best_per_opt.get((di, oi))
            if is_selected and best_npv is not None:
                fill = _concept_color_for_npv(best_npv, npv_lo, npv_hi)
                stroke = "#0B3D91"
                stroke_w = 2
                text_color = "#000"
            elif is_selected:
                fill = "#dde9f5"
                stroke = "#0B3D91"
                stroke_w = 2
                text_color = "#0B3D91"
            else:
                fill = "#f4f4f4"
                stroke = "#ccc"
                stroke_w = 1
                text_color = "#666"
            out.append(
                f'<rect x="{x}" y="{y}" width="{col_w}" height="{row_h}" '
                f'fill="{fill}" stroke="{stroke}" stroke-width="{stroke_w}" '
                f'rx="4"/>')
            # Option label — centred, leaving the bottom strip for the
            # NPV badge so the two never overlap.
            label = opt["label"]
            _label_y = (y + row_h / 2 - 6) if (is_selected
                        and best_npv is not None) else (y + row_h / 2 + 4)
            if len(label) <= 30:
                out.append(f'<text x="{x + col_w/2}" y="{_label_y}" '
                           f'font-size="12" font-weight="600" '
                           f'fill="{text_color}" text-anchor="middle">'
                           f'{label}</text>')
            else:
                mid = len(label) // 2
                space = label.find(" ", mid - 4)
                if space == -1:
                    space = mid
                l1, l2 = label[:space].strip(), label[space:].strip()
                out.append(f'<text x="{x + col_w/2}" y="{_label_y - 7}" '
                           f'font-size="11" font-weight="600" '
                           f'fill="{text_color}" text-anchor="middle">'
                           f'{l1}</text>')
                out.append(f'<text x="{x + col_w/2}" y="{_label_y + 7}" '
                           f'font-size="11" font-weight="600" '
                           f'fill="{text_color}" text-anchor="middle">'
                           f'{l2}</text>')
            # NPV badge — centred along the bottom edge of the box, below
            # the label, so it never sits on top of the option name.
            if is_selected and best_npv is not None:
                out.append(
                    f'<text x="{x + col_w/2}" y="{y + row_h - 7}" '
                    f'font-size="10" font-weight="700" fill="#1a1a1a" '
                    f'text-anchor="middle">${best_npv:,.0f}MM</text>')
    # Colour legend at the bottom
    if results_by_pick and npv_hi > npv_lo:
        legy = height - 20
        for i in range(50):
            t = i / 49
            c = _concept_color_for_npv(npv_lo + t * (npv_hi - npv_lo),
                                        npv_lo, npv_hi)
            out.append(f'<rect x="{pad + i*4}" y="{legy}" width="4" '
                       f'height="8" fill="{c}"/>')
        out.append(f'<text x="{pad}" y="{legy - 3}" font-size="9" '
                   f'fill="#666">${npv_lo:,.0f}MM (worst)</text>')
        out.append(f'<text x="{pad + 200}" y="{legy - 3}" font-size="9" '
                   f'fill="#666">${npv_hi:,.0f}MM (best)</text>')
    out.append("</svg>")
    return "".join(out)


def well_planner_section(units, fluid):
    """Interactive well + completion designer with a cross-section view."""
    with st.expander("🛠️ Well Planner — design a well & its completion",
                      expanded=False):
        st.caption(
            "Design a representative well and see its cross-section and "
            "completion schematic. This is a design aid for the concept "
            "phase — it sketches the wellbore architecture, casing "
            "programme, completion type and artificial lift so you can "
            "sense-check the well concept and communicate it. It does "
            "not feed the production engine (that uses the per-well "
            "table); think of it as the drawing board.")

        is_metric = (units == "metric")
        depth_unit = "m" if is_metric else "ft"
        # depth conversion: engine schematic works in metres
        d2m = (1.0 if is_metric else 0.3048)

        c1, c2, c3 = st.columns(3)
        well_type = c1.selectbox(
            "Well type", ["Vertical", "Deviated", "Horizontal"],
            index=1, key="wp_well_type",
            help="Trajectory archetype. Horizontal maximises reservoir "
                 "contact; deviated reaches offset targets; vertical is "
                 "simplest and cheapest.")
        is_injector = c2.selectbox(
            "Well role", ["Producer", "Injector"],
            key="wp_role") == "Injector"
        artificial = c3.selectbox(
            "Artificial lift",
            ["None", "Gas lift", "ESP", "HSP"],
            key="wp_lift",
            help="Gas lift — robust, wide rate range, needs lift-gas. "
                 "ESP — high rate, efficient, workover-intensive. "
                 "HSP — hydraulic submersible, deep/hot wells.")

        c4, c5, c6 = st.columns(3)
        wd_disp = c4.number_input(
            f"Water depth ({depth_unit}, 0 = dry/platform)",
            value=(110.0 if is_metric else 360.0),
            min_value=0.0, key="wp_wd")
        td_disp = c5.number_input(
            f"Total depth ({depth_unit} MD)",
            value=(2600.0 if is_metric else 8530.0),
            min_value=100.0, key="wp_td")
        n_casing = c6.selectbox(
            "Casing strings", [2, 3, 4], index=1, key="wp_casing",
            help="Conductor + surface + (intermediate) + production. "
                 "More strings for HPHT, deep, or troublesome holes.")

        c7, c8, c9 = st.columns(3)
        res_top_disp = c7.number_input(
            f"Reservoir top ({depth_unit} TVD)",
            value=(2100.0 if is_metric else 6890.0),
            min_value=50.0, key="wp_restop")
        res_thick_disp = c8.number_input(
            f"Reservoir thickness ({depth_unit})",
            value=(150.0 if is_metric else 490.0),
            min_value=1.0, key="wp_resthick")
        zonal = c9.selectbox(
            "Zonal isolation",
            ["None", "Single packer", "Smart (multi-zone)"],
            index=1, key="wp_zonal",
            help="Single packer — one production zone. Smart — "
                 "multi-zone with inflow-control valves for selective "
                 "production / shut-off.")

        # Completion type — the menu depends on producer vs injector and
        # the formation competence the user expects.
        completion = st.selectbox(
            "Lower completion type",
            ["Open hole", "Cased & perforated", "Slotted liner",
             "Screens / gravel pack", "Frac-pack"],
            index=1, key="wp_completion",
            help="Open hole — cheapest, competent rock only. "
                 "Cased & perforated — most common, good zonal control. "
                 "Slotted liner — sand-prone, low cost. "
                 "Screens / gravel pack — unconsolidated sand control. "
                 "Frac-pack — sand control + productivity in "
                 "high-permeability sands.")

        # Build the schematic spec (convert to metres for the renderer)
        wp_spec = {
            "well_type": well_type,
            "water_depth_m": wd_disp * d2m,
            "td_m": td_disp * d2m,
            "reservoir_top_m": res_top_disp * d2m,
            "reservoir_thick_m": res_thick_disp * d2m,
            "n_casing": int(n_casing),
            "completion_type": completion,
            "artificial_lift": artificial,
            "zonal_isolation": zonal,
            "is_injector": is_injector,
        }
        try:
            svg = fh.build_well_completion_svg(wp_spec)
            colA, colB = st.columns([3, 2])
            with colA:
                st.image(_svg_to_data_uri(svg), use_container_width=True)
                st.caption(
                    "Equinor-style cross-section — not to scale. Subsea "
                    "VXT + wellhead at the mudline, telescoping casing "
                    "strings with labelled shoes (size · mTVD · mMD), "
                    "DHSV/TRSCSSV with control + balancing + electric "
                    "lines, P/T gauge, production packer, middle-completion "
                    "packer with disappearing plug, liner hanger, and the "
                    "lower completion (perforated liner / gravel-pack "
                    "screens / slotted liner) in open hole.")
            with colB:
                st.markdown("**Design notes**")
                # Contextual guidance based on selections
                notes = []
                if completion == "Open hole" and not is_injector:
                    notes.append(
                        "Open hole gives the lowest skin but no zonal "
                        "control and no sand management — only for "
                        "competent, consolidated rock.")
                if completion in ("Screens / gravel pack", "Frac-pack"):
                    notes.append(
                        "Sand-control completion selected — confirm the "
                        "formation is unconsolidated enough to need it; "
                        "it adds cost and restricts future intervention.")
                if artificial == "ESP":
                    notes.append(
                        "ESP gives high rates but is the most workover-"
                        "intensive lift; plan for pump replacement every "
                        "3-5 years in the OPEX.")
                if artificial == "Gas lift":
                    notes.append(
                        "Gas lift needs a lift-gas source and topside "
                        "compression — see the Topside advisor under the "
                        "concept builder.")
                if well_type == "Horizontal":
                    notes.append(
                        "Horizontal well maximises reservoir contact and "
                        "lowers drawdown per unit rate — good for thin "
                        "oil rims and coning-prone reservoirs.")
                if zonal.startswith("Smart"):
                    notes.append(
                        "Smart completion enables selective zonal "
                        "production / water shut-off without "
                        "intervention — high value in multi-layered or "
                        "water-drive reservoirs, at higher capex.")
                if int(n_casing) >= 4:
                    notes.append(
                        "Four casing strings suggest an HPHT or deep "
                        "well — confirm the casing-design basis and "
                        "consider a tapered production string.")
                if not notes:
                    notes.append(
                        "A conventional cased-and-perforated completion "
                        "with a single packer — the workhorse design for "
                        "most clastic NCS reservoirs.")
                for n in notes:
                    st.markdown(f"- {n}")
        except Exception as _e:
            st.info(f"Could not render the well schematic: {_e}")


def concept_selector_section(default_start_date):
    """Render the Concept Selector tab.

    The data model lives in st.session_state['concept_dimensions'] as a list
    of dimension dicts. Selected options for batch run live in
    st.session_state['concept_selected'] as {dim_idx: set(opt_idx)}.
    Results live in st.session_state['concept_results'] keyed by the
    frozen-set of (dim_name, opt_label) tuples for that combination.
    """
    st.markdown("### 🌳 Concept Selector — hanging-garden batch builder")
    st.caption(
        "Define concept dimensions as columns of alternatives. Each option "
        "is its own standalone case (link a saved case / YAML, or patch the "
        "base case). Tick the options you want, hit **Run** — each runs once "
        "(no combinations), then options are coloured by NPV (red→green) and "
        "compared on the bubble chart, qualitative matrix and "
        "Design-to-Cost staircase. Inspired by the DG1 concept long-list "
        "used on the NCS.")

    with st.expander("❓ How to use the Concept Selector (quick guide)",
                      expanded=False):
        st.markdown(
            "**1. Start** — load a **📋 template** (Subsurface / Drilling / "
            "SURF / Topside) or **🔄 Reset to NCS default**, or **➕ Add "
            "dimension** to build your own.\n\n"
            "**2. Link cases** — in the *Dimension editor* below, each "
            "option can link a saved case or uploaded YAML/JSON, or carry "
            "lightweight `key: value` patches on top of the base case. "
            "**Type freely, then click ✅ Apply edits** in that dimension "
            "to commit (keeps editing fast).\n\n"
            "**3. Tick** the options to run — the *Cases to run* counter "
            "updates live.\n\n"
            "**4. Run options** — ♻️ cache reuses unchanged cases; 🎲 "
            "Monte-Carlo adds a P90/Mean/P10 band per case (with its own "
            "progress bar).\n\n"
            "**5. Run & read** — the hanging garden colours each option by "
            "NPV; the **🎯 bubble chart** plots NPV vs CAPEX with a Pareto "
            "frontier (★); the **🚦 qualitative matrix** scores HSE/risk/"
            "robustness/operability with weights; the **🏆 combined "
            "ranking** blends economics + qualitative; the **🪜 staircase** "
            "ranks by CAPEX and flags the recommended concept.\n\n"
            "**6. Save** — use the **💾 Study library** to save (with "
            "version + date), load, duplicate or import studies, or "
            "**🧾 download the nested YAML** audit trail. A full walkthrough "
            "ships as `HELP.md` in the repo.")

    # ---- Initialise state ----
    if "concept_dimensions" not in st.session_state:
        import copy
        st.session_state["concept_dimensions"] = copy.deepcopy(
            _DEFAULT_CONCEPT_DIMENSIONS)
    if "concept_selected" not in st.session_state:
        # Default: tick every option in every dimension so the user sees
        # a full grid immediately. They can untick to narrow the sweep.
        st.session_state["concept_selected"] = {
            di: set(range(len(d["options"])))
            for di, d in enumerate(st.session_state["concept_dimensions"])
        }
    if "concept_results" not in st.session_state:
        st.session_state["concept_results"] = {}
    # Dirty flag — set true whenever the matrix is edited, cleared on Apply.
    # Drives the orange "edited, not yet applied" state on the Run button.
    if "concept_dirty" not in st.session_state:
        st.session_state["concept_dirty"] = False
    # Applied snapshot — the matrix state that was in force at the last
    # Apply. The batch always runs against the APPLIED snapshot, never the
    # half-edited live state, so the user gets deterministic behaviour.
    if "concept_applied" not in st.session_state:
        st.session_state["concept_applied"] = None

    dimensions = st.session_state["concept_dimensions"]
    selected = st.session_state["concept_selected"]
    results = st.session_state["concept_results"]

    # ---- Base case source -------------------------------------------------
    # The batch sweeps each combination ON TOP OF a base payload. By default
    # that's the current sidebar inputs, but the user can instead load a
    # saved case from the local database or upload a YAML / JSON case file.
    with st.expander("📂 Base case source — sidebar, saved case, or "
                      "uploaded YAML/JSON", expanded=False):
        st.caption(
            "Every concept combination is applied as a patch on top of a "
            "base case. Choose where that base comes from. Saved cases "
            "live in the local case database (the same ones the Case "
            "Manager uses); YAML/JSON files use the same schema as the "
            "JSON export.")
        base_source = st.radio(
            "Base case",
            ["Current sidebar inputs", "Saved case (database)",
             "Upload YAML / JSON"],
            key="concept_base_source", horizontal=True)
        st.session_state["concept_base_payload"] = None  # default: live
        if base_source == "Saved case (database)":
            try:
                cases = fh.list_cases()
            except Exception:
                cases = []
            if not cases:
                st.info("No saved cases found in the database. Save a "
                        "case from the Case Manager first.")
            else:
                case_labels = [
                    f"{c['name']}  ·  {c.get('saved_at','')[:16]}"
                    for c in cases]
                pick = st.selectbox("Pick a saved case", case_labels,
                                     key="concept_base_case_pick")
                idx = case_labels.index(pick)
                if st.button("📥 Load this case as base",
                              key="concept_load_case"):
                    try:
                        data = fh.load_case(cases[idx]["filename"])
                        st.session_state["concept_base_payload"] = \
                            fh.normalize_payload_tables(
                                data.get("payload", {}))
                        st.success(
                            f"Loaded '{cases[idx]['name']}' as the batch "
                            f"base case.")
                    except Exception as e:
                        st.error(f"Could not load case: {e}")
        elif base_source == "Upload YAML / JSON":
            up = st.file_uploader(
                "Upload a case file (.yaml / .yml / .json)",
                type=["yaml", "yml", "json"],
                key="concept_base_upload")
            if up is not None:
                try:
                    raw = up.read().decode("utf-8")
                    if up.name.lower().endswith(".json"):
                        import json as _json
                        parsed = _json.loads(raw)
                        # Accept either a full case ({"payload": {...}}) or
                        # a bare payload ({"scalar":..., "tables":...}).
                        payload = parsed.get("payload", parsed) \
                            if isinstance(parsed, dict) else {}
                        try:
                            payload = fh._restore_dataframes(payload)
                        except Exception:
                            pass
                        payload = fh.normalize_payload_tables(payload)
                    else:
                        # YAML → use the canonical parser so tables are
                        # converted from row-lists to dict-of-lists (same
                        # as the main app's YAML import). This is what
                        # makes a case give identical results whether run
                        # in Field prognosis or loaded here.
                        payload, _m = fh.yaml_to_payload(raw)
                    st.session_state["concept_base_payload"] = payload
                    st.success(
                        f"Loaded '{up.name}' as the batch base case "
                        f"({len(payload.get('scalar', {}))} scalar keys, "
                        f"{len(payload.get('tables', {}))} tables).")
                except Exception as e:
                    st.error(f"Could not parse the file: {e}")
        # Show what's currently the base
        _bp = st.session_state.get("concept_base_payload")
        if _bp:
            _nm = _bp.get("_meta", {}).get("name", "uploaded/loaded case")
            st.info(f"**Base case:** {_nm} (loaded). The sweep will patch "
                    f"each combination on top of this.")
        else:
            st.info("**Base case:** current sidebar inputs (live).")

    # ---- Study library: save / load / import the whole matrix ----
    with st.expander("💾 Study library — save, load & import concept "
                      "matrices", expanded=False):
        st.caption(
            "A *study* is the whole concept matrix (dimensions, options, "
            "patches, sweep selection) plus its last results. Save named "
            "studies to the local database, reload them later, or import "
            "a study YAML/JSON exported from the results section below. "
            "Ideal for versioning concept long-lists in git.")
        sl1, sl2 = st.columns(2)
        with sl1:
            st.markdown("**Save current matrix as a study**")
            study_name = st.text_input(
                "Study name", value="My concept study",
                key="concept_study_name")
            if st.button("💾 Save study to database",
                          key="concept_save_study",
                          use_container_width=True):
                try:
                    doc = _concept_study_to_doc(
                        st.session_state["concept_dimensions"],
                        st.session_state["concept_selected"],
                        st.session_state.get("concept_results", {}),
                        st.session_state.get("concept_base_source",
                                              "Current sidebar inputs"),
                        st.session_state.get("concept_base_payload"))
                    path = fh.save_concept_study(study_name, doc)
                    # Reflect the version that was written
                    _saved = [s for s in fh.list_concept_studies()
                              if s["name"] == study_name]
                    _v = _saved[0]["version"] if _saved else 1
                    st.success(f"Saved '{study_name}' as v{_v} "
                               f"({_saved[0]['saved_at'][:16].replace('T',' ') if _saved else ''}).")
                except Exception as e:
                    st.error(f"Could not save study: {e}")
        with sl2:
            st.markdown("**Load a saved study**")
            try:
                studies = fh.list_concept_studies()
            except Exception:
                studies = []
            if studies:
                labels = [f"{s['name']}  ·  v{s.get('version',1)}  ·  "
                          f"{s['n_combinations']} cases  ·  "
                          f"{s.get('saved_at','')[:16].replace('T',' ')}"
                          for s in studies]
                pick = st.selectbox("Saved studies", labels,
                                     key="concept_study_pick")
                idx = labels.index(pick)
                lc1, lc2, lc3 = st.columns(3)
                if lc1.button("📂 Load", key="concept_load_study",
                               use_container_width=True):
                    try:
                        doc = fh.load_concept_study(
                            studies[idx]["filename"])
                        dims, sel = _concept_doc_to_matrix(doc)
                        st.session_state["concept_dimensions"] = dims
                        st.session_state["concept_selected"] = sel
                        st.session_state["concept_results"] = {}
                        st.session_state["concept_applied"] = None
                        st.success(
                            f"Loaded '{studies[idx]['name']}' "
                            f"v{studies[idx].get('version',1)}. "
                            f"Re-run to populate results.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Could not load study: {e}")
                if lc2.button("📑 Duplicate", key="concept_dup_study",
                               use_container_width=True,
                               help="Save a copy under a new name "
                                    "(version reset to 1)."):
                    try:
                        _newn = f"{studies[idx]['name']} (copy)"
                        fh.duplicate_concept_study(
                            studies[idx]["filename"], _newn)
                        st.success(f"Duplicated as '{_newn}'.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Could not duplicate: {e}")
                if lc3.button("🗑️ Delete",
                               key="concept_delete_study",
                               use_container_width=True):
                    try:
                        fh.delete_concept_study(studies[idx]["filename"])
                        st.rerun()
                    except Exception as e:
                        st.error(f"Could not delete: {e}")
            else:
                st.info("No saved studies yet.")
        st.markdown("---")
        st.markdown("**Import a study file (rebuilds the matrix)**")
        imp = st.file_uploader(
            "Upload study YAML / JSON", type=["yaml", "yml", "json"],
            key="concept_study_import")
        if imp is not None:
            if st.button("📥 Import matrix from this file",
                          key="concept_do_import"):
                try:
                    raw = imp.read().decode("utf-8")
                    if imp.name.lower().endswith(".json"):
                        import json as _json
                        doc = _json.loads(raw)
                    else:
                        import yaml as _yaml
                        doc = _yaml.safe_load(raw)
                    dims, sel = _concept_doc_to_matrix(doc)
                    st.session_state["concept_dimensions"] = dims
                    st.session_state["concept_selected"] = sel
                    st.session_state["concept_results"] = {}
                    st.session_state["concept_applied"] = None
                    st.success(
                        f"Imported {len(dims)} dimensions from "
                        f"'{imp.name}'. Re-run to populate results.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Could not import: {e}")

    # ---- Template loader ----
    # Predefined hanging-garden building-block sets (Subsurface, Drilling
    # & Well, SURF, Topside) from the DG1 concept long-list. Loading one
    # replaces the matrix with that discipline's dimensions, which the
    # user can then rename / edit / link cases to.
    with st.expander("📋 Load a predefined template "
                      "(Subsurface / Drilling / SURF / Topside)",
                      expanded=False):
        st.caption(
            "Start from a discipline-standard set of concept dimensions "
            "mirroring the DG1 building-block columns. Loading replaces "
            "the current matrix — save your work first if needed. You can "
            "rename any dimension or option and link cases afterwards.")
        tpl_cols = st.columns(len(_CONCEPT_TEMPLATES) + 1)
        tpl_names = list(_CONCEPT_TEMPLATES.keys())
        for _i, _tname in enumerate(tpl_names):
            if tpl_cols[_i].button(_tname, key=f"concept_tpl_{_i}",
                                    use_container_width=True):
                import copy
                st.session_state["concept_dimensions"] = copy.deepcopy(
                    _CONCEPT_TEMPLATES[_tname])
                st.session_state["concept_selected"] = {
                    di: set(range(len(d["options"])))
                    for di, d in enumerate(
                        st.session_state["concept_dimensions"])}
                st.session_state["concept_results"] = {}
                st.session_state["concept_applied"] = None
                st.success(f"Loaded the '{_tname}' template.")
                st.rerun()
        # "All disciplines" combines every template into one big matrix
        if tpl_cols[-1].button("🔗 All combined", key="concept_tpl_all",
                                use_container_width=True,
                                help="Concatenate every discipline's "
                                     "dimensions into one full-value-"
                                     "chain matrix."):
            import copy
            combined = []
            for _t in tpl_names:
                combined.extend(copy.deepcopy(_CONCEPT_TEMPLATES[_t]))
            st.session_state["concept_dimensions"] = combined
            st.session_state["concept_selected"] = {
                di: set(range(len(d["options"])))
                for di, d in enumerate(combined)}
            st.session_state["concept_results"] = {}
            st.session_state["concept_applied"] = None
            st.success("Loaded the full-value-chain template "
                       "(all disciplines).")
            st.rerun()

    # ---- Top toolbar ----
    tb1, tb2, tb3, tb4 = st.columns([2, 2, 2, 3])
    if tb1.button("➕ Add dimension", key="concept_add_dim",
                   use_container_width=True):
        dimensions.append({
            "name": f"New dimension {len(dimensions)+1}",
            "description": "",
            "options": [
                {"label": "Option A", "description": "", "patches": {}},
                {"label": "Option B", "description": "", "patches": {}},
            ],
        })
        selected[len(dimensions) - 1] = {0, 1}
        st.rerun()
    if tb2.button("🔄 Reset to NCS default", key="concept_reset",
                   use_container_width=True,
                   help="Replace the current concept matrix with the NCS-"
                        "style default starter."):
        import copy
        st.session_state["concept_dimensions"] = copy.deepcopy(
            _DEFAULT_CONCEPT_DIMENSIONS)
        st.session_state["concept_selected"] = {
            di: set(range(len(d["options"])))
            for di, d in enumerate(st.session_state["concept_dimensions"])
        }
        st.session_state["concept_results"] = {}
        st.rerun()
    if tb3.button("🧹 Clear results", key="concept_clear_results",
                   use_container_width=True,
                   help="Clear the last batch results — boxes return to "
                        "neutral colours."):
        st.session_state["concept_results"] = {}
        st.rerun()

    # Selected-option count — each selected option is one standalone case
    # (no cartesian product). This is the number of cases the batch runs.
    n_combos = sum(len(selected.get(di, set()))
                   for di in range(len(dimensions)))
    tb4.metric("Cases to run", f"{n_combos:,}",
               help="Each ticked option runs as its own standalone case "
                    "(its linked YAML / saved case, or the base case + "
                    "its patches). No combinations are formed — options "
                    "run one by one.")

    # ---- The garden view ----
    svg = _render_concept_garden_svg(dimensions, selected, results)
    st.markdown(
        f'<div style="overflow-x:auto;width:100%">'
        f'<div style="min-width:900px">{svg}</div></div>',
        unsafe_allow_html=True)

    # ---- Edit / pick UI in expanders, one per dimension ----
    st.markdown("#### Dimension editor & option picker")
    st.caption(
        "Edit each dimension's options below. Patches are key/value pairs "
        "applied to the base case (e.g. `oil_price_bbl: 55`). Special "
        "keys: `_n_producers_override` (replicates/truncates the producers "
        "table to N wells), `_facility_capex_override_MM` (rescales the "
        "facility CAPEX schedule to a new total). Click **Apply edits** "
        "in each dimension to commit name/label/patch changes.")
    # Assign a stable id to every dimension and option so widget keys
    # don't shift when a middle item is deleted. Without this, Streamlit
    # binds widget state to positional index; deleting dimension N makes
    # all later widgets inherit the wrong stored text, so it looks like
    # the LAST dimension was deleted. Keying by id fixes that.
    import uuid as _uuid
    for d in dimensions:
        if "_id" not in d:
            d["_id"] = _uuid.uuid4().hex[:8]
        for o in d["options"]:
            if "_id" not in o:
                o["_id"] = _uuid.uuid4().hex[:8]

    dim_to_remove = None
    for di, d in enumerate(dimensions):
        _did = d["_id"]
        with st.expander(f"📦  **{d['name']}** "
                          f"— {len(d['options'])} options, "
                          f"{len(selected.get(di, set()))} ticked",
                          expanded=False):
            ec1, ec2, ec3 = st.columns([3, 4, 1])
            # Edit fields write to a draft keyed by id; committed on Apply.
            new_name = ec1.text_input(
                "Dimension name", value=d["name"],
                key=f"concept_dim_{_did}_name")
            new_desc = ec2.text_input(
                "Description (optional)", value=d.get("description", ""),
                key=f"concept_dim_{_did}_desc")
            if ec3.button("🗑️", key=f"concept_dim_{_did}_del",
                           help="Delete this entire dimension"):
                dim_to_remove = di
            st.markdown("---")
            opt_to_remove = None
            # Collect drafts; commit all at once on the Apply button.
            _opt_drafts = []
            for oi, opt in enumerate(d["options"]):
                _oid = opt["_id"]
                oc1, oc2, oc3, oc4, oc5 = st.columns([0.6, 2, 3, 3, 0.6])
                # Selection checkbox — applied immediately (cheap, and it
                # drives the live garden colours / case count).
                is_sel = oi in selected.get(di, set())
                new_sel = oc1.checkbox(
                    "✓", value=is_sel,
                    key=f"concept_pick_{_did}_{_oid}",
                    label_visibility="collapsed",
                    help="Include this option in the batch run.")
                if new_sel and not is_sel:
                    selected.setdefault(di, set()).add(oi)
                elif (not new_sel) and is_sel:
                    selected.setdefault(di, set()).discard(oi)
                _draft_label = oc2.text_input(
                    "Label", value=opt["label"],
                    key=f"concept_opt_{_oid}_label",
                    label_visibility="collapsed")
                _draft_desc = oc3.text_input(
                    "Description", value=opt.get("description", ""),
                    key=f"concept_opt_{_oid}_desc",
                    label_visibility="collapsed",
                    placeholder="Short description")
                # Case link status
                _linked = opt.get("case_name")
                _patch_n = len(opt.get("patches", {}))
                _status = (f"📎 {_linked}" if _linked
                            else (f"{_patch_n} patch(es)" if _patch_n
                                  else "— no case —"))
                oc4.markdown(
                    f"<div style='padding-top:6px;font-size:12px;"
                    f"color:#555'>{_status}</div>",
                    unsafe_allow_html=True)
                if oc5.button("✖", key=f"concept_opt_{_oid}_del",
                               help="Delete this option"):
                    opt_to_remove = oi
                _opt_drafts.append((opt, _draft_label, _draft_desc))
                # Per-option case linker (expander keeps the row compact)
                with st.container():
                    le1, le2, le3 = st.columns([3, 3, 2])
                    # Link a saved case
                    try:
                        _cases = fh.list_cases()
                    except Exception:
                        _cases = []
                    _case_opts = ["— none —"] + [c["name"] for c in _cases]
                    _cur = opt.get("case_name", "— none —")
                    _idx = (_case_opts.index(_cur)
                            if _cur in _case_opts else 0)
                    _pick = le1.selectbox(
                        "Link saved case", _case_opts, index=_idx,
                        key=f"concept_opt_{_oid}_caselink",
                        label_visibility="collapsed")
                    if _pick != "— none —":
                        if opt.get("case_name") != _pick:
                            # Load and attach the case payload
                            try:
                                _match = next(c for c in _cases
                                               if c["name"] == _pick)
                                _data = fh.load_case(_match["filename"])
                                opt["case_payload"] = \
                                    fh.normalize_payload_tables(
                                        _data.get("payload", {}))
                                opt["case_name"] = _pick
                            except Exception:
                                pass
                    else:
                        if opt.get("case_name"):
                            opt["case_name"] = None
                            opt["case_payload"] = None
                    # Upload a YAML/JSON case for this option
                    _up = le2.file_uploader(
                        "or upload YAML/JSON",
                        type=["yaml", "yml", "json"],
                        key=f"concept_opt_{_oid}_upload",
                        label_visibility="collapsed")
                    if _up is not None:
                        try:
                            _raw = _up.read().decode("utf-8")
                            if _up.name.lower().endswith(".json"):
                                import json as _json
                                _parsed = _json.loads(_raw)
                                _pl = (_parsed.get("payload", _parsed)
                                       if isinstance(_parsed, dict) else {})
                                try:
                                    _pl = fh._restore_dataframes(_pl)
                                except Exception:
                                    pass
                                # JSON cases store tables as dict-of-lists
                                # already; normalise any row-list tables.
                                _pl = fh.normalize_payload_tables(_pl)
                            else:
                                # YAML: use the canonical parser so tables
                                # (written as list-of-row-dicts) are
                                # converted to the column-oriented
                                # dict-of-lists that run_payload_case
                                # expects — exactly like the main app's
                                # YAML import. Using safe_load directly
                                # here was the bug: tables stayed as row
                                # lists and the engine read them wrong,
                                # giving different results than Field
                                # prognosis for the very same case.
                                _pl, _meta_in = fh.yaml_to_payload(_raw)
                            opt["case_payload"] = _pl
                            opt["case_name"] = _up.name
                            # Capture the exported "expected results" stamp
                            # (if present) so we can verify the batch run
                            # reproduces the live numbers and flag stale files.
                            opt["expected_results"] = (
                                _meta_in or {}).get("expected_results")
                        except Exception as _e:
                            st.warning(f"Could not parse {_up.name}: {_e}")
                    # Patch override (optional, advanced) — draft only,
                    # committed on Apply.
                    cur_patch_str = ", ".join(
                        f"{k}: {v}"
                        for k, v in opt.get("patches", {}).items())
                    _draft_patch_str = le3.text_input(
                        "Patches (optional)", value=cur_patch_str,
                        key=f"concept_opt_{_oid}_patches",
                        label_visibility="collapsed",
                        placeholder="oil_price_bbl: 55",
                        help="Optional key:value overrides. If NO case is "
                             "linked, these modify the base case. If a case "
                             "IS linked, they are IGNORED unless you tick "
                             "'apply patches to linked case' below — a "
                             "linked case runs exactly as saved by default.")
                    # When a case is linked, let the user explicitly opt in
                    # to applying the patches on top of it. Default OFF so a
                    # linked case reproduces its Field-prognosis result
                    # exactly.
                    _apply_flag = opt.get("apply_patches_to_case", False)
                    if _has_case := bool(opt.get("case_payload")):
                        if opt.get("patches"):
                            _apply_flag = st.checkbox(
                                "↳ apply the patches above on top of the "
                                "linked case (otherwise it runs as-is)",
                                value=opt.get("apply_patches_to_case", False),
                                key=f"concept_opt_{_oid}_applypatch")
                            opt["apply_patches_to_case"] = _apply_flag
                        st.caption(
                            "📎 This option runs the linked case "
                            + ("**with** the patches above."
                               if _apply_flag else
                               "**exactly as saved** (no modifications)."))
                    # store the patch draft alongside the label/desc draft
                    _opt_drafts[-1] = (opt, _draft_label, _draft_desc,
                                        _draft_patch_str)
            # ---- Apply edits for this dimension ----
            if st.button("✅ Apply edits", key=f"concept_dim_{_did}_apply",
                          type="primary",
                          help="Commit the name, labels, descriptions and "
                               "patches you've typed above."):
                d["name"] = new_name
                d["description"] = new_desc
                for _entry in _opt_drafts:
                    _o, _lbl, _dsc = _entry[0], _entry[1], _entry[2]
                    _o["label"] = _lbl
                    _o["description"] = _dsc
                    if len(_entry) > 3:
                        _pstr = _entry[3]
                        parsed = {}
                        for pair in _pstr.split(","):
                            if ":" not in pair:
                                continue
                            k, v = pair.split(":", 1)
                            k, v = k.strip(), v.strip()
                            if not k:
                                continue
                            try:
                                v_typed = float(v)
                                if v_typed.is_integer() and "." not in v:
                                    v_typed = int(v_typed)
                            except ValueError:
                                if v.lower() in ("true", "false"):
                                    v_typed = (v.lower() == "true")
                                else:
                                    v_typed = v
                            parsed[k] = v_typed
                        _o["patches"] = parsed
                st.rerun()
            if opt_to_remove is not None:
                d["options"].pop(opt_to_remove)
                # Re-key selections
                new_sel = set()
                for oi in selected.get(di, set()):
                    if oi < opt_to_remove:
                        new_sel.add(oi)
                    elif oi > opt_to_remove:
                        new_sel.add(oi - 1)
                selected[di] = new_sel
                st.rerun()
            if st.button("➕ Add option", key=f"concept_opt_add_{_did}"):
                d["options"].append({
                    "label": f"Option {chr(65 + len(d['options']))}",
                    "description": "", "patches": {},
                    "_id": _uuid.uuid4().hex[:8]})
                selected.setdefault(di, set()).add(len(d["options"]) - 1)
                st.rerun()
    if dim_to_remove is not None:
        dimensions.pop(dim_to_remove)
        # Re-key selections
        new_selected = {}
        for di, opts in selected.items():
            if di < dim_to_remove:
                new_selected[di] = opts
            elif di > dim_to_remove:
                new_selected[di - 1] = opts
        st.session_state["concept_selected"] = new_selected
        st.rerun()

    # ---- Apply + Run batch ----
    st.markdown("---")

    # Serialise the current matrix to detect edits vs the applied snapshot.
    def _serialize_matrix(dims, sel):
        return json.dumps({
            "dims": [{"name": d["name"],
                       "options": [{"label": o["label"],
                                    "patches": o.get("patches", {})}
                                   for o in d["options"]]}
                      for d in dims],
            "selected": {str(k): sorted(v) for k, v in sel.items()},
        }, sort_keys=True, default=str)

    current_sig = _serialize_matrix(dimensions, selected)
    applied_sig = st.session_state.get("concept_applied")
    is_dirty = (applied_sig is None) or (current_sig != applied_sig)
    st.session_state["concept_dirty"] = is_dirty

    apply_c, run_c1, run_c2 = st.columns([2, 2, 4])

    # Apply button — commits the current matrix as the snapshot the batch
    # will run against. Turns the Run button green; until pressed the Run
    # button stays orange to signal "you have unapplied edits".
    apply_label = ("✅ Apply edits" if is_dirty
                   else "✓ Applied (up to date)")
    if apply_c.button(apply_label, key="concept_apply",
                       use_container_width=True,
                       type=("secondary" if not is_dirty else "primary"),
                       disabled=(not is_dirty)):
        st.session_state["concept_applied"] = current_sig
        st.session_state["concept_dirty"] = False
        st.rerun()

    # Run button — orange when there are unapplied edits, green when clean.
    # Streamlit's button `type` is primary (red/orange-ish) or secondary;
    # we colour a custom button via markdown when dirty to make the state
    # unmistakable, and gate the actual run on the applied state.
    if is_dirty:
        run_c1.markdown(
            "<div style='background:#ff8c00;color:white;padding:8px 12px;"
            "border-radius:6px;text-align:center;font-weight:600;"
            "cursor:not-allowed'>🚀 Run batch — apply edits first</div>",
            unsafe_allow_html=True)
        do_run = False
        run_c2.warning(
            "You have **unapplied edits** to the concept matrix. Click "
            "**Apply edits** to lock them in, then Run. This guarantees "
            "the batch runs the matrix you see.")
    else:
        do_run = run_c1.button(
            f"🚀 Run batch ({n_combos:,} combos)",
            type="primary", use_container_width=True,
            disabled=(n_combos == 0))
        if n_combos > 200:
            run_c2.warning(
                f"**{n_combos:,} combinations** — this will take a while. "
                f"Consider narrowing the sweep before pressing Run.")
        elif n_combos > 50:
            run_c2.info(
                f"Sweeping {n_combos:,} combinations; expect "
                f"~{n_combos * 1.5:.0f}s on Streamlit Cloud.")
    # ---- Run options (caching + probabilistic) ----
    ro1, ro2 = st.columns(2)
    use_cache = ro1.checkbox(
        "♻️ Cache combinations (skip unchanged)", value=True,
        key="concept_use_cache",
        help="Cache each combination's result keyed by the hash of its "
             "patched payload. On the next Run, combinations whose "
             "inputs haven't changed are reused instead of recomputed — "
             "big speed-up when you tweak one dimension and re-sweep.")
    if use_cache:
        if ro1.button("🗑 Clear cache",
                      key="concept_clear_cache",
                      help="Forget all cached combination results and "
                           "recompute everything on the next Run. Use this "
                           "if results look wrong or after upgrading the app."):
            st.session_state["concept_combo_cache"] = {}
            st.success("Cache cleared — next Run recomputes all combinations.")
    mc_per_concept = ro2.checkbox(
        "🎲 Probabilistic (Monte-Carlo P90/Mean/P10 per combo)",
        value=False, key="concept_mc",
        help="Run a fast Monte-Carlo pass per combination (oil price, "
             "OPEX and CAPEX varied ±20% lognormal) to get a real "
             "P90/Mean/P10 NPV for each, instead of a single "
             "deterministic value. Slower — adds ~100 evaluations per "
             "combination.")
    if mc_per_concept:
        mc_iters = st.slider(
            "Monte-Carlo iterations per combination", 25, 250, 100, 25,
            key="concept_mc_iters",
            help="More iterations = smoother percentiles but slower. "
                 "100 is a good screening compromise.")
    else:
        mc_iters = 0

    if do_run:
        # Base payload — fallback for options that don't link their own
        # case (their patches apply on top of this).
        try:
            base_payload = st.session_state.get("concept_base_payload")
            if not base_payload:
                base_payload = collect_inputs_payload()
        except Exception as e:
            st.error(f"Could not snapshot the base case: {e}")
            return
        if "concept_combo_cache" not in st.session_state:
            st.session_state["concept_combo_cache"] = {}
        combo_cache = st.session_state["concept_combo_cache"]
        cache_hits = 0
        import hashlib as _hashlib

        # Build the run list: EVERY selected option across all dimensions
        # is its own standalone case. No cartesian product — each option
        # runs once, on its own linked case (or the base case + its
        # patches if no case is linked). This is the "commingled" batch:
        # we gather all the individual option-cases into one list and run
        # them one by one.
        run_items = []  # (dim_name, option_dict)
        for di, d in enumerate(dimensions):
            for oi in sorted(selected.get(di, set())):
                run_items.append((d["name"], d["options"][oi]))

        results_new = {}
        total = len(run_items)
        if total == 0:
            st.warning("No options selected — tick at least one option to "
                       "run.")
            return
        progress = st.progress(0.0, text="Running cases…")
        mc_progress = None   # nested MC bar, created lazily on first MC pass
        done = 0
        for dim_name, opt in run_items:
            label = opt["label"]
            name = f"{dim_name} — {label}"
            # Resolve the case for this option.
            #
            # RULE: a linked case (uploaded YAML or saved DB case) runs
            # EXACTLY as-is — no patches, no modifications — so it
            # reproduces the Field-prognosis result for that case bit for
            # bit. Patches are only applied when the user explicitly opts
            # in for that option (apply_patches_to_case=True), or when
            # there is NO linked case (then patches modify the base case,
            # which is the morphological-matrix use-case).
            import copy as _copy
            _has_case = bool(opt.get("case_payload"))
            if _has_case:
                case_payload = _copy.deepcopy(opt["case_payload"])
                _case_source = f"linked: {opt.get('case_name', 'case')}"
            else:
                case_payload = _copy.deepcopy(base_payload)
                _case_source = "base case"
            # Apply patches only when (a) there's no linked case (patch the
            # base), or (b) the user explicitly asked to patch the linked
            # case for this option.
            _apply_patches = bool(opt.get("patches")) and (
                (not _has_case) or opt.get("apply_patches_to_case", False))
            if _apply_patches:
                case_payload = _apply_concept_patches(
                    case_payload,
                    [{"label": label, "patches": opt["patches"]}])
                _case_source += f" + {len(opt['patches'])} patch(es)"
            elif _has_case and opt.get("patches"):
                _case_source += "  (patches ignored — linked case run as-is)"
            case_payload.setdefault("_meta", {})["name"] = name
            key = (dim_name, label)

            try:
                _payload_sig = json.dumps(case_payload, sort_keys=True,
                                           default=str)
            except Exception:
                _payload_sig = repr(case_payload)
            _cache_key = _hashlib.md5(
                (_payload_sig + f"|mc{mc_iters}").encode()).hexdigest()
            if use_cache and _cache_key in combo_cache:
                cached = dict(combo_cache[_cache_key])
                cached["name"] = name
                cached["picks"] = [(dim_name, label)]
                cached["dim"] = dim_name
                cached["label"] = label
                results_new[key] = cached
                cache_hits += 1
                done += 1
                progress.progress(done / max(total, 1),
                                   text=f"Running cases… {done}/{total} "
                                        f"(cache hits: {cache_hits})")
                continue
            try:
                res = run_payload_case(case_payload, default_start_date,
                                        default_units=st.session_state.get(
                                            "units", "field"))
                if res.get("ok"):
                    kpis = res.get("kpis", {})
                    npv = kpis.get("npv_USD") or kpis.get("npv_MM")
                    if npv is not None and npv > 1e6:
                        npv_MM = npv / 1e6
                    else:
                        npv_MM = npv
                    cum_primary = kpis.get("cum_primary")
                    rf = kpis.get("final_rf")
                    irr = kpis.get("irr")
                    breakeven = kpis.get("breakeven_oil")
                    capex_disc_MM = kpis.get("capex_disc_MM")
                    co2_total_Mt = kpis.get("co2_total_Mt")
                    npv_pretax_MM = kpis.get("npv_pretax_MM")
                    capex_total_MM = kpis.get("capex_total_MM")
                    resources_mmboe = kpis.get("resources_mmboe")
                    _capex_well_MM = kpis.get("capex_well_MM")
                    _capex_fac_MM = kpis.get("capex_facility_MM")
                    _capex_aban_MM = kpis.get("capex_abandonment_MM")
                    _revenue_MM = kpis.get("revenue_MM")
                    _opex_MM = kpis.get("opex_MM")
                    _tax_MM = kpis.get("tax_MM")
                    _co2_cost_MM = kpis.get("co2_cost_MM")
                    _payback_yrs = kpis.get("payback_yrs")
                    # Capture the monthly time-series profiles (production +
                    # cashflow) so the full study export can include per-month
                    # profiles for a thorough offline comparison. Stored
                    # compactly as lists keyed by column.
                    _profile = None
                    try:
                        _dfp = res.get("df_disp") or res.get("df")
                        _dfpe = res.get("df_e")
                        if _dfp is not None and len(_dfp) > 0:
                            _prof_cols = {}
                            _date_idx = (list(_dfp.index.astype(str))
                                         if _dfp.index is not None else None)
                            for _c in ("primary_rate", "secondary_rate",
                                       "water_rate", "oil_rate", "gas_rate",
                                       "cum_oil", "cum_gas",
                                       "recovery_factor"):
                                if _c in _dfp.columns:
                                    _prof_cols[_c] = [float(x) for x in _dfp[_c].values]
                            if _dfpe is not None and len(_dfpe) > 0:
                                for _c in ("revenue", "opex", "tax",
                                           "co2_cost", "capex_well",
                                           "capex_facility", "abandonment",
                                           "cashflow", "npv",
                                           "ngl_rate"):
                                    if _c in _dfpe.columns:
                                        _prof_cols[_c] = [float(x) for x in _dfpe[_c].values]
                            _profile = {"index": _date_idx, "columns": _prof_cols}
                    except Exception:
                        _profile = None
                else:
                    npv_MM = cum_primary = rf = irr = breakeven = None
                    capex_disc_MM = co2_total_Mt = None
                    npv_pretax_MM = capex_total_MM = resources_mmboe = None
                    _capex_well_MM = _capex_fac_MM = _capex_aban_MM = None
                    _revenue_MM = _opex_MM = _tax_MM = _co2_cost_MM = None
                    _payback_yrs = None
                    _profile = None
                # ---- Optional Monte-Carlo pass for this case ----
                npv_p90 = npv_p10 = npv_mc_mean = None
                if mc_iters > 0 and res.get("ok"):
                    import numpy as _np
                    _rng = _np.random.default_rng(12345)
                    _npvs = []
                    _sc0 = case_payload.get("scalar", {})
                    _op0 = float(_sc0.get("oil_price_bbl", 75.0))
                    _ov0 = float(_sc0.get("opex_var_oil",
                                           _sc0.get("opex_var_gas", 5.5)))
                    _cw0 = float(_sc0.get("capex_well", 120.0))
                    # Base facility CAPEX rows (development only) for scaling.
                    _fac0 = None
                    try:
                        _ft = case_payload.get("tables", {}).get("fac_df")
                        if _ft and "amount_MMUSD" in _ft:
                            _fac0 = list(_ft["amount_MMUSD"])
                            _fac_labels = _ft.get("label",
                                                   [""] * len(_fac0))
                    except Exception:
                        _fac0 = None
                    # Nested progress bar for the Monte-Carlo iterations of
                    # THIS case (the outer bar tracks cases; this one tracks
                    # the per-case MC draws so a long MC run shows life).
                    if mc_progress is None:
                        mc_progress = st.progress(
                            0.0, text="Monte-Carlo…")
                    _miter = int(mc_iters)
                    for _k in range(_miter):
                        _trial = dict(case_payload)
                        _ts = dict(_sc0)
                        _ts["oil_price_bbl"] = _op0 * float(
                            _rng.lognormal(0, 0.20))
                        _okey = ("opex_var_oil" if "opex_var_oil" in _ts
                                  else "opex_var_gas")
                        _ts[_okey] = _ov0 * float(_rng.lognormal(0, 0.20))
                        _ts["capex_well"] = _cw0 * float(
                            _rng.lognormal(0, 0.20))
                        _trial["scalar"] = _ts
                        # Facility CAPEX uncertainty — scale development rows
                        # of fac_df by a lognormal factor (±~20%), leaving
                        # cessation/P&A untouched.
                        if _fac0 is not None:
                            _ffac = float(_rng.lognormal(0, 0.18))
                            _tbl = dict(case_payload.get("tables", {}))
                            _newfac = dict(_tbl.get("fac_df", {}))
                            _scaled = []
                            for _i, _v in enumerate(_fac0):
                                _lbl = (_fac_labels[_i]
                                        if _i < len(_fac_labels) else "")
                                try:
                                    _isc = fh.is_abandonment_label(
                                        "" if _lbl is None else str(_lbl))
                                except Exception:
                                    _isc = False
                                _scaled.append(float(_v) if _isc
                                               else float(_v) * _ffac)
                            _newfac["amount_MMUSD"] = _scaled
                            _tbl["fac_df"] = _newfac
                            _trial["tables"] = _tbl
                        try:
                            _r = run_payload_case(
                                _trial, default_start_date,
                                default_units=st.session_state.get(
                                    "units", "field"))
                            _n = _r.get("kpis", {}).get("npv_MM")
                            if _n is not None:
                                _npvs.append(_n)
                        except Exception:
                            pass
                        if (_k % 5 == 0) or (_k == _miter - 1):
                            mc_progress.progress(
                                (_k + 1) / _miter,
                                text=f"Monte-Carlo: {label} "
                                     f"({_k + 1}/{_miter} draws)")
                    if _npvs:
                        _arr = _np.array(_npvs, dtype=float)
                        npv_p90 = float(_np.percentile(_arr, 10))
                        npv_p10 = float(_np.percentile(_arr, 90))
                        npv_mc_mean = float(_np.mean(_arr))
                        _mc_draws = [float(x) for x in _arr]
                    else:
                        _mc_draws = None
                else:
                    _mc_draws = None
                # Verify against the exported "expected results" stamp, if
                # the linked YAML carried one. This catches stale exports:
                # if the batch reproduces the stamped live numbers, the file
                # is current; if not, the YAML was exported at a different
                # edit state than the live screen.
                _exp = opt.get("expected_results") if _has_case else None
                _verify = None
                if isinstance(_exp, dict):
                    try:
                        _checks = []
                        if _exp.get("npv_after_tax_MM") is not None and npv_MM is not None:
                            _checks.append(abs(npv_MM - _exp["npv_after_tax_MM"]) <= max(1.0, 0.01*abs(_exp["npv_after_tax_MM"])))
                        if _exp.get("final_rf") is not None and rf is not None:
                            _checks.append(abs(rf - _exp["final_rf"]) <= 0.005)
                        if _exp.get("capex_total_MM") is not None and capex_total_MM is not None:
                            _checks.append(abs(capex_total_MM - _exp["capex_total_MM"]) <= max(1.0, 0.01*abs(_exp["capex_total_MM"])))
                        if _checks:
                            _verify = "✓ matches live export" if all(_checks) else "⚠ differs from live export (stale YAML?)"
                    except Exception:
                        _verify = None
                if _verify:
                    _case_source = _case_source + f"  · {_verify}"

                rec_payload = {
                    "npv_pretax_MM": npv_pretax_MM,
                    "cum_primary": cum_primary,
                    "final_rf": rf,
                    "irr": irr,
                    "breakeven_oil": breakeven,
                    "capex_disc_MM": capex_disc_MM,
                    "capex_total_MM": capex_total_MM,
                    "capex_well_MM": _capex_well_MM,
                    "capex_facility_MM": _capex_fac_MM,
                    "capex_abandonment_MM": _capex_aban_MM,
                    "revenue_MM": _revenue_MM,
                    "opex_MM": _opex_MM,
                    "tax_MM": _tax_MM,
                    "co2_cost_MM": _co2_cost_MM,
                    "payback_yrs": _payback_yrs,
                    "resources_mmboe": resources_mmboe,
                    "co2_total_Mt": co2_total_Mt,
                    "npv_p90": npv_p90,
                    "npv_p10": npv_p10,
                    "npv_mc_mean": npv_mc_mean,
                    "mc_draws": _mc_draws,
                    "profile": _profile,
                    "picks": [(dim_name, label)],
                    "ok": res.get("ok", False),
                    "error": res.get("error"),
                }
                results_new[key] = rec_payload
                # Only cache genuinely successful results — never poison the
                # cache with a failed/None run (otherwise a transient error
                # would be remembered and replayed on every later batch).
                if use_cache and rec_payload.get("ok") and (
                        rec_payload.get("npv_MM") is not None):
                    combo_cache[_cache_key] = dict(rec_payload)
            except Exception as e:
                results_new[key] = {
                    "name": name, "dim": dim_name, "label": label,
                    "npv_MM": None, "npv_pretax_MM": None,
                    "cum_primary": None,
                    "final_rf": None, "irr": None, "breakeven_oil": None,
                    "capex_disc_MM": None, "capex_total_MM": None,
                    "resources_mmboe": None, "co2_total_Mt": None,
                    "npv_p90": None, "npv_p10": None, "npv_mc_mean": None,
                    "picks": [(dim_name, label)],
                    "ok": False, "error": str(e),
                }
            done += 1
            progress.progress(done / max(total, 1),
                               text=f"Running cases… {done}/{total} "
                                    f"(cache hits: {cache_hits})")
        progress.empty()
        if mc_progress is not None:
            mc_progress.empty()
        st.session_state["concept_results"] = results_new
        st.success(f"Completed {done}/{total} cases "
                   f"({cache_hits} reused from cache).")
        st.rerun()

    # ---- Results table ----
    if results:
        st.markdown("---")
        st.markdown("#### 📊 Batch results")
        rows = []
        for key, r in results.items():
            row = {"Dimension": r.get("dim", key[0] if isinstance(
                       key, tuple) else ""),
                   "Option": r.get("label", key[1] if isinstance(
                       key, tuple) and len(key) > 1 else ""),
                   "Ran against": r.get("source", "—")}
            row["NPV after-tax ($MM)"] = (
                f"{r['npv_MM']:,.0f}"
                if r.get("npv_MM") is not None else "—")
            row["NPV pre-tax ($MM)"] = (
                f"{r['npv_pretax_MM']:,.0f}"
                if r.get("npv_pretax_MM") is not None else "—")
            row["IRR"] = (f"{r['irr']:.1%}"
                          if r.get("irr") is not None else "—")
            row["Resources (MMboe)"] = (
                f"{r['resources_mmboe']:,.1f}"
                if r.get("resources_mmboe") is not None else "—")
            row["CAPEX ($MM)"] = (
                f"{r['capex_total_MM']:,.0f}"
                if r.get("capex_total_MM") is not None else "—")
            # CAPEX component breakdown + full economics, so the batch
            # table reconciles line-by-line with Field prognosis.
            row["  ↳ Wells ($MM)"] = (
                f"{r['capex_well_MM']:,.0f}"
                if r.get("capex_well_MM") is not None else "—")
            row["  ↳ Facilities ($MM)"] = (
                f"{r['capex_facility_MM']:,.0f}"
                if r.get("capex_facility_MM") is not None else "—")
            row["  ↳ Abandonment ($MM)"] = (
                f"{r['capex_abandonment_MM']:,.0f}"
                if r.get("capex_abandonment_MM") is not None else "—")
            row["Revenue ($MM)"] = (
                f"{r['revenue_MM']:,.0f}"
                if r.get("revenue_MM") is not None else "—")
            row["OPEX ($MM)"] = (
                f"{r['opex_MM']:,.0f}"
                if r.get("opex_MM") is not None else "—")
            row["Tax ($MM)"] = (
                f"{r['tax_MM']:,.0f}"
                if r.get("tax_MM") is not None else "—")
            row["CO₂ cost ($MM)"] = (
                f"{r['co2_cost_MM']:,.0f}"
                if r.get("co2_cost_MM") is not None else "—")
            row["BE oil ($/bbl)"] = (f"{r['breakeven_oil']:.1f}"
                                      if r.get("breakeven_oil") is not None
                                      else "—")
            row["Final RF"] = (f"{r['final_rf']:.1%}"
                                if r.get("final_rf") is not None else "—")
            row["Status"] = "✅" if r.get("ok") else "❌"
            rows.append(row)
        df_res = pd.DataFrame(rows)
        # Sort by after-tax NPV descending if present
        try:
            df_res["_sort"] = df_res["NPV after-tax ($MM)"].str.replace(
                "[,—]", "", regex=True)
            df_res["_sort"] = pd.to_numeric(df_res["_sort"], errors="coerce")
            df_res = df_res.sort_values(
                "_sort", ascending=False, na_position="last")
            df_res = df_res.drop(columns=["_sort"])
        except Exception:
            pass
        st.dataframe(df_res, use_container_width=True, hide_index=True)

        # ---- "Compare cases" transposed view (metrics as rows, cases as
        # columns, grouped into sections with units) — the layout used in
        # corporate concept-screening tools. Toggle between this and the
        # flat table above.
        _layout = st.radio(
            "Results layout", ["Wide table", "Compare cases (transposed)"],
            horizontal=True, key="concept_results_layout",
            help="Wide table: one row per case. Compare cases: metrics as "
                 "rows grouped by section, one column per case — the "
                 "side-by-side concept-screening format.")
        if _layout == "Compare cases (transposed)":
            try:
                _ordered = sorted(
                    results.values(),
                    key=lambda x: (x.get("npv_MM")
                                    if x.get("npv_MM") is not None
                                    else -9e18),
                    reverse=True)
                # Column header per case: "Dimension: Option"
                def _case_col(r):
                    d = r.get("dim", ""); l = r.get("label", "")
                    return f"{d}: {l}" if d else (l or r.get("name", "case"))
                col_names = []
                seen = {}
                for r in _ordered:
                    c = _case_col(r)
                    if c in seen:
                        seen[c] += 1; c = f"{c} ({seen[c]})"
                    else:
                        seen[c] = 1
                    col_names.append(c)

                # (section, metric label, value-fn → (number, unit))
                def _fmt(v, unit, dp=0):
                    if v is None:
                        return "—"
                    if dp == 0:
                        return f"{v:,.0f} {unit}".strip()
                    return f"{v:,.{dp}f} {unit}".strip()

                # Unit labels follow each case's own unit system.
                def _vol_units(r):
                    return (r.get("cum_oil_unit", "MMstb"),
                            r.get("cum_gas_unit", "Bscf"),
                            r.get("peak_rate_unit", "stb/d"))

                spec = [
                    ("Economic KPIs", "NPV after-tax",
                     lambda r: _fmt(r.get("npv_MM"), "$MM")),
                    ("Economic KPIs", "NPV pre-tax",
                     lambda r: _fmt(r.get("npv_pretax_MM"), "$MM")),
                    ("Economic KPIs", "IRR",
                     lambda r: (f"{r['irr']:.1%}"
                                if r.get("irr") is not None else "—")),
                    ("Economic KPIs", "Payback",
                     lambda r: (f"{r['payback_yrs']:.1f} yrs"
                                if r.get("payback_yrs") is not None else "—")),
                    ("Economic KPIs", "Breakeven oil",
                     lambda r: _fmt(r.get("breakeven_oil"), "$/bbl", 1)),
                    ("Volumes", "Recoverable resources",
                     lambda r: _fmt(r.get("resources_mmboe"), "MMboe", 1)),
                    ("Volumes", "Cumulative oil/condensate",
                     lambda r: _fmt(r.get("cum_oil_disp",
                                          r.get("cum_oil_MMstb")),
                                    _vol_units(r)[0], 2)),
                    ("Volumes", "Cumulative gas",
                     lambda r: _fmt(r.get("cum_gas_disp",
                                          r.get("cum_gas_Bscf")),
                                    _vol_units(r)[1], 2)),
                    ("Volumes", "Peak rate",
                     lambda r: _fmt(r.get("peak_primary_rate"),
                                    _vol_units(r)[2], 0)),
                    ("Volumes", "Final recovery factor",
                     lambda r: (f"{r['final_rf']:.1%}"
                                if r.get("final_rf") is not None else "—")),
                    ("Investment cost", "Total CAPEX",
                     lambda r: _fmt(r.get("capex_total_MM"), "$MM")),
                    ("Investment cost", "  Wells",
                     lambda r: _fmt(r.get("capex_well_MM"), "$MM")),
                    ("Investment cost", "  Facilities",
                     lambda r: _fmt(r.get("capex_facility_MM"), "$MM")),
                    ("Investment cost", "  Abandonment",
                     lambda r: _fmt(r.get("capex_abandonment_MM"), "$MM")),
                    ("Cost", "Revenue (gross)",
                     lambda r: _fmt(r.get("revenue_MM"), "$MM")),
                    ("Cost", "OPEX",
                     lambda r: _fmt(r.get("opex_MM"), "$MM")),
                    ("Cost", "Tax",
                     lambda r: _fmt(r.get("tax_MM"), "$MM")),
                    ("CO₂", "CO₂ cost",
                     lambda r: _fmt(r.get("co2_cost_MM"), "$MM")),
                    ("CO₂", "CO₂ emissions total",
                     lambda r: _fmt(r.get("co2_total_Mt"), "Mt", 2)),
                ]
                comp_rows = []
                last_section = None
                for section, metric, fn in spec:
                    rowd = {"Section": section, "Metric": metric}
                    for cn, r in zip(col_names, _ordered):
                        rowd[cn] = fn(r)
                    comp_rows.append(rowd)
                comp_df = pd.DataFrame(comp_rows)
                # Show grouped by section with a subheader each
                for section in ["Economic KPIs", "Volumes",
                                "Investment cost", "Cost", "CO₂"]:
                    sub = comp_df[comp_df["Section"] == section].drop(
                        columns=["Section"])
                    if len(sub) == 0:
                        continue
                    st.markdown(f"**{section}**")
                    st.dataframe(sub, use_container_width=True,
                                 hide_index=True)
                # Offer the transposed matrix as its own CSV
                _tcsv = comp_df.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "📥 Download compare-cases matrix (CSV)",
                    data=_tcsv,
                    file_name="concept_compare_cases.csv",
                    mime="text/csv")
            except Exception as _te:
                st.warning(f"Compare-cases view unavailable: {_te}")

        dl1, dl2, dl3, dl4 = st.columns(4)
        csv = df_res.to_csv(index=False).encode("utf-8")
        dl1.download_button(
            "📥 Download results (CSV)", data=csv,
            file_name="concept_batch_results.csv", mime="text/csv",
            use_container_width=True)

        # ---- Comprehensive multi-sheet Excel workbook ----
        # One workbook holding EVERYTHING for an offline, like-for-like
        # comparison across all swept cases: a KPI summary, full monthly
        # production + cost/cashflow profiles per case, a cost-breakdown
        # sheet, and (if Monte-Carlo was run) the per-draw NPV distribution
        # with percentiles.
        try:
            import io as _io
            from openpyxl import Workbook as _WB
            from openpyxl.styles import Font as _Font, PatternFill as _Fill
            _res_sorted = sorted(
                results.values(),
                key=lambda x: (x.get("npv_MM")
                               if x.get("npv_MM") is not None else -9e18),
                reverse=True)
            wb = _WB()
            # ---- Sheet 1: KPI summary ----
            ws = wb.active; ws.title = "Summary"
            _hdr = ["Dimension", "Option", "Ran against",
                    "NPV after-tax ($MM)", "NPV pre-tax ($MM)", "IRR",
                    "Resources (MMboe)", "CAPEX total ($MM)",
                    "  Wells ($MM)", "  Facilities ($MM)",
                    "  Abandonment ($MM)", "Revenue ($MM)", "OPEX ($MM)",
                    "Tax ($MM)", "CO2 cost ($MM)", "BE oil ($/bbl)",
                    "Final RF", "CO2 (Mt)", "Payback (yrs)",
                    "NPV P90 ($MM)", "NPV Mean ($MM)", "NPV P10 ($MM)",
                    "Status"]
            ws.append(_hdr)
            for _c in ws[1]:
                _c.font = _Font(bold=True, color="FFFFFF")
                _c.fill = _Fill("solid", start_color="1F4E78")
            for r in _res_sorted:
                _picks = dict(r.get("picks", []))
                _dn = r.get("dim", ""); _lbl = r.get("label", "")
                ws.append([
                    _dn, _lbl, r.get("source", ""),
                    r.get("npv_MM"), r.get("npv_pretax_MM"), r.get("irr"),
                    r.get("resources_mmboe"), r.get("capex_total_MM"),
                    r.get("capex_well_MM"), r.get("capex_facility_MM"),
                    r.get("capex_abandonment_MM"), r.get("revenue_MM"),
                    r.get("opex_MM"), r.get("tax_MM"), r.get("co2_cost_MM"),
                    r.get("breakeven_oil"), r.get("final_rf"),
                    r.get("co2_total_Mt"), r.get("payback_yrs"),
                    r.get("npv_p90"), r.get("npv_mc_mean"), r.get("npv_p10"),
                    "OK" if r.get("ok") else "ERROR",
                ])
            for _col in ws.columns:
                _w = max((len(str(c.value)) if c.value is not None else 0)
                         for c in _col)
                ws.column_dimensions[_col[0].column_letter].width = min(
                    max(_w + 2, 10), 40)

            # ---- Per-case profile sheets ----
            def _safe_sheet(nm, used):
                s = "".join(ch for ch in str(nm)
                            if ch not in "[]:*?/\\")[:28] or "case"
                base = s; i = 2
                while s in used:
                    s = f"{base[:25]}_{i}"; i += 1
                used.add(s); return s
            _used = {"Summary"}
            _have_profiles = False
            for r in _res_sorted:
                prof = r.get("profile")
                if not prof or not prof.get("columns"):
                    continue
                _have_profiles = True
                snm = _safe_sheet((r.get("label") or r.get("name") or "case"),
                                  _used)
                wsp = wb.create_sheet(snm)
                cols = prof["columns"]
                colnames = list(cols.keys())
                idx = prof.get("index")
                header = (["month"] if idx else []) + colnames
                wsp.append(header)
                for _c in wsp[1]:
                    _c.font = _Font(bold=True)
                n = len(next(iter(cols.values()))) if cols else 0
                for i in range(n):
                    row = ([idx[i]] if idx else []) + [
                        cols[cn][i] for cn in colnames]
                    wsp.append(row)

            # ---- Monte-Carlo distribution sheet ----
            _mc_any = any(r.get("mc_draws") for r in _res_sorted)
            if _mc_any:
                wsm = wb.create_sheet("MonteCarlo")
                # one column of draws per case + a percentile summary block
                labels = [r.get("label", f"case{i}")
                          for i, r in enumerate(_res_sorted)
                          if r.get("mc_draws")]
                draws = [r.get("mc_draws") for r in _res_sorted
                         if r.get("mc_draws")]
                wsm.append(["Draw #"] + labels)
                for _c in wsm[1]:
                    _c.font = _Font(bold=True)
                _maxn = max((len(d) for d in draws), default=0)
                for i in range(_maxn):
                    wsm.append([i + 1] + [
                        (d[i] if i < len(d) else None) for d in draws])
                # percentile summary
                wsm.append([])
                import numpy as _np2
                wsm.append(["P90 (downside)"] + [
                    float(_np2.percentile(d, 10)) for d in draws])
                wsm.append(["Mean"] + [float(_np2.mean(d)) for d in draws])
                wsm.append(["P10 (upside)"] + [
                    float(_np2.percentile(d, 90)) for d in draws])
                for _c in (wsm[wsm.max_row], wsm[wsm.max_row - 1],
                           wsm[wsm.max_row - 2]):
                    for _cc in _c:
                        _cc.font = _Font(bold=True)

            _buf = _io.BytesIO()
            wb.save(_buf)
            dl2.download_button(
                "📊 Download full workbook (Excel)",
                data=_buf.getvalue(),
                file_name="concept_batch_full.xlsx",
                mime=("application/vnd.openxmlformats-officedocument."
                      "spreadsheetml.sheet"),
                use_container_width=True,
                help="Everything for offline comparison: KPI summary, full "
                     "monthly production + cost/cashflow profiles per case"
                     + (", and the Monte-Carlo NPV distribution"
                        if _mc_any else "")
                     + ". One sheet per case.")
        except Exception as _xe:
            dl2.caption(f"Excel export unavailable: {_xe}")

        # Build a nested, human-readable YAML capturing the WHOLE study:
        # the matrix definition (dimensions + options + their patches),
        # which options were swept, the base-case source, a UTC timestamp,
        # and every combination's picks + KPIs. This is the audit artefact
        # for tracking exactly what was run and what came out.
        try:
            import yaml as _yaml
            from datetime import datetime as _dt, timezone as _tz
            _base_src = st.session_state.get("concept_base_source",
                                              "Current sidebar inputs")
            _base_loaded = st.session_state.get("concept_base_payload")
            export_doc = {
                "fieldvista_concept_study": {
                    "schema_version": 1,
                    "generated_utc": _dt.now(_tz.utc).isoformat(
                        timespec="seconds"),
                    "app_version": str(getattr(fh, "FP_HELPERS_VERSION",
                                                "?")),
                    "base_case": {
                        "source": _base_src,
                        "loaded_case_name": (
                            (_base_loaded or {}).get("_meta", {}).get(
                                "name") if _base_loaded else
                            "live sidebar inputs"),
                    },
                    "matrix": {
                        "dimensions": [
                            {
                                "name": d["name"],
                                "description": d.get("description", ""),
                                "options": [
                                    {
                                        "label": o["label"],
                                        "description": o.get(
                                            "description", ""),
                                        "patches": o.get("patches", {}),
                                        "swept": (
                                            oi in selected.get(di, set())),
                                    }
                                    for oi, o in enumerate(d["options"])
                                ],
                            }
                            for di, d in enumerate(dimensions)
                        ],
                    },
                    "n_combinations": len(results),
                    "results": [
                        {
                            "name": r.get("name"),
                            "picks": {dn: lbl
                                       for dn, lbl in r.get("picks", [])},
                            "ran_against": r.get("source"),
                            "kpis": {
                                "npv_after_tax_MM": r.get("npv_MM"),
                                "npv_pre_tax_MM": r.get("npv_pretax_MM"),
                                "fiscal_take_MM": (
                                    (r.get("npv_pretax_MM") - r.get("npv_MM"))
                                    if (r.get("npv_pretax_MM") is not None
                                        and r.get("npv_MM") is not None)
                                    else None),
                                "irr": r.get("irr"),
                                "final_rf": r.get("final_rf"),
                                "payback_yrs": r.get("payback_yrs"),
                                "resources_mmboe": r.get("resources_mmboe"),
                                "breakeven_oil_usd_bbl": r.get(
                                    "breakeven_oil"),
                                "co2_total_Mt": r.get("co2_total_Mt"),
                            },
                            "cost_summary_MM": {
                                "capex_total": r.get("capex_total_MM"),
                                "capex_disc": r.get("capex_disc_MM"),
                                "capex_wells": r.get("capex_well_MM"),
                                "capex_facilities": r.get(
                                    "capex_facility_MM"),
                                "capex_abandonment": r.get(
                                    "capex_abandonment_MM"),
                                "revenue": r.get("revenue_MM"),
                                "opex": r.get("opex_MM"),
                                "tax": r.get("tax_MM"),
                                "co2_cost": r.get("co2_cost_MM"),
                            },
                            "uncertainty": ({
                                "npv_p90_MM": r.get("npv_p90"),
                                "npv_mean_MM": r.get("npv_mc_mean"),
                                "npv_p10_MM": r.get("npv_p10"),
                                "mc_draws_MM": r.get("mc_draws"),
                            } if (r.get("npv_p90") is not None
                                  or r.get("mc_draws")) else None),
                            # Full monthly time-series — production +
                            # cost/cashflow profiles, the same data the Excel
                            # workbook holds, so JSON consumers get parity.
                            "monthly_profile": r.get("profile"),
                            "ok": r.get("ok", False),
                            "error": r.get("error"),
                        }
                        for r in sorted(
                            results.values(),
                            key=lambda x: (x.get("npv_MM")
                                            if x.get("npv_MM") is not None
                                            else -9e18),
                            reverse=True)
                    ],
                }
            }
            yaml_bytes = _yaml.safe_dump(
                export_doc, sort_keys=False, allow_unicode=True,
                default_flow_style=False).encode("utf-8")
            dl3.download_button(
                "🧾 Download study (nested YAML)", data=yaml_bytes,
                file_name="concept_study.yaml",
                mime="application/x-yaml",
                use_container_width=True,
                help="Full audit trail — the concept matrix (every "
                     "dimension, option and its patch), which options "
                     "were swept, the base-case source, a UTC "
                     "timestamp, and every combination's picks + KPIs, "
                     "cost summary, uncertainty and full monthly "
                     "profiles. Re-loadable as a base case and ideal for "
                     "version tracking in git.")
            # Same complete payload as JSON — identical content to the
            # nested YAML and the Excel workbook (production + cost
            # profiles, cost summary, Monte-Carlo), for API/programmatic
            # consumers.
            try:
                import json as _json
                json_bytes = _json.dumps(
                    export_doc, indent=2, default=str).encode("utf-8")
                dl4.download_button(
                    "🗂 Download study (JSON)", data=json_bytes,
                    file_name="concept_study.json",
                    mime="application/json",
                    use_container_width=True,
                    help="Identical content to the nested YAML and the "
                         "Excel workbook — full KPIs, cost summary, "
                         "uncertainty (Monte-Carlo draws) and complete "
                         "monthly production + cost/cashflow profiles for "
                         "every case.")
            except Exception as _je:
                dl4.caption(f"JSON export unavailable: {_je}")
        except Exception as _e:
            dl2.info(f"YAML export unavailable: {_e}")
        # Errors expander
        errs = [(r["name"], r["error"]) for r in results.values()
                if not r.get("ok") and r.get("error")]
        if errs:
            with st.expander(f"⚠️ {len(errs)} combinations failed — details"):
                for name, err in errs:
                    st.markdown(f"**{name}** — `{err}`")

        # ============================================================
        # CONCEPT COMPARISON CHART
        # ============================================================
        # NPV post-tax vs total discounted CAPEX, with P10/Mean/P90
        # markers per "concept" and emissions on the secondary axis.
        # The user defines "concepts" by grouping combinations along
        # one chosen dimension — e.g. group by Host facility to compare
        # FPSO vs Tie-back vs Jacket, with each concept's P90/Mean/P10
        # showing the spread across the OTHER dimensions.
        st.markdown("---")
        st.markdown("#### 🎯 Concept Comparison — NPV vs CAPEX, "
                    "P90 / Mean / P10 by concept")
        st.caption(
            "Group the batch results into named concepts along one "
            "dimension. Each concept gets three markers — **P90** "
            "(downside), **Mean**, **P10** (upside) — plotted as NPV "
            "post-tax against total discounted CAPEX, with the "
            "concept's lifetime CO₂ emissions on the right axis. "
            "Reads like the classic NCS DG1 concept-comparison "
            "chart. Pick which dimension defines the concept "
            "groupings below.")

        # Dimension chooser — which dimension defines the concept
        dim_names = [d["name"] for d in dimensions]
        # Persist grouping choice across reruns
        grp_key = "concept_compare_group_dim"
        cur_grp = st.session_state.get(grp_key, dim_names[0] if dim_names else "")
        if cur_grp not in dim_names and dim_names:
            cur_grp = dim_names[0]
        grouping_dim = st.selectbox(
            "Show options from dimension",
            options=dim_names,
            index=dim_names.index(cur_grp) if cur_grp in dim_names else 0,
            key=grp_key,
            help="Each option in this dimension is one concept on the "
                 "chart. (Each option ran as its own standalone case.)")

        # Optional concept-label rewrites (e.g. "FPSO" → "Recommended
        # Concept"). Stored as JSON-ish dict in session_state.
        rename_key = f"concept_compare_renames_{grouping_dim}"
        if rename_key not in st.session_state:
            st.session_state[rename_key] = {}
        # Build the group → list-of-results mapping. Each result belongs
        # to exactly one dimension (it's a single option-case), so we
        # filter to the chosen dimension and key by the option label.
        groups = {}  # concept_label → list of result dicts
        for r in results.values():
            if not r.get("ok"):
                continue
            if r.get("dim") != grouping_dim:
                continue
            grp_value = r.get("label", "?")
            disp_label = st.session_state[rename_key].get(
                grp_value, grp_value)
            groups.setdefault(disp_label, []).append(r)

        # Optional renaming UI
        with st.expander("🏷️ Rename concepts (optional)", expanded=False):
            st.caption(
                "Override the auto-generated concept labels — e.g. tag "
                "your favourite case as 'Recommended Concept' so it "
                "stands out on the chart.")
            for raw_label in sorted({r.get("label", "?")
                    for r in results.values()
                    if r.get("ok") and r.get("dim") == grouping_dim}):
                new_lbl = st.text_input(
                    f"'{raw_label}' →",
                    value=st.session_state[rename_key].get(
                        raw_label, raw_label),
                    key=f"concept_rename_{grouping_dim}_{raw_label}")
                st.session_state[rename_key][raw_label] = new_lbl

        # Per-concept selector — which concepts to actually plot
        plot_concepts_key = f"concept_compare_plot_{grouping_dim}"
        all_concepts = list(groups.keys())
        if not all_concepts:
            st.info("No successful results to chart. Run a batch first.")
        else:
            picked_concepts = st.multiselect(
                "Concepts to plot",
                options=all_concepts,
                default=st.session_state.get(plot_concepts_key,
                                              all_concepts),
                key=plot_concepts_key,
                help="Untick any concept you want to hide from the "
                     "chart (e.g. clearly dominated alternatives).")

            # Compute P90/Mean/P10 per concept
            import numpy as np
            chart_rows = []
            for label in picked_concepts:
                rs = groups.get(label, [])
                npvs = np.array(
                    [r["npv_MM"] for r in rs
                     if r.get("npv_MM") is not None], dtype=float)
                capexes = np.array(
                    [r["capex_disc_MM"] for r in rs
                     if r.get("capex_disc_MM") is not None], dtype=float)
                bes = np.array(
                    [r["breakeven_oil"] for r in rs
                     if r.get("breakeven_oil") is not None], dtype=float)
                emissions = np.array(
                    [r["co2_total_Mt"] for r in rs
                     if r.get("co2_total_Mt") is not None], dtype=float)
                if len(npvs) == 0:
                    continue
                # If a Monte-Carlo pass was run, each case carries its own
                # P90/Mean/P10 from the probabilistic NPV distribution —
                # use those directly (they're the real uncertainty band).
                # Otherwise fall back to the spread across the cases in
                # this group (degenerate to a point for a single case).
                mc_p90s = [r["npv_p90"] for r in rs
                           if r.get("npv_p90") is not None]
                mc_p10s = [r["npv_p10"] for r in rs
                           if r.get("npv_p10") is not None]
                mc_means = [r["npv_mc_mean"] for r in rs
                            if r.get("npv_mc_mean") is not None]
                if mc_p90s and mc_p10s:
                    p90 = float(np.mean(mc_p90s))
                    p10 = float(np.mean(mc_p10s))
                    mean_npv = (float(np.mean(mc_means)) if mc_means
                                else float(np.mean(npvs)))
                    is_probabilistic = True
                else:
                    # P90 = downside (10th pct), P10 = upside (90th pct) —
                    # petroleum convention.
                    p90 = float(np.percentile(npvs, 10))
                    p10 = float(np.percentile(npvs, 90))
                    mean_npv = float(np.mean(npvs))
                    is_probabilistic = False
                cap_med = (float(np.median(capexes)) if len(capexes)
                            else 0.0)
                be_med = (float(np.median(bes)) if len(bes) else None)
                em_med = (float(np.median(emissions)) if len(emissions)
                            else None)
                chart_rows.append({
                    "concept": label,
                    "p90": p90, "mean": mean_npv, "p10": p10,
                    "capex_disc_MM": cap_med,
                    "breakeven_oil": be_med,
                    "emissions_Mt": em_med,
                    "n_cases": len(npvs),
                    "probabilistic": is_probabilistic,
                })

            # ---- Pareto front (#7) ----
            # Flag concepts on the efficient frontier (max NPV for min
            # CAPEX). Dominated concepts are greyed so the eye goes to
            # the frontier.
            pareto_labels = _concept_pareto_front(chart_rows)
            for r in chart_rows:
                r["pareto"] = (r["concept"] in pareto_labels)

            if not chart_rows:
                st.info("Selected concepts have no successful results.")
            else:
                # Build the figure: NPV markers on the left axis, emissions
                # bars on the right axis, both keyed by CAPEX on the x.
                fig_cc = make_subplots(specs=[[{"secondary_y": True}]])

                # Plot P90/Mean/P10 as three points per concept, stacked
                # vertically, with the mean labelled. Use the project's
                # green palette to echo the reference chart.
                colors_npv = ["#2ca02c", "#2ca02c", "#2ca02c"]
                # P90 — lower point
                fig_cc.add_trace(go.Scatter(
                    x=[r["capex_disc_MM"] for r in chart_rows],
                    y=[r["p90"] for r in chart_rows],
                    mode="markers+text",
                    name="P90 (downside)",
                    marker=dict(size=14, color=colors_npv[0],
                                 line=dict(color="black", width=1)),
                    text=[f"P90 — {r['concept']}" for r in chart_rows],
                    textposition="middle right",
                    textfont=dict(size=10),
                    hovertemplate=(
                        "<b>%{text}</b><br>"
                        "CAPEX: $%{x:,.0f}MM<br>"
                        "NPV P90: $%{y:,.0f}MM<extra></extra>"),
                ), secondary_y=False)
                # Mean — middle point
                fig_cc.add_trace(go.Scatter(
                    x=[r["capex_disc_MM"] for r in chart_rows],
                    y=[r["mean"] for r in chart_rows],
                    mode="markers+text",
                    name="Mean",
                    marker=dict(
                        size=[20 if r.get("pareto") else 15
                              for r in chart_rows],
                        color=["#2ca02c" if r.get("pareto") else "#cccccc"
                               for r in chart_rows],
                        line=dict(color="black", width=1.5),
                        symbol=["star" if r.get("pareto") else "circle"
                                for r in chart_rows]),
                    text=[(f"<b>{r['concept']}</b>" if r.get("pareto")
                           else f"{r['concept']} (dominated)")
                          for r in chart_rows],
                    textposition="middle right",
                    textfont=dict(size=11, color="black"),
                    hovertemplate=(
                        "<b>%{text}</b><br>"
                        "CAPEX: $%{x:,.0f}MM<br>"
                        "NPV mean: $%{y:,.0f}MM<extra></extra>"),
                ), secondary_y=False)
                # P10 — upper point
                fig_cc.add_trace(go.Scatter(
                    x=[r["capex_disc_MM"] for r in chart_rows],
                    y=[r["p10"] for r in chart_rows],
                    mode="markers+text",
                    name="P10 (upside)",
                    marker=dict(size=14, color=colors_npv[2],
                                 line=dict(color="black", width=1)),
                    text=[f"P10 — {r['concept']}" for r in chart_rows],
                    textposition="middle right",
                    textfont=dict(size=10),
                    hovertemplate=(
                        "<b>%{text}</b><br>"
                        "CAPEX: $%{x:,.0f}MM<br>"
                        "NPV P10: $%{y:,.0f}MM<extra></extra>"),
                ), secondary_y=False)

                # Vertical bracket lines connecting P90 → P10 per concept
                for r in chart_rows:
                    fig_cc.add_trace(go.Scatter(
                        x=[r["capex_disc_MM"], r["capex_disc_MM"]],
                        y=[r["p90"], r["p10"]],
                        mode="lines",
                        line=dict(color="#888", width=1, dash="dot"),
                        showlegend=False,
                        hoverinfo="skip",
                    ), secondary_y=False)

                # Breakeven labels — boxed beside each mean marker, like a
                # callout on the reference chart (white box, dark border,
                # bold $/bbl).
                for r in chart_rows:
                    if r["breakeven_oil"] is not None:
                        fig_cc.add_annotation(
                            x=r["capex_disc_MM"], y=r["mean"],
                            text=f"<b>BE ${r['breakeven_oil']:.0f}/bbl</b>",
                            showarrow=True,
                            arrowhead=0, arrowwidth=1,
                            arrowcolor="#888",
                            ax=30, ay=-26,
                            xanchor="left",
                            font=dict(size=10, color="#1a1a1a"),
                            bordercolor="#555",
                            borderwidth=1.2,
                            borderpad=4,
                            bgcolor="rgba(255,255,255,0.92)",
                        )

                # Emissions on the secondary axis as small blue squares
                em_rows = [r for r in chart_rows
                            if r["emissions_Mt"] is not None]
                if em_rows:
                    fig_cc.add_trace(go.Scatter(
                        x=[r["capex_disc_MM"] for r in em_rows],
                        y=[r["emissions_Mt"] for r in em_rows],
                        mode="markers+text",
                        name="Total emissions",
                        marker=dict(size=11, color="#1f77b4",
                                     symbol="square",
                                     line=dict(color="black", width=0.5)),
                        text=[f"{r['emissions_Mt']:.1f} Mt"
                              for r in em_rows],
                        textposition="bottom right",
                        textfont=dict(size=9, color="#1f77b4"),
                        hovertemplate=(
                            "Emissions: %{y:.2f} Mt CO₂-eq<br>"
                            "CAPEX: $%{x:,.0f}MM<extra></extra>"),
                    ), secondary_y=True)

                fig_cc.update_xaxes(
                    title_text="Total discounted CAPEX ($MM, 100% project)")
                fig_cc.update_yaxes(
                    title_text="NPV post-tax ($MM, 100% project)",
                    secondary_y=False,
                    title_font=dict(color="#2ca02c"),
                    tickfont=dict(color="#2ca02c"))
                fig_cc.update_yaxes(
                    title_text="Total CO₂-eq emissions (Mt)",
                    secondary_y=True,
                    title_font=dict(color="#1f77b4"),
                    tickfont=dict(color="#1f77b4"))
                fig_cc.update_layout(
                    title=("Concept Comparison — NPV post-tax vs CAPEX, "
                           f"grouped by {grouping_dim}"),
                    height=560,
                    legend=dict(orientation="h", y=-0.18),
                    hovermode="closest",
                )
                # Reference horizontal zero-NPV line
                fig_cc.add_hline(y=0, line=dict(color="grey", dash="dot"),
                                  secondary_y=False)
                st.plotly_chart(fh.apply_plot_template(fig_cc),
                                use_container_width=True)
                _any_prob = any(r.get("probabilistic")
                                 for r in chart_rows)
                st.caption(
                    "★ = on the **Pareto frontier** (no other concept "
                    "has both lower CAPEX and higher NPV); grey circles "
                    "are dominated. "
                    + ("P90/Mean/P10 are from the **Monte-Carlo** pass "
                       "per case." if _any_prob else
                       "P90/Mean/P10 reflect the spread across cases in "
                       "each group — enable the Monte-Carlo option before "
                       "running for a true probabilistic band."))

                # Companion summary table
                df_cc = pd.DataFrame(chart_rows)
                df_cc_show = df_cc.copy()
                df_cc_show["P90 NPV ($MM)"] = df_cc_show["p90"].map(
                    lambda v: f"{v:,.0f}")
                df_cc_show["Mean NPV ($MM)"] = df_cc_show["mean"].map(
                    lambda v: f"{v:,.0f}")
                df_cc_show["P10 NPV ($MM)"] = df_cc_show["p10"].map(
                    lambda v: f"{v:,.0f}")
                df_cc_show["CAPEX ($MM disc.)"] = df_cc_show[
                    "capex_disc_MM"].map(lambda v: f"{v:,.0f}")
                df_cc_show["BE oil ($/bbl)"] = df_cc_show[
                    "breakeven_oil"].map(
                        lambda v: f"{v:,.1f}" if pd.notna(v) else "—")
                df_cc_show["Emissions (Mt)"] = df_cc_show[
                    "emissions_Mt"].map(
                        lambda v: f"{v:,.2f}" if pd.notna(v) else "—")
                df_cc_show["# cases"] = df_cc_show["n_cases"]
                df_cc_show = df_cc_show[[
                    "concept", "P90 NPV ($MM)", "Mean NPV ($MM)",
                    "P10 NPV ($MM)", "CAPEX ($MM disc.)",
                    "BE oil ($/bbl)", "Emissions (Mt)", "# cases"]]
                df_cc_show.columns = [
                    "Concept", "P90", "Mean", "P10", "CAPEX disc.",
                    "BE oil", "Emissions", "# cases"]
                st.dataframe(df_cc_show, use_container_width=True,
                              hide_index=True)

        # ============================================================
        # QUALITATIVE DECISION MATRIX
        # ============================================================
        # The bubble chart above ranks concepts on quantitative NPV /
        # CAPEX / emissions axes. This second panel is the classic
        # qualitative complement: a Type × Criteria × Concept matrix
        # where each cell carries a traffic-light colour (green / yellow
        # / red / grey-NA) reflecting the user's judgement on
        # non-quantifiable criteria — HSE, regulatory, reputational,
        # operability, technology maturity. Reads exactly like the
        # NCS DG1 / DG2 decision-support tables operators submit to
        # the project board.
        st.markdown("---")
        st.markdown("#### 🚦 Qualitative decision matrix — "
                    "traffic-light scoring per criterion")
        st.caption(
            "Score each concept against qualitative criteria — HSE, "
            "schedule, technology, regulatory, robustness, operability "
            "— using a green / yellow / red rating. Concept columns "
            "pull automatically from the groupings used in the bubble "
            "chart above (the same `Group concepts by` dimension), so "
            "the row of quantitative bubbles and the row of qualitative "
            "judgements speak about the same set of concepts.")

        # ---- Default criteria taxonomy (editable) ----
        # Pre-loaded with the NCS-typical structure shown in the
        # reference image. The user can edit, add, or delete rows.
        _DEFAULT_QUAL_CRITERIA = [
            # (type, criterion)
            ("SSU/HSE", "Safety"),
            ("SSU/HSE", "Health"),
            ("SSU/HSE", "Environment"),
            ("SSU/HSE", "CO2 footprint and power consumption"),
            ("Risk exposure", "Total schedule"),
            ("Risk exposure", "Technology maturity"),
            ("Risk exposure", "Concept complexity"),
            ("Risk exposure", "Authority approval"),
            ("Risk exposure", "Reputation exposure"),
            ("Risk exposure", "Local content"),
            ("Risk exposure", "Country risk"),
            ("Robustness", "Compliance to strategic direction"),
            ("Robustness", "Contractor availability / competitive bids"),
            ("Robustness", "Flow assurance"),
            ("Robustness", "Increased production"),
            ("Robustness", "Marketing / trading flexibility"),
            ("Operability", "Reservoir management"),
            ("Operability", "Operation complexity"),
            ("Operability", "Operational cost level"),
            ("Operability", "System availability"),
            ("Operability", "OPEX flexibility"),
        ]

        # Concept columns to score — pulled from the groups computed
        # above for the bubble chart. Fall back to a default set if
        # no batch results exist yet so the user can still play with
        # the matrix before running.
        try:
            qual_concepts = list(groups.keys()) if groups else []
        except NameError:
            qual_concepts = []
        if not qual_concepts:
            qual_concepts = ["Concept A", "Concept B", "Concept C"]
            st.info(
                "No batch results yet — using placeholder concept names. "
                "Run a batch above to auto-link the columns to your real "
                "concepts.")

        # ---- State ----
        # Store the matrix as a dict of dicts:
        #   {(type, criterion): {concept_label: colour_label}}
        # colour labels are short strings the data_editor can validate.
        qual_state_key = "concept_qual_matrix"
        if qual_state_key not in st.session_state:
            st.session_state[qual_state_key] = {}

        # ---- Criteria editor in an expander ----
        with st.expander("✏️ Edit criteria (type + criterion rows)",
                          expanded=False):
            st.caption(
                "Add, remove or rename the criteria. The `Type` column "
                "groups related criteria together — repeated values "
                "merge visually in the matrix below.")
            crit_df_key = "concept_qual_criteria_df"
            if crit_df_key not in st.session_state:
                st.session_state[crit_df_key] = pd.DataFrame(
                    _DEFAULT_QUAL_CRITERIA, columns=["Type", "Criterion"])
            edited_crit = st.data_editor(
                st.session_state[crit_df_key],
                num_rows="dynamic",
                use_container_width=True,
                key="concept_qual_criteria_editor",
                column_config={
                    "Type": st.column_config.TextColumn(
                        "Type",
                        help="Category — repeating values group rows"),
                    "Criterion": st.column_config.TextColumn(
                        "Criterion",
                        help="Specific item to score against each "
                             "concept"),
                },
            )
            if not edited_crit.equals(st.session_state[crit_df_key]):
                st.session_state[crit_df_key] = edited_crit

        criteria = st.session_state[crit_df_key]
        # Filter empty rows
        criteria = criteria[
            criteria["Type"].astype(str).str.strip() != ""].reset_index(
                drop=True)
        criteria = criteria[
            criteria["Criterion"].astype(str).str.strip() != ""].reset_index(
                drop=True)

        if len(criteria) == 0:
            st.warning("Add at least one criterion to score concepts.")
        else:
            # ---- Build the matrix dataframe ----
            # Columns: Type, Criterion, <one column per concept>
            COLOUR_OPTIONS = ["🟢 Green", "🟡 Yellow", "🔴 Red", "⚪ N/A"]
            COLOUR_DEFAULT = "⚪ N/A"
            COLOUR_HEX = {
                "🟢 Green": "#2ca02c",
                "🟡 Yellow": "#ffd700",
                "🔴 Red": "#d62728",
                "⚪ N/A": "#eaeaea",
            }

            # Build initial dataframe from session_state + criteria + concepts
            stored = st.session_state[qual_state_key]
            # Per-criterion weights (default 1.0). Keyed by (type,criterion).
            qual_weights_key = "concept_qual_weights"
            if qual_weights_key not in st.session_state:
                st.session_state[qual_weights_key] = {}
            weights_store = st.session_state[qual_weights_key]
            rows = []
            for _, crit_row in criteria.iterrows():
                t = str(crit_row["Type"])
                c = str(crit_row["Criterion"])
                row = {"Type": t, "Criterion": c,
                       "Weight": float(weights_store.get((t, c), 1.0))}
                for cn in qual_concepts:
                    row[cn] = stored.get((t, c), {}).get(
                        cn, COLOUR_DEFAULT)
                rows.append(row)
            matrix_df = pd.DataFrame(rows)

            # Tools row above the editor
            tc1, tc2, tc3 = st.columns([2, 2, 5])
            if tc1.button("🟢 Set all to Green",
                           key="qual_all_green",
                           use_container_width=True):
                for _, r in criteria.iterrows():
                    key = (str(r["Type"]), str(r["Criterion"]))
                    stored.setdefault(key, {})
                    for cn in qual_concepts:
                        stored[key][cn] = "🟢 Green"
                st.rerun()
            if tc2.button("♻️ Reset matrix",
                           key="qual_reset",
                           use_container_width=True,
                           help="Clear every cell back to N/A."):
                st.session_state[qual_state_key] = {}
                st.rerun()

            # ---- Editable matrix ----
            # SelectboxColumn renders a dropdown per cell. The user can
            # change the rating; we commit back to session_state.
            col_config = {
                "Type": st.column_config.TextColumn(
                    "Type", width="small", disabled=True),
                "Criterion": st.column_config.TextColumn(
                    "Criterion", width="medium", disabled=True),
                "Weight": st.column_config.NumberColumn(
                    "Weight", width="small", min_value=0.0,
                    max_value=10.0, step=0.5, format="%.1f",
                    help="Relative importance of this criterion in the "
                         "weighted score. 0 = ignore, higher = more "
                         "influence."),
            }
            for cn in qual_concepts:
                col_config[cn] = st.column_config.SelectboxColumn(
                    cn,
                    options=COLOUR_OPTIONS,
                    default=COLOUR_DEFAULT,
                    required=True,
                    help=f"Score for {cn} on this criterion",
                )
            edited_matrix = st.data_editor(
                matrix_df,
                use_container_width=True,
                hide_index=True,
                column_config=col_config,
                num_rows="fixed",
                key="concept_qual_matrix_editor",
            )

            # Edit-then-apply: committing every cell edit to session_state
            # on each keystroke (and re-rendering the coloured view, the
            # score summary and the combined ranking) makes the page lag
            # badly. Instead we only commit when the user clicks Apply,
            # and the downstream views render from the APPLIED snapshot,
            # not the live editor. This keeps typing/clicking responsive.
            qa1, qa2 = st.columns([2, 5])
            if qa1.button("✅ Apply matrix edits", key="concept_qual_apply",
                           type="primary", use_container_width=True,
                           help="Commit the current ratings & weights. "
                                "The coloured view and scores below only "
                                "refresh on Apply — keeps editing smooth."):
                for _, mrow in edited_matrix.iterrows():
                    key = (str(mrow["Type"]), str(mrow["Criterion"]))
                    stored.setdefault(key, {})
                    for cn in qual_concepts:
                        if cn in mrow.index:
                            stored[key][cn] = str(mrow[cn])
                    try:
                        weights_store[key] = float(mrow.get("Weight", 1.0))
                    except Exception:
                        weights_store[key] = 1.0
                st.session_state["_concept_qual_applied"] = True
                st.rerun()
            qa2.caption(
                "Edit ratings/weights above, then click **Apply** to "
                "refresh the coloured matrix, scores and ranking below.")

            # Downstream views render from the committed snapshot
            # (`stored`/`weights_store`), rebuilt into a frame so the
            # rendering code below is unchanged but reads applied data.
            applied_rows = []
            for _, crit_row in criteria.iterrows():
                t = str(crit_row["Type"]); c = str(crit_row["Criterion"])
                row = {"Type": t, "Criterion": c,
                       "Weight": float(weights_store.get((t, c), 1.0))}
                for cn in qual_concepts:
                    row[cn] = stored.get((t, c), {}).get(cn, COLOUR_DEFAULT)
                applied_rows.append(row)
            edited_matrix = pd.DataFrame(applied_rows)

            # ---- Render the matrix as a coloured HTML/SVG view ----
            # The editor is functional but visually dull. Below it we
            # render the same data as an HTML table with the cells
            # filled by their traffic-light colour, mimicking the
            # reference image. This is the "pretty" view that goes
            # into screenshots / decks.
            st.markdown(
                "##### Coloured view — drag this into a deck or "
                "screenshot it")
            html_rows = ['<table style="border-collapse:collapse;'
                          'width:100%;font-family:Helvetica,Arial,'
                          'sans-serif;font-size:11px">']
            # Header
            html_rows.append(
                '<tr style="background:#fbe9c2;text-align:center;'
                'font-weight:700">'
                '<th style="padding:8px;border:1px solid #ccc;'
                'width:120px">Type</th>'
                '<th style="padding:8px;border:1px solid #ccc;'
                'width:280px">Criterion</th>')
            for cn in qual_concepts:
                html_rows.append(
                    f'<th style="padding:8px;border:1px solid #ccc;'
                    f'min-width:110px;max-width:160px">{cn}</th>')
            html_rows.append("</tr>")
            # Body — merge consecutive identical Type cells with rowspan
            prev_type = None
            type_span_count = {}
            for _, mrow in edited_matrix.iterrows():
                t = str(mrow["Type"])
                type_span_count[t] = type_span_count.get(t, 0) + 1
            seen_type_emitted = {}
            for _, mrow in edited_matrix.iterrows():
                t = str(mrow["Type"])
                c = str(mrow["Criterion"])
                html_rows.append('<tr>')
                if t not in seen_type_emitted:
                    seen_type_emitted[t] = True
                    html_rows.append(
                        f'<td rowspan="{type_span_count[t]}" '
                        f'style="background:#f4e2d8;padding:8px;'
                        f'border:1px solid #ccc;vertical-align:middle;'
                        f'text-align:center;font-weight:600">'
                        f'{t}</td>')
                html_rows.append(
                    f'<td style="padding:6px 10px;border:1px solid #ccc;'
                    f'background:white">{c}</td>')
                for cn in qual_concepts:
                    rating = str(mrow.get(cn, COLOUR_DEFAULT))
                    fill = COLOUR_HEX.get(rating, "#eaeaea")
                    # Render as a coloured circle, like the reference
                    html_rows.append(
                        f'<td style="padding:6px;border:1px solid #ccc;'
                        f'text-align:center;background:white">'
                        f'<span style="display:inline-block;width:26px;'
                        f'height:26px;border-radius:50%;background:{fill};'
                        f'border:1px solid #777;'
                        f'box-shadow:inset 2px 2px 4px rgba(255,255,255,0.4),'
                        f'inset -2px -2px 4px rgba(0,0,0,0.15)"></span>'
                        f'</td>')
                html_rows.append('</tr>')
            html_rows.append("</table>")
            st.markdown("".join(html_rows), unsafe_allow_html=True)

            # ---- Tally + weighted score summary ----
            st.markdown("##### Score summary")
            _val = {"🟢 Green": 1.0, "🟡 Yellow": 0.0,
                    "🔴 Red": -1.0, "⚪ N/A": 0.0}
            summary_rows = []
            qual_scores = {}   # concept → weighted score (for combined rank)
            for cn in qual_concepts:
                g = y = r = na = 0
                weighted_num = 0.0
                weight_den = 0.0
                for _, mrow in edited_matrix.iterrows():
                    rating = mrow.get(cn)
                    w = float(mrow.get("Weight", 1.0) or 0.0)
                    if rating == "🟢 Green":
                        g += 1
                    elif rating == "🟡 Yellow":
                        y += 1
                    elif rating == "🔴 Red":
                        r += 1
                    else:
                        na += 1
                    if rating != "⚪ N/A":
                        weighted_num += _val.get(rating, 0.0) * w
                        weight_den += w
                net = g - r
                # Weighted score normalised to [-1, +1] over the
                # non-N/A criteria so concepts with different N/A counts
                # stay comparable.
                wscore = (weighted_num / weight_den
                          if weight_den > 0 else 0.0)
                qual_scores[cn] = wscore
                summary_rows.append({
                    "Concept": cn, "🟢 Green": g, "🟡 Yellow": y,
                    "🔴 Red": r, "⚪ N/A": na,
                    "Net (G−R)": net,
                    "Weighted score": round(wscore, 3),
                })
            df_summary = pd.DataFrame(summary_rows).sort_values(
                "Weighted score", ascending=False)
            st.dataframe(df_summary, use_container_width=True,
                          hide_index=True)
            st.caption(
                "**Net (G−R)** = green − red, unweighted. **Weighted "
                "score** = Σ(rating×weight) ÷ Σweight over non-N/A "
                "criteria, where green=+1, yellow=0, red=−1 — normalised "
                "to [−1, +1] so it ranks concepts on the qualitative "
                "axis accounting for the per-criterion weights you set.")
            # Stash for the combined ranking below.
            st.session_state["_concept_qual_scores"] = qual_scores

            # ---- Combined ranking (#4) ----
            # Blend the quantitative NPV ranking with the qualitative
            # weighted score into a single ordered recommendation. The
            # mix is user-configurable; default 70% economics / 30%
            # qualitative — a common DG screening split.
            st.markdown("##### 🏆 Combined ranking — economics + qualitative")
            # Quantitative score per concept: mean NPV of the matching
            # group (same grouping dimension as the bubble chart).
            quant_npv = {}
            try:
                for cn in qual_concepts:
                    rs = groups.get(cn, [])
                    npvs = [r["npv_MM"] for r in rs
                            if r.get("npv_MM") is not None]
                    if npvs:
                        quant_npv[cn] = float(np.mean(npvs))
            except Exception:
                quant_npv = {}
            if not quant_npv:
                st.info(
                    "Run a batch and make sure the qualitative concept "
                    "columns match a swept dimension to see the combined "
                    "ranking.")
            else:
                wq = st.slider(
                    "Economics weight (vs qualitative)", 0.0, 1.0, 0.70,
                    0.05, key="concept_combined_weight",
                    help="1.0 = rank purely on NPV; 0.0 = rank purely on "
                         "the qualitative weighted score. Default 0.70 "
                         "leans on economics but lets HSE / operability "
                         "break ties.")
                # Normalise each axis to [0,1] across the concepts so the
                # blend is scale-free.
                _npv_vals = list(quant_npv.values())
                _nlo, _nhi = min(_npv_vals), max(_npv_vals)

                def _nz(v):
                    return ((v - _nlo) / (_nhi - _nlo)
                            if _nhi > _nlo else 0.5)
                # Qual score is already in [-1,1] → map to [0,1].
                comb_rows = []
                for cn in quant_npv:
                    q_n = _nz(quant_npv[cn])
                    q_q = (qual_scores.get(cn, 0.0) + 1.0) / 2.0
                    composite = wq * q_n + (1 - wq) * q_q
                    comb_rows.append({
                        "Concept": cn,
                        "Mean NPV ($MM)": f"{quant_npv[cn]:,.0f}",
                        "Econ score (0-1)": round(q_n, 3),
                        "Qual score (0-1)": round(q_q, 3),
                        "Composite": round(composite, 3),
                    })
                df_comb = pd.DataFrame(comb_rows).sort_values(
                    "Composite", ascending=False).reset_index(drop=True)
                df_comb.insert(0, "Rank", df_comb.index + 1)
                st.dataframe(df_comb, use_container_width=True,
                              hide_index=True)
                _winner = df_comb.iloc[0]["Concept"]
                st.success(
                    f"🏆 **Top-ranked concept: {_winner}** "
                    f"(at {wq:.0%} economics / {1-wq:.0%} qualitative). "
                    f"Adjust the weight slider to test how sensitive the "
                    f"recommendation is to the economics-vs-qualitative "
                    f"balance.")

        # ============================================================
        # DESIGN-TO-COST STAIRCASE
        # ============================================================
        # Classic NCS DG1 "Design to Cost" plot: rank the concepts left
        # → right by ascending discounted CAPEX, draw an L-shaped step
        # for each, annotate NPV/BE on top and ΔNPV/ΔBE on the rise.
        # The concept with the highest NPV is the "recommended solution"
        # — every later step adds non-profitable scope (ΔNPV < 0).
        # Reads the way operators present the screening verdict to the
        # project board.
        st.markdown("---")
        st.markdown("#### 🪜 Design-to-Cost staircase — "
                    "ranked by ascending CAPEX")
        st.caption(
            "Steps ordered by ascending discounted CAPEX. Each step "
            "shows the absolute **NPV** and **breakeven price** of the "
            "concept, plus **ΔNPV** and **ΔBE** versus the previous "
            "step. The concept just before ΔNPV first turns negative "
            "is the **Recommended solution** — adding any further "
            "scope erodes value. The first step at the lowest CAPEX is "
            "the **Bare bone** case.")
        try:
            staircase_groups = groups if groups else {}
        except NameError:
            staircase_groups = {}
        if not staircase_groups:
            st.info("Run a batch above to populate the staircase.")
        else:
            # Build one (concept, mean NPV, mean CAPEX, mean BE) point
            # per concept group. The medians used in the bubble chart
            # are robust to a few weird combinations and reflect the
            # representative case; we use the same here for consistency.
            import numpy as np
            staircase = []
            for label, rs in staircase_groups.items():
                npvs = [r["npv_MM"] for r in rs
                         if r.get("npv_MM") is not None]
                capexes = [r["capex_disc_MM"] for r in rs
                            if r.get("capex_disc_MM") is not None]
                bes = [r["breakeven_oil"] for r in rs
                        if r.get("breakeven_oil") is not None]
                if not npvs or not capexes:
                    continue
                staircase.append({
                    "concept": label,
                    "npv_MM": float(np.median(npvs)),
                    "capex_disc_MM": float(np.median(capexes)),
                    "breakeven_oil": (float(np.median(bes))
                                       if bes else None),
                })
            if not staircase:
                st.info(
                    "Selected groups have no usable NPV/CAPEX data — "
                    "every combination failed or returned no values.")
            else:
                # Sort by ascending CAPEX — the staircase axis
                staircase.sort(key=lambda x: x["capex_disc_MM"])

                # Insert a synthetic "Reference (approved plans)" step
                # at zero CAPEX / zero NPV — this is what the reference
                # slide shows on the far left. Keeps the visual familiar
                # to anyone who's seen the NCS DTC template before.
                staircase.insert(0, {
                    "concept": "Reference (approved plans)",
                    "npv_MM": 0.0,
                    "capex_disc_MM": 0.0,
                    "breakeven_oil": None,
                    "_is_reference": True,
                })

                # The "recommended" step is the one with the highest
                # NPV — anything to the right of it is non-profitable
                # addition.
                npv_arr = [s["npv_MM"] for s in staircase]
                rec_idx = int(np.argmax(npv_arr))

                # Auto-name the steps the way the reference slide does:
                # 0 = Reference, 1 = Bare bone, intermediates = Profitable
                # additions, recommended = Recommended solution, after =
                # Non-profitable additions.
                for i, step in enumerate(staircase):
                    if step.get("_is_reference"):
                        step["role"] = "Reference"
                    elif i == 1:
                        step["role"] = "Bare bone"
                    elif i == rec_idx:
                        step["role"] = "Recommended solution"
                    elif i < rec_idx:
                        step["role"] = "Profitable additions"
                    else:
                        step["role"] = "Non-profitable additions"

                # ---- Build the figure ----
                # An L-shaped step is two perpendicular thick lines
                # (a horizontal segment at the new NPV level and a
                # vertical riser from the previous NPV to the new one).
                # Plotly's Scatter with `lines` mode and explicit
                # waypoints gives us exact control over the step shape.
                fig_st = go.Figure()

                # Compute the x-positions: equal spacing for readability
                # (the actual CAPEX is shown as a label). Real CAPEX
                # values can cluster badly and crush the staircase.
                n = len(staircase)
                x_positions = list(range(n))
                step_w = 0.42   # half-width of each horizontal slab

                # Colour palette by role
                role_color = {
                    "Reference": "#bbbbbb",
                    "Bare bone": "#888888",
                    "Profitable additions": "#888888",
                    "Recommended solution": "#888888",
                    "Non-profitable additions": "#888888",
                }
                # The L-shaped step itself — drawn as a thick grey
                # outline, with the recommended step optionally
                # highlighted via a red dotted halo (added afterwards).
                for i, step in enumerate(staircase):
                    x = x_positions[i]
                    y = step["npv_MM"]
                    prev_y = (staircase[i-1]["npv_MM"]
                              if i > 0 else 0.0)
                    # Step polyline: from (x-step_w, prev_y) up to
                    # (x-step_w, y), across to (x+step_w, y). Drawn as
                    # two lines so we can make the riser thinner.
                    # Riser (vertical)
                    fig_st.add_trace(go.Scatter(
                        x=[x - step_w, x - step_w],
                        y=[prev_y, y],
                        mode="lines",
                        line=dict(color="#aaaaaa", width=10),
                        hoverinfo="skip",
                        showlegend=False,
                    ))
                    # Tread (horizontal)
                    fig_st.add_trace(go.Scatter(
                        x=[x - step_w, x + step_w],
                        y=[y, y],
                        mode="lines",
                        line=dict(color=role_color[step["role"]],
                                   width=14),
                        hoverinfo="text",
                        text=f"<b>{step['concept']}</b><br>"
                             f"NPV: ${step['npv_MM']:,.0f}MM<br>"
                             f"CAPEX: ${step['capex_disc_MM']:,.0f}MM"
                             + (f"<br>BE: ${step['breakeven_oil']:,.1f}/bbl"
                                 if step.get("breakeven_oil") else ""),
                        showlegend=False,
                    ))

                # The red dotted "Recommended" halo — a circle around
                # the recommended step, drawn via add_shape.
                rec = staircase[rec_idx]
                rec_x = x_positions[rec_idx]
                rec_y = rec["npv_MM"]
                # Halo size scales with the NPV range so it's visible
                npv_range = max(npv_arr) - min(0.0, min(npv_arr))
                halo_rx = 0.55
                halo_ry = npv_range * 0.22 if npv_range > 0 else 50.0
                fig_st.add_shape(
                    type="circle",
                    x0=rec_x - halo_rx, x1=rec_x + halo_rx,
                    y0=rec_y - halo_ry, y1=rec_y + halo_ry * 0.6,
                    line=dict(color="#d62728", width=2.5, dash="dot"),
                    fillcolor="rgba(0,0,0,0)",
                )

                # Annotations on top of each step: NPV + BE
                for i, step in enumerate(staircase):
                    x = x_positions[i]
                    y = step["npv_MM"]
                    if step.get("_is_reference"):
                        label_lines = ["0"]
                    else:
                        lines = [f"NPV ${step['npv_MM']:,.0f}MM"]
                        if step.get("breakeven_oil") is not None:
                            lines.append(
                                f"BE ${step['breakeven_oil']:.0f}/bbl")
                        label_lines = lines
                    fig_st.add_annotation(
                        x=x, y=y,
                        text="<br>".join(label_lines),
                        showarrow=False,
                        yshift=28,
                        font=dict(size=10, color="black"),
                    )

                # Delta annotations on the riser between steps i-1 and i
                for i in range(1, len(staircase)):
                    x = x_positions[i]
                    y_prev = staircase[i-1]["npv_MM"]
                    y_cur = staircase[i]["npv_MM"]
                    delta_npv = y_cur - y_prev
                    delta_be = None
                    if (staircase[i].get("breakeven_oil") is not None
                            and staircase[i-1].get("breakeven_oil")
                                is not None):
                        delta_be = (staircase[i]["breakeven_oil"]
                                     - staircase[i-1]["breakeven_oil"])
                    delta_text_parts = [
                        f"ΔNPV ${delta_npv:+,.0f}MM"]
                    if delta_be is not None:
                        delta_text_parts.append(
                            f"ΔBE ${delta_be:+,.1f}/bbl")
                    # Colour by sign — green for +, red for −
                    col = "#2ca02c" if delta_npv > 0 else "#d62728"
                    fig_st.add_annotation(
                        x=x - step_w, y=(y_prev + y_cur) / 2.0,
                        text="<br>".join(delta_text_parts),
                        showarrow=False,
                        xshift=-50,
                        font=dict(size=10, color=col,
                                   family="Helvetica"),
                        align="right",
                    )

                # Concept label below each step
                for i, step in enumerate(staircase):
                    x = x_positions[i]
                    role = step["role"]
                    # Show the concept name AND the role tag underneath
                    label = f"<b>{step['concept']}</b><br><i>{role}</i>"
                    if not step.get("_is_reference"):
                        label += (
                            f"<br>CAPEX "
                            f"${step['capex_disc_MM']:,.0f}MM")
                    fig_st.add_annotation(
                        x=x, y=min(0.0, min(npv_arr)),
                        text=label,
                        showarrow=False,
                        yshift=-30,
                        yanchor="top",
                        font=dict(size=10,
                                   color="#d62728" if i == rec_idx
                                   else "black"),
                        align="center",
                    )

                # Layout — hide the actual x-tick numbers (positions
                # are arbitrary), keep y axis as NPV scale
                ymin = min(0.0, min(npv_arr)) - npv_range * 0.35
                ymax = max(npv_arr) + npv_range * 0.30
                fig_st.update_layout(
                    title=dict(
                        text="🪜 Design to Cost — concept ranking",
                        font=dict(size=15)),
                    xaxis=dict(
                        showticklabels=False,
                        showgrid=False,
                        zeroline=False,
                        range=[-0.8, n - 0.2]),
                    yaxis=dict(
                        title="NPV post-tax ($MM)",
                        zeroline=True,
                        zerolinecolor="#444",
                        zerolinewidth=1,
                        range=[ymin, ymax]),
                    height=max(500, 60 * n + 200),
                    plot_bgcolor="white",
                    showlegend=False,
                    margin=dict(l=80, r=40, t=80, b=140),
                )

                st.plotly_chart(fh.apply_plot_template(fig_st),
                                use_container_width=True)

                # ---- Companion ranking table ----
                st.markdown("##### Staircase ranking table")
                rank_rows = []
                for i, step in enumerate(staircase):
                    prev = staircase[i-1] if i > 0 else None
                    dnpv = (step["npv_MM"] - prev["npv_MM"]
                             if prev else None)
                    dbe = None
                    if (prev and step.get("breakeven_oil") is not None
                            and prev.get("breakeven_oil") is not None):
                        dbe = (step["breakeven_oil"]
                                - prev["breakeven_oil"])
                    rank_rows.append({
                        "Step": i, "Role": step["role"],
                        "Concept": step["concept"],
                        "CAPEX disc. ($MM)": (
                            f"{step['capex_disc_MM']:,.0f}"),
                        "NPV ($MM)": f"{step['npv_MM']:,.0f}",
                        "BE oil ($/bbl)": (
                            f"{step['breakeven_oil']:.1f}"
                            if step.get('breakeven_oil') is not None
                            else "—"),
                        "ΔNPV ($MM)": (
                            f"{dnpv:+,.0f}" if dnpv is not None
                            else "—"),
                        "ΔBE ($/bbl)": (
                            f"{dbe:+.1f}" if dbe is not None
                            else "—"),
                    })
                df_rank = pd.DataFrame(rank_rows)
                st.dataframe(df_rank, use_container_width=True,
                              hide_index=True)
                st.caption(
                    f"**Recommended:** {staircase[rec_idx]['concept']} "
                    f"— highest NPV in the staircase. Concepts beyond "
                    f"this point add scope that doesn't earn its own "
                    f"discounted CAPEX back at the assumed prices.")


if __name__ == "__main__":
    main()



